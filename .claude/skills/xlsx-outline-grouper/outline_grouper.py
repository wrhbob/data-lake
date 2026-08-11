"""
outline_grouper.py — 给 quota-csv-finalize xlsx 套自适应分组大纲（纯分组版）

仅负责一件事: 给「定额条目」sheet 套上按段/定额层级组织的 Excel 可折叠
分组大纲。**不**做任何单元格式修改（不加粗 / 不合并 / 不居中 / 不解上游
合并）。如需要那些格式, 请在 quota-csv-finalize 主流水线里做。

输入 / 输出:
    输入: 一份含 "定额条目" sheet 的 xlsx（典型为 <stem>_待审核.xlsx）。
    输出: 原地覆盖（或写入指定路径），其它 sheet（册说明 / 章）保持不动。

自适应层数（关键性质):
    - 段行 group 的 outline_level = sec_id 中点号数 + 1
      ("A"→1, "A.1"→2, "A.1.1"→3, ... 任意深度)
    - 定额 group 的 outline_level = 父段 depth + 1
      这样定额始终嵌套在所属段之下, 段是 5 层 / 6 层也能正确嵌套。

CLI:
    python outline_grouper.py <input.xlsx> [output.xlsx] [--sheet 定额条目]

Python API:
    from outline_grouper import process_xlsx, build_outline_groups
    info = process_xlsx(input_xlsx, output_xlsx=None, sheet_name="定额条目")
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ──────────────────────────────────────────────
# 硬编码: 本 skill 唯一服务的 schema
# ──────────────────────────────────────────────
SECTION_MARKER = "段"     # row[0] == "段"
QUOTA_MARKER = "定"       # row[0] == "定"
DEFAULT_SHEET = "定额条目"
SECTION_ID_COL = 1        # row[1] 是段 ID（A / A.1 / A.1.1）
QUOTA_ID_COL = 1          # row[1] 是定额 ID（H100001 等）
SUMMARY_BELOW = False     # 标题行在上、内容在下折叠（符合阅读习惯）


# ──────────────────────────────────────────────
# 纯函数 — 无 openpyxl 依赖
# ──────────────────────────────────────────────
def get_section_depth(section_id: str) -> int:
    """A → 1, A.1 → 2, A.1.1 → 3, ... 任意深度自适应。

    空字符串视为深度 1（兜底）。
    """
    if not section_id:
        return 1
    return section_id.count(".") + 1


@dataclass
class Section:
    row: int                  # 1-based Excel row
    depth: int                # 1..N 自适应
    id: str
    name: str = ""
    end_row: int = 0          # 由 compute_end_rows 填充


@dataclass
class Quota:
    row: int                  # 1-based
    id: str
    own_depth: int = 1        # 定额 ID 通常不含 ".",但为对齐口径保留
    end_row: int = 0          # 由 compute_end_rows 填充
    parent_row: int = 0       # 所属段行
    parent_depth: int = 0     # 父段 depth（用于 outline_level）


def _read_row_cells(row: Iterable) -> list[str]:
    """把一行 cells 转 str list（None → ""）。"""
    return [str(v) if v is not None else "" for v in row]


def read_sheet_rows(ws: Worksheet) -> list[list[str]]:
    """读 ws 所有数据行（1-based 数组语义保留,对调用者友好）。"""
    return [_read_row_cells(row) for row in ws.iter_rows(values_only=True)]


def scan_sections(rows: list[list[str]]) -> list[Section]:
    """收集所有 段 行, 返回按出现顺序的 Section 列表（end_row 待填充）。"""
    out: list[Section] = []
    for r_idx, row in enumerate(rows, start=1):
        if not row:
            continue
        if str(row[0]).strip() != SECTION_MARKER:
            continue
        sec_id = str(row[1]).strip() if len(row) > 1 else ""
        sec_name = str(row[2]).strip() if len(row) > 2 else ""
        out.append(Section(row=r_idx, depth=get_section_depth(sec_id),
                           id=sec_id, name=sec_name))
    return out


def scan_quotas(rows: list[list[str]]) -> list[Quota]:
    """收集所有 定额 行, 返回按出现顺序的 Quota 列表（end_row 待填充）。"""
    out: list[Quota] = []
    for r_idx, row in enumerate(rows, start=1):
        if not row:
            continue
        if str(row[0]).strip() != QUOTA_MARKER:
            continue
        quota_id = str(row[1]).strip() if len(row) > 1 else ""
        out.append(Quota(row=r_idx, id=quota_id,
                         own_depth=get_section_depth(quota_id)))
    return out


def find_end_row(sections: list[Section], idx: int, total_rows: int) -> int:
    """返回当前段后的最近一个「同级或更高级段」所在行 - 1。

    若其后没有任何同级或更高级段, 则返回 total_rows。
    """
    sec_depth = sections[idx].depth
    for j in range(idx + 1, len(sections)):
        if sections[j].depth <= sec_depth:
            return sections[j].row - 1
    return total_rows


def compute_section_end_rows(sections: list[Section], total_rows: int) -> None:
    """原地填充每个 Section.end_row。"""
    for i in range(len(sections)):
        sections[i].end_row = find_end_row(sections, i, total_rows)


def compute_quota_end_rows(quotas: list[Quota], sections: list[Section],
                          total_rows: int) -> None:
    """对每个定额:

        1. 默认 end_row = 下一个定额所在行 - 1（最后一个则 = total_rows）
        2. 找到它所属的最近一个段, end_row 取 min(默认, 父段 end_row)
        3. 同时记录 parent_row / parent_depth
    """
    for i, q in enumerate(quotas):
        if i + 1 < len(quotas):
            default_end = quotas[i + 1].row - 1
        else:
            default_end = total_rows

        parent: Section | None = None
        for sec in reversed(sections):
            if sec.row <= q.row:
                parent = sec
                break

        if parent is not None:
            q.parent_row = parent.row
            q.parent_depth = parent.depth
            q.end_row = min(default_end, parent.end_row)
        else:
            q.parent_row = 0
            q.parent_depth = 0
            q.end_row = default_end


def build_outline_groups(rows: list[list[str]]) -> list[tuple[int, int, int]]:
    """纯函数: 从行集合计算出分组区间列表。

    Returns:
        list[(start_row, end_row, outline_level)], start/end 1-based.
        已按 (level, start) 排序 —— 低 level 先, 高 level 后(Excel 嵌套安全)。

    自适应性质:
        - 段 group level = sec.depth （A→1, A.1→2, A.1.1→3, ...）
        - 定额 group level = parent_sec.depth + 1（始终比父段深一层）
    """
    total_rows = len(rows)
    sections = scan_sections(rows)
    quotas = scan_quotas(rows)
    compute_section_end_rows(sections, total_rows)
    compute_quota_end_rows(quotas, sections, total_rows)

    groups: list[tuple[int, int, int]] = []
    for sec in sections:
        start = sec.row + 1          # 段行本身不折叠(作为标题行可见)
        end = sec.end_row
        if start <= end:
            groups.append((start, end, sec.depth))

    for q in quotas:
        start = q.row + 1            # 定额行本身可见
        end = q.end_row
        if start <= end:
            level = q.parent_depth + 1 if q.parent_depth >= 1 else 1
            groups.append((start, end, level))

    groups.sort(key=lambda x: (x[2], x[0]))
    return groups


# ──────────────────────────────────────────────
# openpyxl adapter — 仅 outline, 不做任何格式修改
# ──────────────────────────────────────────────
def apply_outline_to_worksheet(ws: Worksheet,
                               groups: list[tuple[int, int, int]],
                               summary_below: bool = SUMMARY_BELOW,
                               ) -> None:
    """把分组区间应用到 ws(只设置 summaryBelow + 各级 row_dimensions.group)。

    不动 cell 值, 不改字体, 不合并, 不解合并 —— 这是一个纯 outline 工具。
    """
    ws.sheet_properties.outlinePr.summaryBelow = summary_below
    for start, end, level in groups:
        ws.row_dimensions.group(start, end, outline_level=level)


def process_xlsx(input_path: Path,
                 output_path: Path | None = None,
                 sheet_name: str = DEFAULT_SHEET,
                 ) -> dict:
    """主入口: xlsx → 带自适应分组大纲的 xlsx。

    Args:
        input_path: 输入 xlsx 路径（必须包含 sheet_name sheet）
        output_path: 输出 xlsx 路径;None → 原地覆盖 input_path
        sheet_name: 目标 sheet 名（默认「定额条目」）

    Returns:
        dict 含 input_path / output_path / sheet / total_rows /
             sections / quotas / groups / other_sheets。

    Raises:
        FileNotFoundError: 输入 xlsx 不存在
        RuntimeError: 目标 sheet 不存在
    """
    if not input_path.exists():
        raise FileNotFoundError(f"xlsx 不存在: {input_path}")

    target = output_path if output_path is not None else input_path

    wb = load_workbook(input_path, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"xlsx 中没有 {sheet_name!r} sheet;现有 sheet: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows = read_sheet_rows(ws)

    if not rows:
        wb.save(target)
        return {
            "input_path": str(input_path),
            "output_path": str(target),
            "sheet": sheet_name,
            "total_rows": 0,
            "sections": 0,
            "quotas": 0,
            "groups": 0,
            "other_sheets": [s for s in wb.sheetnames if s != sheet_name],
            "warning": "empty sheet",
        }

    # 1. 计算分组(纯函数, 与 ws 解耦, 也方便单测)
    groups = build_outline_groups(rows)

    # 2. 仅设置 outline —— 不改 cell 值、字体、合并
    apply_outline_to_worksheet(ws, groups)

    wb.save(target)

    sec_count = sum(1 for r in rows if r and str(r[0]).strip() == SECTION_MARKER)
    quota_count = sum(1 for r in rows if r and str(r[0]).strip() == QUOTA_MARKER)

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "sheet": sheet_name,
        "total_rows": len(rows),
        "sections": sec_count,
        "quotas": quota_count,
        "groups": len(groups),
        "other_sheets": [s for s in wb.sheetnames if s != sheet_name],
    }


# ──────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(
        description="xlsx → 带自适应分组大纲的 Excel(纯分组, 不改格式)。"
                    "硬编码 sheet 默认「定额条目」, 行标记「段」「定」。"
    )
    ap.add_argument("input_xlsx", help="输入 xlsx 路径")
    ap.add_argument(
        "output_xlsx",
        nargs="?",
        default=None,
        help="输出 xlsx 路径(默认 = 原地覆盖输入文件)",
    )
    ap.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"目标 sheet 名(默认 {DEFAULT_SHEET!r})",
    )
    args = ap.parse_args()

    input_path = Path(args.input_xlsx).resolve()
    output_path = Path(args.output_xlsx).resolve() if args.output_xlsx else None

    try:
        info = process_xlsx(input_path, output_path, sheet_name=args.sheet)
    except (FileNotFoundError, RuntimeError) as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return 1

    print(f"[OK] 输入: {info['input_path']}")
    print(f"[OK] 输出: {info['output_path']}")
    print(
        f"[OK] sheet: {info['sheet']} 总行数 {info['total_rows']}, "
        f"段数 {info['sections']}, 定额数 {info['quotas']}"
    )
    print(f"[OK] 分组区间数: {info['groups']}")
    print(
        f"[OK] 其它保留 sheet: {info['other_sheets']}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
