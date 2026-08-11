"""
clear_and_regroup.py — 清空「定额条目」sheet 全部旧 outline + 重新套自适应分组

适用场景: xlsx 已被某工具设过 outline(可能错位/旧版/手动), 用户希望
        「全部取消后重新分组」, 且要求原地覆盖原件。

策略:
  1. 先清: ws.row_dimensions[*].outline_level = 0; hidden = False;
     sheet_properties.outlinePr 重置(summaryBelow=None)
  2. 再做: 调用 skill 自带的 process_xlsx(input, output=None, sheet_name)
     重新生成自适应分组并 save
  3. 不修改 cell 值、不动字体、不合并/解合并、不动其它 sheet
"""
from __future__ import annotations

import sys
from pathlib import Path

# 把 skill 目录加入 path, 直接 import
SKILL_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SKILL_DIR))

from openpyxl import load_workbook  # noqa: E402

from outline_grouper import (  # noqa: E402
    DEFAULT_SHEET,
    apply_outline_to_worksheet,
    build_outline_groups,
    read_sheet_rows,
)


def clear_outline(ws) -> int:
    """清掉 ws 上所有 outline 痕迹, 返回清掉的行数。"""
    cleared = 0
    for rd in ws.row_dimensions.values():
        if rd.outline_level or rd.hidden:
            rd.outline_level = 0
            rd.hidden = False
            cleared += 1
    # 重置 outlinePr
    ws.sheet_properties.outlinePr.summaryBelow = None
    ws.sheet_properties.outlinePr.summaryRight = None
    return cleared


def main(input_path: str, sheet_name: str = DEFAULT_SHEET) -> int:
    p = Path(input_path).resolve()
    if not p.exists():
        print(f"[ERROR] xlsx 不存在: {p}", file=sys.stderr)
        return 1

    wb = load_workbook(p, read_only=False)
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] sheet {sheet_name!r} 不存在; 现有: {wb.sheetnames}", file=sys.stderr)
        return 1

    ws = wb[sheet_name]

    # 1) 清空旧 outline
    cleared = clear_outline(ws)
    print(f"[1/3] 清空旧 outline 行数: {cleared}")

    # 2) 重新计算并套 outline
    rows = read_sheet_rows(ws)
    groups = build_outline_groups(rows)
    apply_outline_to_worksheet(ws, groups)
    print(f"[2/3] 重新生成分组区间数: {len(groups)} (段+定额)")

    # 3) 原地保存
    wb.save(p)

    # 验证
    sec = sum(1 for r in rows if r and str(r[0]).strip() == "段")
    quo = sum(1 for r in rows if r and str(r[0]).strip() == "定")
    print(
        f"[3/3] 保存: {p}\n"
        f"      sheet={sheet_name} 总行数={len(rows)} 段={sec} 定额={quo}"
    )
    print(f"      其它保留 sheet: {[s for s in wb.sheetnames if s != sheet_name]}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: clear_and_regroup.py <input.xlsx> [sheet_name]")
        sys.exit(2)
    sheet = sys.argv[2] if len(sys.argv) >= 3 else DEFAULT_SHEET
    sys.exit(main(sys.argv[1], sheet))
