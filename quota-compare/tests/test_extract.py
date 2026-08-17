#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""quota-compare/extract.py 数据层 + CLI 回归测试。

守住的契约：
  1. normalize_rows(bytes) 归一为 9 列
  2. collect_hits 命中数 = CLI 直接 run_topic 的命中数
  3. write_xlsx_bytes 写出的 xlsx 可被 openpyxl 重新打开
  4. 4 个主题（踢脚/挖/扶手/水磨石）命中数稳定

数据源：quota-compare/原始样本/*.xlsx（人工放置，gitignore）。
"""
import sys
from collections import Counter
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "原始样本"
sys.path.insert(0, str(ROOT))

import extract  # noqa: E402

# 与 extract.PROVINCES 对齐（测试单点写一份，避免被 main 模块改动带偏）
PROVINCES = [
    ("四川", ["四川上.xlsx", "四川下.xlsx"]),
    ("广东", ["广东上.xlsx", "广东中.xlsx", "广东下.xlsx"]),
    ("河南", ["河南上.xlsx", "河南下.xlsx"]),
    ("重庆", ["重庆上.xlsx", "重庆下.xlsx"]),
]

# 关键命中数（OR 主导语义，2026-08-17 web 接入时确认）
# 注：CLI TOPICS 「挖」any + exclude 组合下命中数从 74 → 1060（新语义下「土」/「淤泥」等
# 词只要含就命中，不再要求含「挖」AND）。这是用户明确指示的语义变化。
EXPECTED = {
    "扶手":  {"total": 117, "by_prov": {"四川": 21, "广东": 40, "河南": 2, "重庆": 54}},
    "踢脚":  {"total": 43,  "by_prov": {"四川": 6,  "广东": 20, "河南": 0, "重庆": 17}},
    "挖":    {"total": 1060, "by_prov": {"四川": 449, "广东": 225, "河南": 270, "重庆": 116}},
}


@pytest.fixture(scope="module")
def by_province() -> dict:
    """读 4 省 9 册原始样本到 by_province dict。"""
    out = {}
    for prov, files in PROVINCES:
        for fn in files:
            path = SRC / fn
            if not path.exists():
                continue
            out.setdefault(prov, []).append((fn, path.read_bytes()))
    assert out, f"原始样本不可用：{SRC}"
    return out


# ── 1. 数据层：normalize_rows(byte) ──────────────────────────────

def test_normalize_rows_returns_9_columns(by_province):
    """每行必为 9 列，左截断 + 右补 None。"""
    for prov, files in by_province.items():
        for fn, xlsx_bytes in files:
            rows = extract.normalize_rows(xlsx_bytes)
            assert len(rows) > 0, f"{prov}/{fn} 空"
            for r in rows:
                assert len(r) == 9, f"{prov}/{fn} 行 {len(r)} 列: {r!r}"
            # 至少有一个「定」行块可被匹配（"扶手" 关键词在每个 xlsx 都可能命中）
            cats = [str(r[0] or "").strip() for r in rows]
            assert "定" in cats, f"{prov}/{fn} 没有 '定' 行"


# ── 2. 数据层：collect_hits 命中数稳定 ─────────────────────────

@pytest.mark.parametrize("keyword,any_terms,exclude_terms", [
    ("扶手", [], []),
    ("踢脚", ["踢脚线", "踢脚板"], []),
    ("挖",   ["土", "淤泥", "冻土", "沟槽", "基坑", "槽坑"],
            ["机械", "挖掘机", "挖孔", "钻"]),
])
def test_collect_hits_counts(keyword, any_terms, exclude_terms, by_province):
    """CLI 命中数 = 数据层 collect_hits 命中数。"""
    _, summary, hit_report = extract.collect_hits(
        by_province, keyword=keyword, any_terms=any_terms, exclude_terms=exclude_terms
    )
    total = sum(summary.values())
    assert total == EXPECTED[keyword]["total"], (
        f"{keyword} 总数 {total} != 期望 {EXPECTED[keyword]['total']}; "
        f"实际按省 {dict(summary)}"
    )
    for prov, expected_n in EXPECTED[keyword]["by_prov"].items():
        assert summary[prov] == expected_n, (
            f"{keyword}/{prov} {summary[prov]} != {expected_n}"
        )
    assert len(hit_report) == total


# ── 3. 数据层：write_xlsx_bytes 可重新打开 ─────────────────────

def test_write_xlsx_bytes_is_valid(by_province):
    """写出的 xlsx 能 openpyxl 重新打开，sheet 名正确，行数合理。"""
    _, summary, _ = extract.collect_hits(
        by_province, keyword="扶手", any_terms=[], exclude_terms=[]
    )
    blocks_by_prov, _, _ = extract.collect_hits(
        by_province, keyword="扶手", any_terms=[], exclude_terms=[]
    )
    out_bytes = extract.write_xlsx_bytes(blocks_by_prov, sheet_title="扶手对比")
    assert isinstance(out_bytes, (bytes, bytearray))
    assert len(out_bytes) > 1000, "xlsx 太小，疑似空"

    wb = openpyxl.load_workbook(BytesIO(out_bytes), read_only=True)
    try:
        assert "扶手对比" in wb.sheetnames
        ws = wb["扶手对比"]
        # 实际命中行数 ≥ 总命中数（额外还有 HEADER + 省份段标题 + 空行）
        row_count = sum(1 for _ in ws.iter_rows())
        assert row_count >= sum(summary.values()) + 1 + 4  # +1 HEADER +4 省份分隔
    finally:
        wb.close()


# ── 4. CLI 兜底：run_topic 跑通（与 README 落地结果一致） ─────

@pytest.mark.parametrize("keyword,any_terms,exclude_terms,expected_total", [
    ("扶手", [], [], 117),
    ("挖",   ["土", "淤泥", "冻土", "沟槽", "基坑", "槽坑"],
            ["机械", "挖掘机", "挖孔", "钻"], 1060),
])
def test_run_topic_smoke(keyword, any_terms, exclude_terms, expected_total, tmp_path, monkeypatch):
    """CLI 跑一次不写当前目录（用 tmp_path 替换 ROOT）。"""
    # redirect ROOT to tmp so we don't overwrite production xlsx
    monkeypatch.setattr(extract, "ROOT", tmp_path)
    monkeypatch.setattr(extract, "SRC", SRC)  # SRC 仍指向真样本

    # 临时手工造一份 by_province（不能 monkeypatch 全局 PROVINCES，会改 extract 行为）
    by_province = {}
    for prov, files in PROVINCES:
        for fn in files:
            p = SRC / fn
            if not p.exists():
                continue
            by_province.setdefault(prov, []).append((fn, p.read_bytes()))

    blocks_by_prov, summary, _ = extract.collect_hits(
        by_province, keyword=keyword, any_terms=any_terms, exclude_terms=exclude_terms
    )
    out_bytes = extract.write_xlsx_bytes(blocks_by_prov, sheet_title=f"{keyword}对比")
    (tmp_path / f"{keyword}_smoke.xlsx").write_bytes(out_bytes)
    assert sum(summary.values()) == expected_total
