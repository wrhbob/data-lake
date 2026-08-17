#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""4 省 9 册广联达 xlsx → 跨省定额对比（硬匹配 + 两个 CLI 兜底）。

两种模式：
  ① 单次：python extract.py <keyword> [--any "t1 t2 ..."] [--exclude "b1 b2 ..."]
  ② 批量：python extract.py --all（跑内置 TOPICS 表）

匹配规则（短路求值）：
  1. keyword 必须在 name 里出现
  2. 任一 exclude 词出现 → 排除
  3. 若给了 --any，至少一个 any 词必须出现

数据层（web 后端复用）：
  - normalize_rows(bytes)         9 列归一
  - collect_hits(by_province,...)  按省聚合命中块
  - write_xlsx_bytes(blocks_by_prov, sheet_title) -> bytes

示例：
  # 踢脚线 / 踢脚板（任一命中）
  python extract.py "踢脚" --any "踢脚线 踢脚板"

  # 人工挖土方（宽召回 + 防误伤）
  python extract.py "挖" --any "土 淤泥 冻土 沟槽 基坑 槽坑" \\
                          --exclude "机械 挖掘机 挖孔 钻"

  # 一次性重新生成 TOPICS 表里所有主题
  python extract.py --all
"""
import argparse
import openpyxl
from collections import Counter
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).resolve().parent
SRC = ROOT / "原始样本"

HEADER = ["省份", "序号", "项目编码", "名称", "项目特征", "计量单位",
          "消耗量", "基价", "标准换算", "标准换算来源"]

# 子行分类（挂在「定」后面同属一个定额块）
CHILD_CATS = {"工", "料", "机", "综", "配", "主材"}

# 省×册（共用一份；所有主题都是这 9 个 xlsx）
PROVINCES = [
    ("四川", ["四川上.xlsx", "四川下.xlsx"]),
    ("广东", ["广东上.xlsx", "广东中.xlsx", "广东下.xlsx"]),
    ("河南", ["河南上.xlsx", "河南下.xlsx"]),
    ("重庆", ["重庆上.xlsx", "重庆下.xlsx"]),
]

# 批量模式用：keyword / sheet_title / output_stem / any 空格串 / exclude 空格串
TOPICS = [
    ("踢脚", "踢脚线对比",     "踢脚线_跨省对比",     "踢脚线 踢脚板",                  ""),
    ("挖",   "人工挖土方对比",  "人工挖土方_跨省对比",   "土 淤泥 冻土 沟槽 基坑 槽坑",    "机械 挖掘机 挖孔 钻"),
]


def sanitize_stem(s: str) -> str:
    """Windows 文件名禁用字符替换。"""
    bad = '<>:"/\\|?*'
    out = "".join("_" if ch in bad else ch for ch in s).strip()
    return out or "未命名"


def split_terms(s: str) -> list[str]:
    return [t for t in s.split() if t]


# ── 数据层（web 后端调用这一层）───────────────────────────────────

def normalize_rows(xlsx_bytes: bytes) -> list[list]:
    """读「定额条目」sheet，找表头位置后归一为 9 列 list[list]。

    web 模式：直接传 final_xlsx 的字节流。
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True)
    ws = wb["定额条目"] if "定额条目" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = None
    for i, r in enumerate(rows):
        cells = [str(c) if c is not None else "" for c in r]
        if "项目编码" in cells and "名称" in cells:
            header_idx = i
            break
    data = rows[header_idx + 1:] if header_idx is not None else rows

    out = []
    for r in data:
        cells = list(r[:9])
        cells += [None] * (9 - len(cells))
        out.append(cells)
    return out


def matches(name: str, *, keyword: str, any_terms: list[str], exclude_terms: list[str]) -> bool:
    """硬匹配 + 兜底谓词（OR 主导语义）。

    规则（短路求值）：
      1. exclude 优先：任一 exclude 词出现 → 排除
      2. keyword 与 any_terms 是「补充词」平行关系，二者取并集：
         name 含 keyword  → 命中；
         任一 any 词在 name 里 → 命中；
         keyword 和 any 都没占中 → 不命中。
      3. 都没填（keyword 与 any_terms 都空）→ 任意 name 都命中（理论上不触发，
         keyword 已 Form(..., description=...) 校验非空）。
    """
    text = name or ""
    if not text:
        return False
    if any(ex in text for ex in exclude_terms):
        return False
    candidates = []
    if keyword:
        candidates.append(keyword)
    candidates.extend(any_terms)
    if not candidates:
        return True
    return any(t in text for t in candidates)


def extract_blocks(rows: list[list], *, keyword: str, any_terms: list[str], exclude_terms: list[str]) -> list[list]:
    """返回命中「定」行块列表；每块 = [段祖先..., 定行, 子行...]."""
    blocks = []
    section_path: list[list] = []
    n = len(rows)
    i = 0
    while i < n:
        cat = str(rows[i][0] or "").strip()
        if cat == "段":
            code = str(rows[i][1] or "").strip()
            depth = code.count(".") + 1 if code else 1
            section_path = section_path[: max(0, depth - 1)]
            section_path.append(rows[i])
            i += 1
            continue
        if cat == "定":
            name = rows[i][2] or ""
            feature = rows[i][3] or ""
            if matches(name, keyword=keyword, any_terms=any_terms, exclude_terms=exclude_terms):
                block = list(section_path)
                block.append(rows[i])
                j = i + 1
                while j < n and str(rows[j][0] or "").strip() in CHILD_CATS:
                    block.append(rows[j])
                    j += 1
                blocks.append(block)
                i = j
                continue
        i += 1
    return blocks


def collect_hits(by_province: dict, *, keyword: str, any_terms: list[str], exclude_terms: list[str]):
    """按省聚合命中块 + 摘要 + 命中清单。

    Args:
        by_province: {省份: [(filename, xlsx_bytes), ...]}，每个省多个册。

    Returns:
        (blocks_by_prov, summary, hit_report)
        blocks_by_prov: {省份: [[段祖先..., 定行, 子行...], ...]} 给 write_xlsx_bytes 用
        summary:       Counter(省份 -> 命中定额数)
        hit_report:    [(prov, filename, code, name), ...] 给 CLI verbose 输出用
    """
    blocks_by_prov = {}
    summary = Counter()
    hit_report = []
    for prov, files in by_province.items():
        blocks_by_prov[prov] = []
        seen_codes = set()             # 同省多册 / 同省重复编码去重
        for filename, xlsx_bytes in files:
            rows = normalize_rows(xlsx_bytes)
            blocks = extract_blocks(rows, keyword=keyword,
                                    any_terms=any_terms, exclude_terms=exclude_terms)
            for block in blocks:
                ding = next(b for b in block if str(b[0] or "").strip() == "定")
                code = str(ding[1] or "").strip()
                if code in seen_codes:
                    continue
                seen_codes.add(code)
                blocks_by_prov[prov].append(block)
                summary[prov] += 1
                hit_report.append((prov, filename, code, ding[2]))
    return blocks_by_prov, summary, hit_report


def write_xlsx_bytes(blocks_by_prov: dict, *, sheet_title: str) -> bytes:
    """把 blocks_by_prov 写成一个 xlsx 文件，返回字节流。

    与 CLI 的 run_topic 共享同样的输出格式（HEADER + 省份段 + 定 + 子行）。
    web 后端拿到 bytes 直接 Response(...) 返回。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(HEADER)

    for prov, blocks in blocks_by_prov.items():
        ws.append([f"==== {prov} ===="] + [None] * 9)
        last_section_key = None  # 段祖先去重
        for block in blocks:
            sec_key = tuple(b[1] for b in block if str(b[0] or "").strip() == "段")
            if sec_key != last_section_key:
                for b in block:
                    if str(b[0] or "").strip() == "段":
                        ws.append([prov] + b)
                last_section_key = sec_key
            for b in block:
                if str(b[0] or "").strip() != "段":
                    ws.append([prov] + b)
        # 省之间空 3 行
        for _ in range(3):
            ws.append([None] * 10)

    # = 开头的字符串被 openpyxl 默认当公式写入；强制改回文本
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CLI 层（仅命令行入口使用；web 后端不调）───────────────────────

def run_topic(*, keyword: str, sheet_title: str, output_stem: str,
              any_terms: list[str], exclude_terms: list[str], verbose: bool) -> dict:
    """CLI 入口：读 PROVINCES×原始样本/*.xlsx，调 collect_hits，写文件。"""
    by_province = {}
    for prov, files in PROVINCES:
        for fn in files:
            path = SRC / fn
            if not path.exists():
                continue
            by_province.setdefault(prov, []).append((fn, path.read_bytes()))

    blocks_by_prov, summary, hit_report = collect_hits(
        by_province, keyword=keyword, any_terms=any_terms, exclude_terms=exclude_terms
    )
    out_bytes = write_xlsx_bytes(blocks_by_prov, sheet_title=sheet_title)
    out_path = ROOT / f"{sanitize_stem(output_stem)}.xlsx"
    out_path.write_bytes(out_bytes)

    if verbose:
        print(f"\n=== [{sheet_title}] ===")
        print(f"输出: {out_path.name}")
        print(f"命中: {len(hit_report)} 条  按省: {dict(summary)}")
        cur = None
        for prov, fn, code, name in hit_report:
            if prov != cur:
                print(f"\n[{prov}]")
                cur = prov
            print(f"  [{fn}] {code}  {name}")
    else:
        print(f"[{sheet_title}] → {out_path.name}  ({len(hit_report)} 条, 按省 {dict(summary)})")

    return {"output": out_path, "n": sum(summary.values()), "by_province": dict(summary)}


def main():
    ap = argparse.ArgumentParser(description="跨省定额对比：从 4 省 9 册广联达 xlsx 抽取主题定额")
    ap.add_argument("keyword", nargs="?", help="必须包含的关键词（name 含才命中）")
    ap.add_argument("--any", dest="any_terms", default="",
                    help="扩展命中词（空格分隔，OR 关系）")
    ap.add_argument("--exclude", dest="exclude_terms", default="",
                    help="排除词（空格分隔）")
    ap.add_argument("--title", default="", help="sheet 标题（默认 = <keyword>对比）")
    ap.add_argument("--stem", default="", help="输出文件 stem（默认 = <keyword>_跨省对比）")
    ap.add_argument("--all", action="store_true", help="批量跑内置 TOPICS 表")
    args = ap.parse_args()

    if args.all:
        for keyword, sheet_title, output_stem, any_s, ex_s in TOPICS:
            run_topic(
                keyword=keyword,
                sheet_title=sheet_title,
                output_stem=output_stem,
                any_terms=split_terms(any_s),
                exclude_terms=split_terms(ex_s),
                verbose=False,
            )
        return

    if not args.keyword:
        ap.error("需要 keyword 或 --all")

    any_terms = split_terms(args.any_terms)
    exclude_terms = split_terms(args.exclude_terms)
    sheet_title = args.title or f"{args.keyword}对比"
    output_stem = args.stem or f"{args.keyword}_跨省对比"

    run_topic(
        keyword=args.keyword,
        sheet_title=sheet_title,
        output_stem=output_stem,
        any_terms=any_terms,
        exclude_terms=exclude_terms,
        verbose=True,
    )


if __name__ == "__main__":
    main()
