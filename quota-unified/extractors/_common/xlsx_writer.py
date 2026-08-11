#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_writer.py — 把定额抽取的多源内容合并成一个多 sheet xlsx

设计目标：
  输入:
    - 已生成的 10 列 CSV (<stem>_数值待审核.csv)
    - narrative_parser.parse_narrative 返回的 {preface, chapters[], warnings}
    - 可选:伴生的 _issues.md 路径(供"问题报告" sheet 留 hook;本期暂未启用)

  输出:
    - 多 sheet xlsx,sheet 顺序固定:
        1. 定额条目     (CSV 全文 → 10 列)        ← 用户硬约束：最开头
        2. 册说明       (preface 全文 → A1)
        3..N. A / B / ...(每章一个 sheet;
                          A1 = 说明、A2 = 工程量计算规则)
    - 用户硬约束: 章 sheet 内说明放第一个单元格、计算规则紧贴其下方,
                  不分列、不加分隔单元。

约束（详见 plan mutable-chasing-honey.md）：
  - Sheet 名限 31 字符(Excel 上限) → 章用单字母 code 永远够短
  - 章 sheet 内部只有 1 列(A 列);其他列不写,Excel 也不会显示空列
  - wrap_text=True,vertical=top,列宽自适应
  - 复用 province 子脚本写出的 CSV,不做重复解析
  - 复用 quota-csv-finalize/to_xlsx.py 的 openpyxl 用法(同款依赖)
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


# Excel sheet 名 31 字符上限
_MAX_SHEET_NAME = 31
# HTML <table>...</table> 检测(若叙述段里嵌了表,替换为占位文字避免污染)
_TABLE_RE = re.compile(r"<table.*?</table>", re.DOTALL | re.IGNORECASE)


def write_quota_xlsx(
    csv_path: Path,
    xlsx_path: Path,
    narrative: dict,
    *,
    issues_md_path: Path | None = None,
) -> dict:
    """合并 CSV + narrative → 多 sheet xlsx;返回写入摘要。

    Args:
        csv_path: 已生成的 <stem>_数值待审核.csv 路径(由 province 子脚本产出)
        xlsx_path: 要写入的 <stem>.xlsx 路径(通常 = csv_path.with_suffix(".xlsx"))
        narrative: parse_narrative() 返回的 dict
        issues_md_path: 可选,伴生 _issues.md(本期未写入 xlsx,保留 hook)

    Returns:
        dict 包含:
          - sheet_names: list[str]   写入的所有 sheet 名
          - n_rows_quota: int        "定额条目" sheet 的数据行数(不含表头)
          - n_chapters:  int         章 sheet 数
          - preface_chars: int       册说明 cell 的字符数
    """
    wb = Workbook()
    # Workbook() 默认会创建一个名为 "Sheet" 的空表;先删掉,后面按顺序创建
    default_ws = wb.active
    wb.remove(default_ws)

    sheet_names: list[str] = []

    # ── Sheet 1: 定额条目 (用户硬约束：最开头) ──
    ws_quota = wb.create_sheet(title="定额条目")
    n_rows_quota = _write_csv_to_sheet(ws_quota, csv_path)
    sheet_names.append("定额条目")

    # ── Sheet 2: 册说明 ──
    preface_text = _clean_for_cell(narrative.get("preface", "") or "")
    ws_preface = wb.create_sheet(title="册说明")
    _write_narrative_cell(ws_preface, preface_text, n_cols=1)
    sheet_names.append("册说明")

    # ── Sheet 3..N: 每章一个 sheet ──
    n_chapters = 0
    chapters = narrative.get("chapters", []) or []
    seen_codes: dict[str, int] = {}  # 重复章 code → 追加 _2 / _3 后缀
    for ch in chapters:
        code = ch.get("code", "") or "?"
        title = ch.get("title", "") or ""
        # 命名: 单字母 code(标题舍去,sheet 名简短为先);
        # 重复 code → 加 _2/_3
        sheet_name = code
        if sheet_name in seen_codes:
            seen_codes[sheet_name] += 1
            sheet_name = f"{code}_{seen_codes[sheet_name]}"
        else:
            seen_codes[sheet_name] = 1
        # 31 字符保护
        sheet_name = sheet_name[:_MAX_SHEET_NAME]

        ws_ch = wb.create_sheet(title=sheet_name)
        description_text = _clean_for_cell(ch.get("description", "") or "")
        calc_rules_text = _clean_for_cell(ch.get("calc_rules", "") or "")

        # 用户硬约束: 章说明放 A1,工程量计算规则紧贴其下方(A2)。
        # 不加分隔单元、不分列。只在 A 列;row=1 / row=2。
        if description_text and calc_rules_text:
            _write_narrative_cell(ws_ch, description_text, n_cols=1, row=1)
            _write_narrative_cell(ws_ch, calc_rules_text, n_cols=1, row=2)
        elif description_text:
            _write_narrative_cell(ws_ch, description_text, n_cols=1, row=1)
        elif calc_rules_text:
            _write_narrative_cell(ws_ch, calc_rules_text, n_cols=1, row=1)
        else:
            # 章内容全空:写一个占位说明
            placeholder = f"(本章无叙述内容; 章标题: {title or '无'})"
            _write_narrative_cell(ws_ch, placeholder, n_cols=1, row=1)

        sheet_names.append(sheet_name)
        n_chapters += 1

    # ── 保存 ──
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)

    return {
        "sheet_names":   sheet_names,
        "n_rows_quota":  n_rows_quota,
        "n_chapters":    n_chapters,
        "preface_chars": len(preface_text),
    }


# ─────────────────────────────────────────────────────────────────────
# 私有辅助
# ─────────────────────────────────────────────────────────────────────

def _write_narrative_cell(ws, text: str, *, n_cols: int = 1, row: int = 1) -> None:
    """把叙述文本写入 ws 的 row 行 / col 1 cell。

    - wrap_text=True(自动换行)
    - vertical=top
    - 列宽设宽(叙述类长文本,80 字符宽)
    - n_cols 只决定"列宽设置"个数;叙述 sheet 只用 A 列
    """
    if not text:
        text = ""

    # 写内容
    ws.cell(row=row, column=1, value=text)

    # 应用样式(只对 A1 cell 即可,因为就一个 cell)
    cell = ws.cell(row=row, column=1)
    cell.alignment = Alignment(
        wrap_text=True,
        vertical="top",
        horizontal="left",
    )

    # 列宽
    col_widths = [80] + [16] * max(0, n_cols - 1)
    for i, w in enumerate(col_widths, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w


def _write_csv_to_sheet(ws, csv_path: Path) -> int:
    """把 csv_path 内容写入 ws;返回数据行数(不含表头)。

    - 第一行作为表头加粗
    - 其余行原样写入
    - 列宽自适应(以最长 cell 估算)
    """
    if not csv_path.exists():
        # 不抛错;写一行错误提示占位
        ws.cell(row=1, column=1, value=f"(CSV 缺失: {csv_path})")
        return 0

    # 用 utf-8-sig 读(兼容 BOM)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return 0

    n_data_rows = 0
    header_font = Font(bold=True)

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            # bj / hu / gd: 编码列 (col 2 = 项目编码, 含 PID / 工料机编码)
            #   PID "1-1" / "G1-1" / "C1-4-1" 被 Excel 自动识别为日期 (→ "1月1日")
            #   长数字编码 "3109003301" (10 位+) 被识别为科学计数法 (→ "3.109E+09")
            #   强制文本格式 '@' 防止 Excel 自动转换.
            # bj 实际受害最重 (PID 短 + 编码长), hu/gd PID 含字母也有同样隐患, 统一处理.
            if r_idx > 1 and c_idx == 2:
                cell.number_format = "@"
            if r_idx == 1:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
        # 数据行计数放在 row 循环内、cell 循环外(避免按列重复加)
        if r_idx > 1:
            n_data_rows += 1

    # 列宽自适应(粗略:每列按前 100 行的最长 cell 估算)
    sample = rows[: min(100, len(rows))]
    for c_idx in range(len(rows[0]) if rows else 0):
        max_len = 0
        for row in sample:
            if c_idx < len(row):
                # 中文字符按 2 算宽度,英文按 1
                max_len = max(max_len, _display_width(row[c_idx]))
        col_letter = ws.cell(row=1, column=c_idx + 1).column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)

    return n_data_rows


def _display_width(s: str) -> int:
    """估算 cell 显示宽度(中文字符按 2,其他按 1)。"""
    if not s:
        return 0
    w = 0
    for ch in s:
        if ord(ch) > 0x2E80:  # CJK 区间
            w += 2
        else:
            w += 1
        if w > 100:  # 上限,避免超宽单 cell 把列撑死
            break
    return w


def _clean_for_cell(text: str) -> str:
    """把叙述文本里偶尔出现的 HTML <table>...</table> 替换为占位提示,
    避免 openpyxl 把原始 HTML 标签当成 cell value 写入后 Excel 显示一坨标签。"""
    if not text:
        return ""
    if "<table" not in text.lower():
        return text
    return _TABLE_RE.sub(
        "\n[此处原为表格,详见 PDF 对应章节]\n",
        text,
    )