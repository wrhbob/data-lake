#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""run_hunan.py — 一次性：跑 hu-2018 profile 处理湖南仿古.md → 写 xlsx 到同目录"""
import sys
from pathlib import Path

sys.path.insert(0, r"D:\工程造价学习\data_lake0714\data_lake0714\quota-unified")

from baseline import process_md_file, FeatureContext
from profile_schema import get_preset_profile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

MD = Path(r"D:\工程造价学习\data_lake0714\data_lake0714\quota\parser\tests\hunan\仿古\仿古.md")
OUT = Path(r"D:\工程造价学习\data_lake0714\data_lake0714\quota\parser\tests\hunan\仿古_待审核.xlsx")

profile = get_preset_profile("hu-2018")
ctx = FeatureContext(profile=profile)
all_rows, issues = process_md_file(MD, profile, ctx)

print(f"[run_hunan] rows={len(all_rows)}, issues={len(issues)}")

wb = Workbook()
ws = wb.active
ws.title = "定额条目"

# 列头（10 列）
header = ["类别", "编号", "名称/工作内容", "计量单位", "材料单价", "材料数量", "基价/单价", "验证和", "差值", "备注"]
ws.append(header)
for col_idx in range(1, len(header) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F5597")
    cell.alignment = Alignment(horizontal="center")

# 数据
sec_fill = PatternFill("solid", fgColor="E7E6E6")
for r in all_rows:
    ws.append(list(r) + [""] * (10 - len(r)))  # 补齐到 10 列
    cur_row = ws.max_row
    if r[0] == "段":
        for c in range(1, 11):
            ws.cell(row=cur_row, column=c).fill = sec_fill
            ws.cell(row=cur_row, column=c).font = Font(bold=True)

# 列宽
widths = [6, 10, 60, 10, 12, 12, 12, 12, 10, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

# 冻结首行
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"[run_hunan] saved: {OUT} ({OUT.stat().st_size:,} bytes)")
