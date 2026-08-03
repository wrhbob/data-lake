"""quota_parser 主流程（阶段 A / B / Worker）"""
from __future__ import annotations

import hashlib
import importlib.util as _ilu
import logging
import os
import sys
import time
import uuid
from pathlib import Path
from typing import Any, Callable

from .config import (
    PARSER_VERSION,
    PROVINCE_DEFAULT_KEY,
    PROVINCE_KEYWORDS,
    PROVINCE_NAMES,
    QUOTA_CSV_FINALIZE_DIR,
    QUOTA_MD_TO_CSV_DIR,
    MINERU_SCRIPT_DIR,
    CHUNK_THRESHOLD_PAGES,
    FINALIZE_STEPS,
    get_ocr_api_url,
    get_work_root,
    get_poll_interval,
    get_database_url,
)
from .exceptions import (
    InvalidPageRangeError,
    InvalidXlsxStructureError,
    ProfileExecutionError,
    UnsupportedProvinceError,
    WorkdirNotWritableError,
)
from .result import StageAResult, StageBResult


_logger = logging.getLogger(__name__)


def _resolve_province(province: object) -> str:
    """把任意形态的 province 输入收敛到 PROVINCE_KEYWORDS 里的 short code.

    v0.8 修复（7-31 重庆事故根因）：
        原代码只用 strip().lower(), 任何非 sc/cq 字符串都静默收敛到 'default',
        落到占位分支产出 0 行 empty_candidate.xlsx (用户看到"成功"但无数据).
        修复后: 识别不出就抛 UnsupportedProvinceError, 让 worker 标 failed_user.

    接受:
      - short code:  sc / cq
      - 中文名:      四川 / 重庆 (PROVINCE_NAMES 反向)
      - 关键词:      川建 / 重庆 (PROVINCE_KEYWORDS 元组反查)

    不接受 (抛 UnsupportedProvinceError):
      - None / 空字符串 / 纯空白
      - 英文别名 (Sichuan / chongqing 等)
      - 数字码 (500000 / 510000 等)
      - 任意其他字符串
    """
    if province is None:
        raise UnsupportedProvinceError("province 必填, 得到 None")
    s = str(province).strip()
    if not s:
        raise UnsupportedProvinceError("province 必填, 得到空字符串")

    # 1. 已是 short code (排除 default sentinel)
    if s in PROVINCE_KEYWORDS and s != PROVINCE_DEFAULT_KEY:
        return s

    # 2. 中文名反查 (PROVINCE_NAMES 是 {short_code: 中文名} 单层映射)
    for short_code, cn_name in PROVINCE_NAMES.items():
        if s == cn_name:
            return short_code

    # 3. PROVINCE_KEYWORDS 元组关键词反查 (sc 还含 '川建')
    for short_code, keywords in PROVINCE_KEYWORDS.items():
        if short_code == PROVINCE_DEFAULT_KEY:
            continue
        if s in keywords:
            return short_code

    # 全部不匹配 — 抛错, 不落占位
    accepted = (
        f"short codes: {[k for k in PROVINCE_KEYWORDS.keys() if k != PROVINCE_DEFAULT_KEY]}; "
        f"中文名: {list(PROVINCE_NAMES.values())}; "
        f"关键词: {[(k, v) for k, v in PROVINCE_KEYWORDS.items() if k != PROVINCE_DEFAULT_KEY]}"
    )
    raise UnsupportedProvinceError(
        f"未知 province={province!r}. 接受: {accepted}"
    )


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def _load_mineru_function(module_filename: str, function_name: str):
    """动态加载 external/mineru_pdf_parse/scripts/<module>.py 中的函数。"""
    path = MINERU_SCRIPT_DIR / module_filename
    if not path.exists():
        raise WorkdirNotWritableError(f"找不到 mineru script: {path}")
    spec = _ilu.spec_from_file_location(f"_quota_parser_{module_filename[:-3]}", str(path))
    if spec is None or spec.loader is None:
        raise WorkdirNotWritableError(f"无法创建 mineru spec: {path}")
    mod = _ilu.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return getattr(mod, function_name)


def _import_md_extract_module():
    """import quota_md_to_csv_v2/extract_quota.py 顶层模块。"""
    path = QUOTA_MD_TO_CSV_DIR / "extract_quota.py"
    if not path.exists():
        raise WorkdirNotWritableError(f"找不到 extract_quota: {path}")
    spec = _ilu.spec_from_file_location(
        "_quota_parser_quota_md_to_csv_v2_extract", str(path),
    )
    if spec is None or spec.loader is None:
        raise WorkdirNotWritableError(f"无法创建 extract spec: {path}")
    mod = _ilu.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _import_finalize_step(fname: str):
    """import quota_csv_finalize/<fname>.py 顶层模块（process_xlsx 函数）。"""
    path = QUOTA_CSV_FINALIZE_DIR / fname
    if not path.exists():
        raise WorkdirNotWritableError(f"找不到 finalize 脚本: {path}")
    spec = _ilu.spec_from_file_location(
        f"_quota_parser_quota_csv_finalize_{fname[:-3]}", str(path),
    )
    if spec is None or spec.loader is None:
        raise WorkdirNotWritableError(f"无法创建 finalize spec: {path}")
    mod = _ilu.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _read_pdf_page_count(pdf: Path) -> int:
    try:
        import fitz  # PyMuPDF
        with fitz.open(pdf) as doc:
            return doc.page_count
    except Exception:
        return 0


# ─────────────────────────────────────────────────────────
# 阶段 A: PDF → candidate xlsx
# ─────────────────────────────────────────────────────────
def run_quota_pipeline(
    *,
    pdf_path: str,
    work_dir: str,
    province: str | None = None,
    ocr_api_url: str | None = None,
    profile: str | None = None,
    enable_chunking: bool = True,
    task_id: str | None = None,
    on_chunk_done: Callable[[int, int, str], None] | None = None,
) -> StageAResult:
    """阶段 A: PDF → candidate xlsx（OCR + MD 抽取 + autofinalize 5 步）

    Args:
        pdf_path:     PDF 本地路径（Worker 由 MinIO 临时下载得到）
        work_dir:     任务独占目录（函数内部追加 task_id 子目录）
        province:     'sc' | 'cq' | 'default' | None（None / 未知值 收敛到 'default'）
        ocr_api_url:  覆盖默认 OCR URL（None 用 config.get_ocr_api_url()）
        profile:      显式 Profile 名;None 时按 province 取默认
        enable_chunking: >100 页 PDF 自动分段 OCR

    Returns:
        StageAResult（含 candidate_xlsx_path / sha256 / qa_report paths）

    Raises:
        InvalidPageRangeError: PDF 页数异常
        WorkdirNotWritableError: workspace 创建失败
        ProfileExecutionError: Province 子脚本抛错

    行为变更（v0.4 §9 #15）：
        - province=None / 'default' / 未知省份 → 收敛到 PROVINCE_DEFAULT_KEY,
          跳过省份抽取子步骤,直接写空 candidate.xlsx 占位。
        - 不再抛 UnsupportedProvinceError（仅这一处移除；其余位置不变）。
    """
    # ── v0.8 修复: province 收敛 + 硬校验 ──
    # 旧代码 (v0.4 §9 #15) 用 strip().lower() 软收敛,任何非 sc/cq 都落 'default'
    # 占位 → 0 行 empty_candidate.xlsx (7-31 重庆事故根因).
    # 新代码用 _resolve_province 三阶段反查 (short code / 中文名 / 关键词),
    # 识别不出就抛 UnsupportedProvinceError → worker 标 failed_user.
    try:
        province = _resolve_province(province)
    except UnsupportedProvinceError:
        # pipeline 顶层会包成 ProfileExecutionError 抛给 worker,
        # worker 在 _process_real_job 的 except 里捕获 → _mark_job_failed (failed_user)
        # 这里只需重新抛出,保留原始错误信息.
        raise

    pdf_p = Path(pdf_path).resolve()
    if not pdf_p.exists():
        raise WorkdirNotWritableError(f"PDF 不存在: {pdf_p}")

    page_count = _read_pdf_page_count(pdf_p)
    if page_count <= 0:
        raise InvalidPageRangeError(f"PDF 页数读取失败: {pdf_p}")

    task_id = task_id or f"qp_{uuid.uuid4().hex[:12]}"
    work_root = Path(work_dir).resolve() / task_id
    try:
        work_root.mkdir(parents=True, exist_ok=False)
    except OSError as e:
        raise WorkdirNotWritableError(f"创建 workspace 失败: {work_root}: {e}")

    ocr_dir = work_root / "ocr"
    ocr_dir.mkdir(parents=True, exist_ok=True)
    source_sha = _sha256_file(pdf_p)

    # ── OCR ──
    api = ocr_api_url or get_ocr_api_url()
    md_text = None
    md_path: Path | None = None
    result_json_path: Path | None = None

    if enable_chunking and page_count > CHUNK_THRESHOLD_PAGES:
        parse_chunked = _load_mineru_function("parse_chunked.py", "parse_chunked")
        info = parse_chunked(
            pdf_path=str(pdf_p),
            output_dir=str(ocr_dir),
            # M2 修复 (2026-08-03): 与 parse_pdf 路径对齐, 显式传 api_url,
            # 避免 parse_chunked 落到自己的 DEFAULT_API (= MINERU_API_URL 环境变量或
            # http://172.16.20.23:8000), 与 worker.metadata_payload["ocr_api_url"] / config.
            # get_ocr_api_url() 分叉. worker 用什么 URL, chunked 路径就用什么 URL.
            api_url=api,
            on_chunk_done=on_chunk_done,
        )
        md_path = Path(info["markdown_path"]) if info.get("markdown_path") else None
        result_json_path = Path(info["result_json_path"]) if info.get("result_json_path") else None
        # ── F10 修复: 部分 chunk 失败被当成功的根因 ──
        # parse_chunked 用 zip(chunks, result_paths) 合并 → 失败的 chunk 被静默跳过,
        # merged.md 仍写出但内容残缺; 老代码不看 all_succeeded → 直接当成功.
        # 修复: 任一 chunk 失败 → 抛 ProfileExecutionError → worker 标 failed_permanent.
        if not info.get("all_succeeded", False):
            failed = [
                f"chunk {c['index']} (PDF p.{c['pages'][0]}-{c['pages'][1]}): {c['error']}"
                for c in info.get("chunks", []) if c.get("status") != "succeeded"
            ]
            raise ProfileExecutionError(
                f"OCR 部分 chunk 失败 ({len(failed)}/{len(info.get('chunks', []))}): "
                + "; ".join(failed)
            )
    else:
        # v0.6 §#6 限制: 非 chunked 路径不传 on_chunk_done（parse_pdf 整体 < 2 min,
        # 30 min 首 chunk 超时足够覆盖）。如果未来 parse_pdf 变慢,再补 on_chunk_done。
        parse_pdf = _load_mineru_function("parse_pdf.py", "parse_pdf")
        info = parse_pdf(pdf_path=str(pdf_p), output_dir=str(ocr_dir), api_url=api)
        md_path = Path(info["markdown_path"]) if info.get("markdown_path") else None
        result_json_path = Path(info["result_json_path"]) if info.get("result_json_path") else None

    if not md_path or not md_path.exists():
        raise ProfileExecutionError(f"OCR 失败: {pdf_p}")

    # ── MD → candidate xlsx ──
    # v0.4 §9 #15: default sentinel 跳过省份抽取(extract_quota._load_province_module
    # 对 'default' 抛 ValueError),直接产空 candidate.xlsx 占位。
    if province == PROVINCE_DEFAULT_KEY:
        candidate_xlsx = work_root / "candidate" / "empty_candidate.xlsx"
        candidate_xlsx.parent.mkdir(parents=True, exist_ok=True)
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "定额条目"
        placeholder = ["placeholder"] * 10
        ws.append(placeholder)
        wb.save(candidate_xlsx)
        extract_info = {
            "rows_count": 0, "sections_count": 0, "projects_count": 0,
            "issues_count": 0, "n_chapters": 0, "preface_chars": 0,
            "issues_md_path": None,
            "xlsx_path": str(candidate_xlsx),  # v0.7 fix: 占位路径也要给下游 line 226 读
            "warnings": ["province=default — placeholder candidate"],
        }
    else:
        extract_mod = _import_md_extract_module()
        extract_info = extract_mod.process_md_file(
            md_path=str(md_path),
            work_dir=str(work_root / "candidate"),
            province=province,
            keep_csv=False,
            run_finalize=True,
        )

    candidate_xlsx = Path(extract_info["xlsx_path"])
    candidate_sha = _sha256_file(candidate_xlsx)

    warnings: list[str] = []
    for w in extract_info.get("warnings", []) or []:
        warnings.append(f"narrative: {w}")

    return StageAResult(
        task_id=task_id,
        status="candidate_ready",
        parser_version=PARSER_VERSION,
        profile=profile or province,
        source_sha256=source_sha,
        source_pdf_path=str(pdf_p),
        ocr_markdown_path=str(md_path),
        ocr_result_json_path=str(result_json_path) if result_json_path else None,
        candidate_xlsx_path=str(candidate_xlsx),
        candidate_xlsx_sha256=candidate_sha,
        issues_md_path=extract_info.get("issues_md_path"),
        qa_report_json_path=None,
        qa_report_md_path=None,
        warnings=warnings,
        metrics={
            "pages": page_count,
            "rows": extract_info["rows_count"],
            "sections": extract_info["sections_count"],
            "projects": extract_info["projects_count"],
            "n_chapters": extract_info["n_chapters"],
            "preface_chars": extract_info["preface_chars"],
            "issues": extract_info["issues_count"],
        },
    )


# ─────────────────────────────────────────────────────────
# 阶段 B: reviewed → final
# ─────────────────────────────────────────────────────────
def finalize_reviewed_xlsx(
    *,
    reviewed_xlsx_path: str,
    output_xlsx_path: str,
    source_candidate_xlsx_path: str | None = None,
    task_id: str | None = None,
) -> StageBResult:
    """阶段 B: reviewed.xlsx → final.xlsx（直接保存，不再跑 finalize）。

    v0.2 选项 B: 仅校验 Sheet1 "定额条目" 结构;其他 sheet 完全信任人工结果。
    """
    from openpyxl import load_workbook

    task_id = task_id or f"qp_{uuid.uuid4().hex[:12]}"
    rev_p = Path(reviewed_xlsx_path).resolve()
    out_p = Path(output_xlsx_path).resolve()

    if not rev_p.exists():
        raise WorkdirNotWritableError(f"reviewed xlsx 不存在: {rev_p}")

    try:
        wb = load_workbook(rev_p, read_only=True)
    except Exception as e:
        raise InvalidXlsxStructureError(f"reviewed xlsx 无法打开: {type(e).__name__}: {e}")

    if not wb.sheetnames:
        raise InvalidXlsxStructureError("reviewed xlsx 没有任何 sheet")

    if "定额条目" not in wb.sheetnames:
        raise InvalidXlsxStructureError(
            f"reviewed xlsx 缺 '定额条目' sheet; 现有: {wb.sheetnames}"
        )

    # 第 1 位必须是 定额条目
    if wb.sheetnames[0] != "定额条目":
        raise InvalidXlsxStructureError(
            f"'定额条目' sheet 不在第 1 位; 顺序: {wb.sheetnames}"
        )

    ws = wb["定额条目"]
    # 人工审核补齐后的标准列数 = 9(worker OCR candidate.xlsx 实际产出是 8 列,
    # 人工补齐 1 列后到 9 列;reviewed upload 入口 service.py 也按 9 列校验,
    # finalize 此处保持一致)
    if ws.max_column != 9:
        raise InvalidXlsxStructureError(
            f"'定额条目' sheet 列数 = {ws.max_column}, 应为 9"
        )

    # reviewed → final 原样拷贝（仅 Sheet1 + 其他 sheet 内容）
    out_p.parent.mkdir(parents=True, exist_ok=True)
    import shutil
    shutil.copy2(rev_p, out_p)

    final_sha = _sha256_file(out_p)

    return StageBResult(
        task_id=task_id,
        status="final_ready",
        parser_version=PARSER_VERSION,
        final_xlsx_path=str(out_p),
        final_xlsx_sha256=final_sha,
        qa_report_json_path=None,
        qa_report_md_path=None,
        warnings=[],
        metrics={"reviewed_sheets": len(wb.sheetnames)},
    )


# ─────────────────────────────────────────────────────────
# Worker 常驻
# ─────────────────────────────────────────────────────────
def serve_worker(
    *,
    database_url: str | None = None,
    poll_interval_seconds: float | None = None,
    ocr_api_url: str | None = None,
    work_root: str | None = None,
) -> None:
    """常驻 Worker（轮询数据库 → 调阶段 A / B）— 当前为占位实现。

    实现细节在 quota_parser/worker/main.py（v0.2 后续 P1 阶段实现）。
    本占位仅做环境变量解析 + 健康检查,避免 Worker 误启动时静默失败。
    """
    db_url = database_url or get_database_url()
    if not db_url:
        raise ValueError(
            "QUOTA_PARSER_DATABASE_URL 未设置;"
            "Worker 无法轮询 parse_run 表"
        )

    interval = poll_interval_seconds if poll_interval_seconds is not None else get_poll_interval()
    wr = Path(work_root).resolve() if work_root else get_work_root()
    api = ocr_api_url or get_ocr_api_url()

    print(f"[worker] database_url = {db_url[:30]}...")
    print(f"[worker] work_root    = {wr}")
    print(f"[worker] ocr_api_url  = {api}")
    print(f"[worker] poll_interval = {interval}s")
    print(f"[worker] v0.2 占位: 真实轮询逻辑在 worker/main.py 后续实现")


# ─────────────────────────────────────────────────────────
# 兼容 v0.1 CLI 入口
# ─────────────────────────────────────────────────────────
def main_cli(argv: list[str]) -> int:
    """CLI 入口（兼容老 skill 调用方式）."""
    import argparse
    ap = argparse.ArgumentParser(description="quota_parser CLI (v0.2)")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p_a = sub.add_parser("stage-a", help="阶段 A: PDF → candidate xlsx")
    p_a.add_argument("pdf_path")
    p_a.add_argument("--work-dir", required=True)
    p_a.add_argument("--province", required=True, choices=list(PROVINCE_KEYWORDS.keys()))
    p_a.add_argument("--ocr-url", default=None)
    p_a.add_argument("--no-chunking", action="store_true")

    p_b = sub.add_parser("stage-b", help="阶段 B: reviewed → final xlsx")
    p_b.add_argument("reviewed_xlsx")
    p_b.add_argument("--output", required=True)
    p_b.add_argument("--source-candidate", default=None)

    sub.add_parser("worker", help="启动常驻 Worker")

    args = ap.parse_args(argv)
    if args.cmd == "stage-a":
        result = run_quota_pipeline(
            pdf_path=args.pdf_path,
            work_dir=args.work_dir,
            province=args.province,
            ocr_api_url=args.ocr_url,
            enable_chunking=not args.no_chunking,
        )
        print(f"[OK] task_id: {result.task_id}")
        print(f"[OK] candidate: {result.candidate_xlsx_path}")
        print(f"[OK] sha256:   {result.candidate_xlsx_sha256}")
        print(f"[OK] metrics:  {result.metrics}")
        return 0
    if args.cmd == "stage-b":
        result = finalize_reviewed_xlsx(
            reviewed_xlsx_path=args.reviewed_xlsx,
            output_xlsx_path=args.output,
            source_candidate_xlsx_path=args.source_candidate,
        )
        print(f"[OK] task_id: {result.task_id}")
        print(f"[OK] final:    {result.final_xlsx_path}")
        print(f"[OK] sha256:   {result.final_xlsx_sha256}")
        return 0
    if args.cmd == "worker":
        serve_worker()
        return 0
    return 1


if __name__ == "__main__":
    sys.exit(main_cli(sys.argv[1:]))