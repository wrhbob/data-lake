"""
space_split_materials.py — 在材料名称字段的"汉字 ↔ 字母/数字"边界插入空格（第 4 步 / XLSX 版，v0.2 函数化版本）

输入：<stem>.xlsx（前 3 步 clean / drop_toc / fill 跑完产出）
      仅修改"定额条目" sheet 的数据，其它 sheet 保持不动。

仅处理行类型为 `料` 或 `配` 的行的第 3 列（材料名称字段）。

v0.2 函数级入口（Worker 使用）:
    from external.quota_csv_finalize.space_split_materials import process_xlsx
    info = process_xlsx(input_xlsx, output_xlsx=None)

CLI:
    python space_split_materials.py <input.xlsx> [output.xlsx]

规则（5 条 han↔alnum 规则；与 v0.1 baseline 完全一致，行为零变更）：
  1. 汉字 → 字母/数字：在字母数字前加空格
  2. 字母 → 汉字：在字母后加空格
  3. 数字 → 汉字：不加空格（且会删除现有空格，避免重复）
     "数字"指包含至少一个 0-9 字符的串（如 42.5、E43、M20×100）。
     "字母"指纯 A-Za-z 不含数字（如 EVA、PVC、SBS、PYII）。
  4. "综合" / "综合规格" 在末尾且前面有文字时，在其前加空格。
  5. 希腊字母（δ/Φ/α 等）按混合规则
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

TARGET_SHEET = "定额条目"

_HAN = r"[一-鿿]"
# 希腊字母 Unicode 范围：U+0370-U+03FF（基本）+ U+1F00-U+1FFF（扩展）
_GREEK = r"Ͱ-Ͽἀ-῿"

_TOKEN_RE = re.compile(
    rf"[一-鿿]+|[A-Za-z0-9{_GREEK}]+|[^一-鿿A-Za-z0-9{_GREEK}]+"
)
_HAS_DIGIT_RE = re.compile(r"\d")
_TAIL_COMPREHENSIVE_RE = re.compile(r"^(.+)综合(规格)?$")
_TRANSPARENT_PUNCT = frozenset("()[]{}<>（）【】《》「」")

_PAT_COMPRESS = re.compile(r" {2,}")
_PAT_GREEK_DIGIT = re.compile(rf"([{_GREEK}])\s+(\d)")


def split_han_alnum(name: str) -> str:
    """按 token 级别处理（与 v0.1 baseline 行为完全一致）。"""
    if not name:
        return name

    # 规则 4：末尾"综合" / "综合规格" 前补空格
    m = _TAIL_COMPREHENSIVE_RE.match(name)
    if m:
        suffix = " 综合规格" if m.group(2) else " 综合"
        name = m.group(1) + suffix

    # 规则 1/2/3/5：token 级别处理
    tokens = _TOKEN_RE.findall(name)
    if not tokens:
        return name

    out: list[str] = []
    prev_type: str | None = None
    prev_alnum_is_digit: bool = False

    for token in tokens:
        if re.fullmatch(rf"{_HAN}+", token):
            cur_type = "han"
            cur_alnum_is_digit = False
        elif re.fullmatch(rf"[A-Za-z0-9{_GREEK}]+", token):
            cur_type = "alnum"
            cur_alnum_is_digit = bool(_HAS_DIGIT_RE.search(token))
        elif token in _TRANSPARENT_PUNCT or all(c in _TRANSPARENT_PUNCT for c in token):
            out.append(token)
            prev_type = "transparent"
            continue
        else:
            out.append(token)
            prev_type = None
            continue

        need_space = False
        if prev_type == "han" and cur_type == "alnum":
            need_space = True
        elif prev_type == "alnum" and cur_type == "han":
            if not prev_alnum_is_digit:
                need_space = True

        if need_space:
            out.append(" ")
        out.append(token)

        prev_type = cur_type
        if cur_type == "alnum":
            prev_alnum_is_digit = cur_alnum_is_digit

    s = "".join(out)
    s = _PAT_COMPRESS.sub(" ", s).strip()
    # post-processing 1：删除"数字 + 空格 + 汉字"中的空格
    s = re.sub(rf"(\d)\s+({_HAN})", r"\1\2", s)
    # post-processing 2：删除"希腊字母 + 空格 + 数字"中的空格
    s = _PAT_GREEK_DIGIT.sub(r"\1\2", s)
    return s


def _read_sheet_rows(ws) -> list[list]:
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _write_sheet_rows(ws, rows: list[list]) -> None:
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


# ─────────────────────────────────────────────────────────
# v0.2 函数级入口
# ─────────────────────────────────────────────────────────
def process_xlsx(input_path: Path, output_path: Path | None = None) -> dict:
    """对 料/配 行做 han↔alnum 边界空格规整（v0.2 函数入口；抛异常）。"""
    if not input_path.exists():
        raise FileNotFoundError(f"xlsx 不存在: {input_path}")

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"xlsx 中没有 {TARGET_SHEET!r} sheet；现有 sheet: {wb.sheetnames}"
        )

    ws = wb[TARGET_SHEET]
    rows = _read_sheet_rows(ws)

    total_rows = 0
    processed_rows = 0
    changed_rows = 0
    for row in rows:
        if not row:
            continue
        total_rows += 1
        if len(row) > 2 and row[0] in ("料", "配"):
            processed_rows += 1
            original = row[2]
            transformed = split_han_alnum(original)
            if transformed != original:
                row[2] = transformed
                changed_rows += 1

    _write_sheet_rows(ws, rows)
    target = output_path if output_path is not None else input_path
    wb.save(target)

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "sheet": TARGET_SHEET,
        "total_rows": total_rows,
        "processed_rows": processed_rows,
        "changed_rows": changed_rows,
        "other_sheets": [s for s in wb.sheetnames if s != TARGET_SHEET],
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(
            "用法: python space_split_materials.py <input.xlsx> [output.xlsx]",
            file=sys.stderr,
        )
        return 1
    input_path = Path(argv[1])
    output_path = Path(argv[2]) if len(argv) >= 3 else None
    info = process_xlsx(input_path, output_path)
    print(f"[OK] sheet: {info['sheet']}")
    print(f"[OK] 总行数: {info['total_rows']}")
    print(f"[OK] 处理行数（料/配）: {info['processed_rows']}")
    print(f"[OK] 实际改动行数: {info['changed_rows']}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))