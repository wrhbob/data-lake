"""
finalize_last_step.py — 给 v2 多 sheet xlsx 套上分组大纲 + 段行合并 + 加粗（第 5 步 / XLSX 版，v0.2 函数化版本）

原名 to_xlsx.py；v0.2 起改名 finalize_last_step.py（更准确反映它
"已是 xlsx，做最终格式化"的本质）。

输入：<stem>_待审核.xlsx（前 4 步 clean_empty_qty / drop_toc_sections /
      fill_work_content / space_split_materials 跑完后产出的 xlsx；
      含 "定额条目 / 册说明 / 章" sheet）

行为：
  1. 读「定额条目」sheet 的所有行
  2. 段行加粗 + 段行 C-D 列合并
  3. 自适应层数嵌套分组（v0.3 backport 自 to_xlsx.py）：
     - 段 group outline_level = sec_id 的点数 + 1（A→1, A.1→2, A.1.1→3, A.1.1.1→4 ...）
     - 定额 group outline_level = 父段 depth + 1
       （定额嵌套在所属段之下，把下属工料机折起来）
  4. 把已格式化的「定额条目」sheet 写回；其它 sheet（册说明 / 章）保持不动
  5. 默认原地覆盖写入（与同 pipeline 中其它 4 步一致）

v0.2 函数级入口（Worker 使用）:
    from external.quota_csv_finalize.finalize_last_step import process_xlsx
    info = process_xlsx(input_xlsx, output_xlsx=None)

CLI:
    python finalize_last_step.py <input.xlsx> [output.xlsx]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

TARGET_SHEET = "定额条目"


def get_section_depth(section_id: str) -> int:
    if not section_id:
        return 1
    return section_id.count(".") + 1


def find_end_row(sections: list[dict], idx: int, total_rows: int) -> int:
    sec_depth = sections[idx]["depth"]
    for j in range(idx + 1, len(sections)):
        if sections[j]["depth"] <= sec_depth:
            return sections[j]["row"] - 1
    return total_rows


def _read_sheet_rows(ws) -> list[list]:
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


# ─────────────────────────────────────────────────────────
# v0.2 函数级入口
# ─────────────────────────────────────────────────────────
def process_xlsx(input_path: Path, output_path: Path | None = None) -> dict:
    """给 input xlsx 的「定额条目」sheet 套格式（v0.2 函数入口；抛异常）。"""
    if not input_path.exists():
        raise FileNotFoundError(f"xlsx 不存在: {input_path}")

    target = output_path if output_path is not None else input_path

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"xlsx 中没有 {TARGET_SHEET!r} sheet；现有 sheet: {wb.sheetnames}"
        )

    ws = wb[TARGET_SHEET]
    rows = _read_sheet_rows(ws)
    if not rows:
        wb.save(target)
        return {
            "input_path": str(input_path),
            "output_path": str(target),
            "sheet": TARGET_SHEET,
            "total_rows": 0,
            "sections": 0,
            "quotas": 0,
            "groups": 0,
            "other_sheets": [s for s in wb.sheetnames if s != TARGET_SHEET],
            "warning": "empty sheet",
        }

    ws.sheet_properties.outlinePr.summaryBelow = False
    total_rows = len(rows)

    # 0. 先解除所有已存在的合并区域,避免循环写 cell 时碰到 MergedCell 只读
    #    (前 4 步 finalize 可能合并过 C:D 列等;后面会在第 2 步按"段行 C-D 合并"重新合并)
    #    v0.3 backport 自 to_xlsx.py L100-104
    pre_existing_merges = list(ws.merged_cells.ranges)
    for rng in pre_existing_merges:
        ws.unmerge_cells(str(rng))

    # 1. 数据写回 + 段行加粗
    bold_font = Font(bold=True)
    for r_idx, row in enumerate(rows, start=1):
        is_section = (str(row[0]).strip() if row else "") == "段"
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if is_section:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 2. 段行 C-D 合并
    for r_idx, row in enumerate(rows, start=1):
        if row and str(row[0]).strip() == "段":
            ws.merge_cells(
                start_row=r_idx, start_column=3, end_row=r_idx, end_column=4
            )

    # 3. 收集段与定额
    sections: list[dict] = []
    quotas: list[dict] = []

    for r_idx, row in enumerate(rows, start=1):
        if not row:
            continue
        row_type = str(row[0]).strip()
        if row_type == "段":
            sec_id = str(row[1]).strip() if len(row) > 1 else ""
            sections.append({
                "row": r_idx,
                "depth": get_section_depth(sec_id),
                "id": sec_id,
            })
        elif row_type == "定":
            quotas.append({"row": r_idx})

    # 4. 段结束位置
    for i in range(len(sections)):
        sections[i]["end_row"] = find_end_row(sections, i, total_rows)

    # 5. 定额结束位置 + 父段关联
    #    v0.3 backport 自 to_xlsx.py: 同时记录 parent_sec, 让定额 group
    #    能用 parent_sec.depth + 1 (自适应) 而不是硬编码 4
    for i, q in enumerate(quotas):
        if i + 1 < len(quotas):
            q["end_row"] = quotas[i + 1]["row"] - 1
        else:
            q["end_row"] = total_rows
        parent_sec = None
        for sec in reversed(sections):
            if sec["row"] <= q["row"]:
                parent_sec = sec
                break
        if parent_sec is not None:
            q["parent_row"] = parent_sec["row"]
            q["parent_depth"] = parent_sec["depth"]
            q["end_row"] = min(q["end_row"], parent_sec["end_row"])
        else:
            # 没找到父段,fallback: 当成 level=1 (与 to_xlsx.py 一致)
            q["parent_row"] = None
            q["parent_depth"] = 0

    # 6. 分组区间
    #    段 group level = sec.depth
    #    定额 group level = parent_sec.depth + 1 (自适应, 嵌套在父段之下)
    groups: list[tuple[int, int, int]] = []

    for sec in sections:
        start = sec["row"] + 1
        end = sec["end_row"]
        if start <= end:
            groups.append((start, end, sec["depth"]))

    for q in quotas:
        start = q["row"] + 1
        end = q["end_row"]
        if start <= end:
            groups.append((start, end, q["parent_depth"] + 1))

    # 低 level 先（Excel 嵌套语义）
    groups.sort(key=lambda x: x[2])
    for start, end, level in groups:
        ws.row_dimensions.group(start, end, outline_level=level)

    wb.save(target)

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "sheet": TARGET_SHEET,
        "total_rows": total_rows,
        "sections": len(sections),
        "quotas": len(quotas),
        "groups": len(groups),
        "other_sheets": [s for s in wb.sheetnames if s != TARGET_SHEET],
    }


def main():
    ap = argparse.ArgumentParser(
        description="XLSX → 带分组大纲 + 段行合并的 Excel（quota-csv-finalize 第 5 步）"
    )
    ap.add_argument("input_xlsx", help="输入 xlsx 路径（典型为 <stem>_待审核.xlsx）")
    ap.add_argument(
        "output_xlsx",
        nargs="?",
        default=None,
        help="输出 xlsx 路径（默认 = 原地覆盖输入文件）",
    )
    args = ap.parse_args()

    input_path = Path(args.input_xlsx).resolve()
    output_path = Path(args.output_xlsx).resolve() if args.output_xlsx else None
    info = process_xlsx(input_path, output_path)
    print(f"[OK] 输入: {info['input_path']}")
    print(f"[OK] 输出: {info['output_path']}")
    print(
        f"[OK] sheet: {info['sheet']} 总行数 {info['total_rows']}, "
        f"段数 {info['sections']}, 定额数 {info['quotas']}"
    )
    print(f"[OK] 分组区间数: {info['groups']}")
    print(
        f"[OK] 其它保留 sheet: {info['other_sheets']}"
    )


if __name__ == "__main__":
    main()