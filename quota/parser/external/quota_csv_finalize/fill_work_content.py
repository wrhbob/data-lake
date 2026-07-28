"""
fill_work_content.py — 定字段（第 4 列）工作内容前向传播（第 3 步 / XLSX 版，v0.2 函数化版本）

对 `定` 行（row[0] == "定"），如果其第 4 列（项目特征 / 工作内容）为空，
从「最近一次有非空工作内容的定行」复制第 4 列内容。

v0.2 函数级入口（Worker 使用）:
    from external.quota_csv_finalize.fill_work_content import process_xlsx
    info = process_xlsx(input_xlsx, output_xlsx=None)

CLI:
    python fill_work_content.py <input.xlsx> [output.xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

TARGET_SHEET = "定额条目"


def _read_sheet_rows(ws) -> list[list]:
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _write_sheet_rows(ws, rows: list[list]) -> None:
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def _fill_rows(rows: list[list]) -> tuple[list[list], int, int, int]:
    last_wc = ""
    total_rows = 0
    quota_rows = 0
    quota_filled = 0

    for row in rows:
        if not row:
            continue
        total_rows += 1
        if row[0] == "定":
            quota_rows += 1
            current_wc = row[3] if len(row) > 3 else ""
            if current_wc.strip():
                last_wc = current_wc
            else:
                if last_wc:
                    while len(row) < 4:
                        row.append("")
                    row[3] = last_wc
                    quota_filled += 1

    return rows, total_rows, quota_rows, quota_filled


# ─────────────────────────────────────────────────────────
# v0.2 函数级入口
# ─────────────────────────────────────────────────────────
def process_xlsx(input_path: Path, output_path: Path | None = None) -> dict:
    """对 定行 工作内容列做前向传播（v0.2 函数入口；抛异常而非 sys.exit）。"""
    if not input_path.exists():
        raise FileNotFoundError(f"xlsx 不存在: {input_path}")

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"xlsx 中没有 {TARGET_SHEET!r} sheet；现有 sheet: {wb.sheetnames}"
        )

    ws = wb[TARGET_SHEET]
    rows = _read_sheet_rows(ws)
    filled, total_rows, quota_rows, quota_filled = _fill_rows(rows)

    _write_sheet_rows(ws, filled)
    target = output_path if output_path is not None else input_path
    wb.save(target)

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "sheet": TARGET_SHEET,
        "total_rows": total_rows,
        "quota_rows": quota_rows,
        "quota_filled": quota_filled,
        "other_sheets": [s for s in wb.sheetnames if s != TARGET_SHEET],
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("用法: python fill_work_content.py <input.xlsx> [output.xlsx]", file=sys.stderr)
        return 1
    input_path = Path(argv[1])
    output_path = Path(argv[2]) if len(argv) >= 3 else None
    info = process_xlsx(input_path, output_path)
    print(f"[OK] sheet: {info['sheet']}")
    print(f"[OK] 总行数: {info['total_rows']}")
    print(f"[OK] 定行数: {info['quota_rows']}")
    print(f"[OK] 工作内容填充行数: {info['quota_filled']}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))