"""
normalize_unit.py — 计量单位字面规范化(autofinalize 第 6 步,v0.1)

输入: <stem>_待审核.xlsx(前 5 步 finalize 跑完后产物的 xlsx;
      Sheet1="定额条目",9 列定宽,col 5 = "计量单位")

行为:
  1. 读「定额条目」sheet 所有行
  2. 仅对 col index 4(1-based 第 5 列 = 计量单位)按 8 条严格字面规则替换
  3. 默认原地覆盖写入(同 autofinalize 5 步中其他脚本的约定一致)
  4. 其他 sheet / 其他列一律不动

v0.1 函数级入口(Worker 使用):
    from external.quota_csv_finalize.normalize_unit import process_xlsx
    info = process_xlsx(input_xlsx, output_xlsx=None)

CLI:
    python normalize_unit.py <input.xlsx> [output.xlsx]

规则集(**仅这 8 条**,严禁添加 — 用户原话"严禁自己瞎添加"):
  1. <sup>2</sup>    -> 2
  2. <sup>3</sup>    -> 3
  3. $\\mathbf{m}^2$ -> m2
  4. $\\mathbf{m}^3$ -> m3
  5. m³             -> m3
  6. m²             -> m2
  7. m^3            -> m3
  8. m^2            -> m2

替换语义:
  - 用 str.replace() 字面替换;不用 re
  - 只替换严格匹配上的子串,不触动其他字符(10m³ → 10m3,5m² → 5m2,前后缀原样保留)
  - 8 条规则 LHS 互不重叠(分属 4 种格式:HTML `<sup>`、LaTeX `$\\mathbf{m}...$`、
    Unicode 上标、纯文本 m^2/m^3),所以规则执行顺序不影响结果;
    代码按用户给定 1→8 顺序跑,与规范一致
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

TARGET_SHEET = "定额条目"
UNIT_COL_INDEX = 5  # 1-based;col index 4 in 0-based = 「计量单位」

# 8 条规则常量(顺序固定,与用户约定一致)
# LHS 互不重叠 → 顺序不影响结果;按原序保留便于对照
UNIT_REPLACE_RULES: list[tuple[str, str]] = [
    ("<sup>2</sup>", "2"),
    ("<sup>3</sup>", "3"),
    ("$\\mathbf{m}^2$", "m2"),
    ("$\\mathbf{m}^3$", "m3"),
    ("m³", "m3"),
    ("m²", "m2"),
    ("m^3", "m3"),
    ("m^2", "m2"),
]


def _read_sheet_rows(ws) -> list[list]:
    """沿用 finalize_last_step.py 的实现:把 sheet 所有 cell 拍平成 str | "" 的二维列表。"""
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _normalize_cell(value: str) -> tuple[str, list[int]]:
    """对单个 cell value 跑 8 条规则。

    Returns:
        (new_value, hits_per_rule)
        - new_value: 替换后的字符串
        - hits_per_rule: 与 UNIT_REPLACE_RULES 同序同长的 list,每条规则命中次数
                        (即使 new_value 已被前面规则改了,后面的规则仍按原 LHS 字面匹配;
                         由于 8 条规则 LHS 互不重叠,后续规则在已替换后的字符串里也不会误命中)
    """
    if not value:
        return value, [0] * len(UNIT_REPLACE_RULES)
    new = value
    hits = [0] * len(UNIT_REPLACE_RULES)
    for i, (old, repl) in enumerate(UNIT_REPLACE_RULES):
        if old in new:
            c = new.count(old)
            new = new.replace(old, repl)
            hits[i] = c
    return new, hits


# ─────────────────────────────────────────────────────────
# v0.1 函数级入口
# ─────────────────────────────────────────────────────────
def process_xlsx(input_path: Path, output_path: Path | None = None) -> dict:
    """对 input xlsx 的「定额条目」sheet 第 5 列跑 8 条字面替换(原地覆盖默认)。

    Returns:
        {
            "input_path": str,
            "output_path": str,
            "sheet": "定额条目",
            "total_rows": int,
            "cells_changed": int,                 # 累计改了 N 个 cell
            "rules_applied": [
                {"from": "<sup>2</sup>", "to": "2", "count": N},
                ...
            ],                                   # 按 1→8 规则顺序
            "other_sheets": [str, ...],          # 保留未动的 sheet 名
        }

    Raises:
        FileNotFoundError: xlsx 不存在
        RuntimeError: xlsx 里没有「定额条目」sheet
    """
    if not input_path.exists():
        raise FileNotFoundError(f"xlsx 不存在: {input_path}")

    target = output_path if output_path is not None else input_path

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"xlsx 中没有 {TARGET_SHEET!r} sheet;现有 sheet: {wb.sheetnames}"
        )

    ws = wb[TARGET_SHEET]
    rows = _read_sheet_rows(ws)

    # 按规则索引累计 count(初始 0)
    rule_counts = [0] * len(UNIT_REPLACE_RULES)
    cells_changed = 0

    if rows:
        for r_idx, row in enumerate(rows, start=1):
            # col_index 越界防御(col 数不足 5 列的退化表直接跳过该行)
            if len(row) < UNIT_COL_INDEX:
                continue
            original = row[UNIT_COL_INDEX - 1]
            if not original:
                continue
            new_value, per_rule_hits = _normalize_cell(original)
            if sum(per_rule_hits) > 0:
                ws.cell(row=r_idx, column=UNIT_COL_INDEX, value=new_value)
                cells_changed += 1
                for i, h in enumerate(per_rule_hits):
                    rule_counts[i] += h

    wb.save(target)

    rules_applied = [
        {"from": old, "to": repl, "count": rule_counts[i]}
        for i, (old, repl) in enumerate(UNIT_REPLACE_RULES)
    ]

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "sheet": TARGET_SHEET,
        "total_rows": len(rows),
        "cells_changed": cells_changed,
        "rules_applied": rules_applied,
        "other_sheets": [s for s in wb.sheetnames if s != TARGET_SHEET],
    }


def main():
    ap = argparse.ArgumentParser(
        description="定额条目 sheet 第 5 列(计量单位)字面规范化 — autofinalize 第 6 步"
    )
    ap.add_argument("input_xlsx", help="输入 xlsx 路径(典型为 <stem>_待审核.xlsx)")
    ap.add_argument(
        "output_xlsx",
        nargs="?",
        default=None,
        help="输出 xlsx 路径(默认 = 原地覆盖输入文件)",
    )
    args = ap.parse_args()

    input_path = Path(args.input_xlsx).resolve()
    output_path = Path(args.output_xlsx).resolve() if args.output_xlsx else None
    info = process_xlsx(input_path, output_path)

    print(f"[OK] 输入: {info['input_path']}")
    print(f"[OK] 输出: {info['output_path']}")
    print(f"[OK] sheet: {info['sheet']} 总行数 {info['total_rows']}")
    print(f"[OK] 改了 {info['cells_changed']} 个 cell:")
    for r in info["rules_applied"]:
        marker = "*" if r["count"] else "."
        print(f"  [{marker}] {r['from']!r} -> {r['to']!r}  hit {r['count']}")
    print(f"[OK] 其它保留 sheet: {info['other_sheets']}")


if __name__ == "__main__":
    main()