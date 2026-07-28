"""
clean_empty_qty.py — 人工验证后的最终清理（第 1 步 / XLSX 版，v0.2 函数化版本）

输入：<stem>.xlsx（v2 extract_quota.py 产出的多 sheet xlsx，含"定额条目 / 册说明 / 章"）
      读取"定额条目" sheet，对其它 sheet（册说明 / 章）保持不动。

清理规则（仅作用于"定额条目" sheet）：
  1. 删除消耗量为空的 工/料/配/机/综 行
  2. 删除项目之间的全空分隔行

v0.2 函数级入口（Worker 使用）:
    from external.quota_csv_finalize.clean_empty_qty import process_xlsx
    info = process_xlsx(input_xlsx, output_xlsx=None)  # 默认原地覆盖

    返回 dict（含 total_rows_before/after、removed_blank、removed_qty_empty）。

CLI（开发调试）:
    python clean_empty_qty.py <input.xlsx> [output.xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

COST_TYPES = {"工", "料", "配", "机", "综"}
TARGET_SHEET = "定额条目"


def _is_blank_row(row: list) -> bool:
    return all((str(cell) if cell is not None else "").strip() == "" for cell in row)


def _clean_rows(rows: list[list]) -> tuple[list[list], int, int]:
    cleaned: list[list] = []
    removed_blank = 0
    removed_qty_empty = 0

    for row in rows:
        if _is_blank_row(row):
            removed_blank += 1
            continue
        if len(row) >= 6:
            row_type = (str(row[0]) if row[0] is not None else "").strip()
            qty = (str(row[5]) if row[5] is not None else "").strip()
            if row_type in COST_TYPES and not qty:
                removed_qty_empty += 1
                continue
        cleaned.append(row)

    return cleaned, removed_blank, removed_qty_empty


def _read_sheet_rows(ws) -> list[list]:
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _write_sheet_rows(ws, rows: list[list]) -> None:
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


# ─────────────────────────────────────────────────────────
# v0.2 函数级入口
# ─────────────────────────────────────────────────────────
def process_xlsx(input_path: Path, output_path: Path | None = None) -> dict:
    """清理空消耗量行 + 全空行（v0.2 函数入口；抛异常而非 sys.exit）。"""
    if not input_path.exists():
        raise FileNotFoundError(f"xlsx 不存在: {input_path}")

    target = output_path if output_path is not None else input_path
    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"xlsx 中没有 {TARGET_SHEET!r} sheet；现有 sheet: {wb.sheetnames}"
        )

    ws = wb[TARGET_SHEET]
    original_rows = _read_sheet_rows(ws)
    cleaned, removed_blank, removed_qty_empty = _clean_rows(original_rows)
    _write_sheet_rows(ws, cleaned)
    wb.save(target)

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "sheet": TARGET_SHEET,
        "total_rows_before": len(original_rows),
        "total_rows_after": len(cleaned),
        "removed_blank": removed_blank,
        "removed_qty_empty": removed_qty_empty,
        "other_sheets": [s for s in wb.sheetnames if s != TARGET_SHEET],
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: clean_empty_qty.py <input.xlsx> [output.xlsx]", file=sys.stderr)
        sys.exit(1)
    info = process_xlsx(
        Path(sys.argv[1]),
        Path(sys.argv[2]) if len(sys.argv) > 2 else None,
    )
    print(f"[OK] 输入: {info['input_path']}")
    print(f"[OK] 输出: {info['output_path']}")
    print(
        f"[OK] sheet: {info['sheet']} 原 {info['total_rows_before']} → "
        f"删空行 {info['removed_blank']}, 删空消耗量 {info['removed_qty_empty']}, "
        f"剩 {info['total_rows_after']}"
    )


if __name__ == "__main__":
    main()