"""
drop_toc_sections.py — 删除被 province 抽取器误捕的「目录行」（第 2 步 / XLSX 版，v0.2 函数化版本）

输入：<stem>.xlsx（v2 extract_quota.py / finalize 流水线前几步的产出）
      仅修改"定额条目" sheet 的数据，其它 sheet（册说明 / 章）保持不动。

判定：
  段行（row[0] == "段"）的第 3 列（名称）匹配 TOC 模式：
      `[….]{2,}\s*[\(（]\d+[\)）]`

并且为了与"目录行没有挂子项"的现场观察匹配：
  - 仅当目标段行下方没有任何 定/工/料/配/机/综 行（直到下一个同级或更高级段行）
    才真正删除
  - 若有子项（理论上不该发生），保留该行 + 打 stderr 警告

v0.2 函数级入口（Worker 使用）:
    from external.quota_csv_finalize.drop_toc_sections import process_xlsx
    info = process_xlsx(input_xlsx, output_xlsx=None)

CLI:
    python drop_toc_sections.py <input.xlsx> [output.xlsx]
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

TARGET_SHEET = "定额条目"
COST_TYPES = {"工", "料", "配", "机", "综"}

_TOC_PATTERN = re.compile(r"[….]{2,}\s*[\(（]\d+[\)）]")


def _is_toc_section_name(name: str) -> bool:
    return bool(_TOC_PATTERN.search(name or ""))


def _get_section_depth(sec_id: str) -> int:
    if not sec_id:
        return 1
    return sec_id.count(".") + 1


def _read_sheet_rows(ws) -> list[list]:
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _write_sheet_rows(ws, rows: list[list]) -> None:
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def _scan_toc_rows(rows: list[list]) -> tuple[set[int], int]:
    N = len(rows)
    toc_indices: set[int] = set()
    warnings = 0

    for i, row in enumerate(rows):
        if not row:
            continue
        if str(row[0] or "").strip() != "段":
            continue
        sec_id = str(row[1] or "").strip()
        sec_name = str(row[2] or "")
        if not _is_toc_section_name(sec_name):
            continue

        cur_depth = _get_section_depth(sec_id)
        has_child = False
        for j in range(i + 1, N):
            nxt = rows[j]
            if not nxt:
                continue
            rt = str(nxt[0] or "").strip()
            if rt == "段":
                next_id = str(nxt[1] or "").strip()
                if _get_section_depth(next_id) <= cur_depth:
                    break
            if rt in COST_TYPES or rt == "定":
                has_child = True
                break

        if has_child:
            print(
                f"[WARN] row{i+1} id={sec_id!r} name={sec_name[:50]!r} "
                f"匹配目录模式但有子项 — 跳过删除",
                file=sys.stderr,
            )
            warnings += 1
        else:
            toc_indices.add(i)

    return toc_indices, warnings


def _drop_rows(rows: list[list], drop_idxs: set[int]) -> list[list]:
    return [row for i, row in enumerate(rows) if i not in drop_idxs]


# ─────────────────────────────────────────────────────────
# v0.2 函数级入口
# ─────────────────────────────────────────────────────────
def process_xlsx(input_path: Path, output_path: Path | None = None) -> dict:
    """删除被误捕的目录段行（v0.2 函数入口；抛异常而非 sys.exit）。"""
    if not input_path.exists():
        raise FileNotFoundError(f"xlsx 不存在: {input_path}")

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"xlsx 中没有 {TARGET_SHEET!r} sheet；现有 sheet: {wb.sheetnames}"
        )

    ws = wb[TARGET_SHEET]
    rows = _read_sheet_rows(ws)
    toc_idxs, warning_count = _scan_toc_rows(rows)
    cleaned = _drop_rows(rows, toc_idxs)

    target = output_path if output_path is not None else input_path
    _write_sheet_rows(ws, cleaned)
    wb.save(target)

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "sheet": TARGET_SHEET,
        "total_rows_before": len(rows),
        "total_rows_after": len(cleaned),
        "removed": len(toc_idxs),
        "warnings": warning_count,
        "other_sheets": [s for s in wb.sheetnames if s != TARGET_SHEET],
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(
            "用法: python drop_toc_sections.py <input.xlsx> [output.xlsx]",
            file=sys.stderr,
        )
        return 1
    input_path = Path(argv[1])
    output_path = Path(argv[2]) if len(argv) >= 3 else None
    info = process_xlsx(input_path, output_path)
    print(f"[OK] sheet: {info['sheet']}")
    print(f"[OK] 删目录行: {info['removed']} 行")
    print(f"[OK] 总行数: {info['total_rows_before']} → {info['total_rows_after']}")
    print(f"[OK] 警告: {info['warnings']} 条")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))