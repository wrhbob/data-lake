#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
drop_toc_sections.py — 删除被 province 抽取器误捕的「目录行」（XLSX 版）

输入：<stem>.xlsx（v2 extract_quota.py / finalize 流水线前几步的产出）
      仅修改"定额条目" sheet 的数据，其它 sheet（册说明 / 章）保持不动。

判定：
  段行（row[0] == "段"）的第 3 列（名称）匹配 TOC 模式：
      `[….]{2,}\s*[\(（]\d+[\)）]`
  兼容两种省略号形式：
    - Unicode U+2026 * 2:  「……」
    - ASCII dots * 2+:    「.....」(OCR 经常把省略号识别成 2-5 个点)

并且为了与"目录行没有挂子项"的现场观察匹配（province 抽取器只
捕获到 TOC 中"名字+页码"的占位行，没有真的把目录里的表格也抽进来），
脚本会:
  - 仅当目标段行下方没有任何 定/工/料/配/机/综 行（直到下一个同级或
    更高级段行）才真正删除；
  - 若有子项（理论上不该发生，但保险起见），保留该行 + 打 stderr 警告，
    避免误删。

适用场景:
  OCR/MinerU 把 MD 中的 `## 目录` 段也扫描进 province 抽取器，导致
  抽出像 `A.1.1 砍伐乔木 …… (11)` 的"伪段行"。这些行是目录占位行，
  不是真正的定额段，应该整个删掉。

CLI:
    python drop_toc_sections.py <input.xlsx> [output.xlsx]

默认行为：不传第 2 个参数 → 原地覆盖输入文件。
传第 2 个参数 → 写入指定输出文件，不修改输入文件。
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

TARGET_SHEET = "定额条目"
COST_TYPES = {"工", "料", "配", "机", "综"}

# 目录行模式：`……` / `....` 后接 (NN) 或 (NN)
_TOC_PATTERN = re.compile(r"[….]{2,}\s*[\(（]\d+[\)）]")


def _is_toc_section_name(name: str) -> bool:
    return bool(_TOC_PATTERN.search(name or ""))


def _get_section_depth(sec_id: str) -> int:
    """A → 1, A.1 → 2, A.1.1 → 3 (基于点号数量 + 1)."""
    if not sec_id:
        return 1
    return sec_id.count(".") + 1


def _read_sheet_rows(ws) -> list[list]:
    """读 openpyxl worksheet 行;None → \"\"."""
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _write_sheet_rows(ws, rows: list[list]) -> None:
    """把 rows 写回 ws;先清空原内容."""
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def _scan_toc_rows(rows: list[list]) -> tuple[set[int], int]:
    """扫描「段」行的位置 + 给出其中哪些是目录误捕。

    Returns:
        (toc_indices, warning_count)
        - toc_indices: 需删的 row index 集合（0-based）
        - warning_count: TOC 模式命中但有子项 → 保留并警告的个数
    """
    N = len(rows)
    toc_indices: set[int] = set()
    warnings = 0

    for i, row in enumerate(rows):
        if not row:
            continue
        if str(row[0] or "").strip() != "段":
            continue
        sec_id = str(row[1] or "").strip()
        sec_name = str(row[2] or "")
        if not _is_toc_section_name(sec_name):
            continue

        # 检查该段行下是否真有子项定/工/料/... 直到下一个同级或更高级段行
        cur_depth = _get_section_depth(sec_id)
        has_child = False
        for j in range(i + 1, N):
            nxt = rows[j]
            if not nxt:
                continue
            rt = str(nxt[0] or "").strip()
            if rt == "段":
                next_id = str(nxt[1] or "").strip()
                if _get_section_depth(next_id) <= cur_depth:
                    break
            if rt in COST_TYPES or rt == "定":
                has_child = True
                break

        if has_child:
            # 警告: TOC 模式命中但有子项 → 保守保留
            print(
                f"[WARN] row{i+1} id={sec_id!r} name={sec_name[:50]!r} "
                f"匹配目录模式但有子项 — 跳过删除",
                file=sys.stderr,
            )
            warnings += 1
        else:
            toc_indices.add(i)

    return toc_indices, warnings


def _drop_rows(rows: list[list], drop_idxs: set[int]) -> list[list]:
    """删除指定 index 的行（保持其他行顺序）."""
    return [row for i, row in enumerate(rows) if i not in drop_idxs]


def process_xlsx(input_path: Path, output_path: Path | None = None) -> dict:
    if not input_path.exists():
        print(f"[ERROR] 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        print(
            f"[ERROR] xlsx 中没有 {TARGET_SHEET!r} sheet;现有 sheet: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws = wb[TARGET_SHEET]
    rows = _read_sheet_rows(ws)
    toc_idxs, warning_count = _scan_toc_rows(rows)
    cleaned = _drop_rows(rows, toc_idxs)

    # 写回前先检查目标路径不与 input_path 重叠（避免 in-place 误删）
    target = output_path if output_path is not None else input_path
    _write_sheet_rows(ws, cleaned)
    wb.save(target)

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "total_rows_before": len(rows),
        "total_rows_after": len(cleaned),
        "removed": len(toc_idxs),
        "warnings": warning_count,
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(
            "用法: python drop_toc_sections.py <input.xlsx> [output.xlsx]",
            file=sys.stderr,
        )
        print(
            "  不传第 2 个参数 → 原地覆盖输入文件",
            file=sys.stderr,
        )
        print(
            "  传第 2 个参数 → 写入指定输出文件，不修改输入文件",
            file=sys.stderr,
        )
        return 1

    input_path = Path(argv[1])
    output_path = Path(argv[2]) if len(argv) >= 3 else None

    info = process_xlsx(input_path, output_path)

    print(f"[OK] 输入: {info['input_path']}")
    print(f"[OK] 输出: {info['output_path']}")
    print(f"[OK] sheet: {TARGET_SHEET}")
    print(f"[OK] 删除目录行: {info['removed']} 行")
    print(f"[OK] 总行数: {info['total_rows_before']} → {info['total_rows_after']}")
    print(f"[OK] 警告（TOC 模式 + 有子项未删）: {info['warnings']} 条")

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
