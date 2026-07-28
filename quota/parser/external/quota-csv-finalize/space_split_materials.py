#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""space_split_materials.py — 在材料名称字段的"汉字 ↔ 字母/数字"边界插入空格（XLSX 版）。

输入：<stem>.xlsx（v2 extract_quota.py 产出）
      仅修改"定额条目" sheet 的数据，其它 sheet（册说明 / 章）保持不动。

仅处理行类型为 `料` 或 `配` 的行的第 3 列（材料名称字段）。
其他行类型（段/定/工/机/综）的名称字段不动，其他 9 列全部不动。

CLI:
    python space_split_materials.py <input.xlsx> [output.xlsx]

默认行为：不传第 2 个参数 → 原地覆盖输入文件。
传第 2 个参数 → 写入指定输出文件，不修改输入文件。

规则：
  1. 汉字 → 字母/数字：在字母数字前加空格
  2. 字母 → 汉字：在字母后加空格
  3. 数字 → 汉字：不加空格（且会删除现有空格，避免重复）
     "数字"指包含至少一个 0-9 字符的串（如 42.5、E43、M20×100）。
     "字母"指纯 A-Za-z 不含数字（如 EVA、PVC、SBS、PYII）。
  4. "综合" / "综合规格" 在末尾且前面有文字时，在其前加空格。
  5. 希腊字母（δ/Φ/α 等）按混合规则：
     - 汉字 → 希腊字母：加空格（同英文字母）
     - 希腊字母 → 汉字：加空格（同英文字母）
     - 希腊字母 → 数字：删除已有空格（与"数字→汉字"对称）
     范围：U+0370-U+03FF（基本希腊字母）+ U+1F00-U+1FFF（希腊字母扩展）。
     希腊字母后接数字的处理**不影响**英文字母（EVA 100 → EVA 100 保留空格）。
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

TARGET_SHEET = "定额条目"

_HAN = r"[一-鿿]"
# 希腊字母 Unicode 范围：U+0370-U+03FF（基本）+ U+1F00-U+1FFF（扩展）
_GREEK = r"Ͱ-Ͽἀ-῿"

# token 切分：汉字串 / 字母数字+希腊字母串 / 其他（标点/空白等）
# 注意：第三个分支直接展开 [一-鿿] 而非用 f-string 嵌套，避免字符类嵌套导致 `[` 被当作字面字符
_TOKEN_RE = re.compile(
    rf"[一-鿿]+|[A-Za-z0-9{_GREEK}]+|[^一-鿿A-Za-z0-9{_GREEK}]+"
)
# 判定 token 是否含数字
_HAS_DIGIT_RE = re.compile(r"\d")
# "综合" / "综合规格" 在末尾且前面有非空文字（group(2) 标记是否带"规格"）
_TAIL_COMPREHENSIVE_RE = re.compile(r"^(.+)综合(规格)?$")
# "透明"标点：作为 han/alnum 边界的一部分传递，不触发 han↔alnum 边界判定
# 例：`塑性聚烯烃(TPO)防水卷材P型` 中 `(`/`)` 是透明的，TPO 紧贴括号，不加空格
_TRANSPARENT_PUNCT = frozenset("()[]{}<>（）【】《》「」")

# 归一化：压缩连续空格为 1 个
_PAT_COMPRESS = re.compile(r" {2,}")
# post-processing：删除"希腊字母 + 空格 + 数字"中的空格
# 例：`δ 100` → `δ100`；`Φ 300` → `Φ300`
# 注意：仅匹配希腊字母，不影响英文字母（EVA 100 保留空格）
_PAT_GREEK_DIGIT = re.compile(rf"([{_GREEK}])\s+(\d)")


def split_han_alnum(name: str) -> str:
    """按 token 级别处理：汉字/字母/数字边界智能插入/删除空格；末尾"综合"前补空格。

    >>> split_han_alnum("CSPE嵌缝油膏330mL")
    'CSPE 嵌缝油膏 330mL'
    >>> split_han_alnum("改性沥青化学耐根穿刺防水卷材PYII 4mm")
    '改性沥青化学耐根穿刺防水卷材 PYII 4mm'
    >>> split_han_alnum("PVC聚氯乙烯粘合剂")
    'PVC 聚氯乙烯粘合剂'
    >>> split_han_alnum("塑性聚烯烃(TPO)防水卷材P型")
    '塑性聚烯烃(TPO)防水卷材 P 型'
    >>> split_han_alnum("EVA高分子防水卷材1.5mm")
    'EVA 高分子防水卷材 1.5mm'
    >>> split_han_alnum("EM金属屋面高弹耐候防水涂料")
    'EM 金属屋面高弹耐候防水涂料'
    >>> split_han_alnum("丙烯酸盐PM-I-A")
    '丙烯酸盐 PM-I-A'
    >>> split_han_alnum("水泥42.5")
    '水泥 42.5'
    >>> split_han_alnum("冷底子油30:70")
    '冷底子油 30:70'
    >>> split_han_alnum("商品混凝土C20")
    '商品混凝土 C20'
    >>> split_han_alnum("石油沥青30#")
    '石油沥青 30#'
    >>> split_han_alnum("低合金钢焊条E43系列")
    '低合金钢焊条 E43系列'
    >>> split_han_alnum("镀锌六角螺栓带螺母2平垫1弹垫M20×100以内")
    '镀锌六角螺栓带螺母 2平垫 1弹垫 M20×100以内'
    >>> split_han_alnum("螺栓综合")
    '螺栓 综合'
    >>> split_han_alnum("焊条综合")
    '焊条 综合'
    >>> split_han_alnum("综合管廊")
    '综合管廊'
    >>> split_han_alnum("综合")
    '综合'
    >>> split_han_alnum("合金钢钻头综合")
    '合金钢钻头 综合'
    >>> split_han_alnum("螺栓综合规格")
    '螺栓 综合规格'
    >>> split_han_alnum("镀锌钢板综合规格")
    '镀锌钢板 综合规格'
    >>> split_han_alnum("综合规格")
    '综合规格'
    >>> split_han_alnum("综合规格管廊")
    '综合规格管廊'
    >>> split_han_alnum("预制轻钢龙骨内隔墙板δ 100")
    '预制轻钢龙骨内隔墙板 δ100'
    >>> split_han_alnum("合金钢切割片φ 300")
    '合金钢切割片 φ300'
    >>> split_han_alnum("隔震支座φ 400")
    '隔震支座 φ400'
    >>> split_han_alnum("板Φ")
    '板 Φ'
    >>> split_han_alnum("Φ板")
    'Φ 板'
    >>> split_han_alnum("Φ100板")
    'Φ100板'
    >>> split_han_alnum("EVA 100")
    'EVA 100'
    >>> split_han_alnum("电焊条(Φ10)")
    '电焊条(Φ10)'
    """
    if not name:
        return name

    # 规则 4：末尾"综合" / "综合规格" 前补空格（仅当在末尾，且前面有非空字符时）
    m = _TAIL_COMPREHENSIVE_RE.match(name)
    if m:
        suffix = " 综合规格" if m.group(2) else " 综合"
        name = m.group(1) + suffix

    # 规则 1/2/3/5：token 级别处理
    tokens = _TOKEN_RE.findall(name)
    if not tokens:
        return name

    out: list[str] = []
    prev_type: str | None = None  # 上一个有意义的 token 类型
    prev_alnum_is_digit: bool = False  # 上一次 alnum token 是否含数字

    for token in tokens:
        if re.fullmatch(rf"{_HAN}+", token):
            cur_type = "han"
            cur_alnum_is_digit = False
        elif re.fullmatch(rf"[A-Za-z0-9{_GREEK}]+", token):
            cur_type = "alnum"
            cur_alnum_is_digit = bool(_HAS_DIGIT_RE.search(token))
        elif token in _TRANSPARENT_PUNCT or all(c in _TRANSPARENT_PUNCT for c in token):
            # 透明标点（括号类）：原样输出，prev_type 设为 "transparent"，
            # 这样后续 han/alnum 不视为 han→alnum 边界（即紧贴括号，不加空格）
            out.append(token)
            prev_type = "transparent"
            continue
        else:
            # 其他字符（空白、×、#、:、- 等）：原样输出，重置 prev_type 为 None，
            # 让后续 han/alnum 视为新的开始（不加空格）
            out.append(token)
            prev_type = None
            continue

        need_space = False
        if prev_type == "han" and cur_type == "alnum":
            # 汉字 → 字母/数字/希腊字母：加空格
            need_space = True
        elif prev_type == "alnum" and cur_type == "han":
            # 字母/数字 → 汉字
            if not prev_alnum_is_digit:
                # 字母 → 汉字：加空格
                need_space = True
            # 数字 → 汉字：不加空格

        if need_space:
            out.append(" ")
        out.append(token)

        prev_type = cur_type
        if cur_type == "alnum":
            prev_alnum_is_digit = cur_alnum_is_digit

    s = "".join(out)
    # 压缩连续空格 + 去首尾空白
    s = _PAT_COMPRESS.sub(" ", s).strip()
    # post-processing 1：删除"数字 + 空格 + 汉字"中的空格（处理已存在的空格）
    # 例：`E43 系列` → `E43系列`；`2 平垫` → `2平垫`
    # 注意：仅删除数字后的空格，字母后的空格（EVA 高分子）保留
    s = re.sub(rf"(\d)\s+({_HAN})", r"\1\2", s)
    # post-processing 2：删除"希腊字母 + 空格 + 数字"中的空格（规则 5 续）
    # 例：`δ 100` → `δ100`；`Φ 300` → `Φ300`
    # 注意：仅匹配希腊字母，不影响英文字母（EVA 100 保留空格）
    s = _PAT_GREEK_DIGIT.sub(r"\1\2", s)
    return s


def _read_sheet_rows(ws) -> list[list]:
    """读 openpyxl worksheet 行;None → ""。"""
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _write_sheet_rows(ws, rows: list[list]) -> None:
    """把 rows 写回 ws;先清空原内容。"""
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def _split_rows(rows: list[list]) -> tuple[list[list], int, int, int]:
    """对内存中的行做料/配行的第 3 列空格插入；返回统计。"""
    total_rows = 0
    processed_rows = 0
    changed_rows = 0

    for row in rows:
        if not row:
            continue
        total_rows += 1
        # 仅处理 料 / 配 行的第 3 列（材料名称字段）
        if len(row) > 2 and row[0] in ("料", "配"):
            processed_rows += 1
            original = row[2]
            transformed = split_han_alnum(original)
            if transformed != original:
                row[2] = transformed
                changed_rows += 1

    return rows, total_rows, processed_rows, changed_rows


def process_xlsx(input_path: Path, output_path: Path | None = None) -> dict:
    if not input_path.exists():
        print(f"[ERROR] 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        print(
            f"[ERROR] xlsx 中没有 {TARGET_SHEET!r} sheet；现有 sheet: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws = wb[TARGET_SHEET]
    rows = _read_sheet_rows(ws)
    processed, total_rows, processed_rows, changed_rows = _split_rows(rows)

    _write_sheet_rows(ws, processed)

    target = output_path if output_path is not None else input_path
    wb.save(target)

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "total_rows": total_rows,
        "processed_rows": processed_rows,
        "changed_rows": changed_rows,
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("用法: python space_split_materials.py <input.xlsx> [output.xlsx]", file=sys.stderr)
        print("  不传第 2 个参数 → 原地覆盖输入文件", file=sys.stderr)
        print("  传第 2 个参数 → 写入指定输出文件，不修改输入文件", file=sys.stderr)
        return 1

    input_path = Path(argv[1])
    output_path = Path(argv[2]) if len(argv) >= 3 else None

    info = process_xlsx(input_path, output_path)

    print(f"[OK] 输入: {info['input_path']}")
    print(f"[OK] 输出: {info['output_path']}")
    print(f"[OK] sheet: {TARGET_SHEET}")
    print(f"[OK] 总行数: {info['total_rows']}")
    print(f"[OK] 处理行数（料/配）: {info['processed_rows']}")
    print(f"[OK] 实际改动行数: {info['changed_rows']}")

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))