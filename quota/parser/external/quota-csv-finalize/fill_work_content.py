#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_work_content.py — 定字段（第 4 列）工作内容前向传播（XLSX 版）

输入：<stem>.xlsx（v2 extract_quota.py 产出）
      仅修改"定额条目" sheet 的数据，其它 sheet（册说明 / 章）保持不动。

对 `定` 行（row[0] == "定"），如果其第 4 列（项目特征 / 工作内容）为空，
从「最近一次有非空工作内容的定行」复制第 4 列内容。

适用场景：OCR 输出的 md 中部分定额表丢失了「工作内容」段落（如表格前的工作内容
文本跨页被切到下一页），但同章节下的定额往往工作内容相同或相近，前向传播是
合理的兜底。

CLI:
    python fill_work_content.py <input.xlsx> [output.xlsx]

默认行为：不传第 2 个参数 → 原地覆盖输入文件。
传第 2 个参数 → 写入指定输出文件，不修改输入文件。

规则：
  1. 扫描「定额条目」sheet（按文档顺序），维护变量 `last_wc` = 最近一个"有非空
     工作内容的定行"的第 4 列内容。
  2. 遇到 `定` 行：
     - 若第 4 列非空 → 更新 `last_wc` 为当前第 4 列。
     - 若第 4 列为空（`strip() == ""`）且 `last_wc` 非空 → 用 `last_wc` 填入。
  3. 跨段 / 工 / 料 / 配 / 机 / 综 行：`last_wc` 状态**不重置**——同章节下的
     后续定额仍可继承前一个有内容的定行。
  4. 第一个 `定` 行就空：没有"上一个"，保持空（不报错）。
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

TARGET_SHEET = "定额条目"


def _read_sheet_rows(ws) -> list[list]:
    """读 openpyxl worksheet 行;None → ""。"""
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _write_sheet_rows(ws, rows: list[list]) -> None:
    """把 rows 写回 ws;先清空原内容。"""
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def _fill_rows(rows: list[list]) -> tuple[list[list], int, int, int]:
    """对内存中的行做前向传播；返回 (filled_rows, total_rows, quota_rows, quota_filled)。"""
    last_wc = ""
    total_rows = 0
    quota_rows = 0
    quota_filled = 0

    for row in rows:
        if not row:
            continue
        total_rows += 1
        if row[0] == "定":
            quota_rows += 1
            current_wc = row[3] if len(row) > 3 else ""
            if current_wc.strip():
                last_wc = current_wc
            else:
                if last_wc:
                    # 补足到 4 列，避免索引错误
                    while len(row) < 4:
                        row.append("")
                    row[3] = last_wc
                    quota_filled += 1

    return rows, total_rows, quota_rows, quota_filled


def process_xlsx(input_path: Path, output_path: Path | None = None) -> dict:
    """处理 xlsx：定行工作内容前向传播。

    Args:
        input_path: 输入 xlsx 路径。
        output_path: 输出 xlsx 路径。None 表示原地覆盖 input_path。

    Returns:
        统计字典。
    """
    if not input_path.exists():
        print(f"[ERROR] 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        print(
            f"[ERROR] xlsx 中没有 {TARGET_SHEET!r} sheet；现有 sheet: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws = wb[TARGET_SHEET]
    rows = _read_sheet_rows(ws)
    filled, total_rows, quota_rows, quota_filled = _fill_rows(rows)

    _write_sheet_rows(ws, filled)

    target = output_path if output_path is not None else input_path
    wb.save(target)

    return {
        "input_path": str(input_path),
        "output_path": str(target),
        "total_rows": total_rows,
        "quota_rows": quota_rows,
        "quota_filled": quota_filled,
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("用法: python fill_work_content.py <input.xlsx> [output.xlsx]", file=sys.stderr)
        print("  不传第 2 个参数 → 原地覆盖输入文件", file=sys.stderr)
        print("  传第 2 个参数 → 写入指定输出文件，不修改输入文件", file=sys.stderr)
        return 1

    input_path = Path(argv[1])
    output_path = Path(argv[2]) if len(argv) >= 3 else None

    info = process_xlsx(input_path, output_path)

    print(f"[OK] 输入: {info['input_path']}")
    print(f"[OK] 输出: {info['output_path']}")
    print(f"[OK] sheet: {TARGET_SHEET}")
    print(f"[OK] 总行数: {info['total_rows']}")
    print(f"[OK] 定行数: {info['quota_rows']}")
    print(f"[OK] 工作内容填充行数: {info['quota_filled']}")

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))