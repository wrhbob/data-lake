#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
quota-csv-finalize — 人工验证后的最终清理（XLSX 版）

输入：<stem>.xlsx（v2 extract_quota.py 产出的多 sheet xlsx，含"定额条目 / 册说明 / 章"）
      读取"定额条目" sheet，对其它 sheet（册说明 / 章）保持不动。

清理规则（仅作用于"定额条目" sheet）：
  1. 删除消耗量为空的 工/料/配/机/综 行
  2. 删除项目之间的全空分隔行

CLI:
    python clean_empty_qty.py <input.xlsx> [output.xlsx]

默认行为：不传第 2 个参数 → 原地覆盖输入文件（与其他 4 步同款约定）
传第 2 个参数 → 写入指定输出文件，不修改输入文件。
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

COST_TYPES = {"工", "料", "配", "机", "综"}
TARGET_SHEET = "定额条目"


def _is_blank_row(row: list) -> bool:
    return all((str(cell) if cell is not None else "").strip() == "" for cell in row)


def _clean_rows(rows: list[list]) -> tuple[list[list], int, int]:
    cleaned: list[list] = []
    removed_blank = 0
    removed_qty_empty = 0

    for row in rows:
        if _is_blank_row(row):
            removed_blank += 1
            continue
        # row 至少 6 列才能判断 row_type / qty
        if len(row) >= 6:
            row_type = (str(row[0]) if row[0] is not None else "").strip()
            qty = (str(row[5]) if row[5] is not None else "").strip()
            if row_type in COST_TYPES and not qty:
                removed_qty_empty += 1
                continue
        cleaned.append(row)

    return cleaned, removed_blank, removed_qty_empty


def _read_sheet_rows(ws) -> list[list]:
    """读 openpyxl worksheet 行;None → ""。"""
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _write_sheet_rows(ws, rows: list[list]) -> None:
    """把 rows 写回 ws;先清空原内容。"""
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def process_xlsx(input_path: Path, output_path: Path | None = None) -> Path:
    if output_path is None:
        # 默认原地覆盖（finalize 流水线约定：5 步共读写同一文件）
        output_path = input_path

    if not input_path.exists():
        print(f"[ERROR] 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        print(
            f"[ERROR] xlsx 中没有 {TARGET_SHEET!r} sheet；现有 sheet: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws = wb[TARGET_SHEET]
    original_rows = _read_sheet_rows(ws)
    cleaned, removed_blank, removed_qty_empty = _clean_rows(original_rows)

    # 把清理后的行写回「定额条目」sheet；其它 sheet（册说明 / 章）不动
    _write_sheet_rows(ws, cleaned)
    wb.save(output_path)

    print(f"[OK] 输入: {input_path}")
    print(f"[OK] 输出: {output_path}")
    print(
        f"[OK] sheet: {TARGET_SHEET} 原行数 {len(original_rows)}, "
        f"删空行 {removed_blank}, 删空消耗量 {removed_qty_empty}, "
        f"剩余 {len(cleaned)}"
    )
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: clean_empty_qty.py <input.xlsx> [output.xlsx]", file=sys.stderr)
        sys.exit(1)

    process_xlsx(Path(sys.argv[1]),
                 Path(sys.argv[2]) if len(sys.argv) > 2 else None)


if __name__ == "__main__":
    main()