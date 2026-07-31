#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
to_xlsx.py — 给 v2 多 sheet xlsx 套上分组大纲 + 段行合并 + 加粗（XLSX 版）

输入：<stem>_待审核.xlsx（前 4 步 clean_empty_qty / drop_toc_sections /
      fill_work_content / space_split_materials 跑完后产出的 xlsx；含
      "定额条目 / 册说明 / 章" sheet）

行为：
  1. 读「定额条目」sheet 的所有行
  2. 段行加粗 + 段行 C-D 列合并
  3. 自适应层数嵌套分组：
     - 段 group outline_level = sec_id 的点数 + 1
       (A→1, A.1→2, A.1.1→3, A.1.1.1→4 ...任意深度)
     - 定额 group outline_level = 父段 depth + 1
       (定额嵌套在所属段之下,把下属工料机折起来)
  4. 把已格式化的「定额条目」sheet 写回；其它 sheet（册说明 / 章）保持不动
  5. 默认原地覆盖写入（与同 pipeline 中其它 4 步一致）

CLI:
    python to_xlsx.py <input.xlsx> [output.xlsx]

不传第 2 个参数 → 原地覆盖输入文件。
传第 2 个参数 → 写入指定输出文件，不修改输入文件。

退出码：
  1  文件不存在 / sheet 缺失 / 异常
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

TARGET_SHEET = "定额条目"


def get_section_depth(section_id: str) -> int:
    """A→1, A.1→2, A.1.1→3."""
    if not section_id:
        return 1
    return section_id.count(".") + 1


def find_end_row(sections: list[dict], idx: int, total_rows: int) -> int:
    """找到当前段之后第一个同级/更高级段的位置。"""
    sec_depth = sections[idx]["depth"]
    for j in range(idx + 1, len(sections)):
        if sections[j]["depth"] <= sec_depth:
            return sections[j]["row"] - 1
    return total_rows


def _read_sheet_rows(ws) -> list[list]:
    return [
        [str(v) if v is not None else "" for v in row]
        for row in ws.iter_rows(values_only=True)
    ]


def process_xlsx(input_path: Path, output_path: Path | None = None) -> Path:
    """给 input xlsx 的「定额条目」sheet 套格式;其它 sheet 保持不动。

    Args:
        input_path: 输入 xlsx 路径
        output_path: 输出 xlsx 路径;None → 原地覆盖 input_path

    Returns:
        实际写入的路径
    """
    if output_path is None:
        output_path = input_path

    if not input_path.exists():
        print(f"[ERROR] 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(input_path, read_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        print(
            f"[ERROR] xlsx 中没有 {TARGET_SHEET!r} sheet;现有 sheet: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws = wb[TARGET_SHEET]
    rows = _read_sheet_rows(ws)
    if not rows:
        print("[WARN] 空「定额条目」sheet")
        wb.save(output_path)
        return output_path

    ws.sheet_properties.outlinePr.summaryBelow = False
    total_rows = len(rows)

    # 0. 先解除所有已存在的合并区域,避免循环写 cell 时碰到 MergedCell 只读
    #    (样本 xlsx 可能上游已经合并过 C:D 列;后面会在第 2 步按"段行 C-D 合并"重新合并)
    pre_existing_merges = list(ws.merged_cells.ranges)
    for rng in pre_existing_merges:
        ws.unmerge_cells(str(rng))

    # 1. 把数据写回(确认所有 cell 都在位) + 段行加粗
    bold_font = Font(bold=True)
    for r_idx, row in enumerate(rows, start=1):
        is_section = (str(row[0]).strip() if row else "") == "段"
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if is_section:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 2. 段行 C-D 列合并
    for r_idx, row in enumerate(rows, start=1):
        if row and str(row[0]).strip() == "段":
            ws.merge_cells(
                start_row=r_idx, start_column=3, end_row=r_idx, end_column=4
            )

    # 3. 收集段与定额位置
    sections: list[dict] = []
    quotas: list[dict] = []

    for r_idx, row in enumerate(rows, start=1):
        if not row:
            continue
        row_type = str(row[0]).strip()
        if row_type == "段":
            sec_id = str(row[1]).strip() if len(row) > 1 else ""
            sections.append({
                "row": r_idx,
                "depth": get_section_depth(sec_id),
                "id": sec_id,
            })
        elif row_type == "定":
            quota_id = str(row[1]).strip() if len(row) > 1 else ""
            # 定额行的「自身深度」= 自身 ID 的点数 + 1（H100001 → 1）
            # 但 outline_level 用「父段 depth + 1」，让 group 嵌套在父段之下
            own_depth = get_section_depth(quota_id)
            quotas.append({
                "row": r_idx,
                "id": quota_id,
                "own_depth": own_depth,
            })

    # 4. 计算段结束位置
    for i in range(len(sections)):
        sections[i]["end_row"] = find_end_row(sections, i, total_rows)

    # 5. 给每个定额关联到它所属的最近一个段;定额结束位置 = min(下一定额, 所属段 end_row)
    #    自适应层数:定额 group 的 outline_level = parent_sec.depth + 1
    #    (而不是之前的硬编码 4,这样 5 层 / 6 层定额也能正确嵌套)
    for i, q in enumerate(quotas):
        # 默认到下一定额之前
        if i + 1 < len(quotas):
            q["end_row"] = quotas[i + 1]["row"] - 1
        else:
            q["end_row"] = total_rows
        # 找最近父段,截到父段 end_row
        parent_sec = None
        for sec in reversed(sections):
            if sec["row"] <= q["row"]:
                parent_sec = sec
                break
        if parent_sec is not None:
            q["parent_row"] = parent_sec["row"]
            q["parent_depth"] = parent_sec["depth"]
            q["end_row"] = min(q["end_row"], parent_sec["end_row"])
        else:
            q["parent_row"] = None
            q["parent_depth"] = 0  # 没找到父段,fallback 1

    # 6. 建立分组区间
    #    段 group level = sec.depth
    #    定额 group level = parent_sec.depth + 1 (嵌套在父段之下)
    groups: list[tuple[int, int, int]] = []

    for sec in sections:
        start = sec["row"] + 1
        end = sec["end_row"]
        if start <= end:
            groups.append((start, end, sec["depth"]))

    for q in quotas:
        start = q["row"] + 1
        end = q["end_row"]
        if start <= end:
            groups.append((start, end, q["parent_depth"] + 1))

    # 先低 level(大范围),后高 level(小范围),
    # 这样小范围 group 不会覆盖大范围 group 的端点
    groups.sort(key=lambda x: x[2])

    for start, end, level in groups:
        ws.row_dimensions.group(start, end, outline_level=level)

    wb.save(output_path)

    print(f"[OK] 输入: {input_path}")
    print(f"[OK] 输出: {output_path}")
    print(
        f"[OK] sheet: {TARGET_SHEET} 总行数 {total_rows}, "
        f"段数 {len(sections)}, 定额数 {len(quotas)}"
    )
    print(f"[OK] 分组区间数: {len(groups)}")
    print(
        f"[OK] 其它保留 sheet: {[s for s in wb.sheetnames if s != TARGET_SHEET]}"
    )
    return output_path


def main():
    ap = argparse.ArgumentParser(
        description="XLSX → 带分组大纲 + 段行合并的 Excel（quota-csv-finalize 第 5 步）"
    )
    ap.add_argument("input_xlsx", help="输入 xlsx 路径（典型为 <stem>_待审核.xlsx）")
    ap.add_argument(
        "output_xlsx",
        nargs="?",
        default=None,
        help="输出 xlsx 路径（默认 = 原地覆盖输入文件）",
    )
    args = ap.parse_args()

    input_path = Path(args.input_xlsx).resolve()
    if not input_path.exists():
        print(f"[ERROR] 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output_xlsx).resolve() if args.output_xlsx else None
    process_xlsx(input_path, output_path)


if __name__ == "__main__":
    main()
