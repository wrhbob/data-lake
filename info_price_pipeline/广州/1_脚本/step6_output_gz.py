"""step6_output_gz.py - 广州 xlsx 输出 (2026-08-05 建, Step 1.6)

输入: extract.json (sheet 列表, step3 输出)
输出: 4_输出/<name>.xlsx

设计:
  - 多 sheet 支持 (Step 5 起), 每个 sheet 用自己的 schema
  - 列头 → dict key 映射 (_KEY_MAP) 让 step3 输出 dict 跟 sheet 列头解耦
  - 9 列 schema (金属材料): 序号|名称|规格(mm)|单位|区间值|均值|含税|除税|备注
  - 11 列 schema (砂浆, Step 8.1): 序号|名称|性能指标|强度等级|单位|区间值|均值|含税|除税|适用范围|t/m³系数

复用 (按约束「不改其他城市」, 复制函数体):
  - 武汉 step6: openpyxl write logic

约束:
  - GBK 编码修复 (复武汉)
  - 含税价/除税价空 → 写空 (不是 "0")
  - 名称去 [复制] / 「」 等 PDF 残留字符 (后续 Step 加, 现在不动)
"""
import json
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils_gz import ROOT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")


# 列头 → dict key 映射 (Step 8.1 加, 让多 sheet 多 schema 复用同一 _write_sheet)
#   10 列 schema (金属材料): 序号→seq, 名称→name, 规格(mm)→spec, 单位→unit, ...
#   11 列 schema (砂浆):     性能指标→performance, 强度等级→strength, 适用范围→scope, t/m³系数→coeff
#   6 列 schema (装配式):    项目名称→name, 钢筋含量(kg/m3)→rebar_kg_m3, 钢筋含量→rebar_segment
#   6 列 schema (电线电缆):  材料名称→name, 标称截面(mm²)→spec, 芯数→core, 单位→unit, 税前综合价格(元)→price, 备注→remark
_KEY_MAP = {
    "序号": "seq",
    "名称": "name",                           # 金属材料
    "项目名称": "name",                       # 装配式 (2026-08-06 加, 跟金属材料共用 name key)
    "材料名称": "name",                       # 2026-08-07 加, 电线电缆 + 镀锌桥架 sheet (md[16-26] 表头)
    "规格(mm)": "spec",                       # 金属材料
    "标称截面(mm²)": "spec",                  # 2026-08-07 加, 电线电缆 sheet (md[16-21] 表头)
    "规格": "spec",                           # 2026-08-07 加, 钢管 sheet (无 (mm) 后缀, 是 DN 值)
    "规格（高×宽）": "spec",                  # 2026-08-07 加, 镀锌桥架 sheet (md[23-26] 表头)
    "规格DN": "spec_dn",                      # 2026-08-07 fix: 钢管 sheet (拆出英寸后)
    "规格英寸": "spec_inch",                  # 2026-08-07 fix: 钢管 sheet (解 LaTeX+HTML)
    "钢筋含量(kg/m3)": "rebar_kg_m3",         # 装配式 (2026-08-06 加, anchor idx 2 总范围)
    "钢筋含量": "rebar_segment",              # 装配式 (2026-08-06 加, 本段)
    "壁厚(mm)": "wall_thickness",             # 2026-08-07 加, 镀锌桥架 sheet (表头带 (mm))
    "壁厚": "wall_thickness",                 # 2026-08-07 加, 钢管 sheet
    "芯数": "core",                           # 2026-08-07 加, 电线电缆 sheet (单芯/二芯/三芯/四芯/五芯/-)
    "性能指标": "performance",
    "强度等级": "strength",
    "单位": "unit",
    "税前综合价格(元)区间值": "interval_value",
    "税前综合价格(元)均值": "mean_value",
    "税前综合价格(元)": "price",              # 2026-08-07 加, 钢管 sheet (单值, 无 (区间值) 后缀)
    "含税价(元)": "incl",
    "除税价(元)": "excl",
    "适用范围": "scope",
    "t/m³系数": "coeff",
    "执行标准": "execution_standard",          # 2026-08-06 加, md[7] 蒸压加气 GB/T 15762-2020
    "表面积（m²/m）单面": "area_single",      # 2026-08-07 加, 镀锌桥架 sheet (单面)
    "表面积（m²/m）双面": "area_double",      # 2026-08-07 加, 镀锌桥架 sheet (双面)
    "备注": "remark",                         # 2026-08-07 加, 电线电缆 + 镀锌桥架 sheet (整段说明, 每行复用)
}


def _write_sheet(ws, headers, rows):
    """单 sheet 写入 (Step 8.1: 按 headers 动态列数 + _KEY_MAP 映射 dict key)
    2026-08-07: 加备注列 wrap_text + 列宽自适应 (B 方案)
    """
    # 1) 表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    # 2) 数据行 (按 headers 顺序 + _KEY_MAP 查 dict key)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for i, r in enumerate(rows, 2):
        for col, h in enumerate(headers, 1):
            key = _KEY_MAP.get(h, h)
            cell = ws.cell(row=i, column=col, value=r.get(key, ""))
            # 备注列加 wrap_text (避免几百字横向溢出到右侧空白)
            if key == "remark" and cell.value:
                cell.alignment = wrap_alignment
    # 3) 列宽自适应 (2026-08-07):
    #    - 备注列 50 字符宽 (留足换行空间)
    #    - 其他列按内容最大长度估算, 限 [10, 30]
    for col, h in enumerate(headers, 1):
        col_letter = get_column_letter(col)
        key = _KEY_MAP.get(h, h)
        if key == "remark":
            ws.column_dimensions[col_letter].width = 50
        else:
            # 估算列宽
            max_len = len(str(h))
            for r in rows:
                v = r.get(key, "")
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = max(10, min(30, max_len + 2))


def output(extract_json, xlsx_name="10696342_p6.xlsx", log_lines=None):
    """主入口: extract.json → xlsx"""
    extract_path = Path(extract_json)
    out_path = ROOT / "4_输出" / xlsx_name
    print(f"[step6 广州] 输入: {extract_path}")
    print(f"[step6 广州] 输出: {out_path}")

    with extract_path.open(encoding="utf-8") as f:
        extract_data = json.load(f)

    wb = Workbook()
    # 默认 sheet 删, 按 extract_data 顺序建
    wb.remove(wb.active)

    for sheet_data in extract_data:
        sheet_name = sheet_data["sheet"]
        headers = sheet_data["headers"]
        rows = sheet_data["rows"]
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, headers, rows)
        print(f"[step6 广州] Sheet '{sheet_name}': {len(rows)} 行 × {len(headers)} 列")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"[step6 广州] ✅ 写出: {out_path}")

    if log_lines is not None:
        for sheet_data in extract_data:
            log_lines.append(
                f"[step6 广州] Sheet '{sheet_data['sheet']}': {len(sheet_data['rows'])} 行"
            )

    return str(out_path)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="广州 xlsx 输出 (Step 1.6)")
    parser.add_argument("extract_json", help="step3 输出的 extract.json")
    parser.add_argument("--name", default="10696342_p6.xlsx", help="xlsx 文件名")
    args = parser.parse_args()
    output(args.extract_json, xlsx_name=args.name, log_lines=[])