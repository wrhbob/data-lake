"""step6_output.py - 写 xlsx（3 Sheet，2026-07-29 删 PC sheet）

目的：
  clean.json → 4_输出/{city}_{year}年{period}期_市场信息价.xlsx

3 个 Sheet（2026-07-29 重庆 06 重构）：
  Sheet 1: 市场信息价
    列: 区名 | 材料名称 | 规格型号 | 产地 | 单位 | 除税价格(元) | 含税价(元) | 不含税价(元) | 原始数据
  Sheet 2: 行业推荐（原"新型材料"，重庆 06 PDF 章节名）
    列: 产品名 | 规格 | 单位 | 含税价(元) | 不含税价(元) | 原始数据
  Sheet 3: 元数据
    城市 / 出版期 / 数据月份 / 周期类型 / 提取时间

注：重庆 06 没有 PC section，删 PC sheet 不影响；成都/湖北仍走 step3_extract.py 不读本函数。
"""
import json
import sys
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Windows GBK 编码修复
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import ROOT, city_to_code


# ============================================================
# Sheet 1: 市场信息价
# ============================================================

def write_market_sheet(wb, data):
    """Sheet 1: 市场信息价（2026-07-29 加回"含税价(元)"列）

    列结构（用户原话 2026-07-29）：
      - 「不要删含税价格，除税价格」= 两个都保留
      - 即使成都 06 含税价整列空（3907/3907 空），也保留列结构（其他城市可能非空）
      - 列顺序：含税价(元) → 除税价格(元) → 不含税价(元)
    """
    ws = wb.create_sheet("市场信息价", 0)
    headers = ["区名", "材料名称", "规格型号", "产地", "单位",
               "含税价(元)", "除税价格(元)", "不含税价(元)", "原始数据"]
    ws.append(headers)

    # 表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    for table in data:
        if table.get("section") != "MARKET":
            continue
        for row in table.get("rows", []):
            # raw_cells 兜底：切分失败的二维护表数据存在这里
            raw = row.get("raw_cells", [])
            raw_str = " | ".join(str(c) for c in raw) if raw else ""
            ws.append([
                row.get("district", ""),            # 2026-07-29: clean.json 字段名是 district 不是 区名
                row.get("材料名称", ""),
                row.get("规格型号", ""),
                row.get("产地", ""),
                row.get("单位", ""),
                row.get("含税价(元)", ""),          # 2026-07-29 加回：用户原话"不要删含税价格"
                row.get("除税价格(元)", ""),
                row.get("不含税价(元)", ""),
                raw_str,
            ])
            n_rows += 1

    # 列宽（9 列：区名 + 材料 + 规格 + 产地 + 单位 + 含税价 + 除税价 + 不含税价 + 原始数据）
    widths = [12, 25, 15, 12, 8, 12, 12, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 2: 新型材料（成都/湖北原结构，2026-07-29 保留兼容）
# ============================================================

def write_new_material_sheet(wb, data):
    """Sheet 2: 新型材料（成都/湖北原 11 列结构，含品牌/公司名/电话/执行标准）

    2026-07-29 改：
      - 重庆 06 用 6 列 write_industry_sheet，本函数不调
      - 成都/湖北仍走本函数（11 列）
    """
    ws = wb.create_sheet("新型材料", 1)
    headers = [
        "产品名", "规格", "品牌", "单位",
        "价格(元)", "含税价(元)", "不含税价(元)",
        "公司名", "联系人", "电话", "执行标准",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    for table in data:
        if table.get("section") != "NEW_MATERIAL":
            continue
        for row in table.get("rows", []):
            ws.append([
                row.get("产品名", ""),
                row.get("规格", ""),
                row.get("品牌", ""),
                row.get("单位", ""),
                row.get("价格(元)", ""),
                row.get("含税价(元)", ""),
                row.get("不含税价(元)", ""),
                row.get("公司名", ""),
                row.get("联系人", ""),
                row.get("电话", ""),
                row.get("执行标准", ""),
            ])
            n_rows += 1

    widths = [30, 25, 15, 8, 12, 12, 12, 25, 15, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 2 alt: 行业推荐（重庆 06 专属 6 列，2026-07-29 增）
# ============================================================

def write_industry_sheet(wb, data):
    """Sheet 2 alt: 行业推荐（重庆 PDF 章节名"行业推荐价格信息"，6 列）

    2026-07-29 增：重庆 06 NM 表实际列结构 = 序号 | 材料名 | 规格 | 单位 | 价格
      - 删品牌/公司名/联系人/电话/执行标准（重庆 06 协会表无这些列）
      - 列: 产品名 | 规格 | 单位 | 含税价 | 不含税价 | 原始数据
    """
    ws = wb.create_sheet("行业推荐", 1)
    headers = ["产品名", "规格", "单位", "含税价(元)", "不含税价(元)", "原始数据"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    for table in data:
        if table.get("section") != "NEW_MATERIAL":
            continue
        for row in table.get("rows", []):
            raw = row.get("raw_cells", [])
            raw_str = " | ".join(str(c) for c in raw) if raw else ""
            ws.append([
                row.get("产品名", ""),
                row.get("规格", ""),
                row.get("单位", ""),
                row.get("含税价(元)", ""),
                row.get("不含税价(元)", ""),
                raw_str,
            ])
            n_rows += 1

    widths = [30, 25, 8, 12, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 3: PC 构件基础价格（2026-07-29 增：CD 有 PC 数据，CQ 无，条件写）
# ============================================================

def write_pc_sheet(wb, data):
    """Sheet 3: PC 构件基础价格（仅当 clean.json 有 PC section 时写）

    2026-07-29 修：之前 step6_output.py 因 CQ 06 无 PC section 删了 PC sheet，导致成都 06 的 66 行 PC 数据丢失。
    改为条件写：CD 有 PC 就写，CQ 没有就跳过。
    """
    pc_rows_total = sum(len(t.get("rows", [])) for t in data if t.get("section") == "PC")
    if pc_rows_total == 0:
        return 0  # CQ 无 PC 数据 → 不写 sheet

    ws = wb.create_sheet("PC构件基础价格")
    headers = ["构件名称", "规格", "单位", "钢筋含量",
               "价格类型", "含税价(元)", "不含税价(元)", "说明", "原始数据"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    for table in data:
        if table.get("section") != "PC":
            continue
        for row in table.get("rows", []):
            raw = row.get("raw_cells", [])
            raw_str = " | ".join(str(c) for c in raw) if raw else ""
            ws.append([
                row.get("构件名称", ""),
                row.get("规格", ""),
                row.get("单位", ""),
                row.get("钢筋含量", ""),
                row.get("价格类型", ""),
                row.get("含税价(元)", ""),
                row.get("不含税价(元)", ""),
                row.get("说明", ""),
                raw_str,
            ])
            n_rows += 1

    widths = [25, 18, 8, 12, 20, 12, 12, 15, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 3: 元数据（2026-07-29 改：原 Sheet 4）
# ============================================================

def write_meta_sheet(wb, data, city, period, year, data_month, cycle_type):
    """Sheet 3: 元数据"""
    ws = wb.create_sheet("元数据", 2)
    ws.append(["字段", "值"])
    ws.append(["城市", city])
    ws.append(["年份", year])
    ws.append(["出版期", period])
    ws.append(["数据月份", data_month])
    ws.append(["周期类型", cycle_type])
    ws.append(["提取时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    # 统计（2026-07-29 改：删 PC 行数）
    n_market = sum(len(t["rows"]) for t in data if t.get("section") == "MARKET")
    n_new = sum(len(t["rows"]) for t in data if t.get("section") == "NEW_MATERIAL")
    ws.append(["市场信息价行数", n_market])
    ws.append(["行业推荐行数", n_new])

    # 样式
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")


# ============================================================
# 主流程
# ============================================================

def write_xlsx(clean_json, city, period, year, data_month, cycle_type):
    """clean.json → xlsx

    2026-07-30 北京版：按 yaml sheet_config 分 5+1 sheet（每个章节独立 sheet）
      - 用户原话 2026-07-29：「代号那一列可以全都不要」→ Sheet 1/2 删代号列
      - Sheet 1 工程造价信息价 (MARKET_1)
      - Sheet 2 市场参考价 (MARKET_2)
      - Sheet 3 厂家参考价 (NM_3)  ← 章节三 OCR 实测无表
      - Sheet 4 绿色建材 (NM_4)    ← 章节四 OCR 实测无表
      - Sheet 5 花园城市 (NM_5)
      - Sheet 6 元数据

    2026-07-29 重庆版：sheet_config 有 → 走 3+1 sheet（Sheet 1/2/3 + 元数据）
    2026-07-29 成都/湖北版：sheet_config 无 → 走原 3+1 sheet（市场信息价+新型材料+元数据）
    """
    code = city_to_code(city)
    out_dir = ROOT / "4_输出"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{city}_{year}年{period}期_市场信息价.xlsx"
    # 2026-07-29 改：递归找空闲文件名（v2 / v3 / v4 ...），避免连续 _v2 同时被锁时丢异常
    # 之前 1 版：只检查 _v2，遇到 _v2 也被锁就 PermissionError
    suffix = 1
    while (out_dir / f"~${out_path.name}").exists() or suffix > 50:
        suffix += 1
        out_path = out_dir / f"{city}_{year}年{period}期_市场信息价_v{suffix}.xlsx"
        if suffix > 50:
            raise RuntimeError(f"xlsx 输出文件名冲突超过 50 次，放弃")
    if suffix > 1:
        print(f"[step6] ⚠️ 前缀被占用，改输出到 {out_path.name}")

    with open(clean_json, encoding="utf-8") as f:
        data = json.load(f)

    # 读城市 sheet_config
    sheet_config = _load_sheet_config(city)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_summary = []  # (sheet 名, 行数)

    if sheet_config and sheet_config.get("sheets"):
        # 2026-07-30 北京 schema 优先模式：按 yaml.sheet_config.sheets 顺序建 sheet
        # sheet 顺序由 yaml 决定（保持稳定），table 由 step3 用 schema 匹配后写 table["sheet"] 字段
        sheets_cfg = sheet_config["sheets"]
        for sc in sheets_cfg:
            sheet_name = sc.get("name")
            if sheet_name is None or sheet_name == "元数据":
                continue  # 元数据后面统一写
            # 按 sheet_name 从 data 中筛选 table（table["sheet"] == sheet_name）
            n = _write_chapter_sheet(wb, sheet_name, data)
            sheet_summary.append((sheet_name, n))
            # 2026-07-30 增：空 sheet 跳过（用户原话「绿色建材你有看到内容吗？没有」）
            # 章节四 OCR 实测无表 → 推荐 6 列 schema 未匹配 → Sheet 4 全空 → 删 sheet
            if n == 0:
                empty_ws = wb[sheet_name]
                wb.remove(empty_ws)
                sheet_summary[-1] = (sheet_name, 0)
                print(f"[step6]   ⚠️ {sheet_name} sheet 空（schema 未匹配），已移除")
    else:
        # 旧逻辑（重庆 3+1 / 成都-湖北 3+1）
        n_market = write_market_sheet(wb, data)
        sheet_summary.append(("市场信息价", n_market))
        if sheet_config:
            n_industry = write_industry_sheet(wb, data)
            sheet_summary.append(("行业推荐", n_industry))
        else:
            n_nm = write_new_material_sheet(wb, data)
            sheet_summary.append(("新型材料", n_nm))
        n_pc = write_pc_sheet(wb, data)
        if n_pc > 0:
            sheet_summary.append(("PC构件基础价格", n_pc))

    write_meta_sheet(wb, data, city, period, year, data_month, cycle_type)

    wb.save(out_path)
    print(f"[step6] ✅ 写出: {out_path}")
    for sname, n in sheet_summary:
        print(f"[step6]   Sheet {sname}: {n} 行")
    return out_path


def _write_chapter_sheet(wb, sheet_name, data):
    """北京版：按 table["sheet"] == sheet_name 筛选，写一个 sheet（2026-07-30 schema 优先）

    列结构由 table["schema_id"] 决定（每 schema 字段映射不同）
    """
    # 收集该 sheet 名下的所有 table
    tables = [t for t in data if t.get("sheet") == sheet_name]
    if not tables:
        # 仍写空 sheet（保 schema 顺序稳定）
        pass

    # 列结构按 schema_id 决定（取第一个有数据的 table 的 schema_id）
    schema_id = ""
    for t in tables:
        if t.get("schema_id"):
            schema_id = t["schema_id"]
            break

    # 列结构按 schema_id 区分（PC 归 Sheet 2 市场参考价末尾 → sheet_name=市场参考价, schema_id=pc_9col_bj）
    # 注：_write_chapter_sheet 一次只生成一个 sheet_name，所以 schema_id 唯一
    if schema_id in ("market_5col_bj_p1", "market_5col_bj_p2", "market_5col_bj_p3"):
        headers = ["产品名称", "规格型号及特征", "计量单位", "含税价(元)", "不含税价(元)", "原始数据"]
        widths = [30, 35, 12, 12, 12, 60]
    elif schema_id in ("recommend_6col_manufacturer", "recommend_6col_green"):
        headers = ["序号", "产品名称", "规格型号及特征", "计量单位", "含税价(元)", "不含税价(元)", "原始数据"]
        widths = [8, 30, 35, 12, 12, 12, 60]
    elif schema_id == "market_8col_bj_garden":
        # 2026-07-30 P2-1：union 11 列宽表（5 张表 schema 不同的规格列展开到固定列）
        # 2026-07-30 P2-2：加 "材料名称" 列（章节名/品种名 拼接，step3 row_out["产品名"] 填入）
        # 2026-07-30 P2-3：加 "分枝数量/条长(m)" 列（月季/藤本表独有，其他表为空）
        headers = ["序号", "材料名称", "高度(m)", "胸径(cm)", "地径(cm)", "苗龄",
                   "分枝数量/条长(m)", "盆径(cm)/穴盘规格",
                   "冠幅(m)", "计量单位", "含税价(元)", "不含税价(元)", "备注", "原始数据"]
        widths = [8, 30, 10, 10, 10, 8, 15, 18, 10, 12, 12, 12, 30, 60]
    elif schema_id == "pc_9col_bj":
        # PC 归 Sheet 2 市场参考价末尾，列结构 9 列（构件名称 + 钢筋含量 + 说明）
        headers = ["构件名称", "规格", "单位", "钢筋含量",
                   "价格类型", "含税价(元)", "不含税价(元)", "说明", "原始数据"]
        widths = [25, 18, 8, 12, 20, 12, 12, 15, 60]
    else:
        # 无 schema（NM 章节数据由旧 NEW_MATERIAL 处理）
        # 兜底：行业推荐 6 列
        headers = ["产品名", "规格", "单位", "含税价(元)", "不含税价(元)", "原始数据"]
        widths = [30, 25, 8, 12, 12, 60]

    ws = wb.create_sheet(sheet_name)
    ws.append(headers)

    # 表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w

    n_rows = 0
    for table in tables:
        for row in table.get("rows", []):
            raw = row.get("raw_cells", [])
            raw_str = " | ".join(str(c) for c in raw) if raw else ""

            if schema_id in ("market_5col_bj_p1", "market_5col_bj_p2", "market_5col_bj_p3"):
                ws.append([
                    row.get("材料名称", ""),
                    row.get("规格型号", ""),
                    row.get("单位", ""),
                    row.get("含税价(元)", ""),
                    row.get("不含税价(元)", ""),
                    raw_str,
                ])
            elif schema_id in ("recommend_6col_manufacturer", "recommend_6col_green"):
                ws.append([
                    row.get("序号", ""),
                    row.get("材料名称", "") or row.get("产品名", ""),
                    row.get("规格型号", "") or row.get("规格", ""),
                    row.get("单位", ""),
                    row.get("含税价(元)", ""),
                    row.get("不含税价(元)", ""),
                    raw_str,
                ])
            elif schema_id == "market_8col_bj_garden":
                # 2026-07-30 P2-1：union 10 列宽表（与 headers 对齐）
                # 2026-07-30 P2-2：col[1] = 材料名称（章节名/品种名 拼接，row["产品名"]）
                # 2026-07-30 P2-3：col[6] = 分枝数量/条长(m)（月季/藤本表独有，其他表为空）
                ws.append([
                    row.get("序号", ""),
                    row.get("产品名", ""),
                    row.get("高度(m)", ""),
                    row.get("胸径(cm)", ""),
                    row.get("地径(cm)", ""),
                    row.get("苗龄", ""),
                    row.get("分枝数量/条长(m)", ""),
                    row.get("盆径(cm)/穴盘规格", ""),
                    row.get("冠幅(m)", ""),
                    row.get("计量单位", ""),
                    row.get("含税价(元)", ""),
                    row.get("不含税价(元)", ""),
                    row.get("备注", ""),
                    raw_str,
                ])
            elif schema_id == "pc_9col_bj":
                ws.append([
                    row.get("构件名称", ""),
                    row.get("规格", ""),
                    row.get("单位", ""),
                    row.get("钢筋含量", ""),
                    row.get("价格类型", ""),
                    row.get("含税价(元)", ""),
                    row.get("不含税价(元)", ""),
                    row.get("说明", ""),
                    raw_str,
                ])
            else:
                # NM 兜底（重庆 06 行业推荐 6 列）
                ws.append([
                    row.get("产品名", ""),
                    row.get("规格", ""),
                    row.get("单位", ""),
                    row.get("含税价(元)", ""),
                    row.get("不含税价(元)", ""),
                    raw_str,
                ])
            n_rows += 1
    return n_rows


def _load_sheet_config(city):
    """读 6_配置/城市模板/{city}.yaml 的 sheet_config（2026-07-29 增）

    存在 → 用行业推荐（6 列）
    不存在 → 用新型材料（11 列，成都/湖北原结构）
    """
    try:
        import yaml
        candidates = [
            ROOT / "6_配置" / "城市模板" / f"{city}.yaml",
            ROOT / "6_配置" / "城市模板" / f"{city}市.yaml",
        ]
        for path in candidates:
            if path.exists():
                with path.open(encoding="utf-8") as f:
                    yaml_data = yaml.safe_load(f)
                if yaml_data.get("sheet_config"):
                    return yaml_data["sheet_config"]
                return None
        return None
    except Exception:
        return None


if __name__ == "__main__":
    if len(sys.argv) < 7:
        print("用法: python step6_output.py <clean_json> <city> <period> <year> <data_month> <cycle_type>", file=sys.stderr)
        sys.exit(1)
    write_xlsx(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6])
