"""step6_output.py - 写 xlsx（3 Sheet，2026-07-29 删 PC sheet）

目的：
  clean.json → 4_输出/{city}_{year}年{period}期_市场信息价.xlsx

3 个 Sheet（2026-07-29 重庆 06 重构）：
  Sheet 1: 市场信息价
    列: 区名 | 材料名称 | 规格型号 | 产地 | 单位 | 除税价格(元) | 含税价(元) | 不含税价(元) | 原始数据
  Sheet 2: 行业推荐（原"新型材料"，重庆 06 PDF 章节名）
    列: 产品名 | 规格 | 单位 | 含税价(元) | 不含税价(元) | 原始数据
  Sheet 3: 元数据
    城市 / 出版期 / 数据月份 / 周期类型 / 提取时间

注：重庆 06 没有 PC section，删 PC sheet 不影响；成都/湖北仍走 step3_extract.py 不读本函数。
"""
import json
import sys
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Windows GBK 编码修复
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import ROOT, city_to_code


# ============================================================
# Sheet 说明（2026-07-31 增：每个 sheet 顶部写 1 行说明，有就写）
# ============================================================

SHEET_DESCRIPTIONS = {
    "市场信息价": "成都市每月发布的各区县建筑材料市场价格（钢筋/水泥/砂石/砌块等大宗建材）",
    "新型材料": "行业协会推荐的新型/再生建筑材料价格（不含税价/含税价/品牌/联系人）",
    "行业推荐": "重庆各行业协会推荐材料价格（协会表 6 列结构）",
    "PC构件基础价格": "装配式建筑预制构件到场价格（PC 构件/钢构件/预制构件，含税+不含税）",
    "苗木": "园林绿化苗木价格（乔木/花灌木/棕榈植物/地被）",
}


def _write_description(ws, sheet_name):
    """在 sheet 顶部写 1 行说明（如有定义），无说明则跳过"""
    desc = SHEET_DESCRIPTIONS.get(sheet_name)
    if not desc:
        return  # 没配说明 → 不写
    ws.append([f"说明：{desc}"])
    # 说明行样式：浅灰底 + 斜体
    for cell in ws[ws.max_row]:
        cell.font = Font(italic=True, color="666666")
        cell.fill = PatternFill("solid", fgColor="F0F0F0")
    # 空行分隔
    ws.append([])


# ============================================================
# Sheet 1: 市场信息价
# ============================================================

def _explode_2d_rows(table_rows):
    """2026-08-11 v13 增：把 is_2d_spliced=True 的整张横向表 explode 成纵向行

    触发场景：成都 split_2d_table 切分失败时，整表按行保留，row['raw_cells'] = [c0, c1, c2, ...]
      - 第 1 行 raw_cells = 横向表头（['成都市','新都区',...,'备注']）
      - 后续行 raw_cells = 横向数据（['m³','27.00','26.00',...])
    explode: 找到首个表头行，后续每个数据行 × 表头 N 个区 = N 行

    缺陷：row['材料名称/规格型号'] 都为空（横向表 OCR 没切出材料名）
          输出时只能 raw_cells 单元映射 → xlsx 看到的材料列是空的
          修：保留原始 8-10 个 cell 的 raw_str 让用户能 reverse
    """
    if not table_rows:
        return [], None
    # 找首个表头行：raw_cells 全非空 + 第1 个 cell 包含区名关键字
    districts_hint = ("成都市", "新都区", "温江区", "双流区", "大邑县", "蒲江县",
                      "都江堰市", "彭州市", "金堂县", "崇州市", "邛崃市", "郫都区",
                      "新津区", "青白江区", "龙泉驿区", "简阳市", "东部新区",
                      "天府新区", "高新区", "金牛区", "锦江区", "武侯区", "成华区")
    header_row_idx = -1
    header_cells = None
    for i, r in enumerate(table_rows):
        if not r.get("is_2d_spliced"):
            continue
        raw = r.get("raw_cells", [])
        if len(raw) >= 3 and any(d in str(raw[0]) for d in districts_hint):
            header_row_idx = i
            header_cells = raw
            break
    if header_row_idx < 0:
        return [], None
    # 2026-08-11 v13.1 改：explode 后**按区聚类**（用户原话：成都市的全在一起，新都的全在一起）
    # 旧逻辑：按行 explode → 成/新/温/双/大/蒲/都/彭/成/新/温/... (横向铺开)
    # 新逻辑：先收集每个数据行 unit + prices + remark → 再按区聚类 explode → 所有"成都市"在一起
    data_records = []  # [{unit, prices, remark}, ...] 每个原始数据行一条
    for r in table_rows[header_row_idx + 1:]:
        if not r.get("is_2d_spliced"):
            continue
        raw = r.get("raw_cells", [])
        if len(raw) < 2:
            continue
        unit = raw[0]
        remark = raw[-1] if raw else ""
        prices = raw[1:-1] if len(raw) >= 3 else []
        data_records.append((unit, prices, remark))
    # 按 header 区名聚类：跳过"备注"列
    out = []
    for hi, hcell in enumerate(header_cells):
        if "备注" in str(hcell):
            continue
        for unit, prices, remark in data_records:
            if hi >= len(prices):
                break
            out.append({
                "district": str(hcell),
                "材料名称": "",
                "规格型号": "",
                "产地": "",
                "单位": unit,
                "含税价(元)": "",
                "除税价格(元)": str(prices[hi]),
                "不含税价(元)": "",
                "remark": str(remark),
                "_exploded": True,
            })
    return out, header_cells


def write_market_sheet(wb, data):
    """Sheet 1: 市场信息价（2026-07-29 加回"含税价(元)"列）

    列结构（用户原话 2026-07-29）：
      - 「不要删含税价格，除税价格」= 两个都保留
      - 即使成都 06 含税价整列空（3907/3907 空），也保留列结构（其他城市可能非空）
      - 列顺序：含税价(元) → 除税价格(元) → 不含税价(元)

    2026-08-11 v13 改：MARKET 段先 explode is_2d_spliced=True 的横向表，再写普通行
    """
    ws = wb.create_sheet("市场信息价", 0)
    _write_description(ws, "市场信息价")  # 2026-07-31 加说明
    headers = ["区名", "材料名称", "规格型号", "产地", "单位",
               "含税价(元)", "除税价格(元)", "不含税价(元)", "原始数据"]
    ws.append(headers)

    # 表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    # 2026-08-11 v13.2 改：先收集所有 exploded rows + 普通行，再按区聚类统一排序
    # 用户原话："成都市的全在一起，新都的全在一起"
    # 区顺序按成都行政区划固定序：成都市, 新都区, 温江区, 双流区, 郫都区, 龙泉驿区, 青白江区,
    #                            新津区, 金堂县, 大邑县, 蒲江县, 都江堰市, 彭州市, 邛崃市,
    #                            崇州市, 简阳市, 东部新区, 天府新区成都直管区, 高新区
    district_order = [
        "成都市", "新都区", "温江区", "双流区", "郫都区", "龙泉驿区", "青白江区",
        "新津区", "金堂县", "大邑县", "蒲江县", "都江堰市", "彭州市", "邛崃市",
        "崇州市", "简阳市", "东部新区", "天府新区成都直管区", "高新区",
    ]
    district_rank = {d: i for i, d in enumerate(district_order)}

    # 第一遍：收集所有 (district, order_in_table, row_dict) 三元组
    all_rows = []
    table_seq = 0
    for table in data:
        if table.get("section") != "MARKET":
            continue
        # 横向表 explode
        exploded_rows, _hdr = _explode_2d_rows(table.get("rows", []))
        for er in exploded_rows:
            all_rows.append((er["district"], table_seq, er))
        # 普通行
        for row in table.get("rows", []):
            if row.get("is_2d_spliced") and not row.get("_exploded"):
                continue
            all_rows.append((row.get("district", ""), table_seq, row))
        table_seq += 1

    # 排序：(district_rank, original_order_in_table)
    # 用 stable sort 保留同一区+同一表内的原顺序
    all_rows.sort(key=lambda t: (district_rank.get(t[0], 999), t[1]))

    # 第二遍：写 xlsx
    for _dist, _seq, row in all_rows:
        if row.get("_exploded"):
            raw_str = f"{row['district']} | 单位={row['单位']} | 价={row['除税价格(元)']} | 备注={row['remark']}"
        else:
            raw = row.get("raw_cells", [])
            raw_str = " | ".join(str(c) for c in raw) if raw else ""
        ws.append([
            row.get("district", ""),
            row.get("材料名称", ""),
            row.get("规格型号", ""),
            row.get("产地", ""),
            row.get("单位", ""),
            row.get("含税价(元)", ""),
            row.get("除税价格(元)", ""),
            row.get("不含税价(元)", ""),
            raw_str,
        ])
        n_rows += 1

    # 列宽（9 列：区名 + 材料 + 规格 + 产地 + 单位 + 含税价 + 除税价 + 不含税价 + 原始数据）
    widths = [12, 25, 15, 12, 8, 12, 12, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 2: 新型材料（成都/湖北原结构，2026-07-29 保留兼容）
# ============================================================

def write_new_material_sheet(wb, data):
    """Sheet 2: 新型材料（成都/湖北原 11 列结构，含品牌/公司名/电话/执行标准）

    2026-07-29 改：
      - 重庆 06 用 6 列 write_industry_sheet，本函数不调
      - 成都/湖北仍走本函数（11 列）
    """
    ws = wb.create_sheet("新型材料", 1)
    _write_description(ws, "新型材料")  # 2026-07-31 加说明
    headers = [
        "产品名", "规格", "品牌", "单位",
        "价格(元)", "含税价(元)", "不含税价(元)",
        "公司名", "联系人", "电话", "执行标准",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    for table in data:
        if table.get("section") != "NEW_MATERIAL":
            continue
        for row in table.get("rows", []):
            # 2026-08-03 修：成都 NM 协会推荐表无「价格」单列，价格落在「不含税价(元)」列
            # 抽到档位串「130/140/150」当成字符串。若「价格(元)」空，把档位串补到「价格(元)」列展示
            # 提示用户档位（完整档位在「不含税价(元)」列）
            price = row.get("价格(元)", "")
            tax_incl = row.get("含税价(元)", "")
            tax_excl = row.get("不含税价(元)", "")
            if not str(price).strip() and "/" in str(tax_excl):
                price = f"档位: {tax_excl}"
            ws.append([
                row.get("产品名", ""),
                row.get("规格", ""),
                row.get("品牌", ""),
                row.get("单位", ""),
                price,
                tax_incl,
                tax_excl,
                row.get("公司名", ""),
                row.get("联系人", ""),
                row.get("电话", ""),
                row.get("执行标准", ""),
            ])
            n_rows += 1

    widths = [30, 25, 15, 8, 12, 12, 12, 25, 15, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 2 alt: 行业推荐（重庆 06 专属 6 列，2026-07-29 增）
# ============================================================

def write_industry_sheet(wb, data):
    """Sheet 2 alt: 行业推荐（重庆 PDF 章节名"行业推荐价格信息"，6 列）

    2026-07-29 增：重庆 06 NM 表实际列结构 = 序号 | 材料名 | 规格 | 单位 | 价格
      - 删品牌/公司名/联系人/电话/执行标准（重庆 06 协会表无这些列）
      - 列: 产品名 | 规格 | 单位 | 含税价 | 不含税价 | 原始数据
    """
    ws = wb.create_sheet("行业推荐", 1)
    _write_description(ws, "行业推荐")  # 2026-07-31 加说明
    headers = ["产品名", "规格", "单位", "含税价(元)", "不含税价(元)", "原始数据"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    for table in data:
        if table.get("section") != "NEW_MATERIAL":
            continue
        for row in table.get("rows", []):
            raw = row.get("raw_cells", [])
            raw_str = " | ".join(str(c) for c in raw) if raw else ""
            ws.append([
                row.get("产品名", ""),
                row.get("规格", ""),
                row.get("单位", ""),
                row.get("含税价(元)", ""),
                row.get("不含税价(元)", ""),
                raw_str,
            ])
            n_rows += 1

    widths = [30, 25, 8, 12, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 3: PC 构件基础价格（2026-07-29 增：CD 有 PC 数据，CQ 无，条件写）
# ============================================================

def write_pc_sheet(wb, data):
    """Sheet 3: PC 构件基础价格（仅当 clean.json 有 PC section 时写）

    2026-07-29 修：之前 step6_output.py 因 CQ 06 无 PC section 删了 PC sheet，导致成都 06 的 66 行 PC 数据丢失。
    改为条件写：CD 有 PC 就写，CQ 没有就跳过。
    """
    pc_rows_total = sum(len(t.get("rows", [])) for t in data if t.get("section") == "PC")
    if pc_rows_total == 0:
        return 0  # CQ 无 PC 数据 → 不写 sheet

    ws = wb.create_sheet("PC构件基础价格")
    _write_description(ws, "PC构件基础价格")  # 2026-07-31 加说明
    headers = ["构件名称", "规格", "单位", "钢筋含量",
               "价格类型", "含税价(元)", "不含税价(元)", "说明", "原始数据"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    for table in data:
        if table.get("section") != "PC":
            continue
        for row in table.get("rows", []):
            raw = row.get("raw_cells", [])
            raw_str = " | ".join(str(c) for c in raw) if raw else ""
            ws.append([
                row.get("构件名称", ""),
                row.get("规格", ""),
                row.get("单位", ""),
                row.get("钢筋含量", ""),
                row.get("价格类型", ""),
                row.get("含税价(元)", ""),
                row.get("不含税价(元)", ""),
                row.get("说明", ""),
                raw_str,
            ])
            n_rows += 1

    widths = [25, 18, 8, 12, 20, 12, 12, 15, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 3: 元数据（2026-07-29 改：原 Sheet 4）
# ============================================================

def write_meta_sheet(wb, data, city, period, year, data_month, cycle_type):
    """Sheet 3: 元数据"""
    ws = wb.create_sheet("元数据", 2)
    ws.append(["字段", "值"])
    ws.append(["城市", city])
    ws.append(["年份", year])
    ws.append(["出版期", period])
    ws.append(["数据月份", data_month])
    ws.append(["周期类型", cycle_type])
    ws.append(["提取时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    # 统计（2026-07-29 改：删 PC 行数）
    n_market = sum(len(t["rows"]) for t in data if t.get("section") == "MARKET")
    n_new = sum(len(t["rows"]) for t in data if t.get("section") == "NEW_MATERIAL")
    ws.append(["市场信息价行数", n_market])
    ws.append(["行业推荐行数", n_new])

    # 样式
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")


# ============================================================
# 主流程
# ============================================================

def write_xlsx(clean_json, city, period, year, data_month, cycle_type, pdf_stem):
    """clean.json → xlsx

    2026-07-29 改：Sheet 2 命名按城市分
      - 重庆（sheet_config 有）："行业推荐"（6 列 NM 结构）
      - 成都/湖北（默认）："新型材料"（11 列 NM 结构）
      删 PC sheet（重庆 06 没 PC，成都 step3 不输出 PC）
    2026-07-31 改：加 pdf_stem 参数（第 7 个），输出文件名直接用 PDF stem
    """
    code = city_to_code(city)
    out_dir = ROOT / "4_输出"
    out_dir.mkdir(parents=True, exist_ok=True)
    # 2026-08-14 改：月刊 data_month 强制按出版期推断（出版期-1，1月刊=12）
    # 月刊约定：数据采集月份=出版前一月。不信用户传值（之前传错过 2 次：传==period、传乱数 1）。
    # 季刊/半年刊不强制（保持用户原值）。
    if cycle_type in ("月刊", "月") and period.isdigit():
        expected = str(int(period) - 1) if int(period) > 1 else "12"
        if str(data_month) != expected:
            print(f"[step6] ⚠ 月刊 data_month={data_month!r} ≠ 期望 {expected}（出版期{period}前1月），强制覆盖", flush=True)
            data_month = expected
    # 2026-07-31 改：直接用 PDF stem 作输出文件名（PDF 名 → xlsx 名只换后缀）
    out_path = out_dir / f"{pdf_stem}.xlsx"
    # 2026-07-29 改：递归找空闲文件名（v2 / v3 / v4 ...），避免连续 _v2 同时被锁时丢异常
    # 之前 1 版：只检查 _v2，遇到 _v2 也被锁就 PermissionError
    suffix = 1
    while (out_dir / f"~${out_path.name}").exists() or suffix > 50:
        suffix += 1
        # 2026-07-31 改：冲突重试也用 pdf_stem
        out_path = out_dir / f"{pdf_stem}_v{suffix}.xlsx"
        if suffix > 50:
            raise RuntimeError(f"xlsx 输出文件名冲突超过 50 次，放弃")
    if suffix > 1:
        print(f"[step6] ⚠️ 前缀被占用，改输出到 {out_path.name}")

    with open(clean_json, encoding="utf-8") as f:
        data = json.load(f)

    # 读城市 sheet_config（重庆有，默认走原结构）
    sheet_config = _load_sheet_config(city)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    n_market = write_market_sheet(wb, data)
    # 2026-08-11 v13 改：成都排除"新型材料"sheet（用户原话：丢掉不要）
    # NM 章节数据仍存在 extract.json/clean.json 但不写 xlsx（用户明确要求）
    if sheet_config:
        # 2026-07-29 重庆用 6 列行业推荐
        n_industry = write_industry_sheet(wb, data)
        sheet2_name = "行业推荐"
    else:
        # 成都/湖北原结构（11 列新型材料）— 2026-08-11 v13 跳过（用户排除 NM）
        n_industry = 0
        sheet2_name = "新型材料(已排除)"
        print(f"[step6] ⏭ 跳过新型材料 sheet（用户原话：丢掉不要）")
    n_pc = write_pc_sheet(wb, data)
    write_meta_sheet(wb, data, city, period, year, data_month, cycle_type)

    wb.save(out_path)
    print(f"[step6] ✅ 写出: {out_path}")
    print(f"[step6]   Sheet 1 市场信息价: {n_market} 行")
    print(f"[step6]   Sheet 2 {sheet2_name}: {n_industry} 行")
    if n_pc > 0:
        print(f"[step6]   Sheet 3 PC构件基础价格: {n_pc} 行")
    return out_path


def _load_sheet_config(city):
    """读 6_配置/城市模板/{city}.yaml 的 sheet_config（2026-07-29 增）

    存在 → 用行业推荐（6 列）
    不存在 → 用新型材料（11 列，成都/湖北原结构）
    """
    try:
        import yaml
        candidates = [
            ROOT / "6_配置" / "城市模板" / f"{city}.yaml",
            ROOT / "6_配置" / "城市模板" / f"{city}市.yaml",
        ]
        for path in candidates:
            if path.exists():
                with path.open(encoding="utf-8") as f:
                    yaml_data = yaml.safe_load(f)
                if yaml_data.get("sheet_config"):
                    return yaml_data["sheet_config"]
                return None
        return None
    except Exception:
        return None


# 2026-07-31 改：加第 7 个位置参数 pdf_stem
if __name__ == "__main__":
    if len(sys.argv) < 8:
        print("用法: python step6_output.py <clean_json> <city> <period> <year> <data_month> <cycle_type> <pdf_stem>", file=sys.stderr)
        sys.exit(1)
    write_xlsx(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6], sys.argv[7])
