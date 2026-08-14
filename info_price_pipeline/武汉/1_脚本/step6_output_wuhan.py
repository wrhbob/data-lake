"""step6_output_wuhan.py - 武汉专属 step6 xlsx writer (2026-08-04 建)

差异 vs hb step6_output_hb.py:
  - 3 sheet: 标准材料 (A 7列) + 装配式 (B 7列含备注) + 节能门窗 (C 10列)
  - 按 schema 写不同列结构
  - hb 是「湖北省XX市」前缀, 武汉不需要 (武汉就是「武汉」)

输入: extract.json (step3 输出, [{sheet, headers, rows}, ...])
输出: 4_输出/<pdf_stem>.xlsx
"""
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("[step6 武汉] ❌ 缺 openpyxl, 请 pip install openpyxl", file=sys.stderr)
    raise

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import ROOT, city_to_code


def _write_header(ws, headers):
    """写表头 + 样式"""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")


def _set_widths(ws, widths):
    """设列宽 (widths: list of int)"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w


def _write_rows_a(ws, rows):
    """schema A: 序号|名称|规格型号|单位|含税价|除税价|备注"""
    n = 0
    for r in rows:
        ws.append([
            r.get("seq", ""),
            r.get("name", ""),
            r.get("spec", ""),
            r.get("unit", ""),
            r.get("incl", ""),
            r.get("excl", ""),
            r.get("remark", ""),
        ])
        n += 1
    return n


def _write_rows_b(ws, rows):
    """schema B: 序号|名称|含钢量(kg)|单位|含税价|除税价|备注"""
    n = 0
    for r in rows:
        ws.append([
            r.get("seq", ""),
            r.get("name", ""),
            r.get("spec", ""),  # schema B 字段复用 spec (= 含钢量)
            r.get("unit", ""),
            r.get("incl", ""),
            r.get("excl", ""),
            r.get("remark", ""),
        ])
        n += 1
    return n


def _write_rows_c(ws, rows):
    """schema C: 序号|名称|配用玻璃|传热系数|单位|含税价|除税价|安装费|检测费|备注"""
    n = 0
    for r in rows:
        ws.append([
            r.get("seq", ""),
            r.get("name", ""),
            r.get("glass", ""),
            r.get("k_factor", ""),
            r.get("unit", ""),
            r.get("incl", ""),
            r.get("excl", ""),
            r.get("install_fee", ""),
            r.get("inspect_fee", ""),
            r.get("remark", ""),
        ])
        n += 1
    return n


# Sheet 名 → (写行函数, 列宽)
_SHEET_WRITERS = {
    "标准材料": (_write_rows_a, [8, 32, 28, 8, 12, 12, 22]),
    "装配式": (_write_rows_b, [8, 32, 14, 8, 12, 12, 22]),
    "节能门窗": (_write_rows_c, [8, 32, 18, 12, 8, 12, 12, 10, 10, 22]),
}

_SHEET_ORDER = ["标准材料", "装配式", "节能门窗"]


def _write_meta_sheet(wb, city, period, year, data_month, cycle_type, pdf_stem):
    """元信息 sheet"""
    ws = wb.create_sheet("元信息")
    ws.append(["字段", "值"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    rows = [
        ("城市", city),
        ("期号", period),
        ("年份", year),
        ("数据月份", data_month),
        ("周期类型", cycle_type),
        ("PDF 文件", pdf_stem),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("schema", "武汉 3 sheet (标准材料 7列 + 装配式 7列含备注 + 节能门窗 10列)"),
    ]
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    # 移到最后
    wb.move_sheet(ws, offset=len(wb.sheetnames) - 1)


def write_xlsx_wuhan(extract_json, city, period, year, data_month, cycle_type, pdf_stem):
    """武汉专属 xlsx writer"""
    code = city_to_code(city)
    out_dir = ROOT / "4_输出"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{pdf_stem}.xlsx"
    # 冲突重试
    suffix = 1
    while (out_dir / f"~${out_path.name}").exists() or suffix > 50:
        suffix += 1
        out_path = out_dir / f"{pdf_stem}_v{suffix}.xlsx"
        if suffix > 50:
            raise RuntimeError("xlsx 输出文件名冲突超过 50 次")

    with open(extract_json, encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 按 _SHEET_ORDER 顺序写 3 sheet
    n_total = 0
    for sheet_name in _SHEET_ORDER:
        # 找对应 sheet 数据
        sheet_data = next((t for t in data if t.get("sheet") == sheet_name), None)
        if not sheet_data:
            continue
        ws = wb.create_sheet(sheet_name)
        _write_header(ws, sheet_data["headers"])
        writer, widths = _SHEET_WRITERS[sheet_name]
        n = writer(ws, sheet_data["rows"])
        _set_widths(ws, widths)
        print(f"[step6 武汉]   Sheet '{sheet_name}': {n} 行")
        n_total += n

    _write_meta_sheet(wb, city, period, year, data_month, cycle_type, pdf_stem)

    wb.save(out_path)
    print(f"[step6 武汉] ✅ 写出: {out_path}")
    print(f"[step6 武汉]   总行数: {n_total}")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 7:
        print("用法: python step6_output_wuhan.py <extract_json> <city> <period> <year> <data_month> <cycle_type> <pdf_stem>", file=sys.stderr)
        sys.exit(1)
    write_xlsx_wuhan(
        sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4],
        sys.argv[5], sys.argv[6], sys.argv[7]
    )