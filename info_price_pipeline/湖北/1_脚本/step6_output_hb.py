"""step6_output_hb.py - 湖北专属 step6 xlsx writer (2026-08-04 建)

差异 vs 通用 step6_output.py:
  - Sheet 1 "MARKET" = 7 列 hb 垂直展开 (序号|城市|名称|规格|单位|含税价|除税价)
  - Sheet 2 "PC" = 6 列 hb 装配式武汉 only (序号|名称|含钢量|单位|含税价|除税价)
  - 不写 "新型材料" sheet (hb 没有 NM)
  - 不写 "原始数据" 列 (hb 不需要)

输入: clean.json (step4 输出, [{caption, section, rows, page}, ...])
输出: 4_输出/<pdf_stem>.xlsx
"""
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("[step6_hb] ❌ 缺 openpyxl, 请 pip install openpyxl", file=sys.stderr)
    raise

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import ROOT, city_to_code


def _write_meta_sheet(wb, city, period, year, data_month, cycle_type, pdf_stem):
    """元信息 sheet"""
    ws = wb.create_sheet("元信息", 0)  # 第 1 个
    ws.append(["字段", "值"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    rows = [
        ("城市", city),
        ("期号", period),
        ("年份", year),
        ("数据月份", data_month),
        ("周期类型", cycle_type),
        ("PDF 文件", pdf_stem),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("schema", "湖北 7 列垂直展开 (MARKET) + 6 列武汉装配式 (PC)"),
    ]
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    # 移到最后
    wb.move_sheet(ws, offset=len(wb.sheetnames) - 1)


def write_market_sheet_hb(wb, data):
    """Sheet 1: MARKET (7 列 hb 垂直)

    2026-08-04 改: 城市列拼湖北省前缀 (HB 省级期刊, 「城市」列显示「湖北省武汉市」)
    """
    ws = wb.create_sheet("市场信息价", 0)
    headers = ["序号", "城市", "名称", "规格", "单位", "含税价(元)", "除税价(元)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
    n_rows = 0
    for table in data:
        if table.get("section") != "MARKET":
            continue
        for row in table.get("rows", []):
            city = row.get("city", row.get("城市", ""))
            city_full = f"湖北省{city}" if city and not str(city).startswith("湖北省") else city
            ws.append([
                row.get("seq", row.get("序号", "")),
                city_full,
                row.get("name", row.get("名称", "")),
                row.get("spec", row.get("规格", "")),
                row.get("unit", row.get("单位", "")),
                row.get("incl", row.get("含税价", row.get("含税价(元)", ""))),
                row.get("excl", row.get("除税价", row.get("除税价(元)", ""))),
            ])
            n_rows += 1
    widths = [8, 16, 30, 25, 8, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    return n_rows


def write_pc_sheet_hb(wb, data):
    """Sheet 2: PC 武汉装配式 (6 列 hb)"""
    n_rows_total = sum(
        len(t.get("rows", [])) for t in data if t.get("section") == "PC"
    )
    if n_rows_total == 0:
        return 0
    ws = wb.create_sheet("装配式", 1)
    headers = ["序号", "名称", "含钢量(kg)", "单位", "含税价(元)", "除税价(元)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
    n_rows = 0
    for table in data:
        if table.get("section") != "PC":
            continue
        for row in table.get("rows", []):
            ws.append([
                row.get("seq", row.get("序号", "")),
                row.get("name", row.get("名称", "")),
                row.get("spec", row.get("含钢量(kg)", row.get("steel_kg", ""))),
                row.get("unit", row.get("单位", "")),
                row.get("incl", row.get("含税价", row.get("含税价(元)", ""))),
                row.get("excl", row.get("除税价", row.get("除税价(元)", ""))),
            ])
            n_rows += 1
    widths = [8, 30, 12, 8, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    return n_rows


def write_xlsx_hb(clean_json, city, period, year, data_month, cycle_type, pdf_stem):
    """hb 专属 xlsx writer"""
    code = city_to_code(city)
    out_dir = ROOT / "4_输出"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{pdf_stem}.xlsx"
    # 冲突重试
    suffix = 1
    while (out_dir / f"~${out_path.name}").exists() or suffix > 50:
        suffix += 1
        out_path = out_dir / f"{pdf_stem}_v{suffix}.xlsx"
        if suffix > 50:
            raise RuntimeError("xlsx 输出文件名冲突超过 50 次")

    with open(clean_json, encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    n_market = write_market_sheet_hb(wb, data)
    n_pc = write_pc_sheet_hb(wb, data)
    _write_meta_sheet(wb, city, period, year, data_month, cycle_type, pdf_stem)

    wb.save(out_path)
    print(f"[step6_hb] ✅ 写出: {out_path}")
    print(f"[step6_hb]   Sheet '市场信息价' (MARKET 7列): {n_market} 行")
    if n_pc > 0:
        print(f"[step6_hb]   Sheet '装配式' (PC 6列): {n_pc} 行")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 7:
        print("用法: python step6_output_hb.py <clean_json> <city> <period> <year> <data_month> <cycle_type> <pdf_stem>")
        sys.exit(1)
    write_xlsx_hb(
        sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4],
        sys.argv[5], sys.argv[6], sys.argv[7]
    )
