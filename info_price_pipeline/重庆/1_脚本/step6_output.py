"""step6_output.py - 写 xlsx（重庆 06 重构版，2026-07-29）

目的：
  clean.json → 4_输出/重庆_{year}年{period}期_市场信息价.xlsx

4 个 Sheet（重庆 06 PDF 章节结构）：
  Sheet 1: 市场信息价（区县材料价）
    列: 序号 | 市/区 | 材料名称 | 规格及型号 | 单位 | 不含税价(元) | 备注 | 原始数据
  Sheet 2: 苗木（专项，市/区="苗木" 单独出 sheet）
    列: 序号 | 市/区 | 材料名称 | 规格及型号 | 单位 | 不含税价(元) | 备注 | 原始数据
  Sheet 3: 行业推荐（协会）
    列: 序号 | 市/区 | 材料名称 | 规格型号 | 单位 | 不含税价(元) | 备注 | 填报人 | 联系方式 | 原始数据
  Sheet 4: 元数据
    城市 / 出版期 / 数据月份 / 周期类型 / 提取时间 / Sheet 行数
"""
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import ROOT, city_to_code


# ============================================================
# Sheet 1: 市场信息价
# ============================================================

def write_market_sheet(wb, data):
    """Sheet 1: 市场信息价（区县 6 列 + 备注 + 市/区）

    2026-07-29 改：
      - 列序：序号 | 市/区 | 材料名称 | ...（市/区移到序号后）
      - 苗木行（市/区="苗木"）跳过，由 write_seedling_sheet 单独出 sheet
    2026-07-29 改：删「原始数据」列（用户原话"能不能把原始数据内一列删掉"）
      + 加「含税价(元)」列（用户原话"每个 sheet 里面必须加上含税价格，内容没有就空着"）
    """
    ws = wb.create_sheet("市场信息价", 0)
    headers = ["序号", "市/区", "材料名称", "规格及型号", "单位",
               "不含税价(元)", "含税价(元)", "备注"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    for table in data:
        if table.get("section") != "MARKET":
            continue
        for row in table.get("rows", []):
            # 2026-07-29 改：苗木行市/区="苗木" → 跳过（由 write_seedling_sheet 单独写）
            if row.get("市/区", "") == "苗木":
                continue
            ws.append([
                row.get("序号", ""),
                row.get("市/区", ""),
                row.get("材料名称", ""),
                row.get("规格及型号", ""),
                row.get("单位", ""),
                row.get("不含税价(元)", ""),
                row.get("含税价(元)", ""),     # 用户要求加的列（PDF 无数据则空）
                row.get("备注", ""),
            ])
            n_rows += 1

    widths = [6, 18, 25, 20, 8, 12, 12, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 2: 苗木（专项，市/区="苗木" 单独出 sheet）
# ============================================================

def write_seedling_sheet(wb, data):
    """Sheet 2: 苗木（p91 单独页 192 行苗木表，单出 sheet）

    2026-07-29 改：4 列规格独立（用户原话："PDF 92 四个也必须要"）
    2026-07-29 改：删「原始数据」列 + 加「含税价(元)」列
    2026-07-29 改：删「市/区」列（苗木不是城市，不属于任何市/区）
    2026-07-29 改：加「规格(cm)」列（草本/地被的材质描述，比如结缕草"草块"）
    2026-07-30 P2-5 改：加「地径」列（用户原话："表头不一样肯定要单独列呀"）
    列序：序号 | 材料名称 | 规格(cm) | 株高 | 胸径 | **地径** | 冠幅 | 分枝高 | 单位 | 不含税价(元) | 含税价(元) | 备注
    """
    ws = wb.create_sheet("苗木", 1)
    headers = ["序号", "材料名称", "规格(cm)", "株高", "胸径", "地径", "冠幅", "分枝高",
               "单位", "不含税价(元)", "含税价(元)", "备注"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")  # 绿色标识苗木
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    seq = 1
    for table in data:
        if table.get("section") != "MARKET":
            continue
        for row in table.get("rows", []):
            if row.get("市/区", "") != "苗木":
                continue
            ws.append([
                seq,
                row.get("材料名称", ""),
                row.get("规格及型号", ""),
                row.get("株高", ""),
                row.get("胸径", ""),
                row.get("地径", ""),   # 2026-07-30 P2-5：花灌木行有值，乔木/棕榈/地被空
                row.get("冠幅", ""),
                row.get("分枝高", ""),
                row.get("单位", ""),
                row.get("不含税价(元)", ""),
                row.get("含税价(元)", ""),
                row.get("备注", ""),
            ])
            seq += 1
            n_rows += 1

    widths = [6, 25, 12, 12, 12, 12, 12, 8, 12, 12, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 3: 行业推荐
# ============================================================

def write_industry_sheet(wb, data):
    """Sheet 3: 行业推荐（协会 6 列 + 备注 + 填报人/联系方式 + 市/区）

    2026-07-29 改：列序改为 序号 | 市/区 | ...（市/区移到序号后）
    2026-07-29 改：删「原始数据」列 + 加「含税价(元)」列
    """
    ws = wb.create_sheet("行业推荐", 2)
    headers = ["序号", "市/区", "材料名称", "规格型号", "单位",
               "不含税价(元)", "含税价(元)", "备注", "填报人", "联系方式"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    n_rows = 0
    for table in data:
        if table.get("section") != "NEW_MATERIAL":
            continue
        for row in table.get("rows", []):
            ws.append([
                row.get("序号", ""),
                row.get("市/区", ""),
                row.get("材料名称", ""),
                row.get("规格型号", ""),
                row.get("单位", ""),
                row.get("不含税价(元)", ""),
                row.get("含税价(元)", ""),
                row.get("备注", ""),
                row.get("填报人", ""),
                row.get("联系方式", ""),
            ])
            n_rows += 1

    widths = [6, 18, 25, 20, 8, 12, 12, 30, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return n_rows


# ============================================================
# Sheet 4: 元数据
# ============================================================

def write_meta_sheet(wb, data, city, period, year, data_month, cycle_type):
    """Sheet 4: 元数据"""
    ws = wb.create_sheet("元数据", 3)
    ws.append(["字段", "值"])
    ws.append(["城市", city])
    ws.append(["年份", year])
    ws.append(["出版期", period])
    ws.append(["数据月份", data_month])
    ws.append(["周期类型", cycle_type])
    ws.append(["提取时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    n_market = sum(1 for t in data if t.get("section") == "MARKET"
                   for r in t.get("rows", []) if r.get("市/区", "") != "苗木")
    n_seedling = sum(1 for t in data if t.get("section") == "MARKET"
                     for r in t.get("rows", []) if r.get("市/区", "") == "苗木")
    n_new = sum(len(t["rows"]) for t in data if t.get("section") == "NEW_MATERIAL")
    ws.append(["Sheet 1 市场信息价行数", n_market])
    ws.append(["Sheet 2 苗木行数", n_seedling])
    ws.append(["Sheet 3 行业推荐行数", n_new])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")


# ============================================================
# 主流程
# ============================================================

def write_xlsx(clean_json, city, period, year, data_month, cycle_type):
    """clean.json → xlsx（重庆 3 sheet 无 PC）"""
    code = city_to_code(city)
    out_dir = ROOT / "4_输出"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{city}_{year}年{period}期_市场信息价.xlsx"
    lock_path = out_dir / f"~${out_path.name}"
    # 检测锁文件 OR xlsx 本身被占用 → 用 _v2 后缀
    if lock_path.exists() or (out_path.exists() and not lock_path.exists()):
        # 已经有旧 xlsx（无论是否锁）→ 用 _v2
        version = 2
        while True:
            alt = out_dir / f"{city}_{year}年{period}期_市场信息价_v{version}.xlsx"
            if not alt.exists():
                out_path = alt
                break
            version += 1
        print(f"[step6] ⚠️ 原文件被占用/已存在，改输出到 {out_path.name}")

    with open(clean_json, encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    n_market = write_market_sheet(wb, data)
    n_seedling = write_seedling_sheet(wb, data)
    n_industry = write_industry_sheet(wb, data)
    write_meta_sheet(wb, data, city, period, year, data_month, cycle_type)

    wb.save(out_path)
    print(f"[step6] ✅ 写出: {out_path}")
    print(f"[step6]   Sheet 1 市场信息价: {n_market} 行")
    print(f"[step6]   Sheet 2 苗木: {n_seedling} 行")
    print(f"[step6]   Sheet 3 行业推荐: {n_industry} 行")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 7:
        print("用法: python step6_output.py <clean_json> <city> <period> <year> <data_month> <cycle_type>",
              file=sys.stderr)
        sys.exit(1)
    write_xlsx(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6])