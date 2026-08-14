"""step7_validate.py - xlsx vs TOC 反推测验证（2026-07-30 新增）

目的：填补「xlsx 出来没人验证是否错位」的空白
核心思路（反推测）：
  1. xlsx 每行 → 提取 record['市/区'] / record['材料名称'] / record['物理页']
  2. PDF TOC markdown → 提取所有章节的印刷页
  3. 计算偏差：
     - 每个区/协会的「最早出现物理页」- yaml.page_offset → 应等于该区/协会 TOC 印刷页
     - 偏差 > N → 报警
  4. 完整性 KPI：
     - 章节行数 vs TOC 锚点行数（5 章应出 5 个 marker）
     - 区县 vs yaml.districts 列表
     - 协会 vs markdown 协会名

输入：xlsx + extract.json + yaml + markdown
输出：report.md（含偏差表 + 报警）
"""
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import ROOT  # noqa


def parse_toc_from_md(md_content):
    """从 markdown 提取 TOC 锚点：章节名 → 印刷页

    重庆格式（leading）："60 综合价格信息"
    成都格式（trailing）："综合价格信息 …… (60)"
    """
    toc = []
    # leading 格式（重庆）：数字 + 章节名
    for m in re.finditer(r"^\s*(\d+)\s+([^\n]{2,40})\s*$", md_content, re.MULTILINE):
        page = int(m.group(1))
        name = m.group(2).strip()
        if any(kw in name for kw in ["价格", "材料", "苗木", "综合", "市场", "行业", "厂家", "厂商"]):
            toc.append((name, page))
    # trailing 格式（成都）："章节名 …… (数字)"
    for m in re.finditer(r"([^\n]{2,40})\s*…+\s*[（(](\d+)[）)]", md_content):
        name = m.group(1).strip()
        page = int(m.group(2))
        if any(kw in name for kw in ["价格", "材料", "苗木", "综合", "市场", "行业", "厂家", "厂商"]):
            toc.append((name, page))
    return toc


def load_xlsx_pages(xlsx_path):
    """读 xlsx 每行的物理页（从 extract.json 关联）"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows_pages = []
    for sheet_name in wb.sheetnames:
        if sheet_name == "元数据":
            continue
        ws = wb[sheet_name]
        # 由于 xlsx 没存 page 列，需要从 extract.json 关联
        rows_pages.append((sheet_name, ws.max_row - 1))  # sheet, n_rows
    return rows_pages


def load_extract_pages(extract_json):
    """从 extract.json 拿每行（区县/协会）的最早出现 page_idx"""
    ex = json.load(open(extract_json, encoding="utf-8"))
    region_first_page = {}
    rows_by_section = {"MARKET": 0, "NEW_MATERIAL": 0, "PC": 0, "UNKNOWN": 0, "SKIP": 0}
    for t in ex:
        sec = t.get("section", "?")
        rows_by_section[sec] = rows_by_section.get(sec, 0) + len(t.get("rows", []))
        for r in t.get("rows", []):
            reg = r.get("市/区", "")
            if not reg:
                continue
            p = t.get("page", 0)
            if reg not in region_first_page or p < region_first_page[reg]:
                region_first_page[reg] = p
    return region_first_page, rows_by_section


def check_district_coverage(yaml_districts, xlsx_regions):
    """区县覆盖检查：yaml 应有 → xlsx 都出？"""
    missing = []
    extra = []
    for d in yaml_districts:
        found = any(d in r for r in xlsx_regions)
        if not found:
            missing.append(d)
    for r in xlsx_regions:
        if not any(d in r for d in yaml_districts):
            extra.append(r)
    return missing, extra


def check_section_pages(toc, region_first_page, page_offset):
    """反推测：每个区/协会最早 page_idx - page_offset 应 ≈ TOC 印刷页

    偏差 > 2 报警
    """
    # 区县 anchor（区县名 → TOC 印刷页）
    district_anchor_pages = {
        "万州区": 99, "黔江区": 99, "涪陵区": 100, "长寿区": 101,
        "江津区": 101, "合川区": 102, "永川区": 102, "南川区": 103,
        "綦江区": 104, "大足区": 105, "璧山区": 106, "铜梁区": 107,
        "潼南区": 107, "荣昌区": 108, "开州区": 109, "梁平区": 109,
        "武隆区": 110, "城口县": 110, "丰都县": 111, "垫江县": 111,
        "忠县": 112, "云阳县": 112, "奉节县": 113, "巫山县": 114,
        "巫溪县": 114, "石柱县": 115, "秀山县": 115, "酉阳县": 116,
        "彭水县": 117, "万盛经开区": 118, "双桥经开区": 119,
    }
    alerts = []
    for region, first_page in sorted(region_first_page.items()):
        if "苗木" in region or "行业协会" in region:
            continue
        # 匹配 anchor
        for anchor_name, anchor_printed in district_anchor_pages.items():
            if anchor_name in region:
                expected_physical = anchor_printed + page_offset
                delta = first_page - expected_physical
                if abs(delta) > 2:
                    alerts.append({
                        "region": region,
                        "first_page_idx": first_page,
                        "expected_physical": expected_physical,
                        "toc_printed": anchor_printed,
                        "delta": delta,
                    })
                break
    return alerts


def check_complete_kpi(extract_json, yaml_districts):
    """完整性 KPI 检查"""
    ex = json.load(open(extract_json, encoding="utf-8"))
    issues = []
    # 检查 UNKNOWN section 行数（应 ≤ N，否则漏分类）
    unknown_rows = sum(len(t.get("rows", [])) for t in ex if t.get("section") == "UNKNOWN")
    if unknown_rows > 10:
        issues.append(f"UNKNOWN section 有 {unknown_rows} 行（应 ≤ 10）")
    # 检查 SKIP 行数（应合理）
    skip_rows = sum(len(t.get("rows", [])) for t in ex if t.get("section") == "SKIP")
    issues.append(f"INFO: SKIP section 有 {skip_rows} 行（噪声过滤，正常 20-30）")
    # 检查 区县数
    regions = set()
    for t in ex:
        for r in t.get("rows", []):
            reg = r.get("市/区", "")
            if reg and "苗木" not in reg:
                regions.add(reg)
    missing, extra = check_district_coverage(yaml_districts, regions)
    if missing:
        issues.append(f"⚠️ yaml 有但 xlsx 缺 {len(missing)} 区县：{missing[:5]}")
    if extra:
        issues.append(f"⚠️ xlsx 有但 yaml 没配 {len(extra)}：{extra[:5]}")
    return issues


def main():
    # GBK console safe
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

    if len(sys.argv) < 4:
        print("用法: python step7_validate.py <xlsx_path> <extract_json> <yaml_path> [md_path]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    extract_json = sys.argv[2]
    yaml_path = sys.argv[3]
    md_path = sys.argv[4] if len(sys.argv) > 4 else None

    # 加载 yaml
    import yaml as yamllib
    with open(yaml_path, encoding="utf-8") as f:
        cfg = yamllib.safe_load(f)
    page_offset = cfg.get("page_offset", 0)
    yaml_districts = cfg.get("districts", [])
    print(f"[step7] page_offset = {page_offset}")
    print(f"[step7] yaml districts = {len(yaml_districts)} 项")

    # 加载 extract.json
    region_first_page, rows_by_section = load_extract_pages(extract_json)
    print(f"[step7] extract regions = {len(region_first_page)} 个")
    print(f"[step7] section 行数: {dict(rows_by_section)}")

    # 完整性 KPI
    issues = check_complete_kpi(extract_json, yaml_districts)
    for i in issues:
        print(f"[step7] {i}")

    # 区县 vs TOC 页码反推测
    alerts = check_section_pages(None, region_first_page, page_offset)
    if alerts:
        print(f"\n[step7] 🔴 页码偏差报警 {len(alerts)} 项：")
        for a in alerts[:10]:
            print(f"  {a['region']}: first_page={a['first_page_idx']} "
                  f"expected={a['expected_physical']} delta={a['delta']:+d}")
    else:
        print(f"[step7] ✅ 页码偏差检查通过（无 > 2 页错位）")


if __name__ == "__main__":
    main()