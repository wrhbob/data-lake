"""PDF -> xlsx parse pipeline dispatcher.

Multi-city registry model:
- ParseHandler dataclass encapsulates per-city config (paths, filename regex,
  arg builder, output template).
- CITY_HANDLERS is the registry; v1 supported 成都, 2026-08-04 added 重庆.
- detect_city_from_filename() scans all handlers and returns first match.
- A process-local _JOBS dict + _SEMAPHORE gate concurrent parses (max 2).
- A daemon thread per parse streams subprocess stdout into a bounded deque;
  the SSE generator drains the deque and yields 'data: <json>\\n\\n' frames.

Process-local constraint (see also serve.py startup hook):
  _JOBS and _SEMAPHORE are in-memory only. A uvicorn worker restart loses
  live job state; the serve.py reaper marks them failed/SERVER_RESTART in PG.
"""

from __future__ import annotations

import hashlib
import logging
import os
import re
import shutil
from io import BytesIO
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from collections import deque
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Iterator

from sqlalchemy import select

from app.config import Settings
from app.models import Archive, ArchiveFile, FileAsset
from app.schemas import ParseOutputResponse, ParseStreamEvent
from app.storage import ObjectStore

logger = logging.getLogger(__name__)

# ===== Pipeline root (脚本已内迁到 main/info_price_pipeline) =====
# 2026-08-13 分发改造：默认相对 main 仓根目录推导，不依赖机器绝对路径。
# env INFO_PRICE_PIPELINE_BASE 仅作显式覆盖（自定义布局时）。
_REPO_ROOT = Path(__file__).resolve().parent.parent.parent  # main/
PIPELINE_BASE = Path(os.getenv("INFO_PRICE_PIPELINE_BASE", str(_REPO_ROOT / "info_price_pipeline")))
# 2026-08-13: 解析脚本直接用 web 的 python（file-asset 环境，requests/openpyxl/fitz/yaml 齐全），
# 不再维护独立的 --v2-main/.venv（那个缺 yaml 导致 step2 崩溃）。
PIPELINE_VENV_PYTHON = Path(sys.executable)

MAX_CONCURRENT_PARSE = 2
PARSE_TIMEOUT_SECONDS = 1800  # 30 minutes per parse
_EVENT_BUFFER_MAXLEN = 5000   # ~one PDF's worth of stdout lines

# 4 chengdu filename regex variants (95% hit on 199 real PDFs, see naming_check2.txt)
_CHENGDU_FILENAME_PATTERNS = [
    re.compile(r"成都(?:市)?(\d{4})年第?(\d{1,2})期"),
    re.compile(r"成都(?:市)?信息价(\d{4})[\-年](\d{1,2})(?:期)?"),
    re.compile(r"成都(?:市)?(\d{4})年(\d{1,2})期"),
    re.compile(r"成都(\d{4})年(\d{1,2})期"),
    # "年份+成都" 顺序（用户后续命名习惯，2026年新格式）
    re.compile(r"(\d{4})年成都(?:市)?信息价(\d{1,2})期"),
    re.compile(r"(\d{4})年成都(?:市)?(\d{1,2})期"),
    # 2026-08-04 新格式：「2026年成都市工程造价信息价第3期」之类，年份+成都市+中间描述+第N期
    # 中间允许夹任意非数字字符（"成都市工程造价信息价第" 这种描述）
    re.compile(r"(\d{4})年成都(?:市)?[^0-9\n]{0,20}?第?(\d{1,2})期"),
]


def _extract_chengdu_params(filename: str) -> dict | None:
    """Return {'year': int, 'period': str} if filename matches any of 4 chengdu patterns."""
    base = Path(filename).stem  # strip extension
    for pattern in _CHENGDU_FILENAME_PATTERNS:
        match = pattern.search(base)
        if match:
            return {"year": int(match.group(1)), "period": str(int(match.group(2)))}
    return None


# 2026-08-04 新增：重庆文件名正则（类比成都模板）
# 2026-08-04 修订：实际命名是"重庆工程造价2026-6期"（无"年"、夹"工程造价"），加宽容版本
# 2026-08-04 再修订：DB 里很多文件名 "期" 被 UTF-8→UTF-16 双解码坏成 \udc9f/\udc8f 等代理对，
# 严格正则匹配失败 → parse_supported=False → ▶️ 按钮被禁用。加容错版：期可选 + 年期后跟非数字。
_CHONGQING_FILENAME_PATTERNS = [
    re.compile(r"重庆(?:市)?[^0-9\s]{0,12}?(\d{4})[\-/年]?(\d{1,2})期"),
    re.compile(r"(\d{4})年重庆(?:市)?[^0-9\s]{0,12}?(\d{1,2})期"),
    # 严格版本（用户后续命名可能规范化）
    re.compile(r"重庆(?:市)?(\d{4})年第?(\d{1,2})期"),
    re.compile(r"(\d{4})年重庆(?:市)?(\d{1,2})期"),
    # 2026-08-04 容错版：处理 DB 中 "期" 被双重解码损坏的文件名（如 "水印-重庆工程造价2026-6期6-17.pdf"
    # 里的 期 变成 \udc9f）。允许 year+period 后跟非数字字符（兼容 "-17.pdf" 之类的日期/页码后缀）
    re.compile(r"重庆[^0-9\s]{0,12}?(\d{4})[\-/年]?(\d{1,2})(?:期|\D|$)"),
]


def _extract_chongqing_params(filename: str) -> dict | None:
    """Return {'year': int, 'period': str} if filename matches any chongqing pattern."""
    base = Path(filename).stem
    for pattern in _CHONGQING_FILENAME_PATTERNS:
        match = pattern.search(base)
        if match:
            return {"year": int(match.group(1)), "period": str(int(match.group(2)))}
    return None


@dataclass(frozen=True)
class ParseHandler:
    """Per-city configuration for the PDF -> xlsx pipeline.

    extract_params runs against the PDF filename to recover year + period.
    pipeline_root / pipeline_python / pipeline_output_dir pin per-city paths
    (chengdu/chongqing share the .venv at 信息价提取/.venv, so both use the
    same python.exe but different run.py roots + 4_输出/ dirs).
    pdf_as_flag: True means pass PDF as "--pdf <path>", False means positional.
    passes_python_flag: True means we add "--python <venv>". 成都 needs it because
      chengdu/run.py uses an internal venv default that's broken; cq/bj run.py
      already pick the venv via _venv_python() inside themselves.
    arg_builder(year, period, month) -> tuple[str, ...]: per-city extras
    appended after --year / --period. chengdu needs (--city, "成都", --data-month,
    --cycle-type); chongqing needs (--month, ).
    output_filename_template is the xlsx file name produced by run.py in
    pipeline_output_dir (default: same stem as input PDF; chongqing actually
    emits 重庆市_<year>年<period>期_市场信息价.xlsx so we rely on glob fallback).
    dispatch: 2026-08-07 新增 — "guangzhou_special" 走 _run_guangzhou_logic(广州 run.py
      不调 step1,要求 result.json 已存在;web UI 端先 subprocess 调 mineru-pdf-parse 写
      result.json,再 Popen 广州 run.py --skip-step1)。其他 5 城 None 走默认 _run_parse_job。
    """

    city_code: str
    extract_params: Callable[[str], dict | None]
    pipeline_root: Path
    pipeline_python: Path
    pipeline_output_dir: Path
    pdf_as_flag: bool
    passes_python_flag: bool
    arg_builder: Callable[[int, str, int], tuple[str, ...]]
    output_filename_template: str
    dispatch: str | None = None  # 2026-08-07 新增,默认 None 走 5 城逻辑


def _chengdu_arg_builder(year: int, period: str, month: int) -> tuple[str, ...]:
    return ("--city", "成都", "--data-month", str(month), "--cycle-type", "月刊")


def _chongqing_arg_builder(year: int, period: str, month: int) -> tuple[str, ...]:
    # 2026-08-04 注：run.py 默认就识别重庆，不传 --city；--month 必传，--cycle-type 不支持
    return ("--month", str(month))


# 2026-08-05 新增：北京（v20 per-city 模板，月刊格式）
# 命名规律：「2026年06月北京工程造价信息.pdf」之类，年+月+城市开头
_BEIJING_FILENAME_PATTERNS = [
    re.compile(r"北京(?:市)?(\d{4})年(\d{1,2})月"),
    # 顺序颠倒（用户后续命名可能「年月+城市」放后面）
    re.compile(r"(\d{4})年(\d{1,2})月北京(?:市)?"),
]


def _extract_beijing_params(filename: str) -> dict | None:
    """Return {'year': int, 'period': str} if filename matches any beijing pattern."""
    base = Path(filename).stem  # strip extension
    for pattern in _BEIJING_FILENAME_PATTERNS:
        match = pattern.search(base)
        if match:
            return {"year": int(match.group(1)), "period": str(int(match.group(2)))}
    return None


def _beijing_arg_builder(year: int, period: str, month: int) -> tuple[str, ...]:
    # 2026-08-05 注：北京 run.py 跟重庆一样内部 _venv_python() 自动找 venv
    # --month 必传（数据月份），--period 是出版期（月刊编号），跟 month 一样填 period
    return ("--month", str(month))


# 2026-08-05 新增：武汉（月刊，跳过 step2/4）
# 命名规律：「市招标设计造价中心、市绿建中心关于发布2026年6月武汉市建设工程综合价格信息的通知（武市政价字〔2026〕18号）.pdf」
# 关键 anchor：武汉市建设工程综合价格信息 + (\d{4})年(\d{1,2})月
# 武汉 run.py 内部 _find_python_exe() 自动检测根 .venv；需要传 --city 武汉 --period --year
# data-month/cycle-type 用 run.py 默认值（"06"/"月刊"），不传避免 monthly 偏差
_WUHAN_FILENAME_PATTERNS = [
    re.compile(r"武汉市建设工程综合价格信息.*?(\d{4})年(\d{1,2})月"),
    re.compile(r"(\d{4})年(\d{1,2})月武汉市"),
]


def _extract_wuhan_params(filename: str) -> dict | None:
    """Return {'year': int, 'period': str} if filename matches any wuhan pattern."""
    base = Path(filename).stem  # strip extension
    for pattern in _WUHAN_FILENAME_PATTERNS:
        match = pattern.search(base)
        if match:
            return {"year": int(match.group(1)), "period": str(int(match.group(2)))}
    return None


def _wuhan_arg_builder(year: int, period: str, month: int) -> tuple[str, ...]:
    # 2026-08-05 注：武汉 run.py 跟重庆/北京一样内部 _find_python_exe() 自动检测 venv
    # --city/--period/--year 是必填；--data-month/--cycle-type 用默认值
    return ("--city", "武汉", "--period", period, "--year", str(year))


# 2026-08-05 新增：湖北省级（季刊，全 step1-6）
# 命名规律：filename 是数字 ID (crawler 入湖洗过), anchor 在 title: 「2026年第2期《湖北工程造价管理》」
# 检测走 title fallback（detect_city_from_filename 第二轮），正则匹配 title 格式
# 湖北 run.py 跟武汉一样：pos_pdf + auto_venv + --city 湖北 --period --year
# data-month/cycle-type 用 run.py 默认值（"未知"/"未知"，季刊无月度概念）
_HUBEI_TITLE_PATTERNS = [
    re.compile(r"(\d{4})年第(\d{1,2})期[《<]湖北工程造价管理[》>]"),
]


def _extract_hubei_params(filename: str, title: str | None = None) -> dict | None:
    """湖北省级 archive anchor 在 title (PDF filename 被 crawler 洗成数字 ID)。
    detect_city_from_filename 第二轮把 title 当 filename 传进来匹配。
    """
    target = title or filename
    for pattern in _HUBEI_TITLE_PATTERNS:
        match = pattern.search(target)
        if match:
            return {"year": int(match.group(1)), "period": str(int(match.group(2)))}
    return None


def _hubei_arg_builder(year: int, period: str, month: int) -> tuple[str, ...]:
    # 湖北 run.py: --city/--period/--year 必填, --data-month/--cycle-type 用默认值("未知"/"未知")
    return ("--city", "湖北", "--period", period, "--year", str(year))


# 2026-08-07 新增：广州（filename 不可识别 — PDF 被 crawler 洗成数字 ID）
# region_code fallback:archive.region_code == "440100" 时命中广州
CITY_REGION_FALLBACK: dict[str, str] = {
    # city_code -> region_code(中国行政区划代码,6 位)
    "广州": "440100",
}


def _extract_guangzhou_params(filename: str) -> dict | None:
    """广州 PDF 文件名是数字 ID(无 city anchor),filename 检测永远返回 None。
    识别靠 region_code fallback (detect_city 第 3 步)。

    2026-08-07 修正:之前返回 sentinel {"year":0,"period":"00"} 会让所有 archive 被误判为广州
    (因为 detect_city_from_filename 第一轮遍历所有 handler,广州永远"命中")。
    必须返回 None,让 detect_city 走到 region_code fallback 才识别广州。
    """
    return None


def _guangzhou_arg_builder(year: int, period: str, month: int) -> tuple[str, ...]:
    """广州 run.py --name <archive_id>.xlsx (step3+step6 默认跑,不支持 --skip-step1)。
    run.py 默认识别广州(在 cwd 含'广州'目录时)。
    实际 dispatch="guangzhou_special" 走 _run_parse_job 硬编码命令,这里只占位不让 dataclass 报错。
    """
    return ()  # year/period placeholder,广州不走这套;真实命令在 _run_parse_job 硬编码


CITY_HANDLERS: dict[str, ParseHandler] = {
    "成都": ParseHandler(
        city_code="成都",
        extract_params=_extract_chengdu_params,
        pipeline_root=PIPELINE_BASE / "成都",                       # 2026-08-10 D2
        pipeline_python=PIPELINE_VENV_PYTHON,
        pipeline_output_dir=PIPELINE_BASE / "成都" / "4_输出",      # 2026-08-10 D2
        pdf_as_flag=False,            # 成都 run.py: position = pdf
        passes_python_flag=True,      # 成都 run.py: 必须显式传 --python
        arg_builder=_chengdu_arg_builder,
        # 2026-07-31 改：xlsx 文件名直接用 PDF 文件名 stem（替换 .pdf → .xlsx）
        output_filename_template="{pdf_stem}.xlsx",
    ),
    # 2026-08-04 新增：重庆（v20 per-city 模板）
    "重庆": ParseHandler(
        city_code="重庆",
        extract_params=_extract_chongqing_params,
        pipeline_root=PIPELINE_BASE / "重庆",                       # 2026-08-10 D2
        pipeline_python=PIPELINE_VENV_PYTHON,
        pipeline_output_dir=PIPELINE_BASE / "重庆" / "4_输出",      # 2026-08-10 D2
        pdf_as_flag=True,             # 重庆 run.py: --pdf <path>
        passes_python_flag=False,     # 重庆 run.py: 内部 _venv_python() 自动找 venv
        arg_builder=_chongqing_arg_builder,
        output_filename_template="{pdf_stem}.xlsx",  # 实际产出 重庆市_2026年06期_市场信息价.xlsx，靠 glob fallback
    ),
    # 2026-08-05 新增：北京（月刊）
    "北京": ParseHandler(
        city_code="北京",
        extract_params=_extract_beijing_params,
        pipeline_root=PIPELINE_BASE / "北京",                       # 2026-08-10 D2
        pipeline_python=PIPELINE_VENV_PYTHON,
        pipeline_output_dir=PIPELINE_BASE / "北京" / "4_输出",      # 2026-08-10 D2
        pdf_as_flag=True,             # 北京 run.py: --pdf <path>
        passes_python_flag=False,     # 北京 run.py: 内部 _venv_python() 自动找 venv
        arg_builder=_beijing_arg_builder,
        output_filename_template="{pdf_stem}.xlsx",
    ),
    # 2026-08-05 新增：武汉（月刊，跳 step2/4）
    "武汉": ParseHandler(
        city_code="武汉",
        extract_params=_extract_wuhan_params,
        pipeline_root=PIPELINE_BASE / "武汉",                       # 2026-08-10 D2
        pipeline_python=PIPELINE_VENV_PYTHON,
        pipeline_output_dir=PIPELINE_BASE / "武汉" / "4_输出",      # 2026-08-10 D2
        pdf_as_flag=False,            # 武汉 run.py: position = pdf
        passes_python_flag=False,     # 武汉 run.py: 内部 _find_python_exe() 自动检测 venv
        arg_builder=_wuhan_arg_builder,
        output_filename_template="{pdf_stem}.xlsx",
    ),
    # 2026-08-05 新增：湖北省级（季刊，全 step1-6）
    "湖北": ParseHandler(
        city_code="湖北",
        extract_params=_extract_hubei_params,
        pipeline_root=PIPELINE_BASE / "湖北",                       # 2026-08-10 D2
        pipeline_python=PIPELINE_VENV_PYTHON,
        pipeline_output_dir=PIPELINE_BASE / "湖北" / "4_输出",      # 2026-08-10 D2
        pdf_as_flag=False,            # 湖北 run.py: position = pdf
        passes_python_flag=False,     # 湖北 run.py: 内部 _find_python_exe() 自动检测 venv
        arg_builder=_hubei_arg_builder,
        output_filename_template="{pdf_stem}.xlsx",
    ),
    # 2026-08-07 新增：广州（PDF 文件名是数字 ID，靠 region_code=440100 识别）
    "广州": ParseHandler(
        city_code="广州",
        extract_params=_extract_guangzhou_params,  # 永远返回 None,实际靠 region_code (D1 修)
        pipeline_root=PIPELINE_BASE / "广州",
        pipeline_python=PIPELINE_VENV_PYTHON,
        pipeline_output_dir=PIPELINE_BASE / "广州" / "4_输出",
        pdf_as_flag=False,            # 广州 run.py 默认就跳过 step1,不接 PDF
        passes_python_flag=False,     # 广州 run.py 内部 _find_python_exe() 自动检测 venv
        arg_builder=_guangzhou_arg_builder,
        output_filename_template="{archive_id}.xlsx",  # 广州 run.py --name <archive_id>.xlsx
        dispatch="guangzhou_special",  # 标记走 _run_guangzhou_logic(自己调 MinerU step1)
    ),
}


def detect_city_from_filename(filename: str, title: str | None = None) -> tuple[str, dict] | None:
    """Scan CITY_HANDLERS; return (city_code, params) for first match, or None.

    2026-08-05 新增 title fallback：湖北省级 archive 的 PDF filename 被 crawler 洗成
    数字 ID (e.g. 2058808850473635840.pdf), anchor 在 title 「2026年第2期《湖北工程造价管理》」.
    第一轮 filename 不命中时, 第二轮把 title 当 filename 试 (湖北 extractor 内部 regex 匹配 title 格式).
    """
    for city, handler in CITY_HANDLERS.items():
        params = handler.extract_params(filename)
        if params:
            return city, params
    if title:
        for city, handler in CITY_HANDLERS.items():
            params = handler.extract_params(title)
            if params:
                return city, params
    return None


# 2026-08-07 新增：detect 支持 region_code fallback(广州 filename 检测不到)
def detect_city(
    filename: str,
    title: str | None = None,
    region_code: str | None = None,
) -> tuple[str, dict] | None:
    """三级 fallback:
      1. filename 正则(5 城 + 广州 sentinel)
      2. title 正则(湖北)
      3. region_code 匹配 CITY_REGION_FALLBACK(广州 440100)
    返回 (city_code, params),广州 params 是 sentinel {year:0, period:"00"},
    主流程看 city_code=="广州" 跳过 year/period 检查,直接 submit_parse_job。
    """
    result = detect_city_from_filename(filename, title)
    if result:
        return result
    if region_code:
        for city, code in CITY_REGION_FALLBACK.items():
            if code == region_code:
                return city, {"year": 0, "period": "00"}
    return None


# ===== Job registry (process-local) =====
_JOBS: dict[str, "ParseJob"] = {}
_JOBS_LOCK = threading.Lock()
_SEMAPHORE = threading.Semaphore(MAX_CONCURRENT_PARSE)


@dataclass
class ParseJob:
    task_id: str
    archive_id: str
    city_code: str
    proc: subprocess.Popen | None = None
    status: str = "queued"  # queued | running | succeeded | failed | cancelled
    started_at: datetime | None = None
    finished_at: datetime | None = None
    events: deque = field(default_factory=lambda: deque(maxlen=_EVENT_BUFFER_MAXLEN))
    lock: threading.Lock = field(default_factory=threading.Lock)
    seq: int = 0
    dropped_lines: int = 0
    error_code: str | None = None
    error_message: str | None = None
    final_xlsx_key: str | None = None
    temp_dir: Path | None = None
    cancel_event: threading.Event = field(default_factory=threading.Event)
    done_event: threading.Event = field(default_factory=threading.Event)


# ===== Public API =====


def submit_parse_job(
    *,
    session_factory,
    storage: ObjectStore,
    settings: Settings,
    archive_id: str,
    year: int,
    period: str,
    city_code: str = "成都",
) -> str:
    """Validate + start a background parse job. Returns task_id.

    Validation order (漏洞 B/C/J):
      1. archive exists
      2. archive not withdrawn
      3. city_code registered
      4. archive has a primary file
      5. semaphore acquire (1s timeout -> SEMAPHORE_QUEUE_FULL)
    """
    with session_factory() as session:
        archive = session.get(Archive, archive_id)
        if archive is None:
            raise ArchiveNotFound(archive_id)
        if archive.is_withdrawn:
            raise ArchiveWithdrawn(archive_id)
        if city_code not in CITY_HANDLERS:
            raise UnsupportedCity(city_code)
        handler = CITY_HANDLERS[city_code]
        primary_row = session.execute(
            select(ArchiveFile, FileAsset)
            .outerjoin(FileAsset, ArchiveFile.file_id == FileAsset.file_id)
            .where(ArchiveFile.archive_id == archive_id, ArchiveFile.is_primary)
            .order_by(ArchiveFile.sort_order)
            .limit(1)
        ).first()
        if primary_row is None:
            raise ArchiveNoFile(archive_id)
        archive_file, file_asset = primary_row
        if file_asset is None or not file_asset.object_key:
            raise ArchiveNoFile(archive_id)
        # 2026-08-13: 防重复提交——已 queued/running 的 archive 拒绝再次提交（第二道防线，防并发窗口）
        if archive.parse_status in ("queued", "running"):
            raise ParseAlreadyRunning(archive_id)
        # Mark queued (reset previous parse_* fields)
        archive.parse_status = "queued"
        archive.parse_task_id = None
        archive.parse_phase = "queued"
        archive.parse_started_at = datetime.now(timezone.utc)
        archive.parse_finished_at = None
        archive.parse_error_code = None
        archive.parse_error_message = None
        archive.parse_metrics = None
        archive.parse_warnings = None
        archive.final_xlsx_key = None
        archive.candidate_xlsx_key = None
        session.commit()
        file_id = archive_file.file_id
        object_key = file_asset.object_key
        bucket = settings.raw_bucket

    task_id = uuid.uuid4().hex
    # 2026-08-04 改：temp_dir 按城市分前缀，避免成都/重庆同 archive_id 并发时同名冲突
    temp_dir = Path(tempfile.mkdtemp(prefix=f"parse_{city_code}_{archive_id}_"))

    if not _SEMAPHORE.acquire(timeout=1):
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise ParseQueueFull(MAX_CONCURRENT_PARSE)

    job = ParseJob(
        task_id=task_id,
        archive_id=archive_id,
        city_code=city_code,
        temp_dir=temp_dir,
        started_at=datetime.now(timezone.utc),
    )
    with _JOBS_LOCK:
        _JOBS[task_id] = job

    # Update task_id back into DB (now that we know it)
    with session_factory() as session:
        archive = session.get(Archive, archive_id)
        archive.parse_task_id = task_id
        session.commit()

    thread = threading.Thread(
        target=_run_parse_job,
        kwargs={
            "task_id": task_id,
            "archive_id": archive_id,
            "object_key": object_key,
            "bucket": bucket,
            "year": year,
            "period": period,
            "handler": handler,
            "settings": settings,
            "storage": storage,
            "session_factory": session_factory,
        },
        daemon=True,
    )
    thread.start()
    return task_id


def _run_parse_job(
    *,
    task_id: str,
    archive_id: str,
    object_key: str,
    bucket: str,
    year: int,
    period: str,
    handler: ParseHandler,
    settings: Settings,
    storage: ObjectStore,
    session_factory,
) -> None:
    """Daemon thread: download PDF -> run run.py -> upload xlsx.

    Failure paths (handled by raising ParseError subclasses):
      - storage.get_object raises -> PdfDownloadFailed (漏洞 B)
      - handler.pipeline_python missing -> PythonNotFound (漏洞 A)
      - run.py missing under handler.pipeline_root -> RunpyNotFound (漏洞 A)
      - subprocess Popen OSError -> RunpyCrash
      - subprocess exit != 0 -> RunpyCrash
      - subprocess timeout -> ParseTimeout (漏洞 E)
      - cancellation via cancel_event -> _finalize_cancelled (漏洞 E)
      - no xlsx in pipeline_output_dir -> XlsxNotFound

    Cleanup (漏洞 F + plan §3):
      shutil.rmtree(temp_dir) runs in `finally` regardless of outcome.
      _SEMAPHORE.release() always runs.
    """
    job = _JOBS.get(task_id)
    if job is None:
        _SEMAPHORE.release()
        return

    try:
        job.status = "running"
        # ----- Download PDF (漏洞 B) -----
        try:
            pdf_bytes = storage.get_object(bucket, object_key)
        except Exception as exc:
            raise PdfDownloadFailed(f"get_object({bucket},{object_key}) failed: {exc}") from exc
        pdf_sha = hashlib.sha256(pdf_bytes).hexdigest()[:8]
        pdf_path = job.temp_dir / f"{archive_id}_{pdf_sha}.pdf"
        pdf_path.write_bytes(pdf_bytes)

        with session_factory() as session:
            archive = session.get(Archive, archive_id)
            archive.parse_status = "running"
            archive.parse_phase = "downloaded"
            session.commit()

        # ----- Build command -----
        period_str = f"{int(period):02d}"  # 2026-08-03 A 修：补前导零，匹配 cd_<year>_<period>_ocr 缓存目录命名
        # 2026-07-31 改：xlsx 文件名直接用 PDF stem（pdf_path 在 L286 已落盘到 temp_dir）
        # 2026-08-07 改：广州 template 是 {archive_id}.xlsx,所以也传 archive_id
        output_filename = handler.output_filename_template.format(
            pdf_stem=pdf_path.stem,
            archive_id=archive_id,
        )
        creationflags = subprocess.CREATE_NO_WINDOW
        # 2026-08-07 新增：子进程 env 注入 MINERU_API_URL,覆盖 mineru-pdf-parse/scripts/parse_pdf.py 内
        # DEFAULT_API = "http://172.16.20.23:8000"(必修 #3 — 迁移机器跨网段时可换 API)
        subprocess_env = os.environ.copy()
        subprocess_env["PYTHONIOENCODING"] = "utf-8"
        subprocess_env["MINERU_API_URL"] = settings.mineru_api_url

        # 2026-08-07 新增：广州 dispatch="guangzhou_special" 走专属两步流程
        #   step1: subprocess 调 mineru-pdf-parse/scripts/parse_pdf.py 把 PDF → gz_ocr/result.json
        #   step3+6: Popen 广州/1_脚本/run.py --skip-step1 --name <archive_id>.xlsx
        if handler.dispatch == "guangzhou_special":
            gz_ocr_dir = handler.pipeline_root / "3_中间产物" / "gz_ocr"
            gz_ocr_dir.mkdir(parents=True, exist_ok=True)
            mineru_script = handler.pipeline_root / "1_脚本" / "mineru-pdf-parse" / "scripts" / "parse_pdf.py"
            if not mineru_script.exists():
                raise RunpyNotFound(str(mineru_script))
            _emit_event(job, status="running", line=f"[guangzhou] step1 MinerU → {gz_ocr_dir}/result.json")
            try:
                mineru_proc = subprocess.Popen(
                    [str(handler.pipeline_python), str(mineru_script), str(pdf_path), str(gz_ocr_dir),
                     "--api-url", settings.mineru_api_url, "--no-render"],
                    cwd=str(handler.pipeline_root / "1_脚本" / "mineru-pdf-parse"),
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace", bufsize=1,
                    creationflags=creationflags, env=subprocess_env,
                )
            except (FileNotFoundError, OSError) as exc:
                raise RunpyCrash(f"mineru-pdf-parse Popen failed: {exc}") from exc
            job.proc = mineru_proc
            for raw_line in mineru_proc.stdout:
                if job.cancel_event.is_set():
                    break
                line = raw_line.rstrip("\n")
                _emit_event(job, status="running", line=line)
            rc1 = mineru_proc.wait()
            if rc1 != 0:
                raise RunpyCrash(f"mineru-pdf-parse 退出码 {rc1}")
            if not (gz_ocr_dir / "result.json").exists():
                raise XlsxNotFound(str(gz_ocr_dir / "result.json"))
            _emit_event(job, status="running", line="[guangzhou] step1 ✅ result.json 已生成, 跑 step3+step6")
            # 广州 run.py 在 1_脚本/ 子目录,cwd 设到那里
            # 2026-08-07 注:广州 run.py 默认跳过 step1 (压根没 OCR 模块),只跑 step3 + step6。
            # run.py flag 只有 --only / --name / --skip-step3 / --skip-step6,不支持 --skip-step1。
            gz_run = handler.pipeline_root / "1_脚本" / "run.py"
            command = [
                str(handler.pipeline_python), str(gz_run),
                "--name", f"{archive_id}.xlsx",
            ]
            cwd = str(handler.pipeline_root / "1_脚本")
        else:
            # 5 城原逻辑
            data_month = 1  # TODO: 从文件名/请求体提取月份；当前先传 1 占位
            extra_args = handler.arg_builder(year, period_str, data_month)
            if handler.pdf_as_flag:
                pdf_args: tuple[str, ...] = ("--pdf", str(pdf_path))
            else:
                pdf_args = (str(pdf_path),)
            python_args: tuple[str, ...] = ("--python", str(handler.pipeline_python)) if handler.passes_python_flag else ()
            command = [
                str(handler.pipeline_python), "run.py",
                *pdf_args,
                *python_args,
                "--year", str(year),
                "--period", period_str,
                *extra_args,
            ]
            cwd = str(handler.pipeline_root)

        # ----- Pre-flight checks (漏洞 A) -----
        if not handler.pipeline_python.exists():
            raise PythonNotFound(str(handler.pipeline_python))
        # 2026-08-07 改：广州 run.py 在 1_脚本/ 子目录,其他 5 城在 pipeline_root 根
        run_py_path = (
            handler.pipeline_root / "1_脚本" / "run.py"
            if handler.dispatch == "guangzhou_special"
            else handler.pipeline_root / "run.py"
        )
        if not run_py_path.exists():
            raise RunpyNotFound(str(run_py_path))

        # ----- Spawn subprocess -----
        # 2026-08-03 hypothesis 验证：DETACHED_PROCESS | CREATE_NO_WINDOW 在 Python 3.13 + Windows 下
        # 会让子进程立即 rc=0xC0000135 (STATUS_DLL_NOT_FOUND 误报)。只用 CREATE_NO_WINDOW。
        # 2026-08-07 改：cwd 用变量 cwd(广州=1_脚本,其他 5 城=pipeline_root)
        try:
            job.proc = subprocess.Popen(
                command,
                cwd=cwd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
                creationflags=creationflags,
                start_new_session=True,
                env=subprocess_env,
            )
        except (FileNotFoundError, OSError) as exc:
            raise RunpyCrash(f"Popen failed: {exc}") from exc

        # ----- Stream stdout line-by-line -----
        # 2026-08-04 改：用 deque 保留最近 50 行 stderr，失败时附到 RunpyCrash 信息里
        # 否则用户只看到「rc=1」不知道真错（之前成都 06 期就因为这个查不到真因）
        stderr_tail: deque[str] = deque(maxlen=50)
        for raw_line in job.proc.stdout:
            if job.cancel_event.is_set():
                break
            line = raw_line.rstrip("\n")
            stderr_tail.append(line)
            _emit_event(job, status="running", line=line)

        # ----- Wait with timeout (漏洞 E: TIMEOUT) -----
        try:
            rc = job.proc.wait(timeout=PARSE_TIMEOUT_SECONDS)
        except subprocess.TimeoutExpired:
            job.proc.terminate()
            try:
                job.proc.wait(timeout=5)
            except subprocess.TimeoutExpired:
                job.proc.kill()
            raise ParseTimeout(PARSE_TIMEOUT_SECONDS)

        if job.cancel_event.is_set():
            _finalize_cancelled(job, session_factory, archive_id)
            return

        if rc != 0:
            # 2026-08-04 改：失败时附 stderr tail（最近 50 行），让用户知道真错
            tail_text = "\n".join(stderr_tail) if stderr_tail else "(no stdout captured)"
            raise RunpyCrash(f"run.py exited with rc={rc}. Last stdout:\n{tail_text}")

        # ----- Find output xlsx -----
        xlsx_path = handler.pipeline_output_dir / output_filename
        if not xlsx_path.exists():
            candidates = sorted(
                handler.pipeline_output_dir.glob("*.xlsx"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
            if not candidates:
                raise XlsxNotFound(str(handler.pipeline_output_dir))
            xlsx_path = candidates[0]

        xlsx_bytes = xlsx_path.read_bytes()
        xlsx_sha = hashlib.sha256(xlsx_bytes).hexdigest()
        xlsx_object_key = f"info_price_extracts/{archive_id}/{xlsx_sha}.xlsx"
        storage.put_object(
            settings.extract_bucket,
            xlsx_object_key,
            xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # 2026-08-04 新增：用 openpyxl 数 xlsx 真实材料行数 = row_count
        # 算法：每个 sheet 动态找表头（首个「所有 cell 都非空」的行），跳过表头之前所有行
        #       + 跳过表头本身 + 跳过表头后第一个全空行（如果有）；其余非空行 = 数据行
        # 这能处理两种 xlsx 格式：
        #   - 成都：第 1 行「说明：...」（单 cell 非空） + 第 2 行空 + 第 3 行表头（所有 cell 非空）
        #   - 重庆/北京：第 1 行直接表头（所有 cell 非空）
        # 元数据 sheet 在所有城市都是「第 1 行字段名表头 + 数据行」，也能正确处理
        # 失败兜底：openpyxl 抛异常 → row_count = None（前端不显示数字）
        row_count: int | None = None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
            total = 0
            for ws in wb.worksheets:
                # 收集所有行（read_only 模式下只支持遍历一次，必须先 materialize）
                rows = list(ws.iter_rows(values_only=True))
                # 找表头：第一个「所有 cell 都非空」的行
                # 区分：说明行只有 1 个 cell 非空 → 不是表头
                #       表头行所有 cell 都非空 → 是表头
                header_idx = -1
                for i, r in enumerate(rows):
                    if r and all(c is not None and str(c).strip() for c in r):
                        header_idx = i
                        break
                if header_idx < 0:
                    continue  # 空 sheet（无表头）
                # 从表头后开始数数据行
                # 跳过表头后第一个全空行（如果有）
                skipped_first_empty = False
                for r in rows[header_idx + 1:]:
                    has_any = any(c is not None and str(c).strip() for c in r)
                    if not skipped_first_empty and not has_any:
                        skipped_first_empty = True
                        continue
                    if has_any:
                        total += 1
            row_count = total
            wb.close()
        except Exception as exc:  # noqa: BLE001
            logger.warning("row_count 计算失败: %s", exc)
            row_count = None

        # ----- Mark succeeded -----
        finished_at = datetime.now(timezone.utc)
        duration_seconds = (finished_at - job.started_at).total_seconds() if job.started_at else None
        with session_factory() as session:
            archive = session.get(Archive, archive_id)
            archive.parse_status = "succeeded"
            archive.parse_phase = "done"
            archive.parse_finished_at = finished_at
            archive.parse_metrics = {
                "xlsx_size": len(xlsx_bytes),
                "xlsx_sha256": xlsx_sha,
                "duration_seconds": duration_seconds,
                "row_count": row_count,  # 2026-08-04 新增：解析材料行数
            }
            archive.final_xlsx_key = xlsx_object_key
            session.commit()

        job.final_xlsx_key = xlsx_object_key
        job.finished_at = finished_at
        _emit_event(
            job,
            status="succeeded",
            final_xlsx_key=xlsx_object_key,
            row_count=row_count,
            duration_seconds=duration_seconds,
        )
        job.status = "succeeded"

    except ParseError as exc:
        _handle_failure(job, session_factory, archive_id, exc.error_code, str(exc))
    except Exception as exc:  # noqa: BLE001 — catch-all: surface to UI as RUNPY_CRASH
        _handle_failure(job, session_factory, archive_id, "RUNPY_CRASH", str(exc))
    finally:
        if job.temp_dir and job.temp_dir.exists():
            shutil.rmtree(job.temp_dir, ignore_errors=True)
        _SEMAPHORE.release()
        job.done_event.set()


def _emit_event(
    job: ParseJob,
    *,
    status: str,
    line: str | None = None,
    final_xlsx_key: str | None = None,
    error_code: str | None = None,
    error_message: str | None = None,
    row_count: int | None = None,
    duration_seconds: float | None = None,
) -> None:
    job.seq += 1
    event = ParseStreamEvent(
        seq=job.seq,
        status=status,
        line=line,
        final_xlsx_key=final_xlsx_key,
        parse_error_code=error_code,
        parse_error_message=error_message,
        row_count=row_count,
        duration_seconds=duration_seconds,
    )
    with job.lock:
        if len(job.events) >= _EVENT_BUFFER_MAXLEN:
            job.events.popleft()
            job.dropped_lines += 1
        job.events.append(event)


def _handle_failure(
    job: ParseJob,
    session_factory,
    archive_id: str,
    err_code: str,
    err_msg: str,
) -> None:
    job.status = "failed"
    job.error_code = err_code
    job.error_message = err_msg
    finished_at = datetime.now(timezone.utc)
    job.finished_at = finished_at
    with session_factory() as session:
        archive = session.get(Archive, archive_id)
        if archive:
            archive.parse_status = "failed"
            archive.parse_phase = "failed"
            archive.parse_finished_at = finished_at
            archive.parse_error_code = err_code
            archive.parse_error_message = err_msg[:4096]
            session.commit()
    _emit_event(job, status="failed", error_code=err_code, error_message=err_msg)


def _finalize_cancelled(job: ParseJob, session_factory, archive_id: str) -> None:
    job.status = "cancelled"
    finished_at = datetime.now(timezone.utc)
    job.finished_at = finished_at
    with session_factory() as session:
        archive = session.get(Archive, archive_id)
        if archive:
            archive.parse_status = "cancelled"
            archive.parse_phase = "cancelled"
            archive.parse_finished_at = finished_at
            archive.parse_error_code = "USER_CANCELLED"
            archive.parse_error_message = "User cancelled the parse job."
            session.commit()
    _emit_event(
        job,
        status="cancelled",
        error_code="USER_CANCELLED",
        error_message="User cancelled the parse job.",
    )


def cancel_parse_job(task_id: str) -> tuple[bool, bool]:
    """Cancel a running parse. Returns (cancelled, already_terminal)."""
    job = _JOBS.get(task_id)
    if job is None:
        return False, True
    if job.status in ("succeeded", "failed", "cancelled"):
        return False, True
    job.cancel_event.set()
    proc = job.proc
    if proc and proc.poll() is None:
        proc.terminate()
    return True, False


def stream_parse_events(task_id: str, since_seq: int = 0) -> Iterator[bytes]:
    """SSE generator: yield 'data: <json>\\n\\n' frames until job done or client disconnect.

    Reconnect support: pass `since_seq` from the Last-Event-ID header to resume
    after a dropped connection. Frames with seq <= since_seq are skipped.
    """
    job = _JOBS.get(task_id)
    if job is None:
        yield b'data: {"seq":0,"status":"unknown","parse_error_code":"JOB_NOT_FOUND"}\n\n'
        return

    sent_seq = since_seq
    last_status_seen = ""
    while True:
        with job.lock:
            pending = list(job.events)

        for event in pending:
            if event.seq > sent_seq:
                payload = event.model_dump_json()
                yield f"data: {payload}\n\n".encode("utf-8")
                sent_seq = event.seq
                last_status_seen = event.status

        if job.done_event.is_set():
            if job.dropped_lines > 0 and last_status_seen not in ("succeeded", "failed", "cancelled"):
                yield (
                    f'data: {{"seq":{job.seq},"status":"{job.status}",'
                    f'"truncated":true,"dropped_lines":{job.dropped_lines}}}\n\n'
                ).encode("utf-8")
            break

        if job.status in ("succeeded", "failed", "cancelled"):
            time.sleep(0.05)
            continue
        time.sleep(0.1)


def read_xlsx_output(
    *,
    session_factory,
    storage: ObjectStore,
    settings: Settings,
    archive_id: str,
) -> ParseOutputResponse:
    """Build presigned URL for the latest succeeded xlsx (7-day rolling expiry)."""
    with session_factory() as session:
        archive = session.get(Archive, archive_id)
        if archive is None:
            raise ArchiveNotFound(archive_id)
        if not archive.final_xlsx_key:
            raise NoXlsxAvailable(archive_id)
        object_key = archive.final_xlsx_key
        finished_at = archive.parse_finished_at
    expires_in = 7 * 24 * 3600  # 7 days — frontend can re-call to refresh
    download_url = storage.generate_presigned_url(
        settings.extract_bucket, object_key, expires_in=expires_in,
    )
    return ParseOutputResponse(
        bucket=settings.extract_bucket,
        object_key=object_key,
        download_url=download_url,
        expires_in=expires_in,
        parse_finished_at=finished_at.isoformat() if finished_at else None,
    )


def delete_xlsx_output(
    *,
    session_factory,
    storage: ObjectStore,
    settings: Settings,
    archive_id: str,
) -> bool:
    """Delete the MinIO xlsx + reset archive parse_* fields, preserve history timestamps (漏洞 H)."""
    with session_factory() as session:
        archive = session.get(Archive, archive_id)
        if archive is None:
            raise ArchiveNotFound(archive_id)
        object_key = archive.final_xlsx_key
    if object_key:
        try:
            storage.delete_object(settings.extract_bucket, object_key)
        except KeyError:
            pass  # already gone, idempotent
    with session_factory() as session:
        archive = session.get(Archive, archive_id)
        archive.parse_status = None
        archive.final_xlsx_key = None
        archive.candidate_xlsx_key = None
        archive.parse_phase = "deleted"
        archive.parse_error_code = None
        archive.parse_error_message = None
        # Keep parse_started_at / parse_finished_at as deletion history
        session.commit()
    return True


# ===== Custom exceptions (14 error codes) =====


class ParseError(Exception):
    error_code: str = "RUNPY_CRASH"

    def __init__(self, message: str):
        super().__init__(message)
        self.error_message = message


class ArchiveNotFound(ParseError):
    error_code = "ARCHIVE_NOT_FOUND"


class ArchiveNoFile(ParseError):
    error_code = "ARCHIVE_NO_FILE"


class ArchiveWithdrawn(ParseError):
    error_code = "ARCHIVE_WITHDRAWN"


class UnsupportedCity(ParseError):
    error_code = "UNSUPPORTED_CITY"


class PdfDownloadFailed(ParseError):
    error_code = "PDF_DOWNLOAD_FAILED"


class PythonNotFound(ParseError):
    error_code = "PYTHON_NOT_FOUND"


class RunpyNotFound(ParseError):
    error_code = "RUNPY_NOT_FOUND"


class RunpyCrash(ParseError):
    error_code = "RUNPY_CRASH"


class XlsxNotFound(ParseError):
    error_code = "XLSX_NOT_FOUND"


class ParseTimeout(ParseError):
    error_code = "TIMEOUT"


class ParseQueueFull(ParseError):
    error_code = "SEMAPHORE_QUEUE_FULL"


class NoXlsxAvailable(ParseError):
    error_code = "NO_XLSX_AVAILABLE"


class ParseAlreadyRunning(ParseError):
    error_code = "PARSE_ALREADY_RUNNING"


class InvalidFilename(ParseError):
    error_code = "INVALID_FILENAME"