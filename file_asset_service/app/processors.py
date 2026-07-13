from collections.abc import Mapping, Sequence
from io import BytesIO

REAL_V01_PROCESSORS = {"unzip", "pdf_extract", "xls_parse", "info_price_parse"}
PLACEHOLDER_V01_PROCESSORS = {"dwg_render", "dwg_extract"}

INFO_PRICE_FIELDS = [
    "原始名称",
    "名称",
    "规格型号",
    "单位",
    "综合单价",
    "不含税价",
    "税率",
    "province",
    "city",
    "district",
    "time",
    "startTime",
    "endTime",
    "来源",
]


def normalize_ext(file_ext: str | None) -> str:
    if not file_ext:
        return ""
    ext = file_ext.lower().strip()
    return ext if ext.startswith(".") else f".{ext}"


def derive_processors(file_ext: str | None, source_type: str, channel_type: str | None = None) -> list[str]:
    ext = normalize_ext(file_ext)
    if ext == ".zip":
        return ["unzip"]
    if (
        source_type == "info_price"
        and channel_type == "manual_upload"
        and ext in {".pdf", ".xls", ".xlsx"}
    ):
        # L0 入湖默认不给信息价建解析任务（见下方 return []）；仅 manual_upload 通道入队，
        # 由下游 info_price_parse 消费者处理（xlsx 由 parse_info_price_xlsx 直读，不叠 xls_parse）。
        return ["info_price_parse"]
    if source_type == "info_price":
        return []
    if ext == ".pdf":
        return ["pdf_extract"]
    if ext in {".xls", ".xlsx"}:
        processors = ["xls_parse"]
        if source_type == "info_price_governance":
            processors.append("info_price_parse")
        return processors
    if ext == ".dwg":
        return ["dwg_render", "dwg_extract"]
    return []


def build_info_price_extract(source_file_id: str, rows: Sequence[Mapping[str, object]]) -> dict[str, object]:
    normalized_rows = []
    for row in rows:
        normalized_rows.append({field: row.get(field, "") for field in INFO_PRICE_FIELDS})
    return {
        "schema": "info_price_extract.v1",
        "source_file_id": source_file_id,
        "rows": normalized_rows,
        "row_count": len(normalized_rows),
    }


def parse_info_price_xlsx(content: bytes) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    parsed_rows: list[dict[str, object]] = []
    for worksheet in workbook.worksheets:
        header_row_index, columns = _find_info_price_header(worksheet)
        if header_row_index is None:
            continue
        parsed_rows.extend(_extract_info_price_rows(worksheet, header_row_index, columns))
    return parsed_rows


def _find_info_price_header(worksheet) -> tuple[int | None, dict[str, int]]:
    max_column = min(worksheet.max_column, 50) if worksheet.max_column else 50
    max_row = min(worksheet.max_row, 80) if worksheet.max_row else 80
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True),
        start=1,
    ):
        headers = [_normalize_header(value) for value in row]
        columns = _map_info_price_columns(headers)
        if columns.get("name") is not None and columns.get("unit") is not None:
            if columns.get("price_with_tax") is not None or columns.get("price_without_tax") is not None:
                return row_index, columns
    return None, {}


def _map_info_price_columns(headers: Sequence[str]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        if "序号" in header and "sequence" not in columns:
            columns["sequence"] = index
        if ("代号" in header or "编码" in header) and "code" not in columns:
            columns["code"] = index
        if _is_name_header(header) and "name" not in columns:
            columns["name"] = index
        if "规格" in header and not _is_name_header(header) and "spec" not in columns:
            columns["spec"] = index
        if "单位" in header and "unit" not in columns:
            columns["unit"] = index
        if ("不含税" in header or "除税" in header) and "price_without_tax" not in columns:
            columns["price_without_tax"] = index
        elif "含税" in header and "price_with_tax" not in columns:
            columns["price_with_tax"] = index
    return columns


def _is_name_header(header: str) -> bool:
    return any(token in header for token in ("材料名称", "产品名称", "设备名称", "设备材料名称"))


def _extract_info_price_rows(worksheet, header_row_index: int, columns: Mapping[str, int]) -> list[dict[str, object]]:
    max_column = min(worksheet.max_column, 50) if worksheet.max_column else 50
    rows: list[dict[str, object]] = []
    for row in worksheet.iter_rows(
        min_row=header_row_index + 1,
        max_row=worksheet.max_row,
        max_col=max_column,
        values_only=True,
    ):
        raw_name = _cell(row, columns.get("name"))
        if raw_name is None:
            continue
        raw_name_text = _format_cell(raw_name)
        if not raw_name_text or raw_name_text.startswith(("备注", "注：", "注:")):
            continue

        unit = _format_cell(_cell(row, columns.get("unit")))
        price_with_tax = _format_cell(_cell(row, columns.get("price_with_tax")))
        price_without_tax = _format_cell(_cell(row, columns.get("price_without_tax")))
        if not unit and not price_with_tax and not price_without_tax:
            continue

        spec = _format_cell(_cell(row, columns.get("spec")))
        row_data = {
            "原始名称": " ".join(part for part in [raw_name_text, spec] if part),
            "名称": raw_name_text,
            "规格型号": spec,
            "单位": unit,
            "综合单价": price_with_tax,
            "不含税价": price_without_tax,
            "税率": _infer_tax_rate(price_with_tax, price_without_tax),
            "province": "",
            "city": "",
            "district": "",
            "time": "",
            "startTime": "",
            "endTime": "",
            "来源": worksheet.title,
        }
        rows.append(row_data)
    return rows


def _cell(row: Sequence[object], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def _normalize_header(value: object) -> str:
    return "".join(_format_cell(value).split())


def _format_cell(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _infer_tax_rate(price_with_tax: str, price_without_tax: str) -> str:
    try:
        with_tax = float(price_with_tax)
        without_tax = float(price_without_tax)
    except (TypeError, ValueError):
        return ""
    if without_tax == 0:
        return ""
    rate = with_tax / without_tax - 1
    if rate <= 0:
        return ""
    return f"{round(rate * 100, 2)}%"
