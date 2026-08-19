"""信息价分析端点（最小接入，2026-08-18）。

参考 quota_compare.py 结构但只暴露 capabilities + 占位 endpoint。
后续接入 phase0_5/phase1/... 时再扩展。

路径规划：
  前缀: /api/data-lake/info-price
  端点:
    GET /capabilities  - 能力探测（前端 quota-api.js FEATURES 数组查这个）
    GET /ping          - 最小连通性测试

sys.path 注入：把 info_price_analysis/6_脚本 加到 sys.path，
  后续调 phase0_5_manifest 等脚本能 import。
"""
from __future__ import annotations

import logging
import sys
from pathlib import Path

from fastapi import APIRouter, HTTPException, Response
from fastapi.responses import HTMLResponse, FileResponse

logger = logging.getLogger(__name__)

# ── 0. 把 info_price_analysis/6_脚本 加到 sys.path ────────────────
_REPO_ROOT = Path(__file__).resolve().parents[2]  # F:/data-lake-main/
_INFO_PRICE_SCRIPTS = _REPO_ROOT / "info_price_analysis" / "6_脚本"
if _INFO_PRICE_SCRIPTS.is_dir() and str(_INFO_PRICE_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_INFO_PRICE_SCRIPTS))
    logger.info(f"[info_price_analysis] sys.path + {_INFO_PRICE_SCRIPTS}")

# ── 1. Router ──────────────────────────────────────────────────────
router = APIRouter(prefix="/api/data-lake/info-price", tags=["info-price"])

# ── 2. FEATURES 字典（参考 quota_api._FEATURES 模式）────────────────
_SCHEMA_VERSION = "1"
_FEATURES: dict[str, str] = {
    # 2026-08-18 接入：信息价分析后端最小化上线（router + capabilities + ping）
    "infoPriceAnalysis": "ready",
    # V1 限制：成都单城 + 同城分析 + 跨城分析骨架（待真实 phase0_5_manifest 接入）
    "sameCityAnalysis": "ready",
    "crossCityAnalysis": "ready",  # 2026-08-18 V2 上线：phase5_cross_city 骨架
}

MIME = {"Content-Type": "application/json; charset=utf-8"}


def _json_ok(response: Response, body: object) -> object:
    response.headers.update(MIME)
    return body


# ── 3. 端点 ────────────────────────────────────────────────────────
@router.get("/capabilities")
def get_capabilities(response: Response) -> dict:
    """能力探测。前端 quota-api.js 的 FEATURES 数组会查这个。"""
    return _json_ok(response, {
        "schema_version": _SCHEMA_VERSION,
        "features": dict(_FEATURES),
    })


@router.get("/ping")
def ping() -> dict:
    """最小连通性测试。"""
    return {"status": "ok", "scripts_path_exists": _INFO_PRICE_SCRIPTS.is_dir()}


# ── 5. GET /reports/ 报告列表 + GET /reports/{filename} 文件浏览（2026-08-18）──────
# 用户场景：在网站内一个路径专门放信息价分析报告，浏览器新 tab 直接打开 md
@router.get("/reports-list")
def list_reports_json() -> dict:
    """列所有报告（JSON 格式，供前端 SPA 用）。"""
    if not INFO_PRICE_OUTPUT_DIR.exists():
        return {"reports": []}
    files = sorted(
        [f for f in INFO_PRICE_OUTPUT_DIR.glob("*") if f.suffix == ".xlsx"],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return {
        "reports": [
            {
                "filename": f.name,
                "ext": f.suffix.lstrip("."),
                "size_bytes": f.stat().st_size,
                "mtime": f.stat().st_mtime,
            }
            for f in files
        ]
    }


@router.get("/reports/", response_class=HTMLResponse)
def list_reports() -> str:
    """列所有信息价分析报告 .md / .xlsx（按 mtime 倒序），HTML 列表带链接。"""
    if not INFO_PRICE_OUTPUT_DIR.exists():
        return "<h1>暂无报告目录</h1>"
    files = sorted(
        [f for f in INFO_PRICE_OUTPUT_DIR.glob("*") if f.suffix == ".xlsx"],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    items = "".join(
        f'<li><a href="/api/data-lake/info-price/reports/{f.name}">{f.name}</a>'
        f' <small>({f.stat().st_size:,} bytes, '
        f'{dt.datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M")})</small></li>'
        for f in files
    )
    return (
        '<!DOCTYPE html><html><head><meta charset="utf-8"><title>信息价分析报告</title>'
        '<style>body{font-family:system-ui,sans-serif;max-width:900px;margin:24px auto;}'
        'h1{color:#1a202c;}li{margin:6px 0;}a{color:#4f6df5;text-decoration:none;}'
        'a:hover{text-decoration:underline;}small{color:#6b7280;}</style>'
        f'</head><body><h1>信息价分析报告（{len(files)} 个）</h1><ul>{items}</ul>'
        '</body></html>'
    )


@router.get("/reports/{filename}")
def get_report(filename: str):
    """返 md / xlsx 原始文件，浏览器新 tab 直接看。防越权：限制后缀 + 不含路径分隔符。"""
    if "/" in filename or "\\" in filename or filename.startswith("."):
        raise HTTPException(status_code=400, detail="非法文件名")
    path = INFO_PRICE_OUTPUT_DIR / filename
    if not path.exists() or not path.is_file() or path.suffix not in (".md", ".xlsx"):
        raise HTTPException(status_code=404, detail=f"未找到 {filename}")
    media_type = (
        "text/markdown; charset=utf-8" if path.suffix == ".md"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return FileResponse(path, media_type=media_type, filename=filename)


# 2026-08-18: 信息价分析报告本地 xlsx 删除（前端文件名旁的删除按钮）
@router.delete("/reports/{filename}")
def delete_report(filename: str) -> dict:
    """删除本地 xlsx 报告（硬删本地文件）。防越权：限制后缀 + 不含路径分隔符 + 不以 . 开头。"""
    if "/" in filename or "\\" in filename or filename.startswith("."):
        raise HTTPException(status_code=400, detail="非法文件名")
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持删除 .xlsx 文件")
    path = INFO_PRICE_OUTPUT_DIR / filename
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail=f"未找到 {filename}")
    path.unlink()
    return {"status": "ok", "filename": filename}


@router.get("/reports/{filename}/preview", response_class=HTMLResponse)
def preview_report(filename: str) -> str:
    """xlsx → HTML 表格预览。读 openpyxl 转 <table>。"""
    if "/" in filename or "\\" in filename or filename.startswith("."):
        raise HTTPException(status_code=400, detail="非法文件名")
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅 .xlsx 支持预览")
    path = INFO_PRICE_OUTPUT_DIR / filename
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail=f"未找到 {filename}")

    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    css = (
        "<style>"
        "body{font-family:system-ui,sans-serif;margin:16px;}"
        "h2{margin-top:24px;color:#1a202c;}"
        "table{border-collapse:collapse;margin:8px 0;}"
        "th,td{border:1px solid #ccc;padding:4px 8px;text-align:left;}"
        "th{background:#f5f5f5;}"
        "</style>"
    )
    parts = [
        f"<!DOCTYPE html><html><head><meta charset='utf-8'><title>{filename} 预览</title>{css}</head><body>",
        f"<p class='fname'>{filename}</p>",
    ]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(values_only=True))
        if not rows_data:
            continue
        # 2026-08-18: 每个 sheet 包 <section data-sheet="...">，前端 sheet tab 切换用
        parts.append(f'<section data-sheet="{sheet_name}">')
        parts.append(f"<h2>📋 {sheet_name}</h2>")
        parts.append("<table>")
        for i, row in enumerate(rows_data):
            tag = "th" if i == 0 else "td"
            cells = "".join(
                f"<{tag}>{(str(c).replace('<', '&lt;').replace('>', '&gt;') if c is not None else '')}</{tag}>"
                for c in row
            )
            parts.append(f"<tr>{cells}</tr>")
        parts.append("</table>")
        parts.append("</section>")
    parts.append("</body></html>")
    return "".join(parts)


# ── 4. POST /analyze（真实业务端点，2026-08-18）────────────────────────────
# 数据流：
#   1) SELECT archive WHERE city region_code + period → final_xlsx_key
#   2) storage.get_object() → bytes → 写 info_price_analysis/1_输入/{name}.xlsx
#      文件名必须让 phase0_5_manifest L1 正则匹配 "年" + "第N期"
#   3) subprocess 串行 phase0_5/phase1/phase2/phase3/phase4（跳 phase2_5_pending）
#   4) glob 4_输出/reports/*.md 最新 → 读前 200 行作 preview
# 任一 phase returncode != 0 → HTTPException 500 + 已完成 phase log + 中间 csv 路径
import subprocess
import shutil
import datetime as dt
import re
import glob
from typing import Optional

from fastapi import Depends, HTTPException
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.database import get_db_session
from app.models import Archive
from app.storage import ObjectStore, get_object_store
from app.config import get_settings


# city → region_code 硬编码字典（V1；后续按 R10 城市模板原则动态化）
CITY_REGION_CODE: dict[str, str] = {
    "成都": "510100",
    "重庆": "500100",
    "北京": "110000",
    "广州": "440100",
    "武汉": "420100",
    "湖北": "420000",
}

# phase 链（V1 跳过 phase2_5_pending 人审）
PHASE_CHAIN: list[str] = [
    "phase0_5_manifest",
    "phase1_columns",
    "phase2_match",
    "phase3_query",
    "phase4_report",
]

# 临时输入目录（每次启动删旧文件）
INFO_PRICE_INPUT_DIR = _INFO_PRICE_SCRIPTS.parent / "1_输入"
INFO_PRICE_OUTPUT_DIR = _INFO_PRICE_SCRIPTS.parent / "4_输出" / "reports"

# Python 解释器：与 web 服务同环境（conda file-asset，含 openpyxl 等依赖）
# 历史曾写死 .venv/Scripts/python.exe，但本机 .venv 是占位文件夹，未装 openpyxl
#   → phase0_5_manifest subprocess 跑时 ModuleNotFoundError 失败
# 现与 info_price_parse.py / package.py 一致，用 sys.executable
VENV_PYTHON = Path(sys.executable)


class AnalyzeRequest(BaseModel):
    """POST /analyze 请求体。

    mode:
      sameCity   - 同城分析（单 city 多期，V1 默认）
      crossCity  - 跨城分析（N 城同期；2026-08-18 V2 上线，phase5_cross_city.py）
    """
    mode: str = Field("sameCity", description='分析模式："sameCity" 或 "crossCity"')
    city: list[str] = Field(..., description="城市列表（同城=1 个，跨城=N 个）")
    periods: list[str] = Field(..., description='期数列表，格式 "YYYY-N" / "YY.N"')
    keyword: str = Field(..., min_length=1, description="分析关键词（必填，逗号分隔，对应 phase4_report_v3.py 的 --filter）")
    note: str = Field("", description="可选备注，写入报告头部")


# 把 "2026-1" / "2026-01" 解析成 (year, period, publish_date)
# 支持 4 种格式：2026-2 / 2026.2 / 26-2 / 26.2 / 26/2 / 26 2（分隔符 - . / 可选）
# 简写年（2 位）自动补 2000 前缀
_PERIOD_RE = re.compile(r"^(?:(?:20(?P<yy>\d{2}))|(?P<y2>\d{2}))[-./]?(?P<period>\d{1,2})$")

# 2026-08-18: 中文格式 + 6 位无分隔符（用户放宽期数输入）
#   "2026年3月" / "2026年3期" / "2026年第3期" / "26年第3期" / "26年3月"
_PERIOD_CN_RE = re.compile(r"^(?:20(?P<yy_cn>\d{2})|(?P<y2_cn>\d{2}))年(?:第(?P<p_qi>\d{1,2})期|(?P<p_mq>\d{1,2})(?:月|期))$")
# 6 位无分隔符 "202603" = 2026-3
_PERIOD_6DIGIT_RE = re.compile(r"^(?P<y6>20\d{2})(?P<p6>\d{2})$")


def _parse_period(token: str) -> dict:
    """解析期数，支持：
      - 数字格式：2026-2 / 2026.2 / 26-2 / 26.2 / 2026/2 / 26 2 / 263
      - 中文格式：2026年3月 / 2026年3期 / 2026年第3期 / 26年3月
      - 6 位无分隔符：202603 (= 2026-3)

    简写年（2 位）自动补 2000 前缀（26 → 2026）。
    """
    s = token.strip()

    # 1. 中文格式优先（年/月/期）
    m = _PERIOD_CN_RE.match(s)
    if m:
        yy = m.group("yy_cn")
        y2 = m.group("y2_cn")
        year = 2000 + int(yy) if yy else (2000 + int(y2) if y2 else None)
        period = int(m.group("p_qi") or m.group("p_mq"))
    else:
        # 2. 6 位无分隔符
        m6 = _PERIOD_6DIGIT_RE.match(s)
        if m6:
            year = int(m6.group("y6"))
            period = int(m6.group("p6"))
        else:
            # 3. 原数字格式
            m = _PERIOD_RE.match(s)
            if not m:
                raise ValueError(f"无效期数格式: {token!r}（期望 2026-2 / 26.3 / 2026年3月 / 2026年第3期 / 202603）")
            yy = m.group("yy")
            y2 = m.group("y2")
            year = 2000 + int(yy) if yy else (2000 + int(y2) if y2 else None)
            period = int(m.group("period"))

    if year is None or year < 2000 or year > 2099:
        raise ValueError(f"年份超出范围: {token!r}")
    if period < 1 or period > 12:
        raise ValueError(f"期号超出 1-12: {token!r}")
    return {"year": year, "period": period, "publish_date": dt.date(year, period, 1)}


# 从 archive.title 推断 (year, period) — 兼容 publish_date 未回填的档案
# 匹配："2026年第5期" / "成都市2026年第5期信息价.pdf"（成都"第N期"格式）
#      "2026年06月" / "2026年6月"（北京/武汉/重庆等"年N月"格式）
_TITLE_PERIOD_RE = re.compile(r"(?P<year>20\d{2})\s*年(?:第\s*(?P<period_qi>\d{1,2})\s*期|(?P<period_month>\d{1,2})\s*月)")


def _infer_period_from_title(title: str) -> tuple[int, int] | None:
    """从 archive.title 推断 (year, period)；失败返回 None。

    支持两种格式：
      "2026年第5期"   → (2026, 5)
      "2026年06月"    → (2026, 6)
    """
    if not title:
        return None
    m = _TITLE_PERIOD_RE.search(title)
    if not m:
        return None
    period = m.group("period_qi") or m.group("period_month")
    if not period:
        return None
    return int(m.group("year")), int(period)


def _select_archives(session: Session, region_codes, periods: list[dict]) -> list[Archive]:
    """查指定 N 城 + periods 下 parse_status=succeeded 的 archive 列表。

    region_codes: str（单城，向后兼容）或 list[str]（跨城 N 城）。
    按 year+period 匹配（成都 publish_date 未回填，从 title 推断；其他城市也兼容）。
    """
    rc_list = [region_codes] if isinstance(region_codes, str) else list(region_codes)
    target_set = {(p["year"], p["period"]) for p in periods}
    rows = (
        session.query(Archive)
        .filter(
            Archive.domain_type == "cost_info",
            Archive.region_code.in_(rc_list),
            Archive.parse_status == "succeeded",
            Archive.final_xlsx_key.isnot(None),
        )
        .all()
    )
    # Python 端过滤（按 title 推断 year/period）
    matched = []
    for r in rows:
        inferred = _infer_period_from_title(r.title)
        if inferred and inferred in target_set:
            matched.append(r)
    return matched


def _prepare_input_dir(blobs: list[dict]) -> list[dict]:
    """清空 1_输入/，把 xlsx bytes 写到 {city}{year}年第{period}期.xlsx。

    返回 [{archive_id, file_path, year, period}]。
    """
    if INFO_PRICE_INPUT_DIR.exists():
        shutil.rmtree(INFO_PRICE_INPUT_DIR)
    INFO_PRICE_INPUT_DIR.mkdir(parents=True, exist_ok=True)
    written = []
    for blob in blobs:
        y = blob["year"]
        p = blob["period"]
        city = blob["city"]
        # phase0_5_manifest L1 正则需要 "年" + "第N期"
        filename = f"{city}{y}年第{p}期.xlsx"
        path = INFO_PRICE_INPUT_DIR / filename
        path.write_bytes(blob["xlsx_bytes"])
        written.append({
            "archive_id": blob["archive_id"],
            "file_path": str(path),
            "year": y,
            "period": p,
        })
    return written


def _run_phase_subprocess(phase_module: str, cwd: Path, extra_args: list[str] | None = None, timeout: int = 600) -> dict:
    """调 6_脚本/{phase_module}.py 跑一个 phase；返回 log dict。extra_args 透传给脚本。"""
    phase_script = cwd / f"{phase_module}.py"
    start = dt.datetime.now()
    try:
        cmd = [str(VENV_PYTHON), str(phase_script)] + (extra_args or [])
        proc = subprocess.run(
            cmd,
            cwd=str(cwd),
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding="utf-8",
            errors="replace",
        )
        duration_ms = int((dt.datetime.now() - start).total_seconds() * 1000)
        stdout_tail = "\n".join(proc.stdout.splitlines()[-20:])
        return {
            "phase": phase_module,
            "returncode": proc.returncode,
            "duration_ms": duration_ms,
            "stdout_tail": stdout_tail,
            "stderr_tail": "\n".join(proc.stderr.splitlines()[-10:]),
        }
    except subprocess.TimeoutExpired:
        return {
            "phase": phase_module,
            "returncode": -1,
            "duration_ms": timeout * 1000,
            "stdout_tail": "",
            "stderr_tail": f"timeout after {timeout}s",
        }


def _find_latest_report() -> Optional[Path]:
    """glob 4_输出/reports/*.md 最新文件。"""
    if not INFO_PRICE_OUTPUT_DIR.exists():
        return None
    files = list(INFO_PRICE_OUTPUT_DIR.glob("*.md"))
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


@router.post("/analyze")
def analyze(
    body: AnalyzeRequest,
    session: Session = Depends(get_db_session),
    storage: ObjectStore = Depends(get_object_store),
) -> dict:
    """信息价分析端到端执行。

    流程：DB 查 xlsx → 拉 S3 → 写 1_输入/ → 串行 5 phase → 读 md 报告。
    """
    started_at = dt.datetime.now()

    # 1) 模式校验
    if body.mode not in ("sameCity", "crossCity"):
        raise HTTPException(status_code=400, detail=f"mode 必须是 sameCity/crossCity，当前 {body.mode!r}")
    if body.mode == "crossCity" and len(body.city) < 2:
        raise HTTPException(status_code=400, detail="跨城分析至少 2 个城市")
    if body.mode == "sameCity" and len(body.city) != 1:
        raise HTTPException(status_code=400, detail=f"同城分析只接受 1 个城市，当前 {len(body.city)} 个")

    # 2) city → region_code 映射（支持 N 城）
    region_codes: list[str] = []
    for c in body.city:
        rc = CITY_REGION_CODE.get(c)
        if not rc:
            raise HTTPException(
                status_code=400,
                detail=f"暂不支持城市 {c!r}；V1 支持: {', '.join(CITY_REGION_CODE.keys())}",
            )
        region_codes.append(rc)

    # 2) 解析期数
    if not body.periods:
        raise HTTPException(
            status_code=400,
            detail="请填期数（如 2026-2，多个用逗号分隔）",
        )
    try:
        periods = [_parse_period(t) for t in body.periods]
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    # 3) 查 archive
    archives = _select_archives(session, region_codes, periods)
    if not archives:
        # 提示友好：列出指定 N 城实际 succeeded 的所有 archive 标题供排查
        all_succeeded = (
            session.query(Archive)
            .filter(
                Archive.domain_type == "cost_info",
                Archive.region_code.in_(region_codes),
                Archive.parse_status == "succeeded",
                Archive.final_xlsx_key.isnot(None),
            )
            .limit(20)
            .all()
        )
        titles = [a.title for a in all_succeeded if a.title][:5]
        raise HTTPException(
            status_code=404,
            detail=(
                f"未找到 city={body.city} + periods={body.periods} 的 succeeded 解析产物。"
                f"已 succeeded 档案示例：{titles}。"
                "请确认期数匹配（格式 YYYY-N 或 YY.N）。"
            ),
        )

    # 4) S3 拉 xlsx
    settings = get_settings()
    blobs = []
    for a in archives:
        # 2026-08-18: 成都 publish_date 未回填，从 title 推断 (year, period)；其他城市 fallback publish_date
        inferred = _infer_period_from_title(a.title)
        if inferred:
            y, m = inferred
        elif a.publish_date:
            y, m = a.publish_date.year, a.publish_date.month
        else:
            continue
        try:
            xlsx_bytes = storage.get_object(settings.extract_bucket, a.final_xlsx_key)
        except Exception as exc:
            raise HTTPException(
                status_code=502,
                detail=f"拉取 xlsx 失败 {a.archive_id}: {exc}",
            )
        # 2026-08-18: 反查 archive.region_code → city 名（用于写文件名 + phase5 city 维度）
        city_name = next((c for c, rc in CITY_REGION_CODE.items() if rc == a.region_code), "未知")
        blobs.append({
            "archive_id": a.archive_id,
            "xlsx_bytes": xlsx_bytes,
            "city": city_name,
            "year": y,
            "period": m,
        })

    if not blobs:
        raise HTTPException(status_code=404, detail="xlsx 字节拉取为空，请检查 publish_date")

    # 5) 写输入
    prepared = _prepare_input_dir(blobs)

    # 6) 串行 phase chain（跨城 mode 时追加 phase5_cross_city）
    active_chain = list(PHASE_CHAIN)
    if body.mode == "crossCity":
        active_chain = PHASE_CHAIN + ["phase5_cross_city"]
    phase_logs = []
    for phase_module in active_chain:
        # 2026-08-18: phase4 透传 keyword 给 phase4_report_v3.py 的 --filter
        extra_args = ["--filter", body.keyword] if phase_module == "phase4_report" else None
        log = _run_phase_subprocess(phase_module, _INFO_PRICE_SCRIPTS, extra_args=extra_args)
        phase_logs.append(log)
        if log["returncode"] != 0:
            raise HTTPException(
                status_code=500,
                detail={
                    "status": "failed",
                    "phase_failed": phase_module,
                    "phase_logs": phase_logs,
                    "intermediate_dir": str(_INFO_PRICE_SCRIPTS.parent / "3_中间产物"),
                    "duration_seconds": round((dt.datetime.now() - started_at).total_seconds(), 2),
                    "message": f"phase {phase_module} 失败 (returncode={log['returncode']})",
                },
            )

    # 7) 找最新 md
    report_path = _find_latest_report()
    if report_path is None:
        raise HTTPException(status_code=500, detail="phase 全跑完但未生成 md 报告")

    report_text = report_path.read_text(encoding="utf-8")
    preview_lines = report_text.splitlines()[:200]

    return {
        "status": "ok",
        "mode": body.mode,
        "city": body.city,
        "periods": body.periods,
        "report_path": str(report_path),
        "report_filename": report_path.name,
        "report_size_bytes": len(report_text.encode("utf-8")),
        "report_preview": "\n".join(preview_lines),
        "files_analyzed": len(prepared),
        "phase_logs": phase_logs,
        "started_at": started_at.isoformat(),
        "duration_seconds": round((dt.datetime.now() - started_at).total_seconds(), 2),
    }
