"""DB ↔ MinIO ↔ quota_parser 桥（web 端 adapter 主逻辑）

核心函数：
  - trigger_parse(archive, profile) 写 parse_status='parsing' + parse_task_id
  - upload_reviewed(archive_id, reviewed_xlsx_path) 写 final_xlsx_key
  - build_parse_section(archive) 把 Archive.parse_* 打包成前端 dict
  - validate_reviewed_xlsx(path) 跑 openpyxl 校验 Sheet1 结构

环境变量：
  - QUOTA_PARSE_MOCK=1 → mock 模式（adapter 层不直接判断，quota_api.py 端点处切）
"""
from __future__ import annotations

import hashlib
import logging
import os
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from app.models import Archive
from app.storage import get_object_store

logger = logging.getLogger(__name__)

# === 常量（quota/INTEGRATION_PLAN.md §2.3，与 cost 域复用 bucket） ===

# quota 解析产物复用 cost 域的 cost-extract / cost-report 桶（quota/README.md §12.G 决策 4）
# key 前缀用 `quota/` 与 cost 域产物区分（cost 用 `cost/...`），共享同一桶命名空间
PARSE_BUCKET_CANDIDATE = "cost-extract"   # candidate.xlsx / final.xlsx
PARSE_BUCKET_REPORT = "cost-report"        # manifest.json / qa_report.{json,md}
PROFILES = ("sichuan", "chongqing")
PARSE_MOCK_ENV_VAR = "QUOTA_PARSE_MOCK"


def is_parse_mock() -> bool:
    """读取 QUOTA_PARSE_MOCK 环境变量。"""
    return os.environ.get(PARSE_MOCK_ENV_VAR, "0") == "1"


# === Manifest 字段常量（与 parser/SPEC.md §6.2 + web-frontend SPEC §6.1 对齐） ===

MANIFEST_SCHEMA = "quota-parser-result/v1"
QA_REPORT_SCHEMA = "quota-parser-qa/v1"


# === 时间戳 / ID 工具 ===

def _now() -> datetime:
    return datetime.now(UTC)


def _ts() -> str:
    return datetime.now(UTC).strftime("%Y%m%d_%H%M%S")


def _make_task_id(archive_id: str) -> str:
    """构造 quota_parser 任务的 task_id。"""
    return f"qp_{_ts()}_{archive_id[:8]}"


def _sha256_bytes(data: bytes) -> str:
    h = hashlib.sha256()
    h.update(data)
    return h.hexdigest()


# === 状态机辅助 ===

_VALID_PARSE_STATUSES = {
    "pending",
    "parsing",
    "parsed",
    "qa_passed",
    "usable",
    "failed_user",
    "failed_permanent",
    "transient",
}


def _check_parse_status(value: str | None) -> str | None:
    if value is None:
        return None
    if value not in _VALID_PARSE_STATUSES:
        raise ValueError(
            f"非法 parse_status {value!r}；合法值：{sorted(_VALID_PARSE_STATUSES)}"
        )
    return value


# === 阶段 A：触发解析 ===

def trigger_parse(archive: Archive, *, profile: str | None = None,
                  province: str | None = None) -> Archive:
    """触发阶段 A：写 parse_status='parsing' + parse_task_id。

    Args:
        archive:  已加载的 Archive 实例（caller 负责 session.commit()）
        profile:  'sichuan' / 'chongqing' / None（None = 用档案已有的 profile）
        province: v0.4 §9 #15 入参；可由 caller 从 metadata_payload 读回。
                  若 metadata_payload 还没有 province,补一个 audit cell（不影响 profile）。

    Returns:
        更新后的 Archive 实例

    Raises:
        ValueError: archive 不在 quota 域；profile 不合法；已有 parsing 状态（不可重入）

    注意：
      - parse_status 不加 CheckConstraint（INTEGRATION_PLAN.md §2.1 方案 A）；
        但仍做白名单校验，避免 UI bug 写出非法值。
      - 重入规则：parse_status='parsing' 时拒绝重入（避免 Worker 重复处理）。
      - 其余状态（parsed / qa_passed / usable / failed_*）允许重入 → 走"重新解析"动线。
    """
    if archive.domain_type != "quota":
        raise ValueError(
            f"archive {archive.archive_id} 不是 quota 域（{archive.domain_type}）"
        )
    if profile is not None and profile not in PROFILES:
        raise ValueError(f"profile {profile!r} 不在注册表 {list(PROFILES)}")
    if archive.parse_status == "parsing":
        raise ValueError(f"parse_status 已是 parsing，不可重入")

    archive.parse_status = "parsing"
    archive.parse_profile = profile or archive.parse_profile
    archive.parse_task_id = _make_task_id(archive.archive_id)
    archive.parse_phase = "stage_a"
    archive.parse_started_at = _now()
    archive.parse_finished_at = None
    archive.parse_metrics = None
    archive.parse_warnings = None
    archive.parse_error_code = None
    archive.parse_error_message = None
    # 重新解析时清掉旧 candidate / final key（web-frontend SPEC §3.1.3 决策）
    archive.candidate_xlsx_key = None
    archive.final_xlsx_key = None

    # ── v0.4 §9 #15：province 透传 — 若 metadata_payload 还没有 province,补 audit cell ──
    # idempotent：上传时已写过的不会覆盖；旧档案（无 province）这里补上后端可见性。
    if province and isinstance(archive.metadata_payload, dict):
        if "province" not in archive.metadata_payload:
            from app.archive_rules import metadata_cell as _mc
            archive.metadata_payload = {
                **archive.metadata_payload,
                "province": _mc(
                    province,
                    source_level="manual", tagged_by="api:trigger-parse",
                ),
            }
    return archive


# === 阶段 B：上传 reviewed → final ===

def validate_reviewed_xlsx(reviewed_path: Path) -> None:
    """校验 reviewed.xlsx 的 Sheet1 结构（parser/SPEC.md §6.2 选项 B）。

    Raises:
        InvalidReviewedXlsxError: 结构不符（前端可定位原因）
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(reviewed_path, read_only=True)
    except Exception as e:
        raise InvalidReviewedXlsxError(
            "INVALID_XLSX",
            f"无法用 openpyxl 打开文件：{e}",
            sheet=None,
            column_index=None,
        ) from e

    try:
        if not wb.sheetnames:
            raise InvalidReviewedXlsxError(
                "MISSING_SHEET_QUOTA_ENTRIES",
                "文件不含任何 sheet",
                sheet=None,
                column_index=None,
            )
        if wb.sheetnames[0] != "定额条目":
            raise InvalidReviewedXlsxError(
                "SHEET_ORDER_WRONG",
                f"Sheet1 必须是「定额条目」，当前是「{wb.sheetnames[0]}」",
                sheet=wb.sheetnames[0],
                column_index=None,
            )
        ws = wb["定额条目"]
        # 校验列数 = 10
        if ws.max_column is None or ws.max_column != 10:
            raise InvalidReviewedXlsxError(
                "COLUMN_COUNT_WRONG",
                f"「定额条目」必须 10 列，当前 {ws.max_column}",
                sheet="定额条目",
                column_index=ws.max_column,
            )
    finally:
        wb.close()


def upload_reviewed(
    archive: Archive,
    reviewed_xlsx_bytes: bytes,
) -> Archive:
    """阶段 B：把 reviewed.xlsx bytes 校验 + 调 finalize_reviewed_xlsx → 写 final_xlsx_key。

    Args:
        archive: 已加载的 Archive 实例（caller 负责 session.commit()）
        reviewed_xlsx_bytes: 用户上传的 reviewed.xlsx 字节内容

    Returns:
        更新后的 Archive 实例（parse_status='qa_passed', final_xlsx_key=...）

    Raises:
        InvalidReviewedXlsxError: 结构不符
        QuotaParserStageBError: finalize_reviewed_xlsx 抛错（落 failed_user）
    """
    from app.models import Archive
    # v0.5 fix: 显式 import 真 pipeline 子模块，避免 sys.path 把 web adapter quota_parser 包当 quota_parser 真包
    from quota_parser.pipeline import finalize_reviewed_xlsx
    from quota_parser.exceptions import (
        InvalidXlsxStructureError,
        ProfileExecutionError,
    )

    if archive.domain_type != "quota":
        raise ValueError(
            f"archive {archive.archive_id} 不是 quota 域（{archive.domain_type}）"
        )

    # 1. 写 reviewed 临时文件（validator 需要 path）
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(reviewed_xlsx_bytes)
        reviewed_path = Path(tmp.name)

    final_xlsx_path: Path | None = None  # 提前初始化,避免 validate 失败时 finally UnboundLocalError

    try:
        # 2. 结构校验
        try:
            validate_reviewed_xlsx(reviewed_path)
        except InvalidReviewedXlsxError:
            archive.parse_status = "parsed"
            archive.parse_error_code = "failed_user"
            archive.parse_error_message = "reviewed.xlsx 结构不符（详见 422 返回）"
            raise

        # 3. 调 finalize_reviewed_xlsx（quota_parser 包）
        final_xlsx_path = reviewed_path.with_name("final.xlsx")
        try:
            finalize_reviewed_xlsx(
                reviewed_xlsx_path=str(reviewed_path),
                output_xlsx_path=str(final_xlsx_path),
            )
        except (InvalidXlsxStructureError, ProfileExecutionError) as e:
            archive.parse_status = "parsed"
            archive.parse_error_code = "failed_user"
            archive.parse_error_message = f"finalize 失败：{e}"
            raise QuotaParserStageBError(str(e)) from e

        # 4. 上传 final.xlsx 到 MinIO（cost-extract 桶，key 前缀 quota/）
        final_bytes = final_xlsx_path.read_bytes()
        final_key = f"quota/{archive.archive_id}/final.xlsx"
        store = get_object_store()
        store.put_object(
            PARSE_BUCKET_CANDIDATE,
            final_key,
            final_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # 5. 写 Archive.parse_* 字段
        archive.parse_status = "qa_passed"
        archive.parse_phase = "stage_b"
        archive.parse_finished_at = _now()
        archive.parse_error_code = None
        archive.parse_error_message = None
        archive.final_xlsx_key = final_key
        archive.parse_metrics = archive.parse_metrics or {}
        archive.parse_metrics["final_xlsx_sha256"] = _sha256_bytes(final_bytes)
        return archive
    finally:
        # 清理 reviewed_path / final_xlsx_path
        for p in (reviewed_path, final_xlsx_path):
            try:
                if p and p.exists():
                    p.unlink()
            except OSError:
                pass


# === 序列化：Archive.parse_* → API dict ===

def build_parse_section(archive: Archive) -> dict[str, Any] | None:
    """把 Archive.parse_* 字段打包成 web-frontend SPEC §6.1 的 `parse` 子对象。

    Args:
        archive: Archive 实例

    Returns:
        dict（前端可直接 JSON 序列化），或 None（parse_status 还没初始化时）
    """
    if archive.parse_status is None:
        return None

    return {
        "profile": archive.parse_profile,
        "task_id": archive.parse_task_id,
        "phase": archive.parse_phase,
        "status": archive.parse_status,
        "started_at": _iso(archive.parse_started_at),
        "finished_at": _iso(archive.parse_finished_at),
        "parser_version": archive.parse_parser_version,
        "metrics": archive.parse_metrics,
        "warnings": archive.parse_warnings,
        "error_code": archive.parse_error_code,
        "error_message": archive.parse_error_message,
        "candidate_xlsx_key": archive.candidate_xlsx_key,
        "final_xlsx_key": archive.final_xlsx_key,
        "manifest": _reconstruct_manifest(archive),
    }


def _iso(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.isoformat()


def _reconstruct_manifest(archive: Archive) -> dict[str, Any] | None:
    """从 Archive.parse_* 字段反推 Manifest 对象（前端零翻译）。

    与 parser/SPEC.md §6.2 schema 对齐。
    """
    if archive.parse_status is None:
        return None
    return {
        "$schema": MANIFEST_SCHEMA,
        "task_id": archive.parse_task_id,
        "phase": archive.parse_phase,
        "status": archive.parse_status,
        "parser_version": archive.parse_parser_version,
        "profile": archive.parse_profile,
        "province": archive.parse_profile,  # Profile == province（v0.3 1:1 对齐）
        "ocr_api_url": "http://172.16.20.23:8000",
        "source_pdf_sha256": None,  # archive 未存原 PDF sha256（FileAsset.sha256 已有，跨表 join）
        "candidate_xlsx_sha256": (
            archive.parse_metrics.get("candidate_xlsx_sha256")
            if archive.parse_metrics
            else None
        ),
        "final_xlsx_sha256": (
            archive.parse_metrics.get("final_xlsx_sha256")
            if archive.parse_metrics
            else None
        ),
        "artifacts": _reconstruct_artifacts(archive),
        "metrics": archive.parse_metrics or {},
        "warnings": archive.parse_warnings or [],
    }


def _reconstruct_artifacts(archive: Archive) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    if archive.candidate_xlsx_key:
        out.append({"kind": "candidate_xlsx", "key": archive.candidate_xlsx_key})
    if archive.final_xlsx_key:
        out.append({"kind": "final_xlsx", "key": archive.final_xlsx_key})
    return out


# === 自定义异常 ===

class InvalidReviewedXlsxError(Exception):
    """reviewed.xlsx 结构校验失败（前端可定位 sheet / column_index）。"""

    def __init__(self, code: str, message: str, *, sheet: str | None, column_index: int | None) -> None:
        super().__init__(message)
        self.code = code
        self.sheet = sheet
        self.column_index = column_index


class QuotaParserStageBError(Exception):
    """finalize_reviewed_xlsx 阶段 B 抛错（落 failed_user）。"""


# === v0.5 新增 — worker 接入 helpers（INTEGRATION_PLAN §1.2.3 / §2.4） ===

# 4 类 parse_* file_role — 与 ck_archive_file_role CheckConstraint + DB_SCHEMA.md §3.10 对齐
PARSE_ARTIFACT_ROLES = (
    "parse_markdown",
    "parse_html",
    "parse_candidate_xlsx",
    "parse_final_xlsx",
)


def register_parse_artifact(
    session,
    archive_id: str,
    role: str,
    *,
    bucket: str,
    key: str,
    sha256: str,
    content_type: str,
    size: int,
):
    """把 worker 产出的 md/html/xlsx 注册成 archive_file 一行。

    Args:
        session: SQLAlchemy session（caller 负责 commit）
        archive_id: 档案 ID
        role: 必须是 PARSE_ARTIFACT_ROLES 之一
        bucket / key: MinIO 位置（与 Archive.candidate_xlsx_key 列同 key — 冗余存）
        sha256: 产物 SHA-256（hex）
        content_type / size: 给 FileAsset 行用

    Returns:
        ArchiveFile 行（已 flush，未 commit）

    实现：按 sha256 复用 FileAsset（去重）+ 建/更新 ArchiveFile 行。
    """
    if role not in PARSE_ARTIFACT_ROLES:
        raise ValueError(
            f"register_parse_artifact: role={role!r} not in {PARSE_ARTIFACT_ROLES}"
        )
    from sqlalchemy import select

    from app.models import Archive, ArchiveFile, FileAsset

    # 0. 拿 archive.tenant_code（FileAsset.tenant_code NOT NULL，幂等 dedup key 也用得上）
    archive = session.get(Archive, archive_id)
    if archive is None:
        raise ValueError(f"register_parse_artifact: archive {archive_id} 不存在")
    tenant_code = archive.tenant_code

    # 1. 建/复用 FileAsset（按 (tenant_code, sha256) 去重 — 与 UQ constraint 对齐）
    fa = session.execute(
        select(FileAsset).where(
            FileAsset.tenant_code == tenant_code,
            FileAsset.sha256 == sha256,
        )
    ).scalar_one_or_none()
    if fa is None:
        fa = FileAsset(
            tenant_code=tenant_code,
            sha256=sha256,
            bucket=bucket,
            object_key=key,
            file_name=key.split("/")[-1],
            mime_type=content_type,
            file_size=size,
        )
        session.add(fa)
        session.flush()  # 拿 fa.file_id

    # 2. 建/更新 ArchiveFile 行（page_range='' 让 (archive_id, file_role) 唯一）
    af = session.execute(
        select(ArchiveFile).where(
            ArchiveFile.archive_id == archive_id,
            ArchiveFile.file_role == role,
            ArchiveFile.page_range == "",
        )
    ).scalar_one_or_none()
    if af is None:
        af = ArchiveFile(
            archive_id=archive_id,
            file_id=fa.file_id,
            file_role=role,
            page_range="",
            is_primary=False,
            sort_order=200,  # 排在 main_document 之后
            link_source="worker",
            linked_by="quota_parser_worker",
        )
        session.add(af)
    else:
        af.file_id = fa.file_id
        af.linked_by = "quota_parser_worker"
    session.flush()
    return af


def enqueue_parse_job(
    session,
    archive_id: str,
    *,
    profile: str,
    created_by: str | None = None,
    mock: bool = False,
    ocr_api_url: str | None = None,
    province: str | None = None,
):
    """INSERT 一条 quota_parse_job (status='queued')。

    Args:
        session: SQLAlchemy session（caller 负责 commit）
        archive_id: 档案 ID
        profile: 'sichuan' / 'chongqing'
        created_by: 触发用户
        mock: 是否 mock 模式（worker 据此分流）
        ocr_api_url / province: 透传到 worker 进程的 metadata

    Returns:
        QuotaParseJob（已 flush，未 commit）

    Raises:
        ValueError: 同 archive_id 已有 active job（status='queued'/'running'）
                    → trigger 端点转 409
    """
    from sqlalchemy import select

    from app.models import QuotaParseJob

    if profile not in PROFILES:
        raise ValueError(f"profile {profile!r} 不在注册表 {list(PROFILES)}")

    active = session.execute(
        select(QuotaParseJob).where(
            QuotaParseJob.archive_id == archive_id,
            QuotaParseJob.status.in_(["queued", "running"]),
        )
    ).scalar_one_or_none()
    if active is not None:
        raise ValueError(
            f"archive {archive_id} 已有 active job {active.job_id} (status={active.status})"
        )

    job = QuotaParseJob(
        archive_id=archive_id,
        profile=profile,
        status="queued",
        created_by=created_by,
        metadata_payload={
            "mock": mock,
            "ocr_api_url": ocr_api_url,
            "province": province,
        },
    )
    session.add(job)
    session.flush()
    return job


def cancel_active_jobs(session, archive_id: str) -> int:
    """把该 archive 的所有 active job 标 cancelled（POST /parse/delete 调）。

    Returns:
        受影响行数
    """
    from sqlalchemy import update

    from app.models import QuotaParseJob

    result = session.execute(
        update(QuotaParseJob)
        .where(
            QuotaParseJob.archive_id == archive_id,
            QuotaParseJob.status.in_(["queued", "running"]),
        )
        .values(status="cancelled", finished_at=_now())
    )
    return result.rowcount or 0