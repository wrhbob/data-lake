"""
P0-4A · Quota domain API endpoints.

capabilities / stats / dictionaries / facets / archives / reconciliation.

All counts computed via SQL aggregation; no all-table Python-side counting.
"""

from datetime import date as date_type, datetime, timezone
from pathlib import Path
import os
import re
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Response, UploadFile
from pydantic import BaseModel
from sqlalchemy import delete, func, select, case, desc, asc
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.archive_rules import build_quota_business_key, metadata_cell, now_iso
from app.archive_service import attach_file as _attach_archive_file, create_archive, UI_EXCLUDED_FILE_ROLES
from app.assets import register_asset
from app.database import get_db_session
from app.models import (
    AdministrativeDivision,
    Archive,
    ArchiveEvent,
    ArchiveFile,
    CrawlLineage,
    DataSource,
    FileAsset,
    IngestEvent,
    Outbox,
    QuotaArchiveProfile,
    QuotaDictionary,
    QuotaProjectionCandidate,
    QuotaPublicationSet,
    QuotaPublicationRelation,
)
from app.storage import ObjectStore, get_object_store

router = APIRouter(prefix="/api/data-lake/quota", tags=["quota"])

# ── helpers ─────────────────────────────────────────────────────────────

_SCHEMA_VERSION = "quota-r1"
_FEATURES = {
    "stats": "ready",
    "facets": "ready",
    "archives": "ready",
    "reconciliation": "ready",
    "dictionaries": "ready",
    "jurisdictions": "ready",
    "compose": "ready",
    "compare": "ready",
    "archiveFiles": "unavailable",
    # 2026-08-17: 覆盖矩阵接入真实接口（/api/quota/coverage-matrix），前端 quota-ui.js 覆盖矩阵 tab 已接线
    "coverage": "ready",
    "publicationSets": "unavailable",
}

MIME = {"Content-Type": "application/json; charset=utf-8"}


def _json_ok(response: Response, body: object) -> object:
    response.headers.update(MIME)
    return body


def _safe_int(val) -> int:
    return int(val) if val is not None else 0


# ── 1. capabilities ─────────────────────────────────────────────────────

@router.get("/capabilities")
def get_capabilities(response: Response) -> dict:
    return _json_ok(response, {
        "schema_version": _SCHEMA_VERSION,
        "features": dict(_FEATURES),
    })


# ── 2. stats ────────────────────────────────────────────────────────────

@router.get("/stats")
def get_stats(response: Response, session: Session = Depends(get_db_session)) -> dict:
    systems = _safe_int(session.scalar(select(func.count()).select_from(QuotaPublicationSet)))
    archived = _safe_int(session.scalar(
        select(func.count()).select_from(Archive).where(Archive.domain_type == "quota")
    ))

    # raw_total: DISTINCT file_asset_id via two-leg UNION
    leg1 = select(FileAsset.file_id).join(
        IngestEvent, IngestEvent.file_id == FileAsset.file_id
    ).join(
        DataSource, DataSource.source_id == IngestEvent.source_id
    ).where(DataSource.data_domain == "quota")

    leg2 = select(FileAsset.file_id).join(
        ArchiveFile, ArchiveFile.file_id == FileAsset.file_id
    ).join(
        Archive, ArchiveFile.archive_id == Archive.archive_id
    ).where(Archive.domain_type == "quota")

    union_cte = leg1.union(leg2).cte("quota_file_ids")
    raw_total = _safe_int(session.scalar(
        select(func.count()).select_from(select(union_cte.c.file_id).distinct().subquery())
    ))

    # projection status counts
    proj_rows = dict(session.execute(
        select(
            QuotaProjectionCandidate.projection_status,
            func.count(),
        ).group_by(QuotaProjectionCandidate.projection_status)
    ).all())

    pending = _safe_int(proj_rows.get("pending", 0))
    linked = _safe_int(proj_rows.get("linked", 0))
    duplicate = _safe_int(proj_rows.get("duplicate", 0))
    invalid = _safe_int(proj_rows.get("invalid", 0))
    ignored = _safe_int(proj_rows.get("ignored", 0))

    return _json_ok(response, {
        "systems": systems,
        "archived": archived,
        "raw_total": raw_total,
        "accessible": None,
        "pending": pending,
        "linked": linked,
        "duplicate": duplicate,
        "invalid": invalid,
        "ignored": ignored,
        "storage_audit_at": None,
    })


# ── 3. dictionaries ─────────────────────────────────────────────────────

@router.get("/dictionaries")
def get_dictionaries(
    response: Response,
    type: str | None = Query(None, alias="type"),
    session: Session = Depends(get_db_session),
) -> dict:
    stmt = select(QuotaDictionary).where(QuotaDictionary.enabled.is_(True))
    if type:
        stmt = stmt.where(QuotaDictionary.dict_type == type)
    stmt = stmt.order_by(QuotaDictionary.sort_order, QuotaDictionary.label)
    rows = list(session.execute(stmt).scalars().all())
    return _json_ok(response, {
        "items": [{
            "dict_type": r.dict_type,
            "code": r.code,
            "label": r.label,
            "sort_order": r.sort_order,
        } for r in rows],
    })


# ── 4. facets ───────────────────────────────────────────────────────────

def _facet_years(session: Session, primary: str | None = None, jurisdiction_code: str | None = None) -> list[dict]:
    stmt = select(
        QuotaPublicationSet.edition_year,
        func.count().label("cnt"),
    )
    # v0.9 (2026-08-18): primary 改用 book_category 单字段 (生成列)。
    #   旧 3 分支 (quota_system_type + material_type) 已废弃, 见 scripts/migration_2026_08_18_book_category.sql。
    if primary and primary != "all":
        stmt = stmt.where(QuotaPublicationSet.book_category == primary)
    if jurisdiction_code:
        stmt = stmt.where(QuotaPublicationSet.jurisdiction_code == jurisdiction_code)
    stmt = stmt.where(QuotaPublicationSet.edition_year.isnot(None))
    stmt = stmt.group_by(QuotaPublicationSet.edition_year).order_by(desc(QuotaPublicationSet.edition_year))
    rows = session.execute(stmt).all()
    years = []
    for year, cnt in rows:
        y = str(year)
        # fetch editions for this year
        ed_stmt = select(
            QuotaPublicationSet.edition_label,
            func.count().label("ed_cnt"),
        ).where(
            # 注意: edition_year 是 int 列, 必须用整数 year 比较
            # 用 str(year)==y 会触发 psycopg 'integer = character varying' 500
            QuotaPublicationSet.edition_year == year,
            QuotaPublicationSet.edition_label.isnot(None),
        )
        if jurisdiction_code:
            ed_stmt = ed_stmt.where(QuotaPublicationSet.jurisdiction_code == jurisdiction_code)
        ed_stmt = ed_stmt.group_by(QuotaPublicationSet.edition_label)
        ed_rows = session.execute(ed_stmt).all()
        editions = [{"value": str(el), "label": str(el), "count": int(ec)} for el, ec in ed_rows]
        years.append({"value": y, "label": y, "count": int(cnt), "editions": editions})
    return years


@router.get("/facets")
def get_facets(
    response: Response,
    primary: str | None = Query(None, description="all | boq_standard | construction_quota | industry_quota (book_category)"),
    jurisdiction_code: str | None = Query(None, description="filter years by jurisdiction (construction_quota only)"),
    session: Session = Depends(get_db_session),
) -> dict:
    jurisdictions: list[dict] = []
    industries: list[dict] = []
    scopes: list[dict] = []
    years: list[dict] = []

    # v0.9 (2026-08-18): primary 改用 book_category 单字段 (生成列)。
    #   旧 3 分支 (quota_system_type + material_type) 已废弃, 见 scripts/migration_2026_08_18_book_category.sql。
    if primary == "construction_quota":
        j_rows = session.execute(
            select(
                QuotaPublicationSet.jurisdiction_code,
                func.count().label("cnt"),
            ).where(
                QuotaPublicationSet.book_category == "construction_quota",
                QuotaPublicationSet.jurisdiction_code.isnot(None),
            ).group_by(QuotaPublicationSet.jurisdiction_code)
        ).all()
        for code, cnt in j_rows:
            div = session.scalar(
                select(AdministrativeDivision.name).where(AdministrativeDivision.code == code)
            )
            label = div or str(code)
            jurisdictions.append({"code": str(code), "label": label, "count": int(cnt)})
        years = _facet_years(session, "construction_quota", jurisdiction_code=jurisdiction_code)

    elif primary == "industry_quota":
        i_rows = session.execute(
            select(
                QuotaPublicationSet.industry_sector_code,
                func.count().label("cnt"),
            ).where(
                QuotaPublicationSet.book_category == "industry_quota",
                QuotaPublicationSet.industry_sector_code.isnot(None),
            ).group_by(QuotaPublicationSet.industry_sector_code)
        ).all()
        for code, cnt in i_rows:
            label_row = session.scalar(
                select(QuotaDictionary.label).where(
                    QuotaDictionary.dict_type == "industry_sector",
                    QuotaDictionary.code == code,
                )
            )
            label = label_row or str(code)
            industries.append({"code": str(code), "label": label, "count": int(cnt)})
        years = _facet_years(session, "industry_quota")

    elif primary == "boq_standard":
        # v0.9.3 (2026-08-18): 与 construction_quota 同形, 按 jurisdiction_code group 出 jurisdictions。
        #   之前按 title group 出的 scopes 只有 1 选项 ("深圳市2019定额..."), 实际等于无筛选。
        #   前端 jurisdiction chip 走静态 PROVINCE_REGIONS 不读 fx.jurisdictions (HOTFIX-QA-CHIP-001),
        #   此处返回 jurisdictions 仅保持 3 类 facets 返回结构对齐 (便于审计 / 调试)。
        j_rows = session.execute(
            select(
                QuotaPublicationSet.jurisdiction_code,
                func.count().label("cnt"),
            ).where(
                QuotaPublicationSet.book_category == "boq_standard",
                QuotaPublicationSet.jurisdiction_code.isnot(None),
            ).group_by(QuotaPublicationSet.jurisdiction_code)
        ).all()
        for code, cnt in j_rows:
            div = session.scalar(
                select(AdministrativeDivision.name).where(AdministrativeDivision.code == code)
            )
            label = div or str(code)
            jurisdictions.append({"code": str(code), "label": label, "count": int(cnt)})
        years = _facet_years(session, "boq_standard")

    else:
        years = _facet_years(session)

    return _json_ok(response, {
        "jurisdictions": jurisdictions,
        "industries": industries,
        "scopes": scopes,
        "years": years,
        "disciplines": [],
        "materialTypes": [],
        "cities": [],
        "issuers": [],
        "metadataStatuses": [],
        "archiveStatuses": [],
        "sourceChannels": [],
        "fileFormats": [],
    })


# ── 5. archives ─────────────────────────────────────────────────────────

@router.get("/archives")
def list_quota_archives(
    response: Response,
    q: str | None = Query(None, description="search title"),
    primary: str | None = Query(None, description="all | boq_standard | construction_quota | industry_quota (book_category)"),
    jurisdiction_code: str | None = Query(None),
    industry_sector_code: str | None = Query(None),
    edition_year: str | None = Query(None),
    edition_label: str | None = Query(None),
    discipline_code: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort_by: str | None = Query(None),
    sort_order: str | None = Query("desc"),
    session: Session = Depends(get_db_session),
) -> dict:
    # base: Archive with domain_type=quota
    stmt = (
        select(Archive, QuotaArchiveProfile, QuotaPublicationSet)
        .join(QuotaArchiveProfile, QuotaArchiveProfile.archive_id == Archive.archive_id, isouter=True)
        .join(QuotaPublicationSet, QuotaPublicationSet.publication_set_id == QuotaArchiveProfile.publication_set_id, isouter=True)
        .where(Archive.domain_type == "quota")
    )

    if q:
        stmt = stmt.where(Archive.title.ilike(f"%{q}%"))
    # v0.9 (2026-08-18): primary 改用 book_category 单字段 (生成列)。
    if primary and primary != "all":
        stmt = stmt.where(QuotaPublicationSet.book_category == primary)
    if jurisdiction_code:
        stmt = stmt.where(QuotaPublicationSet.jurisdiction_code == jurisdiction_code)
    if industry_sector_code:
        stmt = stmt.where(QuotaPublicationSet.industry_sector_code == industry_sector_code)
    if edition_year:
        stmt = stmt.where(QuotaPublicationSet.edition_year == edition_year)
    if edition_label:
        stmt = stmt.where(QuotaPublicationSet.edition_label == edition_label)
    if discipline_code:
        stmt = stmt.where(QuotaArchiveProfile.discipline_code == discipline_code)

    # sort
    sort_col = Archive.created_at
    if sort_by == "title":
        sort_col = Archive.title
    elif sort_by == "edition_year":
        sort_col = QuotaPublicationSet.edition_year
    stmt = stmt.order_by(desc(sort_col) if sort_order == "desc" else asc(sort_col))

    # count (与 /stats 同源:默认不 join profile/pubset;filter 含相关字段时按需 join)
    needs_pubset = bool(primary or jurisdiction_code or industry_sector_code or edition_year or edition_label)
    needs_profile = bool(discipline_code) or needs_pubset
    count_stmt = select(func.count()).select_from(Archive).where(Archive.domain_type == "quota")
    if needs_profile:
        count_stmt = count_stmt.join(
            QuotaArchiveProfile, QuotaArchiveProfile.archive_id == Archive.archive_id, isouter=True
        )
    if needs_pubset:
        count_stmt = count_stmt.join(
            QuotaPublicationSet,
            QuotaPublicationSet.publication_set_id == QuotaArchiveProfile.publication_set_id,
            isouter=True,
        )
    if q:
        count_stmt = count_stmt.where(Archive.title.ilike(f"%{q}%"))
    # v0.9 (2026-08-18): 与主查询一致, count_stmt 也用 book_category 单字段。
    if primary and primary != "all":
        count_stmt = count_stmt.where(QuotaPublicationSet.book_category == primary)
    if jurisdiction_code:
        count_stmt = count_stmt.where(QuotaPublicationSet.jurisdiction_code == jurisdiction_code)
    if industry_sector_code:
        count_stmt = count_stmt.where(QuotaPublicationSet.industry_sector_code == industry_sector_code)
    if edition_year:
        count_stmt = count_stmt.where(QuotaPublicationSet.edition_year == edition_year)
    if edition_label:
        count_stmt = count_stmt.where(QuotaPublicationSet.edition_label == edition_label)
    if discipline_code:
        count_stmt = count_stmt.where(QuotaArchiveProfile.discipline_code == discipline_code)
    total = _safe_int(session.scalar(count_stmt))

    offset = (page - 1) * page_size
    rows = session.execute(stmt.offset(offset).limit(page_size)).all()

    items = []
    for archive, profile, pubset in rows:
        archive_file_count = _safe_int(session.scalar(
            select(func.count())
            .select_from(ArchiveFile)
            .where(
                ArchiveFile.archive_id == archive.archive_id,
                # 2026-08-03: UI 过滤 parse_markdown / parse_html 两个中间产物 role,
                # 它们是 worker 写出的 OCR/MD 中间产物, 不进用户可见 file_count。
                # 同时排除 parse_candidate_xlsx / parse_final_xlsx —— 它们与 archive 表的
                # candidate_xlsx_key / final_xlsx_key 是同一份文件, 已在 parse_artifact_count
                # 里计数 (避免重复)。用户期望: 已完成=3 (PDF+candidate+final),
                # 待审核=2 (PDF+candidate), 解析中=1 (PDF only)。
                ArchiveFile.file_role.notin_([
                    "parse_markdown", "parse_html",
                    "parse_candidate_xlsx", "parse_final_xlsx",
                ]),
            )
        ))
        # 同 archive_service.py:_archive_summary_rows 一致:把 parse 产物 (candidate.xlsx / final.xlsx)
        # 也算进 file_count(它们落 MinIO 但不挂 ArchiveFile)。quota 档案期望:1 PDF + 1 candidate + 1 final = 3。
        parse_artifact_count = (
            (1 if archive.candidate_xlsx_key else 0)
            + (1 if archive.final_xlsx_key else 0)
        )
        file_count = archive_file_count + parse_artifact_count
        items.append({
            "archive_id": archive.archive_id,
            "title": archive.title,
            "domain_type": archive.domain_type,
            "status": archive.status,
            # v0.4 Bug#1 修：list 端点必须透出 parse_* 字段,前端 resolveUiStatus
            # 直接读 parse_status(原本读 row.status 命中的是 archive.status='pending_tag',
            # 不在 PARSE_STATUS_VARIANT 表里,导致所有行永远渲染"未解析")。
            # 保留 status 字段以兼容旧客户端;parse_status 才是驱动 5 态徽章的真值。
            "parse_status": archive.parse_status,
            "parse_phase": archive.parse_phase,
            "parse_task_id": archive.parse_task_id,
            "parse_started_at": archive.parse_started_at.isoformat() if archive.parse_started_at else None,
            "parse_finished_at": archive.parse_finished_at.isoformat() if archive.parse_finished_at else None,
            "parse_error_code": archive.parse_error_code,
            "candidate_xlsx_key": archive.candidate_xlsx_key,
            "final_xlsx_key": archive.final_xlsx_key,
            "business_key": archive.business_key,
            "region_code": archive.region_code,
            "publication_set_title": pubset.title if pubset else None,
            "quota_system_type": pubset.quota_system_type if pubset else None,
            "jurisdiction_code": pubset.jurisdiction_code if pubset else None,
            "industry_sector_code": pubset.industry_sector_code if pubset else None,
            "edition_year": pubset.edition_year if pubset else None,
            "edition_label": pubset.edition_label if pubset else None,
            "standard_or_quota_code": pubset.standard_or_quota_code if pubset else None,
            "volume_title": profile.volume_title if profile else None,
            "discipline_code": profile.discipline_code if profile else None,
            "document_role": profile.document_role if profile else None,
            "metadata_status": profile.metadata_status if profile else None,
            "completeness_score": profile.completeness_score if profile else None,
            "file_count": file_count,
        })

    response.headers.update({
        **MIME,
        "X-Total-Count": str(total),
        "X-Page": str(page),
        "X-Page-Size": str(page_size),
    })
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/archives/{archive_id}")
def get_quota_archive_detail(
    response: Response,
    archive_id: str,
    session: Session = Depends(get_db_session),
) -> dict:
    """Read-only archive detail: archive + publication_set + volumes + files.

    Does NOT advance the state machine. Status transitions happen elsewhere.
    """
    main_row = session.execute(
        select(Archive, QuotaArchiveProfile, QuotaPublicationSet)
        .join(
            QuotaArchiveProfile,
            QuotaArchiveProfile.archive_id == Archive.archive_id,
            isouter=True,
        )
        .join(
            QuotaPublicationSet,
            QuotaPublicationSet.publication_set_id == QuotaArchiveProfile.publication_set_id,
            isouter=True,
        )
        .where(Archive.archive_id == archive_id, Archive.domain_type == "quota")
    ).first()
    if main_row is None:
        raise HTTPException(status_code=404, detail="档案不存在或已删除")

    archive, profile, pubset = main_row

    archive_file_count = _safe_int(session.scalar(
        select(func.count())
        .select_from(ArchiveFile)
        .where(
            ArchiveFile.archive_id == archive.archive_id,
            # 2026-08-03: 同 list 端点, 过滤 parse_markdown / parse_html 中间产物,
            # 同时排除 parse_candidate_xlsx / parse_final_xlsx 避免与 candidate/final_key 重复。
            ArchiveFile.file_role.notin_([
                "parse_markdown", "parse_html",
                "parse_candidate_xlsx", "parse_final_xlsx",
            ]),
        )
    ))
    # 同 list_quota_archives 注释:把 parse 产物 (candidate/final xlsx) 也算进 file_count。
    parse_artifact_count = (
        (1 if archive.candidate_xlsx_key else 0)
        + (1 if archive.final_xlsx_key else 0)
    )
    file_count = archive_file_count + parse_artifact_count

    archive_obj = {
        "archive_id": archive.archive_id,
        "title": archive.title,
        "domain_type": archive.domain_type,
        "status": archive.status,
        # v0.4 Bug#1 第四层修:detail 端点也补 8 个 parse_* 字段(与 list 端点 L389-405 对齐)。
        # 之前只有 status='pending_tag' 不在 PARSE_STATUS_VARIANT 表里,前端 resolveUiStatus
        # 永远渲染"未解析"。修完后 detail 页能直接驱动 5 态徽章 + 文件下载按钮启用。
        "parse_status": archive.parse_status,
        "parse_phase": archive.parse_phase,
        "parse_task_id": archive.parse_task_id,
        "parse_started_at": archive.parse_started_at.isoformat() if archive.parse_started_at else None,
        "parse_finished_at": archive.parse_finished_at.isoformat() if archive.parse_finished_at else None,
        "parse_error_code": archive.parse_error_code,
        "candidate_xlsx_key": archive.candidate_xlsx_key,
        "final_xlsx_key": archive.final_xlsx_key,
        "business_key": archive.business_key,
        "region_code": archive.region_code,
        "publication_set_id": profile.publication_set_id if profile else None,
        "volume_code": profile.volume_code if profile else None,
        "volume_title": profile.volume_title if profile else None,
        "discipline_code": profile.discipline_code if profile else None,
        "document_role": profile.document_role if profile else None,
        "metadata_status": profile.metadata_status if profile else None,
        "completeness_score": profile.completeness_score if profile else None,
        "file_count": file_count,
        "created_at": archive.created_at.isoformat() if archive.created_at else None,
        "updated_at": archive.updated_at.isoformat() if archive.updated_at else None,
    }

    publication_set_obj = None
    if pubset is not None:
        publication_set_obj = {
            "publication_set_id": pubset.publication_set_id,
            "biz_key": pubset.biz_key,
            "title": pubset.title,
            "material_type": pubset.material_type,
            "quota_system_type": pubset.quota_system_type,
            "jurisdiction_level": pubset.jurisdiction_level,
            "jurisdiction_code": pubset.jurisdiction_code,
            "industry_sector_code": pubset.industry_sector_code,
            "issuer_name": pubset.issuer_name,
            "standard_or_quota_code": pubset.standard_or_quota_code,
            "edition_label": pubset.edition_label,
            "edition_year": pubset.edition_year,
            "publish_date": pubset.publish_date.isoformat() if pubset.publish_date else None,
            "effective_date": pubset.effective_date.isoformat() if pubset.effective_date else None,
            "legal_status": pubset.legal_status,
            "metadata_status": pubset.metadata_status,
        }

    # ── volumes (sibling archives sharing the same publication_set_id) ──
    volumes: list[dict] = []
    if profile is not None and profile.publication_set_id:
        vol_rows = session.execute(
            select(Archive, QuotaArchiveProfile)
            .join(
                QuotaArchiveProfile,
                QuotaArchiveProfile.archive_id == Archive.archive_id,
                isouter=True,
            )
            .where(QuotaArchiveProfile.publication_set_id == profile.publication_set_id)
            .order_by(asc(QuotaArchiveProfile.part_no), asc(Archive.title))
        ).all()
        for a, p in vol_rows:
            fc = _safe_int(session.scalar(
                select(func.count())
                .select_from(ArchiveFile)
                .where(ArchiveFile.archive_id == a.archive_id)
            ))
            volumes.append({
                "archive_id": a.archive_id,
                "title": a.title,
                "status": a.status,
                "volume_code": p.volume_code if p else None,
                "volume_title": p.volume_title if p else None,
                "discipline_code": p.discipline_code if p else None,
                "document_role": p.document_role if p else None,
                "file_count": fc,
                "is_current": a.archive_id == archive.archive_id,
            })

    # ── files (ArchiveFile LEFT JOIN FileAsset) ────────────────────────
    # 2026-08-04: 应用 UI_EXCLUDED_FILE_ROLES 过滤, 修复左侧栏多出 3 个 parse_* 按钮.
    # parse_markdown/parse_html/parse_candidate_xlsx 是 worker 写出的中间产物,
    # parse_final_xlsx 与 archive.final_xlsx_key 同一份 (前端 archiveFiles() 已 virtual 合成).
    # 这 4 行 archive_file 都不应在 UI 显示. 与 archive_service.py:list_archives() 的
    # file_count 计算保持一致, 避免列表/详情两个 API 行为分裂.
    files: list[dict] = []
    file_rows = session.execute(
        select(ArchiveFile, FileAsset)
        .join(FileAsset, FileAsset.file_id == ArchiveFile.file_id, isouter=True)
        .where(ArchiveFile.archive_id == archive.archive_id)
        .order_by(asc(ArchiveFile.sort_order), asc(ArchiveFile.added_at))
    ).all()
    for af, fa in file_rows:
        if af.file_role in UI_EXCLUDED_FILE_ROLES:
            continue
        files.append({
            "archive_file_id": af.archive_file_id,
            "file_id": af.file_id,
            "display_name": af.display_name or (fa.file_name if fa else None),
            "file_role": af.file_role,
            "is_primary": af.is_primary,
            "fetch_status": af.fetch_status,
            "representation_role": af.representation_role,
            "page_range": af.page_range,
            "mime_type": fa.mime_type if fa else None,
            "size_bytes": fa.file_size if fa else None,
            "file_name": fa.file_name if fa else None,
        })

    # ── parse 子对象（quota_parser integration, 2026-07-28 · web-frontend SPEC §6.1） ──
    from app.quota_parser.service import build_parse_section

    parse_obj = build_parse_section(archive)

    return _json_ok(response, {
        "archive": archive_obj,
        "publication_set": publication_set_obj,
        "volumes": volumes,
        "files": files,
        "parse": parse_obj,
    })


# ── 6. reconciliation ──────────────────────────────────────────────────

@router.get("/reconciliation")
def get_reconciliation(
    response: Response,
    projection_status: str | None = Query(None, description="pending|linked|duplicate|invalid|ignored"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    session: Session = Depends(get_db_session),
) -> dict:
    # summary counts
    count_rows = dict(session.execute(
        select(
            QuotaProjectionCandidate.projection_status,
            func.count(),
        ).group_by(QuotaProjectionCandidate.projection_status)
    ).all())
    pending = _safe_int(count_rows.get("pending", 0))
    linked = _safe_int(count_rows.get("linked", 0))
    duplicate = _safe_int(count_rows.get("duplicate", 0))
    invalid = _safe_int(count_rows.get("invalid", 0))
    ignored = _safe_int(count_rows.get("ignored", 0))

    # detail query: join file_asset, ingest_event, data_source
    stmt = (
        select(
            QuotaProjectionCandidate,
            FileAsset.file_name,
            FileAsset.file_size,
            FileAsset.sha256,
            FileAsset.mime_type,
            DataSource.name,
            DataSource.source_type,
            IngestEvent.ingested_at,
        )
        .join(FileAsset, FileAsset.file_id == QuotaProjectionCandidate.file_id, isouter=True)
        .join(IngestEvent, IngestEvent.file_id == FileAsset.file_id, isouter=True)
        .join(DataSource, DataSource.source_id == IngestEvent.source_id, isouter=True)
    )
    if projection_status:
        stmt = stmt.where(QuotaProjectionCandidate.projection_status == projection_status)
    stmt = stmt.order_by(desc(QuotaProjectionCandidate.suggestion_confidence))

    # count
    count_q = select(func.count()).select_from(QuotaProjectionCandidate)
    if projection_status:
        count_q = count_q.where(QuotaProjectionCandidate.projection_status == projection_status)
    total = _safe_int(session.scalar(count_q))

    offset = (page - 1) * page_size
    rows = session.execute(stmt.offset(offset).limit(page_size)).all()

    items = []
    for cand, filename, filesize, filehash, mimetype, src_name, src_type, ingested_at in rows:
        items.append({
            "candidate_id": cand.candidate_id,
            "file_id": cand.file_id,
            "file_name": filename,
            "file_size": filesize,
            "file_hash": filehash,
            "file_type": mimetype,
            "projection_status": cand.projection_status,
            "suggested_metadata": cand.suggested_metadata,
            "suggestion_confidence": float(cand.suggestion_confidence or 0),
            "matched_rules": cand.matched_rules,
            "duplicate_of_file_id": cand.duplicate_of_file_id,
            "source_name": src_name,
            "source_type": src_type,
            "ingested_at": ingested_at.isoformat() if ingested_at else None,
        })

    return _json_ok(response, {
        "pending": pending,
        "linked": linked,
        "duplicate": duplicate,
        "invalid": invalid,
        "ignored": ignored,
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": items,
    })


# ── 7. compose (POST) ──────────────────────────────────────────────────


class _ComposeVolume(BaseModel):
    volume_title: str = ""
    discipline_code: str = ""
    files: list[dict] = []


class ComposeRequest(BaseModel):
    action: str
    material_type: str = "quota_base"
    systemType: str = ""
    path: dict = {}
    set: dict = {}
    relation: dict = {}
    targetKind: str = ""
    targetId: str = ""
    targetLabel: str = ""
    volumes: list[_ComposeVolume] = []
    unassignedFiles: list[dict] = []
    supplementFiles: list[dict] = []
    tenant: str = "platform_public"


@router.post("/compose")
def compose_quota(
    response: Response,
    body: ComposeRequest,
    session: Session = Depends(get_db_session),
) -> dict:
    try:
        # ── determine material_type from compose state ────────────────
        material_type = body.set.get("material_type") or body.material_type or "quota_base"
        quota_system_type = body.systemType or None

        # ── validate jurisdiction for construction_regional ───────────
        jurisdiction_code = body.path.get("jurisdiction_code", "") if body.path else ""
        jurisdiction_level = body.path.get("jurisdiction_level", "") if body.path else ""
        if quota_system_type == "construction_regional" and not jurisdiction_code:
            return _json_ok(response, {
                "ok": False,
                "error": "REGIONAL_REQUIRES_JURISDICTION",
                "detail": "地方定额必须选择省份",
            })

        # ── create or resolve QuotaPublicationSet ─────────────────────
        set_data = body.set or {}
        edition_year_raw = set_data.get("edition_year", "")
        edition_year = int(edition_year_raw) if edition_year_raw else None
        title = set_data.get("title", "") or "未命名定额体系"

        # Build biz_key: quota:{material_type}:{system_type}:{jurisdiction}:{sector}:{family}:{edition}
        family_code = _make_publication_family_code(body)
        biz_key_parts = [
            "quota",
            material_type,
            quota_system_type or "general",
            jurisdiction_code or "national",
            body.path.get("industry_sector_code", "") or "general",
            family_code,
            set_data.get("edition_label", edition_year_raw or ""),
        ]
        biz_key = ":".join(biz_key_parts)

        # Upsert publication set
        existing_pub_set = session.scalar(
            select(QuotaPublicationSet).where(QuotaPublicationSet.biz_key == biz_key)
        )
        if existing_pub_set is not None:
            pub_set = existing_pub_set
        else:
            pub_set = QuotaPublicationSet(
                publication_set_id=str(uuid4()),
                biz_key=biz_key,
                publication_family_code=family_code,
                title=title,
                material_type=material_type,
                quota_system_type=quota_system_type,
                jurisdiction_level=jurisdiction_level or "province",
                jurisdiction_code=jurisdiction_code or "000000",
                industry_sector_code=body.path.get("industry_sector_code") or None,
                issuer_name=set_data.get("issuer_name", "") or _default_issuer(jurisdiction_code, set_data),
                standard_or_quota_code=set_data.get("standard_or_quota_code") or None,
                edition_label=set_data.get("edition_label", "") or (str(edition_year) if edition_year else ""),
                edition_year=edition_year,
                publish_date=_parse_date(set_data.get("publish_date")),
                effective_date=_parse_date(set_data.get("effective_date")),
                legal_status=set_data.get("legal_status", "unknown") or "unknown",
                metadata_status="partial",
                tenant_code=body.tenant,
                visibility_scope="public",
            )
            session.add(pub_set)
            session.flush()

        # ── handle relations ──────────────────────────────────────────
        rel_data = body.relation or {}
        related_id = rel_data.get("related_publication_set_id")
        rel_type = rel_data.get("relation_type")
        if related_id and rel_type:
            existing_rel = session.scalar(
                select(QuotaPublicationRelation).where(
                    QuotaPublicationRelation.source_publication_set_id == pub_set.publication_set_id,
                    QuotaPublicationRelation.target_publication_set_id == related_id,
                    QuotaPublicationRelation.relation_type == rel_type,
                )
            )
            if existing_rel is None:
                session.add(QuotaPublicationRelation(
                    relation_id=str(uuid4()),
                    source_publication_set_id=pub_set.publication_set_id,
                    target_publication_set_id=related_id,
                    relation_type=rel_type,
                ))

        # ── create Archive + QuotaArchiveProfile per volume ───────────
        source_id = f"quota-compose-{pub_set.publication_set_id[:8]}"

        # Ensure DataSource exists before create_archive (it checks DataSource）
        _ensure_compose_source(session, source_id, body.tenant)
        session.flush()

        created_archives: list[dict] = []

        for i, vol in enumerate(body.volumes):
            vol_title = vol.volume_title or f"第{i + 1}分册"
            business_key = build_quota_business_key(
                source_id=source_id,
                title=f"{title}—{vol_title}",
                region=jurisdiction_code or None,
                quota_code=set_data.get("standard_or_quota_code") or None,
                version=str(edition_year) if edition_year else None,
                effective_date=set_data.get("effective_date") or None,
                standard_no=set_data.get("standard_or_quota_code") or None,
            )
            # create_archive internally calls session.commit() — service layer contract
            archive = create_archive(
                session,
                domain_type="quota",
                channel_type="manual_upload",
                business_key=business_key,
                title=f"{title}—{vol_title}",
                source_id=source_id,
                tenant_code=body.tenant,
                visibility_scope="public",
                status="pending_tag",
                metadata={"_compose_provenance": _provenance_cell(material_type)},
                field_sources=_build_archive_field_sources(
                    title=f"{title}—{vol_title}",
                    business_key=business_key,
                    region_code=jurisdiction_code or "",
                ),
            )
            # Upsert profile (session was committed by create_archive, so add needs fresh flush)
            existing_profile = session.get(QuotaArchiveProfile, archive.archive_id)
            if existing_profile is None:
                profile = QuotaArchiveProfile(
                    archive_id=archive.archive_id,
                    publication_set_id=pub_set.publication_set_id,
                    document_role=_map_action_to_document_role(body.action),
                    discipline_code=vol.discipline_code or None,
                    volume_code="",
                    volume_title=vol_title,
                    metadata_status="partial",
                    completeness_score=0,
                )
                session.add(profile)

            created_archives.append({
                "archive_id": archive.archive_id,
                "title": archive.title,
                "volume_title": vol_title,
                "file_count": len(vol.files),
            })

        session.commit()

        return _json_ok(response, {
            "ok": True,
            "publication_set_id": pub_set.publication_set_id,
            "biz_key": biz_key,
            "archives": created_archives,
        })
    except IntegrityError as e:
        session.rollback()
        return _json_ok(response, {
            "ok": False,
            "error": "INTEGRITY_ERROR",
            "detail": str(e.orig) if e.orig else str(e),
        })
    except Exception as e:
        session.rollback()
        return _json_ok(response, {
            "ok": False,
            "error": "COMPOSE_FAILED",
            "detail": str(e),
        })


# ── compose helpers ────────────────────────────────────────────────────


def _make_publication_family_code(body: ComposeRequest) -> str:
    """Derive a stable family code from jurisdiction + system_type + material_type."""
    jc = (body.path or {}).get("jurisdiction_code", "") or "national"
    st = body.systemType or "general"
    mt = ((body.set or {}).get("material_type") or body.material_type or "quota_base")
    return f"{jc}-{st}-{mt}"


def _default_issuer(jurisdiction_code: str, set_data: dict) -> str:
    """Guess issuer name when not provided."""
    name = (set_data.get("issuer_name") or "").strip()
    if name:
        return name
    if jurisdiction_code == "510000":
        return "四川省住房和城乡建设厅"
    return ""


def _parse_date(val: object) -> date_type | None:
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return date_type.fromisoformat(s[:10])
    except (ValueError, TypeError):
        return None


def _build_archive_field_sources(
    *, title: str, business_key: str, region_code: str = ""
) -> dict:
    """Build field_sources for create_archive C-class fields."""
    ts = now_iso()
    sources: dict = {}
    for field_name in ("domain_type", "channel_type"):
        sources[field_name] = metadata_cell(
            "quota" if field_name == "domain_type" else "manual_upload",
            source_level="manual", tagged_by="system", tagged_at=ts,
        )
    if title:
        sources["title"] = metadata_cell(title, source_level="manual", tagged_by="system", tagged_at=ts)
    if business_key:
        sources["business_key"] = metadata_cell(business_key, source_level="manual", tagged_by="system", tagged_at=ts)
    if region_code:
        sources["region_code"] = metadata_cell(region_code, source_level="manual", tagged_by="system", tagged_at=ts)
    return sources


def _provenance_cell(value: object) -> dict:
    return {
        "source_level": "manual",
        "tagged_by": "system",
        "tagged_at": datetime.now(timezone.utc).isoformat(),
        "value": value,
    }


def _map_action_to_document_role(action: str) -> str:
    if action in ("new_set", "new_boq"):
        return "main_volume"
    if action == "add_volume":
        return "main_volume"
    if action == "supplement":
        return "other"
    return "main_volume"


def _ensure_compose_source(session: Session, source_id: str, tenant_code: str) -> None:
    existing = session.get(DataSource, source_id)
    if existing is not None:
        return
    ds = DataSource(
        source_id=source_id,
        source_scope="platform_public",
        managed_by="platform",
        source_type="manual",
        connector_type="manual_upload",
        name=f"Quota Compose ({source_id[-8:]})",
        data_domain="quota",
        format="pdf",
        config={},
        status="active",
        created_by="system",
        asset_tenant_code=tenant_code,
    )
    session.add(ds)


# ── 8. upload (POST /upload) · HOTFIX-QA-UPLOAD-001 ────────────────────
#
# 极简上传闭环：PDF → MinIO → FileAsset → Archive(domain=quota) →
# ArchiveFile(main_document, is_primary=True)。
#
# 设计约束（任务单明令）：
# - file_role 恒为 main_document，is_primary=True（资料分类 category 是另一维度，
#   只写入 Archive.metadata_payload，绝不映射到 file_role）。
# - business_key 基于 sha256 而非标题：同名异内容 → 不同 Archive；
#   异名同内容 → 同一 Archive（幂等）。
# - 严禁把 industry_quota（前端 category）写成 industry_specialty
#   （quota_system_type）；二者语义不同。
# - 单文件粒度事务：一文件失败不影响其它，逐项返回。
# - 防空 Archive：attach_file 失败时若 Archive 是本次新建的，回滚掉。

_QUOTA_UPLOAD_SOURCE_ID = "quota-manual-upload"

QUOTA_UPLOAD_CATEGORY_LABELS = {
    "construction_quota": "建筑工程定额",
    "industry_quota":     "专业工程定额",
    "boq_standard":       "清单规范",
}

_QUOTA_UPLOAD_VALID_CATEGORIES = set(QUOTA_UPLOAD_CATEGORY_LABELS.keys())

# v0.9 (2026-08-18): category 驱动 (material_type, quota_system_type) 派生。
# 历史 bug: category 进了 metadata 但后端硬编码 material_type=quota_base+system=construction_regional,
# 三个选项之间没有任何 DB 列区分。新版用这张表决定写什么列。
_CATEGORY_TYPE_MAP: dict[str, dict] = {
    "boq_standard": {
        "material_type": "boq_standard",
        "quota_system_type": None,
        "needs_industry": False,
    },
    "construction_quota": {
        "material_type": "quota_base",
        "quota_system_type": "construction_regional",
        "needs_industry": False,
    },
    "industry_quota": {
        "material_type": "quota_base",
        "quota_system_type": "industry_specialty",
        "needs_industry": True,
    },
}

# v0.9 (2026-08-18): 专业工程定额 v1 受控词表 (12 项)
#   短码 → 中文名. UI 行业下拉数据源;后端校验填了行业必须在词表内。
#   词表维护流程: 走 quota_taxonomy.py + 同步 ADMIN 文档, 不在本表改。
INDUSTRY_SECTOR_LABELS: dict[str, str] = {
    "water_conservancy":  "水利工程",
    "electric_power":     "电力工程",
    "power_grid":         "电网工程",
    "railway":            "铁路工程",
    "highway":            "公路工程",
    "petroleum":          "石油工程",
    "petrochemical":      "石化工程",
    "coal":               "煤炭工程",
    "solar_power":        "光伏发电工程",
    "waterway_port":      "水运港口工程",
    "non_ferrous_metal":  "有色金属工业",
    "ict":                "信息通信工程",
}
_VALID_INDUSTRY_SECTOR_CODES = set(INDUSTRY_SECTOR_LABELS.keys())


# ── 省份简称白名单（手动上传 v0.3.2 引入） ──────────────────────────────
# 格式：short_code → (jurisdiction_code, jurisdiction_label, default_profile, jurisdiction_level)
#
# default_profile 含义：
#   - "sichuan"/"chongqing"：已实现解析器，上传后立即可用
#   - None：尚无专用解析器；上传会创建 archive + profile，但 parse 阶段需人工指定 profile
#
# 当前有解析器的省：sc（四川）、cq（重庆）。
# 其余 30 个省级单位 + 深圳（副省级）均为 None。
# short code 命名约定：
#   - 沿用车牌/拼音简称
#   - 河南=yu（豫）、湖北=hu（鄂）、湖南=xi（湘）— 单字母按车牌简称区分
#   - 山西=sx（晋）、陕西=snx（陕）— 避开 shanxi/shanxi 冲突
#   - 深圳=sz，副省级，jurisdiction_level=city
_UPLOAD_PROVINCE_MAP: dict[str, tuple[str, str, str, str]] = {
    # tuple = (jurisdiction_code, label, profile, jurisdiction_level)
    # profile 是拼音长名（GB/T 28039），与 province 短码 1:1。
    # 完整语义见 quota/README.md §8（Profile 字段语义）。
    # 注意：陕西 shaanxi ≠ 山西 shanxi（GB/T 28039 拼写）；湖南 hunan（湘 xi 是省内简称，不进 profile）。
    # v0.9 (2026-08-18): 加 nat = 全国 (清单规范特殊省份) — jurisdiction_code=000000, level=national.
    #                    profile=None 因为全国清单无单一省解析器。
    "nat": ("000000", "全国",              None,           "national"),
    "bj":  ("110000", "北京市",            "beijing",      "province"),
    "tj":  ("120000", "天津市",            "tianjin",      "province"),
    "hb":  ("130000", "河北省",            "hebei",        "province"),
    "sx":  ("140000", "山西省",            "shanxi",       "province"),
    "nm":  ("150000", "内蒙古自治区",      "neimenggu",    "province"),
    "ln":  ("210000", "辽宁省",            "liaoning",     "province"),
    "jl":  ("220000", "吉林省",            "jilin",        "province"),
    "hl":  ("230000", "黑龙江省",          "heilongjiang", "province"),
    "sh":  ("310000", "上海市",            "shanghai",     "province"),
    "js":  ("320000", "江苏省",            "jiangsu",      "province"),
    "zj":  ("330000", "浙江省",            "zhejiang",     "province"),
    "ah":  ("340000", "安徽省",            "anhui",        "province"),
    "fj":  ("350000", "福建省",            "fujian",       "province"),
    "jx":  ("360000", "江西省",            "jiangxi",      "province"),
    "sd":  ("370000", "山东省",            "shandong",     "province"),
    "yu":  ("410000", "河南省",            "henan",        "province"),  # 豫
    "hu":  ("420000", "湖北省",            "hubei",        "province"),  # 鄂
    "xi":  ("430000", "湖南省",            "hunan",        "province"),  # 湘
    "gd":  ("440000", "广东省",            "guangdong",    "province"),
    "gx":  ("450000", "广西壮族自治区",    "guangxi",      "province"),
    "hi":  ("460000", "海南省",            "hainan",       "province"),
    "cq":  ("500000", "重庆市",            "chongqing",    "province"),
    "sc":  ("510000", "四川省",            "sichuan",      "province"),
    "gz":  ("520000", "贵州省",            "guizhou",      "province"),  # 黔
    "yn":  ("530000", "云南省",            "yunnan",       "province"),  # 滇
    "xz":  ("540000", "西藏自治区",        "xizang",       "province"),  # 藏
    "snx": ("610000", "陕西省",            "shaanxi",      "province"),  # 区别于山西 shanxi
    "gs":  ("620000", "甘肃省",            "gansu",        "province"),  # 陇
    "qh":  ("630000", "青海省",            "qinghai",      "province"),
    "nx":  ("640000", "宁夏回族自治区",    "ningxia",      "province"),
    "xj":  ("650000", "新疆维吾尔自治区",  "xinjiang",     "province"),
    "sz":  ("440300", "深圳市",            "shenzhen",     "city"),      # 副省级
}

_VALID_PROVINCE_CODES = set(_UPLOAD_PROVINCE_MAP.keys())
# v0.9 (2026-08-18): "全国" (nat) profile=None, 不进 profile 白名单。
#   否则 quota_parser/service.py:_load_profiles() 用 sorted() 排时会因
#   None vs str 不可比而 TypeError 退出。
_VALID_PROFILES = {v[2] for v in _UPLOAD_PROVINCE_MAP.values() if v[2] is not None}

# v0.8 删闸口: 「入库 ≠ 解析」。
# - 之前 v0.7 在 upload 入口拒绝非 sc/cq 省 (理由: pipeline 会落 empty_candidate.xlsx 占位),
#   但这违反了入库纪律 — 档案能不能入只该看元数据是否齐全, 不该看后续能不能解析。
# - 现在 32 省都允许入库, profile 字段全员预填 (上方 _UPLOAD_PROVINCE_MAP 第三列)。
#   是否有 extractor 是 worker 的事, 由 has_parser_for(province) 守卫,
#   缺失时标 skipped_no_parser 终态 + warning, 不报 failed。
# - 详见 quota/README.md §8 铁律 3。
# _EXTRACTABLE_PROVINCES 已删除 (历史遗留硬白名单, 与"入库 ≠ 解析"原则冲突)。


def _decode_upload_filename(raw: str | None) -> str:
    """修复 Windows curl 把 UTF-8 文件名以 GBK 字节送进 Content-Disposition 后的乱码。

    FastAPI 把 multipart filename 字段按 latin-1 解码成 str（HTTP/1.1 multipart 协议限制），
    但 Windows curl 在中国区默认按 GBK 编码原始字节，导致中文字符显示成 ¡¶ËÄ´¨... 这类 mojibake。

    检测启发式：
      1. 字符串能干净通过 UTF-8 验证，且不含中文，但落在 latin-1 高位（¡-ÿ）— 大概率是 GBK
         被错当成 UTF-8 解码后又当成 latin-1 反解码回 bytes 的结果。
      2. 否则若整体 UTF-8 解码失败，按 latin-1 编码 + GBK 解码还原。

    返回：修复后的 UTF-8 字符串；无法修复时原样返回。
    """
    if not raw:
        return "(unnamed)"
    try:
        raw.encode("utf-8").decode("utf-8")
        if any("¡" <= ch <= "ÿ" for ch in raw) and not any("一" <= ch <= "鿿" for ch in raw):
            try:
                return raw.encode("latin-1").decode("gbk")
            except (UnicodeDecodeError, UnicodeEncodeError):
                pass
        return raw
    except UnicodeError:
        pass
    try:
        return raw.encode("latin-1").decode("gbk")
    except (UnicodeDecodeError, UnicodeEncodeError):
        return raw


def _ensure_quota_publication_set(
    session: Session,
    *,
    archive: Archive,
    province: str | None,
    year: int | None,
    title: str,
    category: str,
    industry_sector_code: str | None = None,
) -> QuotaPublicationSet:
    """为手动上传的 archive 创建配套 pubset + profile（v0.3.2 引入）。

    - province 必填 → 决定 jurisdiction_code / jurisdiction_level
    - year 必填 → 决定 edition_year / edition_label
    - category 必填 → 决定 material_type + quota_system_type (v0.9 接入)
    - industry_sector_code 仅 industry_quota 必填,其余  必填 None
    - 同一 (province, year) 同一档案：复用已有 pubset（幂等）
    """
    if year is None:
        raise HTTPException(
            status_code=422,
            detail="MISSING_YEAR: 必须传 year（定额年份）",
        )
    if category not in _CATEGORY_TYPE_MAP:
        raise HTTPException(
            status_code=422,
            detail=f"INVALID_CATEGORY_IN_PUBSET: {category}",
        )
    cat_cfg = _CATEGORY_TYPE_MAP[category]

    # industry_quota 必须带 industry_sector_code; 其它类型必须不带 (CHECK 强制)
    # v0.9 fix: 必须先解析 category 与 industry_sector_code 互斥校验, 再做 province 校验。
    #   否则 industry_quota 的 empty province 会在 province 校验 422, 后面的 province="nat"
    #   覆盖永远到不了。已实测: industry_quota 不传 province 会撞 MISSING_PROVINCE。
    if cat_cfg["needs_industry"]:
        if not industry_sector_code:
            raise HTTPException(
                status_code=422,
                detail="MISSING_INDUSTRY: industry_quota 必须传 industry_sector_code",
            )
        if industry_sector_code not in _VALID_INDUSTRY_SECTOR_CODES:
            raise HTTPException(
                status_code=422,
                detail=f"INVALID_INDUSTRY: {industry_sector_code}. 必须为 {sorted(_VALID_INDUSTRY_SECTOR_CODES)} 之一",
            )
        # industry 类型不依赖省份: jurisdiction_code 强制 000000, level=national.
        # 前端允许不传 province, 后端强制覆盖为 "nat" 后再做 _UPLOAD_PROVINCE_MAP 查表。
        province = "nat"
    else:
        if industry_sector_code is not None:
            raise HTTPException(
                status_code=422,
                detail="UNEXPECTED_INDUSTRY: 非 industry_quota 类型不能传 industry_sector_code",
            )
        # 非 industry 类型: province 必填 (前段省份或全国 nat)。
        if not province:
            raise HTTPException(
                status_code=422,
                detail="MISSING_PROVINCE: 必须传 province（省份简称或全国）",
            )

    jurisdiction_code, jurisdiction_label, _, jurisdiction_level = _UPLOAD_PROVINCE_MAP[province]
    edition_label = str(year)
    title_slug = re.sub(r"[^0-9A-Za-z]+", "-", title)[:32].strip("-") or "untitled"

    # biz_key 稳定可幂等:quota:upload:{province}:{year}:{title_slug}
    # 行业工程用 industry_sector_code 隔离 (避免"电网2025"与"电力2025"撞 biz_key)。
    if cat_cfg["needs_industry"]:
        biz_key = f"quota:upload:{industry_sector_code}:{year}:{title_slug}"
    else:
        biz_key = f"quota:upload:{province}:{year}:{title_slug}"

    existing = session.scalar(
        select(QuotaPublicationSet).where(QuotaPublicationSet.biz_key == biz_key)
    )
    if existing is not None:
        pub_set = existing
    else:
        # 标题生成: 全国清单特殊;省级通用模板;专业工程用行业名。
        if cat_cfg["quota_system_type"] == "industry_specialty":
            industry_label = INDUSTRY_SECTOR_LABELS[industry_sector_code]
            pub_title = f"{industry_label}{year}定额（{title}）"
        elif jurisdiction_code == "000000":
            pub_title = f"全国{year}清单规范（{title}）"
        else:
            pub_title = f"{jurisdiction_label}{year}定额（{title}）"

        pub_set = QuotaPublicationSet(
            publication_set_id=str(uuid4()),
            biz_key=biz_key,
            publication_family_code=(
                f"upload-{industry_sector_code}-{year}" if cat_cfg["needs_industry"]
                else f"upload-{province}-{year}"
            ),
            title=pub_title,
            material_type=cat_cfg["material_type"],
            quota_system_type=cat_cfg["quota_system_type"],
            jurisdiction_level=jurisdiction_level,
            jurisdiction_code=jurisdiction_code,
            industry_sector_code=industry_sector_code if cat_cfg["needs_industry"] else None,
            issuer_name="manual_upload",
            standard_or_quota_code=None,
            edition_label=edition_label,
            edition_year=year,
            publish_date=None,
            effective_date=None,
            legal_status="effective",
            metadata_status="complete",
            tenant_code="platform_public",
            visibility_scope="public",
        )
        session.add(pub_set)
        session.flush()

    existing_profile = session.get(QuotaArchiveProfile, archive.archive_id)
    if existing_profile is None:
        session.add(QuotaArchiveProfile(
            archive_id=archive.archive_id,
            publication_set_id=pub_set.publication_set_id,
            document_role="main_volume",
            discipline_code=None,
            metadata_status="complete",
            completeness_score=80,
            completeness_blockers=[],
        ))
    elif existing_profile.publication_set_id != pub_set.publication_set_id:
        existing_profile.publication_set_id = pub_set.publication_set_id

    session.flush()
    # v0.3.3：补 commit。upload_quota_files 此前只 flush 不 commit，
    # FastAPI dep get_db_session (database.py:761) 关 session 时会回滚未提交事务，
    # 导致 PubSet + Profile 永久丢失。helper 自洽 commit 与 _ensure_quota_upload_source 一致。
    session.commit()
    return pub_set


class QuotaUploadItem(BaseModel):
    filename: str
    status: str            # "created" | "duplicate" | "failed"
    file_id: str | None = None
    archive_id: str | None = None
    title: str | None = None
    already_exists: bool = False
    error: str | None = None


class QuotaUploadResponse(BaseModel):
    count: int
    succeeded: int
    failed: int
    items: list[QuotaUploadItem]


def _ensure_quota_upload_source(session: Session) -> str:
    """查询或创建稳定的 quota 手工上传 DataSource，返回 source_id。

    不假设 database.py 的启动种子化一定跑过；端点入口自洽查询/创建。
    """
    existing = session.get(DataSource, _QUOTA_UPLOAD_SOURCE_ID)
    if existing is not None:
        return _QUOTA_UPLOAD_SOURCE_ID
    ds = DataSource(
        source_id=_QUOTA_UPLOAD_SOURCE_ID,
        source_scope="platform_public",
        managed_by="platform",
        source_type="quota_manual_upload",
        connector_type="manual_upload",
        name="quota-manual-upload",
        data_domain="quota",
        format="pdf",
        config={},
        status="active",
        created_by="system",
        asset_tenant_code="platform_public",
    )
    session.add(ds)
    session.commit()
    return _QUOTA_UPLOAD_SOURCE_ID


def _cleanup_empty_archive(session: Session, archive_id: str) -> None:
    """attach_file 失败时删除本次新建的空 Archive（及其 ArchiveFile/ArchiveEvent/
    QuotaArchiveProfile 残骸）。

    FileAsset 由 register_asset 独立 commit，保留为原件（用户可重试或自行归档）。
    本补偿仅在"本次 create_archive 成功 + 后续 attach_file 失败"时触发，
    不触及历史数据。
    """
    session.execute(delete(ArchiveFile).where(ArchiveFile.archive_id == archive_id))
    session.execute(delete(ArchiveEvent).where(ArchiveEvent.archive_id == archive_id))
    session.execute(delete(QuotaArchiveProfile).where(QuotaArchiveProfile.archive_id == archive_id))
    session.execute(delete(Archive).where(Archive.archive_id == archive_id))
    session.commit()


@router.post("/upload", response_model=QuotaUploadResponse)
async def upload_quota_files(
    response: Response,
    files: list[UploadFile] = File(...),
    category: str = Form(..., description="boq_standard | construction_quota | industry_quota"),
    province: str | None = Form(None, description="省简称（sc/cq/bj/tj/nat...）；boq_standard+construction_quota 必填，决定 jurisdiction_code"),
    year: int | None = Form(None, description="定额版本年（1900-2100）；必填，决定 edition_year"),
    profile: str | None = Form(None, description="解析 profile 名；缺省按 province 自动映射"),
    industry_sector_code: str | None = Form(None, description="仅 industry_quota 必填；12 项行业 v1 词表短码（water_conservancy / electric_power / ...）"),
    session: Session = Depends(get_db_session),
    storage: ObjectStore = Depends(get_object_store),
) -> QuotaUploadResponse:
    """极简定额 PDF 上传：PDF → FileAsset → Archive(quota) → ArchiveFile → PubSet/Profile。

    v0.3.2 起新增必填 Form 字段：province、year。可选：profile（缺省按 province 自动映射）。
    v0.9 起 category 真正驱动 (material_type, quota_system_type) — 3 种资料（清单规范 / 建筑工程定额 /
        专业工程定额）走不同 DB 列，不再硬编码 quota_base+construction_regional。

    单文件粒度处理，一文件失败不影响其它。重复文件（同 tenant + 同 sha256）幂等返回。
    """
    # category 校验：非法值直接 422（Pydantic response_model 不约束 Form 参数）
    if category not in _QUOTA_UPLOAD_VALID_CATEGORIES:
        raise HTTPException(
            status_code=422,
            detail=f"INVALID_CATEGORY: {category}. 必须为 {sorted(_QUOTA_UPLOAD_VALID_CATEGORIES)} 之一",
        )
    cat_cfg = _CATEGORY_TYPE_MAP[category]
    # province 校验（按 category 分流）
    #   - industry_quota: province 占位但后端忽略(强制 nat)。允许传/不传;不传记占位。
    #   - 其它: province 必填且在 _VALID_PROVINCE_CODES (32 省 + 深圳 + nat)。
    if not cat_cfg["needs_industry"]:
        if not province:
            raise HTTPException(422, detail="MISSING_PROVINCE: 必须传 province（省份简称或全国）")
        if province not in _VALID_PROVINCE_CODES:
            raise HTTPException(
                422,
                detail=f"INVALID_PROVINCE: {province}. 必须为 {sorted(_VALID_PROVINCE_CODES)} 之一",
            )
    else:
        # industry_quota: 强制要求 industry_sector_code; province 占位即可,允许传/不传
        if province and province not in _VALID_PROVINCE_CODES:
            raise HTTPException(
                422,
                detail=f"INVALID_PROVINCE: {province}. 必须为 {sorted(_VALID_PROVINCE_CODES)} 之一（或不传）",
            )
        if not industry_sector_code:
            raise HTTPException(
                422,
                detail="MISSING_INDUSTRY: industry_quota 必须传 industry_sector_code",
            )
        if industry_sector_code not in _VALID_INDUSTRY_SECTOR_CODES:
            raise HTTPException(
                422,
                detail=f"INVALID_INDUSTRY: {industry_sector_code}. 必须为 {sorted(_VALID_INDUSTRY_SECTOR_CODES)} 之一",
            )
    if year is None or year < 1900 or year > 2100:
        raise HTTPException(422, detail=f"INVALID_YEAR: {year}（必须 1900-2100 整数）")
    # industry_quota 时 province 强制 nat, 不再走 _UPLOAD_PROVINCE_MAP 的 province 查询;
    # _ensure_quota_publication_set 内部会覆盖。这里 jurisdiction_code 仅用于 metadata/region_code。
    jurisdiction_code, jurisdiction_label, default_profile, jurisdiction_level = (
        _UPLOAD_PROVINCE_MAP["nat"] if cat_cfg["needs_industry"]
        else _UPLOAD_PROVINCE_MAP[province]
    )
    # profile 校验/缺省：v0.8 起 32 省 profile 全员预填, _VALID_PROFILES 与 _UPLOAD_PROVINCE_MAP
    # 第三列同步. 用户不传 → 沿用 province 对应 profile (即每个省自动有合法 profile).
    # 用户传值 → 必须落在 _VALID_PROFILES 内 (历史 sichuan/chongqing 等都还在).
    # 字段合法性 ≠ 有 extractor (铁律 3): 没 extractor 的省照样入库, 由 worker has_parser_for 守卫.
    if profile is None:
        profile = default_profile
    if profile is not None and profile not in _VALID_PROFILES:
        raise HTTPException(
            422,
            detail=f"INVALID_PROFILE: {profile}. 必须为 {sorted(_VALID_PROFILES)} 之一（或不传）",
        )
    if not files:
        raise HTTPException(status_code=422, detail="NO_FILES_RECEIVED")

    source_id = _ensure_quota_upload_source(session)
    category_cell = metadata_cell(category, source_level="manual", tagged_by="ui:quota-upload")

    items: list[QuotaUploadItem] = []
    succeeded = 0
    failed = 0

    for f in files:
        filename = _decode_upload_filename(f.filename)

        # ── 1. 扩展名校验（仅接受 .pdf）──────────────────────────────
        ext = Path(filename).suffix.lower()
        if ext != ".pdf":
            items.append(QuotaUploadItem(
                filename=filename, status="failed", error="NON_PDF_REJECTED",
            ))
            failed += 1
            continue

        # ── 2. 读取内容 ────────────────────────────────────────────────
        try:
            content = await f.read()
        except Exception as read_exc:  # noqa: BLE001
            items.append(QuotaUploadItem(
                filename=filename, status="failed",
                error=f"READ_FAILED: {read_exc}",
            ))
            failed += 1
            continue

        # ── 3. register_asset（按 tenant+sha256 幂等）─────────────────
        try:
            result = register_asset(
                session,
                storage,
                tenant_code="platform_public",
                source_type="quota_manual_upload",
                batch_id=f"quota-upload-{uuid4().hex[:8]}",
                file_name=filename,
                content=content,
                source_id=source_id,
                create_ingest_event=True,
                derive_tasks=False,           # hotfix 不触发 OCR/解析任务
                mime_type="application/pdf",
            )
        except Exception as reg_exc:  # noqa: BLE001
            items.append(QuotaUploadItem(
                filename=filename, status="failed",
                error=f"REGISTER_FAILED: {reg_exc}",
            ))
            failed += 1
            continue

        # ── 4. 查 ArchiveFile JOIN Archive 是否已挂载到 quota 域 ────────
        existing_mount = session.scalar(
            select(ArchiveFile)
            .join(Archive, ArchiveFile.archive_id == Archive.archive_id)
            .where(
                ArchiveFile.file_id == result.file_id,
                Archive.domain_type == "quota",
            )
        )
        if existing_mount is not None:
            existing_archive = session.get(Archive, existing_mount.archive_id)
            items.append(QuotaUploadItem(
                filename=filename,
                status="duplicate",
                file_id=result.file_id,
                archive_id=existing_mount.archive_id,
                title=existing_archive.title if existing_archive else None,
                already_exists=True,
            ))
            succeeded += 1
            continue

        # ── 5. 构造 title 与 province+year+sha256-based business_key ──
        title = Path(filename).stem or filename
        # business_key 含省份 + 年份 + sha256 前 12 位：让 quota 域的 dedup 同时区分
        # 同文件 + 同省同年的「重复」 vs 同文件跨省跨年的「迁移」（属同一文件但不同档案归属）
        business_key = f"quota:manual-upload:{province}:{year}:{result.sha256[:12]}"

        # ── 6. create_archive + attach_file（含空 Archive 补偿）────────
        archive_created_by_us = False
        try:
            try:
                archive = create_archive(
                    session,
                    domain_type="quota",
                    channel_type="manual_upload",
                    collection_method="manual_denovo",
                    business_key=business_key,
                    title=title,
                    source_id=source_id,
                    tenant_code="platform_public",
                    visibility_scope="public",
                    status="pending_tag",
                    region_code=jurisdiction_code,    # v0.3.2：写入省份 code
                    metadata={
                        "category": category_cell,
                        # v0.4 §9 #15: 把 province 短码透传到 metadata_payload,
                        # 下游 trigger_quota_parse 从这里读回后传给 worker。
                        "province": metadata_cell(
                            province,
                            source_level="manual", tagged_by="ui:quota-upload",
                        ),
                    },
                    field_sources=_build_archive_field_sources(
                        title=title, business_key=business_key,
                        region_code=jurisdiction_code),
                )
                archive_created_by_us = True
            except ValueError as ve:
                if "ARCHIVE_BUSINESS_KEY_EXISTS" in str(ve):
                    # 极小概率：同 province+year+sha256[:12] 已存在（理论不应发生，防御）
                    archive = session.scalar(
                        select(Archive).where(Archive.business_key == business_key)
                    )
                    if archive is None:
                        raise
                else:
                    raise

            _attach_archive_file(
                session,
                archive_id=archive.archive_id,
                file_id=result.file_id,
                file_role="main_document",   # 恒为 main_document，不按 category 映射
                is_primary=True,
                sort_order=10,
                display_name=filename,
                actor_type="user",
                actor_id="ui:quota-upload",
            )

            # ── 6b. 配套 pubset + profile（v0.3.2 引入；v0.9 接入 category + industry）──
            if archive_created_by_us:
                _ensure_quota_publication_set(
                    session,
                    archive=archive,
                    province=province,
                    year=year,
                    title=title,
                    category=category,
                    industry_sector_code=industry_sector_code,
                )
        except Exception as attach_exc:  # noqa: BLE001
            # 防空 Archive 补偿：attach 失败时若 Archive 是本次新建的，回滚掉
            if archive_created_by_us:
                try:
                    _cleanup_empty_archive(session, archive.archive_id)
                except Exception:  # noqa: BLE001
                    session.rollback()
            items.append(QuotaUploadItem(
                filename=filename, status="failed",
                error=f"ATTACH_FAILED: {attach_exc}",
            ))
            failed += 1
            continue

        items.append(QuotaUploadItem(
            filename=filename,
            status="created",
            file_id=result.file_id,
            archive_id=archive.archive_id,
            title=title,
            already_exists=False,
        ))
        succeeded += 1

    body = QuotaUploadResponse(
        count=len(items),
        succeeded=succeeded,
        failed=failed,
        items=items,
    )
    return _json_ok(response, body)


# ── P0-4B · quota_parser integration (2026-07-28, web-frontend SPEC §5) ─────
# 7 个新端点：
#   POST   /archives/{archive_id}/parse           触发解析（阶段 A）
#   GET    /archives/{archive_id}/candidate.xlsx  下载 candidate
#   POST   /archives/{archive_id}/reviewed        上传 reviewed（阶段 B）
#   GET    /archives/{archive_id}/final.xlsx      下载 final
#   GET    /archives/{archive_id}/manifest        拿 Manifest JSON
#   GET    /archives/{archive_id}/qa-report       拿 QA 报告（json/md）
#   POST   /archives/{archive_id}/parse/delete    删除解析结果（清 MinIO + parse_* 字段）
#
# Mock 模式（QUOTA_PARSE_MOCK=1）下：解析引擎走 app.mock_parse_runner（异步后台）；
# 否则走真 quota_parser worker（未来批次）。


@router.post("/archives/{archive_id}/parse")
async def trigger_quota_parse(
    response: Response,
    archive_id: str,
    body: dict | None = None,
    session: Session = Depends(get_db_session),
):
    """触发阶段 A 解析（v0.5：入队不调度）。

    行为：
      - 写 archive.parse_status='parsing' + parse_task_id
      - INSERT 一条 quota_parse_job (status='queued', mock=is_parse_mock())
      - 独立 worker 进程消费该 job（mock / real 按 metadata_payload["mock"] 分流）
      - 不再 schedule_mock_a()（避免 asyncio 阻塞 8010 端口）
      - 不再 501（worker 已上线,无论 mock/real 都成功入队）
    """
    from app.quota_parser import is_parse_mock, trigger_parse as _trigger
    from app.quota_parser.service import enqueue_parse_job

    archive = session.get(Archive, archive_id)
    if archive is None:
        raise HTTPException(404, detail="ARCHIVE_NOT_FOUND")
    if archive.domain_type != "quota":
        raise HTTPException(400, detail="NOT_A_QUOTA_ARCHIVE")

    profile = (body or {}).get("profile")

    # ── v0.4 §9 #15：province 透传 + body.profile 白名单校验 ──
    # (a) 从 archive.metadata_payload 读回 province；旧档案 fallback None
    province_value: str | None = None
    mp = archive.metadata_payload or {}
    prov_cell = mp.get("province")
    if isinstance(prov_cell, dict) and "value" in prov_cell:
        province_value = prov_cell.get("value")
    elif isinstance(prov_cell, str):  # 防御：历史裸字符串
        province_value = prov_cell

    # (a') v0.5 fix: 若 archive 没有 province,按 profile 推一个默认 short code
    # (避免 pipeline.placeholder 分支因缺 xlsx_path 报 KeyError)
    # profile='sichuan' → 'sc', 'chongqing' → 'cq'
    if province_value is None:
        _profile_for_prov = (profile or archive.parse_profile or "").lower()
        if _profile_for_prov == "sichuan":
            province_value = "sc"
        elif _profile_for_prov == "chongqing":
            province_value = "cq"
        # 仍为 None 时,worker 走 run_quota_pipeline 的 default 分支（由 pipeline 决定行为）

    # (a'') v0.7 fix: 多层兜底, 即使 metadata_payload.province 缺失 + profile=None,
    # 也能从 archive.biz_key / archive.region_code 推出 short code.
    # 2026-07-31 教训: upload 路径 (quota_api.py:/upload) 只构造了 category_cell,
    # 没构造 province_cell, 导致历史 archive (如 a5f716b0 重庆装饰工程第二册 2018)
    # metadata 里 province 永远是 None, 走 pipeline default 分支 → 0 行 empty xlsx.
    # 而 biz_key 形如 'quota:manual-upload:cq:2018:...' 自带省份简写;
    # region_code = GB/T 2260 (500000=重庆, 510000=四川), 也可反查 _UPLOAD_PROVINCE_MAP.
    if province_value is None and archive.business_key:
        _biz_parts = (archive.business_key or "").split(":")
        # 期望 ['quota', 'manual-upload', 'cq', '2018', '<slug>']
        if (len(_biz_parts) >= 3
                and _biz_parts[0] == "quota"
                and _biz_parts[1] == "manual-upload"):
            _candidate = _biz_parts[2]
            if _candidate in _VALID_PROVINCE_CODES:
                province_value = _candidate

    if province_value is None and archive.region_code:
        # region_code 是 GB/T 2260 jurisdiction_code (6 位), 反查 _UPLOAD_PROVINCE_MAP
        # tuple: (jurisdiction_code, label, default_profile, jurisdiction_level)
        for _k, _v in _UPLOAD_PROVINCE_MAP.items():
            if _v[0] == archive.region_code:
                province_value = _k
                break

    # (a''') v0.13.1 fix: profile 缺省时按 province 推 default profile (替代历史硬编码 "sichuan" fallback,
    #         让非四川省份不再被错误 fallback 成 sichuan → 跑错 extractor)
    if not (profile or archive.parse_profile):
        if province_value and province_value in _UPLOAD_PROVINCE_MAP:
            profile = _UPLOAD_PROVINCE_MAP[province_value][2]  # tuple 第 3 列 = default profile
        else:
            raise HTTPException(
                status_code=422,
                detail=f"PROVINCE_NO_PROFILE: 无法从 province={province_value!r} 推 profile (不在 _UPLOAD_PROVINCE_MAP)",
            )

    # (b) profile 白名单：用 _VALID_PROFILES (32 省, 从 _UPLOAD_PROVINCE_MAP 派生,
    #     已含 guangdong/xi/hu/yu 等) — 与 quota_parser_worker._guard_has_parser + DB CHECK 保持单一真源
    if profile is not None and profile not in _VALID_PROFILES:
        raise HTTPException(
            status_code=422,
            detail=f"INVALID_PROFILE: {profile!r}. 必须为 {sorted(_VALID_PROFILES)} 之一或省略",
        )

    # (1) 推 archive.parse_status='parsing'（已有 _trigger）—— 必须在入队前 commit,
    #     否则 enqueue_parse_job 找不到状态机上下文。
    try:
        archive = _trigger(archive, profile=profile, province=province_value)
    except ValueError as e:
        msg = str(e)
        if "profile" in msg:
            raise HTTPException(422, detail=f"INVALID_PROFILE: {msg}")
        raise HTTPException(409, detail=f"ALREADY_PARSING_OR_DONE: {msg}")

    # (2) INSERT quota_parse_job —— 同 archive 已有 active job → 409
    try:
        job = enqueue_parse_job(
            session,
            archive_id,
            profile=profile or archive.parse_profile,  # v0.13.1: 已按 province 推 default profile (L1583-1592)
            created_by=(body or {}).get("created_by"),
            mock=is_parse_mock(),
            ocr_api_url=os.environ.get("OCR_URL", "http://171.212.159.15:8000"),
            province=province_value,
        )
    except ValueError as e:
        # 回滚 archive.parse_status,避免状态机卡 'parsing'
        session.rollback()
        raise HTTPException(409, detail=f"ALREADY_PARSING_OR_DONE: {e}")

    session.commit()
    session.refresh(archive)

    from app.quota_parser.service import build_parse_section

    return _json_ok(response, {
        "archive_id": archive.archive_id,
        "status": archive.status,
        "parse": build_parse_section(archive),
        "job_id": job.job_id,  # v0.5 新增：前端可轮询
    })


@router.get("/archives/{archive_id}/candidate.xlsx")
def get_candidate_xlsx(
    archive_id: str,
    session: Session = Depends(get_db_session),
):
    """下载 candidate.xlsx（QUOTA_PARSE_MOCK=1 时返回 mock fixture）。"""
    from urllib.parse import quote

    from app.quota_parser import PARSE_BUCKET_CANDIDATE
    from app.storage import get_object_store

    archive = session.get(Archive, archive_id)
    if archive is None:
        raise HTTPException(404, detail="ARCHIVE_NOT_FOUND")
    if not archive.candidate_xlsx_key:
        raise HTTPException(404, detail="CANDIDATE_NOT_READY")
    store = get_object_store()
    data = store.get_object(PARSE_BUCKET_CANDIDATE, archive.candidate_xlsx_key)
    # 用 archive_id 作 ascii-safe filename，中文 title 用 RFC 5987 扩展
    filename_ascii = f"{archive_id}-candidate.xlsx"
    filename_utf8 = quote(f"{archive.title}-candidate.xlsx")
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{filename_ascii}\"; "
                f"filename*=UTF-8''{filename_utf8}"
            ),
        },
    )


@router.post("/archives/{archive_id}/reviewed")
async def upload_reviewed_xlsx(
    response: Response,
    archive_id: str,
    file: UploadFile = File(...),
    session: Session = Depends(get_db_session),
):
    """上传 reviewed.xlsx → 阶段 B → 写 final_xlsx_key。

    QUOTA_PARSE_MOCK=1 → 后台 mock runner；否则调 quota_parser.finalize_reviewed_xlsx。
    """
    from app.quota_parser import is_parse_mock
    from app.quota_parser.service import (
        InvalidReviewedXlsxError,
        QuotaParserStageBError,
        build_parse_section,
        upload_reviewed as _upload,
    )

    archive = session.get(Archive, archive_id)
    if archive is None:
        raise HTTPException(404, detail="ARCHIVE_NOT_FOUND")
    if archive.domain_type != "quota":
        raise HTTPException(400, detail="NOT_A_QUOTA_ARCHIVE")
    if archive.parse_status not in ("parsed", "candidate_ready"):
        raise HTTPException(
            409,
            detail=f"ARCHIVE_NOT_READY_FOR_REVIEW: parse_status={archive.parse_status}",
        )

    # ── 文件类型防御:filename 后缀必须 .xlsx(mime 放宽,允许 octet-stream / None) ──
    # 静态审计 D 修复:之前没校验,POST 一个 text/html 也能进 mock_b,最终落 MinIO 当 xlsx
    # 给前端,SheetJS / openpyxl 读不到内容。
    fname = (file.filename or "").lower()
    if not fname.endswith(".xlsx"):
        raise HTTPException(
            422,
            detail=f"INVALID_FILE_TYPE: filename={file.filename!r}, 必须为 .xlsx 后缀",
        )
    if file.content_type not in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",  # curl -F / 某些浏览器兜底
        None,  # 缺失 mime(防御性,部分 multipart client 不带)
    ):
        raise HTTPException(
            422,
            detail=f"INVALID_FILE_MIME: content_type={file.content_type!r}, "
                   f"必须为 spreadsheetml.sheet / octet-stream",
        )

    reviewed_bytes = await file.read()
    if not reviewed_bytes:
        raise HTTPException(422, detail="EMPTY_FILE")

    if is_parse_mock():
        # mock 模式：跳过结构校验，2-3s 后推到 qa_passed
        from app.mock_parse_runner import schedule_mock_b

        schedule_mock_b(archive_id, reviewed_bytes=reviewed_bytes)
        return _json_ok(response, {
            "archive_id": archive_id,
            "scheduled": True,
            "note": "mock mode — review 校验已跳过；2-3s 后 status=qa_passed",
        })

    # 真模式：调 quota_parser.finalize_reviewed_xlsx
    try:
        archive = _upload(archive, reviewed_bytes)
    except InvalidReviewedXlsxError as e:
        session.rollback()
        raise HTTPException(
            status_code=422,
            detail={
                "code": e.code,
                "reason": str(e),
                "sheet": e.sheet,
                "column_index": e.column_index,
            },
        )
    except QuotaParserStageBError as e:
        session.rollback()
        raise HTTPException(status_code=422, detail=f"FINALIZE_FAILED: {e}")

    session.commit()
    session.refresh(archive)
    return _json_ok(response, {
        "archive_id": archive.archive_id,
        "parse": build_parse_section(archive),
    })


@router.get("/archives/{archive_id}/final.xlsx")
def get_final_xlsx(
    archive_id: str,
    session: Session = Depends(get_db_session),
):
    """下载 final.xlsx。"""
    from urllib.parse import quote

    from app.quota_parser import PARSE_BUCKET_CANDIDATE
    from app.storage import get_object_store

    archive = session.get(Archive, archive_id)
    if archive is None:
        raise HTTPException(404, detail="ARCHIVE_NOT_FOUND")
    if not archive.final_xlsx_key:
        raise HTTPException(404, detail="FINAL_NOT_READY")
    store = get_object_store()
    data = store.get_object(PARSE_BUCKET_CANDIDATE, archive.final_xlsx_key)
    filename_ascii = f"{archive_id}-final.xlsx"
    filename_utf8 = quote(f"{archive.title}-final.xlsx")
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{filename_ascii}\"; "
                f"filename*=UTF-8''{filename_utf8}"
            ),
        },
    )


@router.get("/archives/{archive_id}/artifacts/{role}/preview")
def get_artifact_preview(
    archive_id: str,
    role: str,
    sheet: int = Query(0, ge=0),
    offset: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=2000),
    session: Session = Depends(get_db_session),
):
    """预览 candidate / final xlsx（HTML 单 sheet，可分页）。

    - role ∈ {parse_candidate_xlsx, parse_final_xlsx}
    - sheet: 0-based sheet index，默认 0
    - offset: 从 sheet 第 offset+1 行开始（0-based skip）
    - limit: 返回几行（默认 200，已知规格：每张表 ≤ 10000 行）
    - 响应头 X-Sheet-Count / X-Sheet-Index 给前端渲染 sheet tab
    - sheet meta JSON 注入 HTML body 顶部 <script type="application/json" id="xlsx-sheet-meta">
    """
    from io import BytesIO

    from app.file_preview import render_excel_preview_html
    from app.quota_parser import PARSE_BUCKET_CANDIDATE
    from app.storage import get_object_store
    from openpyxl import load_workbook

    if role not in ("parse_candidate_xlsx", "parse_final_xlsx"):
        raise HTTPException(400, detail=f"INVALID_ROLE: {role}")

    archive = session.get(Archive, archive_id)
    if archive is None:
        raise HTTPException(404, detail="ARCHIVE_NOT_FOUND")

    if role == "parse_candidate_xlsx":
        key = archive.candidate_xlsx_key
        if not key:
            raise HTTPException(404, detail="CANDIDATE_NOT_READY")
        display_name = f"{archive.title}-candidate.xlsx"
    else:
        key = archive.final_xlsx_key
        if not key:
            raise HTTPException(404, detail="FINAL_NOT_READY")
        display_name = f"{archive.title}-final.xlsx"

    store = get_object_store()
    data = store.get_object(PARSE_BUCKET_CANDIDATE, key)

    # 取所有 sheet 名 + 校验 sheet index
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
    finally:
        wb.close()
    if not sheet_names:
        raise HTTPException(404, detail="XLSX_HAS_NO_SHEETS")
    if sheet >= len(sheet_names):
        raise HTTPException(
            400,
            detail=f"SHEET_OUT_OF_RANGE: requested={sheet}, max={len(sheet_names) - 1}",
        )

    html = render_excel_preview_html(
        file_name=display_name,
        content=data,
        file_ext="xlsx",
        sheet_index=sheet,
        row_offset=offset,
        row_limit=limit,
    )
    # 方案 E：sheet 名是中文,HTTP header 强制 latin-1 编码会 UnicodeEncodeError。
    # 改用 HTML body 顶部 <script type="application/json"> 注入,前端用 JSON.parse 读。
    # X-Sheet-Count / X-Sheet-Index 仍是 ASCII 数字,保留为 header。
    import json as _json
    sheet_meta_script = (
        '<script id="xlsx-sheet-meta" type="application/json">'
        + _json.dumps(
            {
                "names": sheet_names,
                "index": sheet,
                "row_offset": offset,
                "row_limit": limit,
            },
            ensure_ascii=False,
        )
        + "</script>"
    )
    body = sheet_meta_script + html
    return Response(
        content=body,
        media_type="text/html; charset=utf-8",
        headers={
            "X-Sheet-Count": str(len(sheet_names)),
            "X-Sheet-Index": str(sheet),
            "X-Row-Offset": str(offset),
            "X-Row-Limit": str(limit),
        },
    )


@router.get("/archives/{archive_id}/manifest")
def get_manifest(
    response: Response,
    archive_id: str,
    session: Session = Depends(get_db_session),
):
    """返回 Manifest JSON（quota-parser-result/v1 schema）。

    v0.4 §8 #14：mock 模式先读 MinIO manifest.json，读不到 fallback fake_manifest()。
    真模式（未启用 worker 前 501）继续走 _reconstruct_manifest 反推。
    """
    from app.mock_parse_runner import fake_manifest as _fake_manifest
    from app.quota_parser import is_parse_mock
    from app.quota_parser.service import (
        PARSE_BUCKET_REPORT,
        _reconstruct_manifest,
    )
    from app.storage import get_object_store
    import json as _json

    archive = session.get(Archive, archive_id)
    if archive is None:
        raise HTTPException(404, detail="ARCHIVE_NOT_FOUND")
    if archive.parse_status is None:
        raise HTTPException(404, detail="MANIFEST_NOT_FOUND")

    if is_parse_mock():
        # v0.4 §8 #14：先读 MinIO 落盘的 manifest.json;读不到/读坏 fallback fake_manifest()
        manifest: dict | None = None
        try:
            store = get_object_store()
            body = store.get_object(
                PARSE_BUCKET_REPORT, f"quota/{archive_id}/manifest.json",
            )
            manifest = _json.loads(body.decode("utf-8"))
        except (KeyError, _json.JSONDecodeError):
            manifest = _fake_manifest(archive_id, parse_status=archive.parse_status)
        return _json_ok(response, manifest)

    return _json_ok(response, _reconstruct_manifest(archive))


@router.get("/archives/{archive_id}/qa-report")
def get_qa_report(
    response: Response,
    archive_id: str,
    format: str = Query("json", pattern="^(json|md)$"),
    session: Session = Depends(get_db_session),
):
    """返回 QA 报告（json / md 切换）。

    QUOTA_PARSE_MOCK=1 → 返回 mock fixture。
    """
    from app.mock_parse_runner import fake_qa_report_json, fake_qa_report_md
    from app.quota_parser import is_parse_mock

    archive = session.get(Archive, archive_id)
    if archive is None:
        raise HTTPException(404, detail="ARCHIVE_NOT_FOUND")
    if archive.parse_status is None:
        raise HTTPException(404, detail="QA_REPORT_NOT_FOUND")

    if format == "md":
        body = fake_qa_report_md(archive_id) if is_parse_mock() else _format_qa_report_md(archive)
        return Response(content=body, media_type="text/markdown; charset=utf-8")

    body = fake_qa_report_json(archive_id) if is_parse_mock() else _format_qa_report_json(archive)
    return _json_ok(response, body)


@router.post("/archives/{archive_id}/parse/delete")
def delete_parse_result(
    response: Response,
    archive_id: str,
    scope: str = "all",  # all | reviewed_only
    session: Session = Depends(get_db_session),
):
    """删除解析结果（v0.5 双语义）

    scope=reviewed_only: 撤回审核（仅删 reviewed xlsx + manifest.json,保留 candidate）
                         仅在 qa_passed / usable / done 状态可用
    scope=all（默认）:    删除全部（清 MinIO 4 对象 + 4 类 archive_file + parse_* 13 列）
                         任何状态可用(parse_status 非 NULL 时)

    web-frontend SPEC §3.1.4 危险操作 — 前端 modal 必填档案标题确认 + 审计。
    """
    if scope not in ("all", "reviewed_only"):
        raise HTTPException(
            422,
            detail=f"INVALID_SCOPE: scope={scope!r} not in ['all', 'reviewed_only']",
        )

    archive = session.get(Archive, archive_id)
    if archive is None:
        raise HTTPException(404, detail="ARCHIVE_NOT_FOUND")
    if archive.parse_status is None:
        raise HTTPException(404, detail="NOTHING_TO_DELETE")

    # 1. 写审计日志（best-effort）
    try:
        from app.models import AuditLog
        audit = AuditLog(
            actor_type="api",
            action=f"quota_parse_result_deleted_{scope}",
            target_type="archive",
            target_id=archive.archive_id,
        )
        session.add(audit)
    except Exception:
        pass

    # 2. v0.5: 调 helper
    from app.quota_parser.service import (
        delete_all_parse_results,
        delete_reviewed_only,
    )

    try:
        if scope == "reviewed_only":
            result = delete_reviewed_only(session, archive)
        else:
            result = delete_all_parse_results(session, archive)
    except ValueError as e:
        session.rollback()
        # reviewed_only 在不允许状态被调用 → 409
        raise HTTPException(409, detail=str(e))

    session.commit()

    return _json_ok(response, {
        "deleted": True,
        "archive_id": archive_id,
        "scope": scope,
        "deleted_archive_file": result["deleted_archive_file"],
        "deleted_minio_objects": result["deleted_minio"],
        "new_status": result["new_status"],
    })


def _format_qa_report_md(archive: Archive) -> str:
    """真模式下的 QA 报告 markdown（占位 — 真 worker 上线后由 quota_parser 包生成）。"""
    return (
        f"# QA Report\n\n"
        f"- archive_id: `{archive.archive_id}`\n"
        f"- parser_version: {archive.parse_parser_version or 'unknown'}\n"
        f"- status: {archive.parse_status}\n\n"
        f"_尚未实现真模式 QA 报告生成；待 quota_parser worker 上线。_\n"
    )


def _format_qa_report_json(archive: Archive) -> dict:
    """真模式下的 QA 报告 JSON（占位 — 真 worker 上线后由 quota_parser 包生成）。"""
    return {
        "$schema": "quota-parser-qa/v1",
        "task_id": archive.parse_task_id,
        "summary": "ok",
        "checks": [],
        "note": "尚未实现真模式 QA 报告生成；待 quota_parser worker 上线。",
    }


# ── §5.10 DELETE /archives/{archive_id} ────────────────────────────────
# web-frontend SPEC §3.1.4 危险操作 #10 — 整档删除（不可恢复）。
# 行为契约：
#   1. 校验 archive 存在；domain_type 必须为 quota（防止误删其他域）
#   2. 写审计日志 action=quota_archive_deleted（best-effort，失败不阻断）
#   3. SQLAlchemy cascade 自动删除 archive_file（all, delete-orphan 已在模型上声明）
#   4. 手动删 quota_archive_profile 1:1 行（无级联 FK）
#   5. 删 Archive 主表行
#   6. MinIO 上的 PDF 原件 + 解析产物 留待运维 GC（与 parse/delete 同样限制，
#      ObjectStore Protocol v1 暂未加 delete_object，避免破坏 S3ObjectStore）
#
# 与 #9 (parse/delete) 的区别：#9 保留原 PDF 与 archive 元数据；#10 全删。
@router.delete("/archives/{archive_id}")
def delete_archive(
    response: Response,
    archive_id: str,
    session: Session = Depends(get_db_session),
):
    """整档删除（不可恢复）。

    仅允许 quota 域；其他域走各自域 API。
    """
    archive = session.get(Archive, archive_id)
    if archive is None:
        raise HTTPException(404, detail="ARCHIVE_NOT_FOUND")
    if archive.domain_type != "quota":
        raise HTTPException(403, detail="WRONG_DOMAIN")

    # 1. 审计（best-effort）
    try:
        from app.models import AuditLog
        session.add(AuditLog(
            actor_type="api",
            action="quota_archive_deleted",
            target_type="archive",
            target_id=archive.archive_id,
        ))
    except Exception:
        pass

    # 2. 先清依赖行（FK 没有 ON DELETE CASCADE，必须手动）
    #    链路：outbox → archive_event → crawl_lineage → archive
    #    - outbox：发件箱，按 event_id 关联 archive_event；不删会拒删 archive_event
    #    - archive_event：状态机事件历史
    #    - crawl_lineage：爬虫血缘，archive_id 直接 FK 到 archive
    #    - quota_archive_profile：1:1 profile 行
    #    archive_file 由 ORM cascade="all, delete-orphan" 在 session.delete 时自动清
    session.execute(
        delete(Outbox).where(
            Outbox.event_id.in_(
                select(ArchiveEvent.event_id).where(ArchiveEvent.archive_id == archive_id)
            )
        )
    )
    session.execute(
        delete(ArchiveEvent).where(ArchiveEvent.archive_id == archive_id)
    )
    session.execute(
        delete(CrawlLineage).where(CrawlLineage.archive_id == archive_id)
    )
    session.execute(
        delete(QuotaArchiveProfile).where(QuotaArchiveProfile.archive_id == archive_id)
    )

    # 3. 删主表行 — 触发 ORM cascade 清 archive_file；DB 层面无其他引用
    session.delete(archive)
    session.commit()

    return _json_ok(response, {
        "deleted": True,
        "archive_id": archive_id,
        "note": "MinIO 上的原件 + 解析产物留待运维 GC",
    })
