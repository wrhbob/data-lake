from __future__ import annotations

from dataclasses import dataclass, field
from html import escape
from html.parser import HTMLParser
from io import BytesIO
from numbers import Number
from pathlib import Path
from urllib.parse import urlparse
from zipfile import BadZipFile, ZipFile

import mammoth
from openpyxl import load_workbook
import xlrd

EXCEL_EXTENSIONS = {"xls", "xlsx"}
HTML_EXTENSIONS = {"html", "htm"}
DOCX_EXTENSIONS = {"docx"}
PDF_EXTENSIONS = {"pdf"}
DOWNLOAD_ONLY_EXTENSIONS = {"zip", "rar", "7z", "cdz"}
ZIP_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif"}

MAX_PREVIEW_ROWS = 200
MAX_PREVIEW_COLS = 60
NOTICE_BODY_CLASSES = ("text", "notice-info-content", "right-content", "article-content")


@dataclass(frozen=True)
class ExcelPreviewData:
    sheet_name: str
    rows: list[list[object]]
    source_rows: int
    source_cols: int
    rendered_cols: int
    merge_spans: dict[tuple[int, int], tuple[int, int]]
    covered_cells: set[tuple[int, int]]
    col_widths: list[float] = field(default_factory=list)
    row_styles: dict[int, str] = field(default_factory=dict)
    cell_styles: dict[tuple[int, int], str] = field(default_factory=dict)
    display_values: dict[tuple[int, int], str] = field(default_factory=dict)
    display_html: dict[tuple[int, int], str] = field(default_factory=dict)


def render_excel_preview_html(
    *,
    file_name: str,
    content: bytes,
    file_ext: str,
    max_rows: int = MAX_PREVIEW_ROWS,
    max_cols: int = MAX_PREVIEW_COLS,
    sheet_index: int = 0,
    row_offset: int = 0,
    row_limit: int | None = None,
) -> str:
    # 方案 B：分页加载。后端一次解析 [row_offset, row_offset+row_limit) 这段。
    # row_limit None 走 max_rows(向后兼容)。
    effective_max = row_offset + (row_limit if row_limit is not None else max_rows)
    preview = _excel_rows(
        content=content,
        file_ext=file_ext,
        max_rows=effective_max,
        max_cols=max_cols,
        sheet_index=sheet_index,
    )
    truncated = preview.source_rows > len(preview.rows) or preview.source_cols > max_cols
    table_rows = "\n".join(
        _render_row(
            row_index,
            row,
            rendered_cols=preview.rendered_cols,
            merge_spans=preview.merge_spans,
            covered_cells=preview.covered_cells,
            row_styles=preview.row_styles,
            cell_styles=preview.cell_styles,
            display_values=preview.display_values,
            display_html=preview.display_html,
        )
        for row_index, row in enumerate(preview.rows, start=1)
        if row_index > row_offset
    )
    # 方案 B：分页加载时不再显示"已截取"的旧提示。整段加载完成后,
    # 前端通过 data-total-rows / data-row-offset 自行决定是否显示 sentinel。
    col_truncated = preview.source_cols > max_cols
    truncated_note = " · 已截取前60列，下载看完整" if col_truncated else ""
    has_merged = "1" if preview.merge_spans else "0"
    source_layout = "1" if preview.col_widths or preview.cell_styles else "0"
    colgroup = _render_colgroup(preview.col_widths)
    table_style = _table_style(preview)
    # 当前段渲染的行数（1-based 全局 row_index, 用来给前端续接 DOM）
    actual_rows = sum(
        1 for row_index in range(1, len(preview.rows) + 1) if row_index > row_offset
    )
    return f"""
<article class="excel-preview" data-preview-mode="ephemeral" data-file-processing="0" data-source-cols="{preview.source_cols}" data-rendered-cols="{preview.rendered_cols}" data-merged="{has_merged}" data-source-layout="{source_layout}" data-row-offset="{row_offset}" data-row-count="{actual_rows}" data-source-rows="{preview.source_rows}">
  <div class="formula-bar">{escape(file_name)} · {escape(preview.sheet_name)}{truncated_note}</div>
  <div class="excel-table-wrap">
    <table{table_style}>
      {colgroup}
      <tbody>
        {table_rows}
      </tbody>
    </table>
  </div>
</article>
""".strip()


def render_excel_preview_multi_sheet_html(
    *,
    file_name: str,
    content: bytes,
    file_ext: str,
    max_rows: int = MAX_PREVIEW_ROWS,
    max_cols: int = MAX_PREVIEW_COLS,
) -> str:
    """2026-08-12 新增（apply 自 info）：多 sheet xlsx 预览。遍历所有 sheet 输出
    sheet tab + 每个 sheet 一个 <article class="excel-sheet-pane">。

    使用场景：信息价域（cost_info）xlsx 含元数据 sheet + 各市州 sheet，
    前端需要 tab 切换。定额域（quota）xlsx 只有 1 个 sheet「定额条目」，
    用方案 B 分页的 render_excel_preview_html 即可，不需要本函数。

    单 sheet 时退化为旧行为（无 tab）。"""
    previews = _excel_sheets(
        content=content,
        file_ext=file_ext,
        max_rows=max_rows,
        max_cols=max_cols,
    )
    if not previews:
        return '<article class="excel-preview" data-preview-mode="ephemeral"><div class="formula-bar">空 Excel 文件</div></article>'
    articles: list[str] = []
    sheet_tabs: list[str] = []
    for sheet_index, preview in enumerate(previews):
        truncated = preview.source_rows > len(preview.rows) or preview.source_cols > max_cols
        table_rows = "\n".join(
            _render_row(
                row_index,
                row,
                rendered_cols=preview.rendered_cols,
                merge_spans=preview.merge_spans,
                covered_cells=preview.covered_cells,
                row_styles=preview.row_styles,
                cell_styles=preview.cell_styles,
                display_values=preview.display_values,
                display_html=preview.display_html,
            )
            for row_index, row in enumerate(preview.rows, start=1)
        )
        truncated_note = " · 已截取前200行/60列，下载看完整" if truncated else ""
        has_merged = "1" if preview.merge_spans else "0"
        source_layout = "1" if preview.col_widths or preview.cell_styles else "0"
        colgroup = _render_colgroup(preview.col_widths)
        table_style = _table_style(preview)
        is_first = sheet_index == 0
        article = f"""
<article class="excel-sheet-pane excel-preview" data-sheet-index="{sheet_index}" data-active="{1 if is_first else 0}" data-preview-mode="ephemeral" data-file-processing="0" data-source-cols="{preview.source_cols}" data-rendered-cols="{preview.rendered_cols}" data-merged="{has_merged}" data-source-layout="{source_layout}"{(" hidden" if not is_first else "")}>
  <div class="formula-bar">{escape(file_name)} · {escape(preview.sheet_name)}{truncated_note}</div>
  <div class="excel-table-wrap">
    <table{table_style}>
      {colgroup}
      <tbody>
        {table_rows}
      </tbody>
    </table>
  </div>
</article>
""".strip()
        articles.append(article)
        sheet_tabs.append(
            f'<button type="button" class="excel-sheet-tab{" active" if is_first else ""}" data-sheet-index="{sheet_index}" data-action="show-sheet">{escape(preview.sheet_name)}</button>'
        )
    if len(previews) == 1:
        # 单 sheet 不显示 tab
        return articles[0]
    tab_bar = (
        '<div class="excel-sheet-tabs" role="tablist">'
        + "".join(sheet_tabs)
        + "</div>"
    )
    return tab_bar + "\n" + "\n".join(articles)


def render_html_notice_preview_html(
    *,
    file_name: str,
    content: bytes,
    title: str | None = None,
) -> str:
    html = _decode_html_bytes(content)
    fragment = _notice_body_fragment(html) or "<p>未识别到公告正文。</p>"
    heading = title or file_name
    return f"""
<article class="notice-html-preview" data-preview-mode="ephemeral" data-file-processing="0">
  <header class="notice-html-preview-header">
    <strong>{escape(heading)}</strong>
  </header>
  <div class="notice-html-preview-body">
    {fragment}
  </div>
</article>
""".strip()


def render_docx_preview_html(
    *,
    file_name: str,
    content: bytes,
) -> str:
    converted = mammoth.convert_to_html(BytesIO(content))
    fragment = _sanitized_html_fragment(converted.value) or "<p>未识别到Word正文。</p>"
    return f"""
<article class="docx-preview" data-preview-mode="ephemeral" data-file-processing="0">
  <header class="notice-html-preview-header">
    <strong>{escape(file_name)}</strong>
  </header>
  <div class="docx-preview-body">
    {fragment}
  </div>
</article>
""".strip()


def zip_image_preview_entries(*, file_id: str, content: bytes) -> list[dict[str, object]]:
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries: list[dict[str, object]] = []
            for index, info in enumerate(archive.infolist()):
                if info.is_dir():
                    continue
                file_ext = Path(info.filename).suffix.lower().lstrip(".")
                if file_ext not in ZIP_IMAGE_EXTENSIONS:
                    continue
                entries.append(
                    {
                        "index": index,
                        "name": info.filename,
                        "file_ext": file_ext,
                        "file_size": info.file_size,
                        "preview_kind": "image",
                        "preview_url": f"/api/file-assets/{file_id}/zip-preview/{index}",
                    }
                )
            return entries
    except BadZipFile as exc:
        raise ValueError("ZIP_PREVIEW_BAD_ZIP") from exc


def read_zip_image_preview_entry(*, content: bytes, entry_index: int) -> tuple[str, bytes, str]:
    try:
        with ZipFile(BytesIO(content)) as archive:
            infos = archive.infolist()
            if entry_index < 0 or entry_index >= len(infos):
                raise IndexError(entry_index)
            info = infos[entry_index]
            if info.is_dir():
                raise ValueError("ZIP_PREVIEW_ENTRY_NOT_IMAGE")
            file_ext = Path(info.filename).suffix.lower().lstrip(".")
            if file_ext not in ZIP_IMAGE_EXTENSIONS:
                raise ValueError("ZIP_PREVIEW_ENTRY_NOT_IMAGE")
            return info.filename, archive.read(info), _image_content_type(file_ext)
    except BadZipFile as exc:
        raise ValueError("ZIP_PREVIEW_BAD_ZIP") from exc


def _image_content_type(file_ext: str) -> str:
    if file_ext in {"jpg", "jpeg"}:
        return "image/jpeg"
    if file_ext == "webp":
        return "image/webp"
    if file_ext == "gif":
        return "image/gif"
    return "image/png"


def _decode_html_bytes(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _notice_body_fragment(html: str) -> str:
    for class_name in NOTICE_BODY_CLASSES:
        parser = _NoticeHtmlPreviewParser(match_class=class_name)
        parser.feed(html)
        parser.close()
        fragment = parser.fragment()
        if fragment:
            return fragment

    parser = _NoticeHtmlPreviewParser(match_tag="body")
    parser.feed(html)
    parser.close()
    return parser.fragment()


def _sanitized_html_fragment(html: str) -> str:
    parser = _NoticeHtmlPreviewParser(match_tag="body")
    parser.feed(f"<body>{html}</body>")
    parser.close()
    return parser.fragment()


class _NoticeHtmlPreviewParser(HTMLParser):
    _VOID_TAGS = {"area", "base", "br", "col", "embed", "hr", "img", "input", "link", "meta", "param", "source", "track", "wbr"}
    _SKIP_TAGS = {"script", "style", "noscript", "iframe", "object", "embed", "head", "form", "button"}
    _ALLOWED_TAGS = {
        "a",
        "b",
        "blockquote",
        "br",
        "code",
        "col",
        "colgroup",
        "dd",
        "div",
        "dl",
        "dt",
        "em",
        "h1",
        "h2",
        "h3",
        "h4",
        "h5",
        "h6",
        "i",
        "li",
        "ol",
        "p",
        "pre",
        "small",
        "span",
        "strong",
        "table",
        "tbody",
        "td",
        "tfoot",
        "th",
        "thead",
        "tr",
        "u",
        "ul",
    }

    def __init__(self, *, match_class: str | None = None, match_tag: str | None = None):
        super().__init__(convert_charrefs=True)
        self.match_class = match_class
        self.match_tag = match_tag
        self._capturing = False
        self._capture_depth = 0
        self._skip_depth = 0
        self._done = False
        self._parts: list[str] = []

    def fragment(self) -> str:
        value = "".join(self._parts).strip()
        return value if self._has_visible_text(value) else ""

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        normalized = tag.lower()
        if self._done:
            return
        if not self._capturing:
            if self._matches(normalized, attrs):
                self._capturing = True
                if normalized not in self._VOID_TAGS:
                    self._capture_depth = 1
                return
            return

        if self._skip_depth:
            if normalized not in self._VOID_TAGS:
                self._skip_depth += 1
            return

        if normalized in self._SKIP_TAGS:
            if normalized not in self._VOID_TAGS:
                self._skip_depth = 1
            return

        if normalized not in self._VOID_TAGS:
            self._capture_depth += 1
        if normalized in self._ALLOWED_TAGS:
            self._parts.append(self._render_start_tag(normalized, attrs))

    def handle_endtag(self, tag: str) -> None:
        normalized = tag.lower()
        if self._done or not self._capturing:
            return
        if self._skip_depth:
            if normalized not in self._VOID_TAGS:
                self._skip_depth -= 1
            return
        if normalized in self._ALLOWED_TAGS and normalized not in self._VOID_TAGS:
            self._parts.append(f"</{normalized}>")
        if normalized not in self._VOID_TAGS:
            self._capture_depth -= 1
        if self._capture_depth <= 0:
            self._done = True
            self._capturing = False

    def handle_data(self, data: str) -> None:
        if self._capturing and not self._skip_depth and not self._done:
            self._parts.append(escape(data))

    def _matches(self, tag: str, attrs: list[tuple[str, str | None]]) -> bool:
        if self.match_tag and tag == self.match_tag:
            return True
        if not self.match_class:
            return False
        class_value = dict(attrs).get("class") or ""
        return self.match_class in class_value.split()

    def _render_start_tag(self, tag: str, attrs: list[tuple[str, str | None]]) -> str:
        if tag == "br":
            return "<br>"
        if tag == "a":
            href = self._safe_href(dict(attrs).get("href"))
            if not href:
                return "<a>"
            return f'<a href="{escape(href, quote=True)}" target="_blank" rel="noreferrer">'
        if tag in {"td", "th"}:
            return f"<{tag}{self._span_attrs(attrs)}>"
        return f"<{tag}>"

    @staticmethod
    def _span_attrs(attrs: list[tuple[str, str | None]]) -> str:
        rendered = []
        for name, value in attrs:
            if name.lower() not in {"colspan", "rowspan"}:
                continue
            if value and value.isdigit():
                rendered.append(f' {name.lower()}="{escape(value, quote=True)}"')
        return "".join(rendered)

    @staticmethod
    def _safe_href(value: str | None) -> str | None:
        if not value:
            return None
        parsed = urlparse(value.strip())
        if parsed.scheme.lower() in {"http", "https", "mailto"}:
            return value.strip()
        return None

    @staticmethod
    def _has_visible_text(value: str) -> bool:
        parser = _NoticeTextProbe()
        parser.feed(value)
        parser.close()
        return bool(parser.text.strip())


class _NoticeTextProbe(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.text = ""

    def handle_data(self, data: str) -> None:
        self.text += data


def _excel_rows(
    *,
    content: bytes,
    file_ext: str,
    max_rows: int,
    max_cols: int,
    sheet_index: int = 0,
) -> ExcelPreviewData:
    if file_ext == "xlsx":
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
        if sheet_index < 0 or sheet_index >= len(workbook.worksheets):
            sheet = workbook.worksheets[0]
        else:
            sheet = workbook.worksheets[sheet_index]
        source_rows = sheet.max_row or 0
        source_cols = sheet.max_column or 0
        row_count = min(source_rows, max_rows)
        col_count = min(source_cols, max_cols)
        merge_spans, covered_cells = _xlsx_merge_spans(sheet, max_row=row_count, max_col=col_count)
        rendered_cols = _rendered_column_count(
            values=lambda row_index, col_index: sheet.cell(row=row_index, column=col_index).value,
            row_count=row_count,
            col_count=col_count,
            merge_spans=merge_spans,
        )
        rows = [
            [sheet.cell(row=row_index, column=col_index).value for col_index in range(1, rendered_cols + 1)]
            for row_index in range(1, row_count + 1)
        ]
        display_values = _xlsx_display_values(sheet, row_count=row_count, rendered_cols=rendered_cols)
        workbook.close()
        return ExcelPreviewData(
            sheet_name=sheet.title,
            rows=rows,
            source_rows=source_rows,
            source_cols=source_cols,
            rendered_cols=rendered_cols,
            merge_spans=merge_spans,
            covered_cells=covered_cells,
            display_values=display_values,
        )
    if file_ext == "xls":
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True, formatting_info=True)
        sheet = workbook.sheet_by_index(0)
        row_count = min(sheet.nrows, max_rows)
        col_count = min(sheet.ncols, max_cols)
        merge_spans, covered_cells = _xlrd_merge_spans(sheet, max_row=row_count, max_col=col_count)
        rendered_cols = _rendered_column_count(
            values=lambda row_index, col_index: sheet.cell_value(row_index - 1, col_index - 1),
            row_count=row_count,
            col_count=col_count,
            merge_spans=merge_spans,
        )
        rows = [sheet.row_values(row_index, 0, rendered_cols) for row_index in range(row_count)]
        col_widths = _xlrd_column_widths(sheet, rendered_cols=rendered_cols)
        row_styles = _xlrd_row_styles(sheet, row_count=row_count)
        cell_styles = _xlrd_cell_styles(workbook, sheet, row_count=row_count, rendered_cols=rendered_cols, merge_spans=merge_spans)
        display_values = _xlrd_display_values(workbook, sheet, row_count=row_count, rendered_cols=rendered_cols)
        display_html = _xlrd_display_html(workbook, sheet, row_count=row_count, rendered_cols=rendered_cols, display_values=display_values)
        workbook.release_resources()
        return ExcelPreviewData(
            sheet_name=sheet.name,
            rows=rows,
            source_rows=sheet.nrows,
            source_cols=sheet.ncols,
            rendered_cols=rendered_cols,
            merge_spans=merge_spans,
            covered_cells=covered_cells,
            col_widths=col_widths,
            row_styles=row_styles,
            cell_styles=cell_styles,
            display_values=display_values,
            display_html=display_html,
        )
    raise ValueError(f"UNSUPPORTED_EXCEL_EXT: {file_ext}")


def _excel_sheets(
    *,
    content: bytes,
    file_ext: str,
    max_rows: int,
    max_cols: int,
) -> list[ExcelPreviewData]:
    """2026-08-12 新增（apply 自 info）：遍历所有 sheet 生成预览数据列表。
    render_excel_preview_multi_sheet_html 专用。定额域单 sheet 用 _excel_rows 即可。"""
    if file_ext == "xlsx":
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
        results: list[ExcelPreviewData] = []
        for sheet in workbook.worksheets:
            source_rows = sheet.max_row or 0
            source_cols = sheet.max_column or 0
            row_count = min(source_rows, max_rows)
            col_count = min(source_cols, max_cols)
            merge_spans, covered_cells = _xlsx_merge_spans(sheet, max_row=row_count, max_col=col_count)
            rendered_cols = _rendered_column_count(
                values=lambda row_index, col_index: sheet.cell(row=row_index, column=col_index).value,
                row_count=row_count,
                col_count=col_count,
                merge_spans=merge_spans,
            )
            rows = [
                [sheet.cell(row=row_index, column=col_index).value for col_index in range(1, rendered_cols + 1)]
                for row_index in range(1, row_count + 1)
            ]
            display_values = _xlsx_display_values(sheet, row_count=row_count, rendered_cols=rendered_cols)
            results.append(ExcelPreviewData(
                sheet_name=sheet.title,
                rows=rows,
                source_rows=source_rows,
                source_cols=source_cols,
                rendered_cols=rendered_cols,
                merge_spans=merge_spans,
                covered_cells=covered_cells,
                display_values=display_values,
            ))
        workbook.close()
        return results
    if file_ext == "xls":
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True, formatting_info=True)
        results: list[ExcelPreviewData] = []
        for sheet in workbook.sheets():
            row_count = min(sheet.nrows, max_rows)
            col_count = min(sheet.ncols, max_cols)
            merge_spans, covered_cells = _xlrd_merge_spans(sheet, max_row=row_count, max_col=col_count)
            rendered_cols = _rendered_column_count(
                values=lambda row_index, col_index: sheet.cell_value(row_index - 1, col_index - 1),
                row_count=row_count,
                col_count=col_count,
                merge_spans=merge_spans,
            )
            rows = [sheet.row_values(row_index, 0, rendered_cols) for row_index in range(row_count)]
            col_widths = _xlrd_column_widths(sheet, rendered_cols=rendered_cols)
            row_styles = _xlrd_row_styles(sheet, row_count=row_count)
            cell_styles = _xlrd_cell_styles(workbook, sheet, row_count=row_count, rendered_cols=rendered_cols, merge_spans=merge_spans)
            display_values = _xlrd_display_values(workbook, sheet, row_count=row_count, rendered_cols=rendered_cols)
            display_html = _xlrd_display_html(workbook, sheet, row_count=row_count, rendered_cols=rendered_cols, display_values=display_values)
            results.append(ExcelPreviewData(
                sheet_name=sheet.name,
                rows=rows,
                source_rows=sheet.nrows,
                source_cols=sheet.ncols,
                rendered_cols=rendered_cols,
                merge_spans=merge_spans,
                covered_cells=covered_cells,
                col_widths=col_widths,
                row_styles=row_styles,
                cell_styles=cell_styles,
                display_values=display_values,
                display_html=display_html,
            ))
        workbook.release_resources()
        return results
    raise ValueError(f"UNSUPPORTED_EXCEL_EXT: {file_ext}")


def _render_row(
    row_index: int,
    row: list[object],
    *,
    rendered_cols: int,
    merge_spans: dict[tuple[int, int], tuple[int, int]],
    covered_cells: set[tuple[int, int]],
    row_styles: dict[int, str],
    cell_styles: dict[tuple[int, int], str],
    display_values: dict[tuple[int, int], str],
    display_html: dict[tuple[int, int], str],
) -> str:
    row_style = _style_attr(row_styles.get(row_index))
    rendered_cells = [f'<th scope="row">{row_index}</th>']
    for col_index in range(1, rendered_cols + 1):
        if (row_index, col_index) in covered_cells:
            continue
        value = row[col_index - 1] if col_index <= len(row) else None
        class_name = " class=\"cell-num\"" if isinstance(value, Number) else ""
        span_attrs = _span_attrs(merge_spans.get((row_index, col_index)))
        style_attr = _style_attr(cell_styles.get((row_index, col_index)))
        display_value = display_values.get((row_index, col_index), _display_value(value))
        cell_html = display_html.get((row_index, col_index), escape(display_value))
        rendered_cells.append(f"<td{class_name}{span_attrs}{style_attr}>{cell_html}</td>")
    return f"<tr{row_style}>{''.join(rendered_cells)}</tr>"


def _xlsx_merge_spans(sheet, *, max_row: int, max_col: int) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    merge_spans: dict[tuple[int, int], tuple[int, int]] = {}
    covered_cells: set[tuple[int, int]] = set()
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row > max_row or merged_range.min_col > max_col:
            continue
        row_span = min(merged_range.max_row, max_row) - merged_range.min_row + 1
        col_span = min(merged_range.max_col, max_col) - merged_range.min_col + 1
        if row_span <= 1 and col_span <= 1:
            continue
        top_left = (merged_range.min_row, merged_range.min_col)
        merge_spans[top_left] = (row_span, col_span)
        for row_index in range(merged_range.min_row, min(merged_range.max_row, max_row) + 1):
            for col_index in range(merged_range.min_col, min(merged_range.max_col, max_col) + 1):
                if (row_index, col_index) != top_left:
                    covered_cells.add((row_index, col_index))
    return merge_spans, covered_cells


def _xlsx_display_values(sheet, *, row_count: int, rendered_cols: int) -> dict[tuple[int, int], str]:
    values: dict[tuple[int, int], str] = {}
    for row_index in range(1, row_count + 1):
        for col_index in range(1, rendered_cols + 1):
            cell = sheet.cell(row=row_index, column=col_index)
            values[(row_index, col_index)] = _xlsx_display_value(cell)
    return values


def _xlsx_display_value(cell) -> str:
    value = cell.value
    if isinstance(value, Number):
        decimals = _decimal_places(str(cell.number_format or ""))
        if decimals is not None:
            return f"{float(value):.{decimals}f}"
    return _display_value(value)


def _xlrd_merge_spans(sheet, *, max_row: int, max_col: int) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    merge_spans: dict[tuple[int, int], tuple[int, int]] = {}
    covered_cells: set[tuple[int, int]] = set()
    for row_start, row_end, col_start, col_end in getattr(sheet, "merged_cells", []):
        start_row = row_start + 1
        start_col = col_start + 1
        end_row = min(row_end, max_row)
        end_col = min(col_end, max_col)
        if start_row > max_row or start_col > max_col:
            continue
        row_span = end_row - start_row + 1
        col_span = end_col - start_col + 1
        if row_span <= 1 and col_span <= 1:
            continue
        top_left = (start_row, start_col)
        merge_spans[top_left] = (row_span, col_span)
        for row_index in range(start_row, end_row + 1):
            for col_index in range(start_col, end_col + 1):
                if (row_index, col_index) != top_left:
                    covered_cells.add((row_index, col_index))
    return merge_spans, covered_cells


def _xlrd_column_widths(sheet, *, rendered_cols: int) -> list[float]:
    widths = []
    for col_index in range(rendered_cols):
        info = sheet.colinfo_map.get(col_index)
        width = float(getattr(info, "width", 2048) or 2048)
        widths.append(max(width, 1.0))
    total = sum(widths) or 1.0
    return [width / total * 100 for width in widths]


def _xlrd_row_styles(sheet, *, row_count: int) -> dict[int, str]:
    styles: dict[int, str] = {}
    for row_index in range(1, row_count + 1):
        info = sheet.rowinfo_map.get(row_index - 1)
        height = getattr(info, "height", None)
        if height:
            styles[row_index] = f"height: {_twips_to_px(height):.1f}px"
    return styles


def _xlrd_cell_styles(
    workbook,
    sheet,
    *,
    row_count: int,
    rendered_cols: int,
    merge_spans: dict[tuple[int, int], tuple[int, int]],
) -> dict[tuple[int, int], str]:
    styles: dict[tuple[int, int], str] = {}
    for row_index in range(1, row_count + 1):
        for col_index in range(1, rendered_cols + 1):
            styles[(row_index, col_index)] = _xlrd_cell_style(
                workbook,
                sheet,
                row_index=row_index,
                col_index=col_index,
                span=merge_spans.get((row_index, col_index)),
            )
    return styles


def _xlrd_cell_style(workbook, sheet, *, row_index: int, col_index: int, span: tuple[int, int] | None) -> str:
    xf = workbook.xf_list[sheet.cell(row_index - 1, col_index - 1).xf_index]
    font = workbook.font_list[xf.font_index]
    declarations: list[str] = []
    if align := _xlrd_text_align(xf.alignment.hor_align):
        declarations.append(f"text-align: {align}")
    if valign := _xlrd_vertical_align(xf.alignment.vert_align):
        declarations.append(f"vertical-align: {valign}")
    if xf.alignment.text_wrapped:
        declarations.append("white-space: normal")
    if font.height:
        declarations.append(f"font-size: {_twips_to_px(font.height):.1f}px")
    if font.bold or "黑体" in str(font.name):
        declarations.append("font-weight: 600")
    declarations.extend(_xlrd_border_styles(workbook, sheet, row_index=row_index, col_index=col_index, span=span))
    return "; ".join(declarations)


def _xlrd_text_align(value: int) -> str | None:
    return {
        2: "center",
        3: "right",
        4: "center",
        5: "justify",
        6: "center",
    }.get(value)


def _xlrd_vertical_align(value: int) -> str | None:
    return {
        0: "top",
        1: "middle",
        2: "bottom",
        3: "middle",
    }.get(value)


def _xlrd_border_styles(workbook, sheet, *, row_index: int, col_index: int, span: tuple[int, int] | None) -> list[str]:
    row_span, col_span = span or (1, 1)
    top_row = row_index - 1
    left_col = col_index - 1
    bottom_row = min(top_row + row_span - 1, sheet.nrows - 1)
    right_col = min(left_col + col_span - 1, sheet.ncols - 1)
    top_left = workbook.xf_list[sheet.cell(top_row, left_col).xf_index].border
    top_right = workbook.xf_list[sheet.cell(top_row, right_col).xf_index].border
    bottom_left = workbook.xf_list[sheet.cell(bottom_row, left_col).xf_index].border
    return [
        _border_decl("left", top_left.left_line_style),
        _border_decl("right", top_right.right_line_style),
        _border_decl("top", top_left.top_line_style),
        _border_decl("bottom", bottom_left.bottom_line_style),
    ]


def _border_decl(side: str, line_style: int) -> str:
    if line_style:
        return f"border-{side}: 1px solid var(--excel-border, #262626)"
    return f"border-{side}: none"


def _xlrd_display_values(workbook, sheet, *, row_count: int, rendered_cols: int) -> dict[tuple[int, int], str]:
    values: dict[tuple[int, int], str] = {}
    for row_index in range(1, row_count + 1):
        for col_index in range(1, rendered_cols + 1):
            cell = sheet.cell(row_index - 1, col_index - 1)
            values[(row_index, col_index)] = _xlrd_display_value(workbook, cell)
    return values


def _xlrd_display_html(
    workbook,
    sheet,
    *,
    row_count: int,
    rendered_cols: int,
    display_values: dict[tuple[int, int], str],
) -> dict[tuple[int, int], str]:
    html_values: dict[tuple[int, int], str] = {}
    rich_text_runs = getattr(sheet, "rich_text_runlist_map", {}) or {}
    for (zero_row, zero_col), runlist in rich_text_runs.items():
        row_index = zero_row + 1
        col_index = zero_col + 1
        if row_index > row_count or col_index > rendered_cols:
            continue
        text = display_values.get((row_index, col_index), "")
        if not text:
            continue
        html_values[(row_index, col_index)] = _xlrd_rich_text_html(
            workbook,
            sheet,
            zero_row=zero_row,
            zero_col=zero_col,
            text=text,
            runlist=runlist,
        )
    return html_values


def _xlrd_rich_text_html(workbook, sheet, *, zero_row: int, zero_col: int, text: str, runlist: list[tuple[int, int]]) -> str:
    cell = sheet.cell(zero_row, zero_col)
    xf = workbook.xf_list[cell.xf_index]
    runs = [(0, xf.font_index)]
    runs.extend((offset, font_index) for offset, font_index in runlist if 0 <= offset < len(text))
    runs = sorted(dict(runs).items())
    parts: list[str] = []
    for run_index, (start, font_index) in enumerate(runs):
        end = runs[run_index + 1][0] if run_index + 1 < len(runs) else len(text)
        segment = text[start:end]
        if not segment:
            continue
        style = _xlrd_font_style(workbook.font_list[font_index])
        segment_html = _html_text_with_line_breaks(segment)
        if style:
            parts.append(f'<span style="{escape(style, quote=True)}">{segment_html}</span>')
        else:
            parts.append(segment_html)
    return "".join(parts)


def _xlrd_font_style(font) -> str:
    declarations: list[str] = []
    if font.height:
        declarations.append(f"font-size: {_twips_to_px(font.height):.1f}px")
    if font.bold or "黑体" in str(font.name):
        declarations.append("font-weight: 600")
    return "; ".join(declarations)


def _html_text_with_line_breaks(value: str) -> str:
    return escape(value).replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")


def _xlrd_display_value(workbook, cell) -> str:
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        xf = workbook.xf_list[cell.xf_index]
        number_format = workbook.format_map.get(xf.format_key)
        format_text = str(number_format.format_str if number_format else "")
        decimals = _decimal_places(format_text)
        if decimals is not None:
            return f"{float(cell.value):.{decimals}f}"
    return _display_value(cell.value)


def _decimal_places(format_text: str) -> int | None:
    if "." not in format_text:
        return None
    right = format_text.split(".", 1)[1]
    digits = 0
    for char in right:
        if char in {"0", "#"}:
            digits += 1
        else:
            break
    return digits if digits else None


def _render_colgroup(col_widths: list[float]) -> str:
    if not col_widths:
        return ""
    cols = ['<col class="excel-row-index-col" />']
    cols.extend(f'<col style="width: {width:.4f}%;" />' for width in col_widths)
    return f"<colgroup>{''.join(cols)}</colgroup>"


def _table_style(preview: ExcelPreviewData) -> str:
    if not preview.col_widths:
        return ""
    min_width = max(44 + preview.rendered_cols * 60, 360)
    return f' style="min-width: {min_width}px;"'


def _style_attr(style: str | None) -> str:
    if not style:
        return ""
    return f' style="{escape(style, quote=True)}"'


def _twips_to_px(twips: int | float) -> float:
    return float(twips) / 20 * 96 / 72


def _rendered_column_count(*, values, row_count: int, col_count: int, merge_spans: dict[tuple[int, int], tuple[int, int]]) -> int:
    last_content_col = 0
    for row_index in range(1, row_count + 1):
        for col_index in range(1, col_count + 1):
            if _display_value(values(row_index, col_index)):
                last_content_col = max(last_content_col, col_index)
    for (row_index, col_index), (_row_span, col_span) in merge_spans.items():
        if _display_value(values(row_index, col_index)):
            last_content_col = max(last_content_col, min(col_count, col_index + col_span - 1))
    if row_count > 0 and col_count > 0:
        return max(1, last_content_col)
    return 0


def _span_attrs(span: tuple[int, int] | None) -> str:
    if span is None:
        return ""
    row_span, col_span = span
    attrs = []
    if col_span > 1:
        attrs.append(f'colspan="{col_span}"')
    if row_span > 1:
        attrs.append(f'rowspan="{row_span}"')
    return f" {' '.join(attrs)}" if attrs else ""


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)
