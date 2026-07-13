from io import BytesIO
from pathlib import Path
import re

from openpyxl import Workbook
import pytest

from app.file_preview import render_excel_preview_html


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def project_path(*parts: str) -> Path:
    return PROJECT_ROOT.joinpath(*parts)


def workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_excel_preview_trims_trailing_empty_columns_and_preserves_long_chinese_headers():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "乌鲁木齐"
    sheet.cell(row=1, column=1, value="材料名称及规格型号")
    sheet.cell(row=1, column=2, value="单位")
    sheet.cell(row=2, column=1, value="HRB400E 螺纹钢 Φ20")
    sheet.cell(row=2, column=2, value="t")
    sheet.cell(row=1, column=60, value="")

    html = render_excel_preview_html(file_name="乌鲁木齐信息价.xlsx", content=workbook_bytes(workbook), file_ext="xlsx")

    assert "材料名称及规格型号" in html
    assert "已截取前200行/60列，下载看完整" not in html
    assert html.count("<td") == 4
    assert "data-source-cols=\"60\"" in html
    assert "data-rendered-cols=\"2\"" in html


def test_excel_preview_outputs_xlsx_merged_cells_as_colspan_and_rowspan():
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "乌鲁木齐市2026年4月份建设工程综合价格信息"
    sheet.merge_cells("A2:A3")
    sheet["A2"] = "序号"
    sheet["B2"] = "材料名称及规格型号"
    sheet["C2"] = "单位"
    sheet["D2"] = "含税价格"
    sheet["B3"] = "名称"
    sheet["C3"] = "计量单位"
    sheet["D3"] = "元"

    html = render_excel_preview_html(file_name="乌鲁木齐信息价.xlsx", content=workbook_bytes(workbook), file_ext="xlsx")

    assert '<td colspan="4">乌鲁木齐市2026年4月份建设工程综合价格信息</td>' in html
    assert '<td rowspan="2">序号</td>' in html
    assert "data-merged=\"1\"" in html
    assert html.count("乌鲁木齐市2026年4月份建设工程综合价格信息") == 1


def test_excel_preview_formats_xlsx_numbers_with_cell_number_format():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "信息价"
    sheet.cell(row=1, column=1, value="除税价格（元）")
    sheet.cell(row=1, column=2, value="含税价格（元）")
    sheet.cell(row=2, column=1, value=3991.83890712321)
    sheet.cell(row=2, column=1).number_format = "0.00_);[Red]\\(0.00\\)"
    sheet.cell(row=2, column=2, value=4500)
    sheet.cell(row=2, column=2).number_format = "0.00_);\\(0.00\\)"

    html = render_excel_preview_html(file_name="临夏信息价.xlsx", content=workbook_bytes(workbook), file_ext="xlsx")

    assert "3991.84" in html
    assert "4500.00" in html
    assert "3991.83890712321" not in html


def test_excel_preview_outputs_real_luzhou_xls_merged_rows_as_colspan():
    path = Path("/Users/tms/Downloads/住建网泸州市市区材料价格信息2025年11月.xls")
    if not path.exists():
        pytest.skip("local Luzhou xls fixture is not available")

    html = render_excel_preview_html(file_name=path.name, content=path.read_bytes(), file_ext="xls")

    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S)
    assert 'data-merged="1"' in html
    assert 'colspan="5"' in rows[0]
    assert rows[0].count("<td") == 1
    assert "使用须知" in rows[0]
    assert 'colspan="5"' in rows[2]
    assert rows[2].count("<td") == 1
    assert "黑色及有色金属" in rows[2]


def test_excel_preview_uses_real_luzhou_xls_source_layout_for_display():
    path = Path("/Users/tms/Downloads/住建网泸州市市区材料价格信息2025年11月.xls")
    if not path.exists():
        pytest.skip("local Luzhou xls fixture is not available")

    html = render_excel_preview_html(file_name=path.name, content=path.read_bytes(), file_ext="xls")

    rows = re.findall(r"<tr(?P<tr_attrs>[^>]*)>(?P<body>.*?)</tr>", html, re.S)
    assert "<colgroup>" in html
    assert "data-source-layout=\"1\"" in html
    assert "min-width:" in html
    assert "3150.00" in rows[3][1]
    assert "text-align: center" in rows[1][1]
    assert "white-space: normal" in rows[0][1]
    assert "height:" in rows[0][0]
    assert "border-bottom: 1px solid var(--excel-border" in rows[1][1]


def test_excel_preview_preserves_real_luzhou_xls_rich_text_font_runs():
    path = Path("/Users/tms/Downloads/住建网泸州市市区材料价格信息2025年11月.xls")
    if not path.exists():
        pytest.skip("local Luzhou xls fixture is not available")

    html = render_excel_preview_html(file_name=path.name, content=path.read_bytes(), file_ext="xls")

    first_row = re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S)[0]
    assert "使用须知" in first_row
    assert "主要材料信息价" in first_row
    assert '<span style="font-size: 26.7px">                 使用须知' in first_row
    assert '<br>' in first_row
    assert re.search(r'<span style="font-size: 13\.3px">主要材料信息价', first_row)


def test_excel_preview_css_uses_content_width_horizontal_scroll_and_nowrap():
    css = project_path("app", "ui", "styles.css").read_text(encoding="utf-8")

    table_block = re.search(r"\.excel-preview table \{(?P<body>.*?)\}", css, re.S).group("body")
    cell_block = re.search(r"\.excel-preview th,\n\.excel-preview td \{(?P<body>.*?)\}", css, re.S).group("body")
    wrap_block = re.search(r"\.excel-table-wrap \{(?P<body>.*?)\}", css, re.S).group("body")

    assert "table-layout: auto;" in table_block
    assert "width: max-content;" in table_block
    assert "white-space: nowrap;" in cell_block
    assert "min-width: 60px;" in cell_block
    assert "overflow-x: auto;" in wrap_block
    assert '.excel-preview[data-source-layout="1"] table' in css
    assert "table-layout: fixed;" in css
    assert '.excel-preview[data-source-layout="1"] td' in css
