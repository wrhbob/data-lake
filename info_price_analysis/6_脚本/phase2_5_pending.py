"""
phase2_5_pending.py - Phase 2.5 pending_pool 闭环

用户审 pending_pool.csv → 回写 YAML 库 → 移出 pending

决策：
  ✅ (yes)    → 写 material_synonyms.yaml（标准名 → 变体）
  ❌ (no)     → 写 known_different.yaml（黑名单）
  ✏️ (edit)   → 用户输入标准名，自定义映射
  ⏭️ (skip)    → 跳过（保留在 pending_pool）

输入:
  5_日志/pending_pool.csv (Phase 2 产出)
  2_库/material_synonyms.yaml
  2_库/known_different.yaml

输出:
  5_日志/decisions_history.csv  ← 审计（不回填 pending）
  2_库/*.yaml                   ← 库更新
"""

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PENDING_PATH = PROJECT_ROOT / "5_日志" / "pending_pool.csv"
DECISIONS_PATH = PROJECT_ROOT / "5_日志" / "decisions_history.csv"
LIBRARY_DIR = PROJECT_ROOT / "2_库"

# 决策 emoji → 内部值
DECISION_MAP = {
    '✅': 'yes', '✔': 'yes', 'yes': 'yes', 'y': 'yes',
    '❌': 'no', '✖': 'no', 'no': 'no', 'n': 'no',
    '✏️': 'edit', 'edit': 'edit', 'e': 'edit',
    '⏭️': 'skip', 'skip': 'skip', 's': 'skip',
}

# decision_history 列（审计）
DECISION_COLUMNS = PENDING_COLUMNS = [
    'pending_id', 'file_a', 'spec_a', 'file_b', 'spec_b',
    'jaccard', 'material_hint', 'reason', 'decision',
    'decided_at', 'user_note', 'writeback_target', 'writeback_done',
]


def load_yaml(path: Path) -> dict:
    if not path.exists():
        return {}
    with open(path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f) or {}


def save_yaml(path: Path, data: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        yaml.dump(data, f, allow_unicode=True, sort_keys=False)


def writeback_material_synonym(canonical: str, variant: str):
    """写 material_synonyms.yaml"""
    path = LIBRARY_DIR / 'material_synonyms.yaml'
    data = load_yaml(path)
    # data: {标准: [变体]}
    if canonical not in data:
        data[canonical] = []
    if variant not in data[canonical]:
        data[canonical].append(variant)
    save_yaml(path, data)


def writeback_known_different(mat_a: str, spec_a: str, mat_b: str, spec_b: str):
    """写 known_different.yaml"""
    path = LIBRARY_DIR / 'known_different.yaml'
    data = load_yaml(path)
    if 'pairs' not in data:
        data['pairs'] = []
    new_pair = [mat_a, spec_a, mat_b, spec_b]
    if new_pair not in data['pairs']:
        data['pairs'].append(new_pair)
    save_yaml(path, data)


def append_decision_history(row: dict):
    """追加到 decisions_history.csv"""
    DECISIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    file_exists = DECISIONS_PATH.exists()
    with open(DECISIONS_PATH, 'a', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=DECISION_COLUMNS)
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)


def prompt_decision(pending_row: dict) -> tuple:
    """
    提示用户决策
    返回 (decision, writeback_target, user_note)
      decision: yes / no / edit / skip
    """
    print(f'\n--- {pending_row["pending_id"]} ---')
    print(f'  Jaccard: {pending_row["jaccard"]}')
    print(f'  材料提示: {pending_row["material_hint"]}')
    print(f'  原因: {pending_row["reason"]}')
    print(f'  A: {pending_row["file_a"]} / spec={pending_row["spec_a"]}')
    print(f'  B: {pending_row["file_b"]} / spec={pending_row["spec_b"]}')

    while True:
        choice = input('\n  决策 [y/n/e/s] (y=同物✅, n=不同❌, e=自定义✏️, s=跳过⏭️): ').strip().lower()
        if choice in ('y', 'yes', '✅'):
            return 'yes', 'material_synonyms.yaml', ''
        elif choice in ('n', 'no', '❌'):
            return 'no', 'known_different.yaml', ''
        elif choice in ('e', 'edit', '✏️'):
            canonical = input('  请输入标准材料名: ').strip()
            if canonical:
                return 'edit', 'material_synonyms.yaml', canonical
            print('  ✗ 标准名不能为空')
        elif choice in ('s', 'skip', '⏭️'):
            return 'skip', '', ''
        else:
            print('  ✗ 请输入 y/n/e/s')


def process_pending(pending_rows: list, args) -> tuple:
    """
    处理 pending_pool
    返回 (kept_rows, decisions_made)
    """
    kept = []
    decisions = []

    for i, row in enumerate(pending_rows, 1):
        print(f'\n[{i}/{len(pending_rows)}]')
        decision, target, note = prompt_decision(row)

        if decision == 'skip':
            # 跳过，保留在 pending
            kept.append(row)
            continue

        # 决策时间
        decided_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row['decision'] = decision
        row['decided_at'] = decided_at
        row['user_note'] = note
        row['writeback_target'] = target
        row['writeback_done'] = False

        # 回写 YAML
        try:
            if decision == 'no':
                # ❌ 不同物 → known_different
                writeback_known_different(
                    row['material_hint'], row['spec_a'],
                    row['material_hint'], row['spec_b'],
                )
                row['writeback_done'] = True
                print(f'  ✓ 已写入 known_different.yaml')
            elif decision in ('yes', 'edit'):
                # ✅ / ✏️ 同物 → material_synonyms
                if decision == 'edit':
                    # 用户输入的标准名作为 canonical
                    canonical = note
                    variant = row['material_hint']
                else:
                    # yes: 取 material_hint 作为标准名
                    canonical = row['material_hint']
                    variant = row['material_hint']  # 简化：自映射
                writeback_material_synonym(canonical, variant)
                row['writeback_done'] = True
                print(f'  ✓ 已写入 material_synonyms.yaml')
        except Exception as e:
            print(f'  ✗ 回写失败: {e}')
            row['writeback_done'] = False

        # 追加审计
        append_decision_history(row)
        decisions.append(row)
        print(f'  ✓ 决策已记录到 decisions_history.csv')

    return kept, decisions


def apply_from_xlsx(xlsx_path: Path) -> tuple:
    """
    从 xlsx（Phase 4_v2 产出）读 decision 列，回写 yaml
    返回 (resolved_ids, unresolved_rows)
    """
    print(f'读取 xlsx: {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    # 找含 'pending_id' 列的 sheet（扫描所有 sheet，所有行）
    ws = None
    for name in wb.sheetnames:
        candidate = wb[name]
        for row in candidate.iter_rows(values_only=True):
            if row and any('pending_id' in str(c).lower() if c else False for c in row):
                ws = candidate
                print(f'  ✓ 找到 pending sheet: {name}')
                break
        if ws is not None:
            break
    if ws is None:
        print('  ✗ xlsx 未找到 pending sheet')
        return [], []

    # 找表头行（扫所有行）
    headers = None
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and any('pending_id' in str(c).lower() if c else False for c in row):
            headers = list(row)
            header_row = row_idx
            break
    if not headers:
        print('  ✗ 未找到 pending_id 列')
        return [], []

    print(f'  表头行: {header_row}')
    print(f'  列: {headers}')

    # 列索引
    col_idx = {h: i for i, h in enumerate(headers)}
    pending_id_col = col_idx.get('pending_id')
    decision_col = col_idx.get('decision(待填)') if 'decision(待填)' in col_idx else col_idx.get('decision')
    user_note_col = col_idx.get('user_note(待填)') if 'user_note(待填)' in col_idx else col_idx.get('user_note')
    material_col = col_idx.get('材料提示') if '材料提示' in col_idx else None
    spec_a_col = col_idx.get('specA')
    spec_b_col = col_idx.get('specB')
    file_a_col = col_idx.get('文件A')
    file_b_col = col_idx.get('文件B')
    reason_col = col_idx.get('原因')
    jaccard_col = col_idx.get('Jaccard')
    writeback_col = col_idx.get('writeback_done')

    if pending_id_col is None or decision_col is None:
        print(f'  ✗ 缺关键列: pending_id={pending_id_col}, decision={decision_col}')
        return [], []

    # 同步 pending_pool.csv 加载（用于精确字段匹配 + 后续更新）
    with open(PENDING_PATH, 'r', encoding='utf-8') as f:
        pending_pool = list(csv.DictReader(f))
    pending_dict = {r['pending_id']: r for r in pending_pool}

    resolved_ids = []
    unresolved_rows = []

    # 逐行处理
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if not row or row[pending_id_col] is None:
            continue
        pid = str(row[pending_id_col]).strip()
        if not pid:
            continue
        decision_raw = row[decision_col] if decision_col < len(row) else None
        if not decision_raw:
            continue  # 未填，跳过

        decision_raw = str(decision_raw).strip()
        decision = DECISION_MAP.get(decision_raw.lower(), DECISION_MAP.get(decision_raw))
        if not decision:
            print(f'  ⚠️ {pid}: 未知 decision "{decision_raw}"，跳过')
            continue

        if decision == 'skip':
            # 跳过，不处理
            continue

        # 从 pending_dict 取完整信息
        pr = pending_dict.get(pid)
        if not pr:
            print(f'  ⚠️ {pid}: 在 pending_pool.csv 中找不到，跳过')
            continue

        user_note = (str(row[user_note_col]).strip() if user_note_col is not None
                     and user_note_col < len(row) and row[user_note_col] else '')
        # 如果是 edit 模式但 user_note 空，提示
        if decision == 'edit' and not user_note:
            print(f'  ⚠️ {pid}: edit 模式但 user_note 空，跳过')
            continue

        # 回写 yaml
        ok = False
        try:
            if decision == 'no':
                # ❌ 不同物 → known_different
                writeback_known_different(
                    pr['material_hint'], pr['spec_a'],
                    pr['material_hint'], pr['spec_b'],
                )
                ok = True
                print(f'  ✓ {pid}: ❌ → known_different.yaml')
            elif decision in ('yes', 'edit'):
                # ✅ / ✏️ 同物 → material_synonyms
                if decision == 'edit':
                    canonical = user_note
                    variant = pr['material_hint']
                else:
                    canonical = pr['material_hint']
                    variant = pr['material_hint']
                writeback_material_synonym(canonical, variant)
                ok = True
                print(f'  ✓ {pid}: {decision_raw} → material_synonyms.yaml ({canonical})')
        except Exception as e:
            print(f'  ✗ {pid}: 回写失败: {e}')

        # 写审计
        decided_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        audit_row = {
            'pending_id': pid,
            'file_a': pr.get('file_a', ''),
            'spec_a': pr.get('spec_a', ''),
            'file_b': pr.get('file_b', ''),
            'spec_b': pr.get('spec_b', ''),
            'jaccard': pr.get('jaccard', ''),
            'material_hint': pr.get('material_hint', ''),
            'reason': pr.get('reason', ''),
            'decision': decision,
            'decided_at': decided_at,
            'user_note': user_note,
            'writeback_target': 'material_synonyms.yaml' if decision in ('yes', 'edit') else 'known_different.yaml',
            'writeback_done': ok,
        }
        append_decision_history(audit_row)

        # 回写 xlsx（标记 writeback_done）
        if writeback_col is not None and writeback_col < len(row):
            ws.cell(row=row_idx, column=writeback_col + 1, value='✓' if ok else '✗')
        # 同时填回 decided_at 列（如果存在）
        decided_at_col = col_idx.get('decided_at')
        if decided_at_col is not None and decided_at_col < len(row):
            ws.cell(row=row_idx, column=decided_at_col + 1, value=decided_at)

        if ok:
            resolved_ids.append(pid)
        else:
            unresolved_rows.append(pr)

    # 保存 xlsx
    wb.save(xlsx_path)
    print(f'\n  ✓ xlsx 已更新: {xlsx_path}')

    return resolved_ids, unresolved_rows


def update_pending_pool(resolved_ids: list):
    """从 pending_pool.csv 移除已解决的"""
    if not resolved_ids:
        return
    with open(PENDING_PATH, 'r', encoding='utf-8') as f:
        rows = list(csv.DictReader(f))
    kept = [r for r in rows if r['pending_id'] not in resolved_ids]
    with open(PENDING_PATH, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=PENDING_COLUMNS)
        writer.writeheader()
        for r in kept:
            writer.writerow(r)
    print(f'  ✓ pending_pool.csv: 移除 {len(resolved_ids)} 条，保留 {len(kept)} 条')


def main():
    parser = argparse.ArgumentParser(description='Phase 2.5 pending_pool 闭环')
    parser.add_argument('--pending', default=str(PENDING_PATH))
    parser.add_argument('--interactive', action='store_true', help='交互式逐个审')
    parser.add_argument('--status', action='store_true', help='只看状态不动')
    parser.add_argument('--apply-xlsx', type=str, metavar='XLSX_PATH',
                        help='从 xlsx（Phase 4_v2 产出）读 decision 列回写 yaml')
    args = parser.parse_args()

    # 读 pending
    pending_path = Path(args.pending)
    if not pending_path.exists():
        print(f'无 pending_pool.csv: {pending_path}')
        sys.exit(0)

    with open(pending_path, 'r', encoding='utf-8') as f:
        pending_rows = list(csv.DictReader(f))

    print(f'=== pending_pool.csv 共 {len(pending_rows)} 条 ===')

    # 状态模式
    if args.status:
        # 按 reason 分组
        from collections import Counter
        reasons = Counter(r['reason'] for r in pending_rows)
        for reason, cnt in reasons.most_common():
            print(f'  {cnt:4}  {reason}')
        print(f'\n  前 5 条:')
        for r in pending_rows[:5]:
            print(f'    {r["pending_id"]}: j={r["jaccard"]}, {r["material_hint"][:30]}')
        return

    # xlsx 回写模式
    if args.apply_xlsx:
        xlsx_path = Path(args.apply_xlsx)
        if not xlsx_path.exists():
            print(f'✗ xlsx 不存在: {xlsx_path}')
            sys.exit(1)
        resolved, unresolved = apply_from_xlsx(xlsx_path)
        print(f'\n=== xlsx 回写完成 ===')
        print(f'已解决: {len(resolved)}')
        print(f'未解决: {len(unresolved)}')
        # 更新 pending_pool
        update_pending_pool(resolved)
        return

    # 交互模式
    if not args.interactive:
        print('提示：')
        print('  --interactive           交互式逐个审（CLI）')
        print('  --status                查看统计')
        print('  --apply-xlsx <path>     从 xlsx 读 decision 列回写 yaml')
        return

    # 处理
    kept, decisions = process_pending(pending_rows, args)

    # 写回 pending_pool.csv（保留 kept，去除 decisions）
    if decisions:
        with open(pending_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=PENDING_COLUMNS)
            writer.writeheader()
            for row in kept:
                writer.writerow(row)
        print(f'\n✓ pending_pool.csv 已更新: {len(kept)}/{len(pending_rows)} 保留')

    print(f'\n=== Phase 2.5 完成 ===')
    print(f'决策: {len(decisions)}')
    print(f'  ✅ yes: {sum(1 for d in decisions if d["decision"]=="yes")}')
    print(f'  ❌ no: {sum(1 for d in decisions if d["decision"]=="no")}')
    print(f'  ✏️ edit: {sum(1 for d in decisions if d["decision"]=="edit")}')
    print(f'  ⏭️ skip: {sum(1 for d in decisions if d["decision"]=="skip")}')
    print(f'剩余 pending: {len(kept)}')
    print(f'审计: {DECISIONS_PATH}')


if __name__ == '__main__':
    main()
