"""
phase1_columns.py - Phase 1 智能列识别

自动识别主 sheet 的列角色：
  district / material / spec / unit / price_with_tax / price_without_tax /
  price_excl_tax / origin / notes / raw_data

输入: 3_中间产物/manifest.csv
输出: 3_中间产物/column_mapping.csv
"""

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = Path(__file__).resolve().parent.parent
MANIFEST_PATH = PROJECT_ROOT / "3_中间产物" / "manifest.csv"
OUTPUT_PATH = PROJECT_ROOT / "3_中间产物" / "column_mapping.csv"

# === 表头关键词映射 ===
# 注意：'price_without_tax' (含"不含税") 必须排在 'price_with_tax' (含"含税") 之前
# 因为"不含税价"是"含税价"的子串，正则匹配时更具体的先匹配
HEADER_PATTERNS = {
    'district': [r'区名', r'区县', r'地区', r'行政区域'],
    'material': [r'材料名?称', r'物料名?称', r'商品名?称', r'品名', r'材料'],
    'spec': [r'规格型?号', r'规格', r'型号', r'规?格?/?型?号'],
    'unit': [r'单位'],
    'origin': [r'产地', r'厂家', r'厂地', r'生产地', r'来源'],
    'notes': [r'备注', r'说明', r'原始数据', r'附注'],
    'price_without_tax': [r'不含税价', r'不含税价格', r'不含税单价'],
    'price_with_tax': [r'含税价', r'含税价格', r'含税单价'],
    'price_excl_tax': [r'除税价', r'除税价格', r'除税单价'],
}

# 输出列
OUTPUT_COLUMNS = [
    'file_path',
    'sheet_name',
    'header_row',
    'district_col',
    'material_col',
    'spec_col',
    'unit_col',
    'origin_col',
    'notes_col',
    'price_with_tax_col',
    'price_without_tax_col',
    'price_excl_tax_col',
    'price_count_filled',  # 实际有数据的价列数
    'confidence',           # 识别置信度 high/medium/low
    'unrecognized',         # 未识别列（逗号分隔）
]


def detect_header_row(ws, max_scan: int = 10) -> int:
    """
    检测表头行
    策略：前 max_scan 行内，找含 ≥3 个非空单元格的行作为候选；
          若该行含'区名'/'材料名称'/'规格'等关键词 → 命中表头
    返回行号（1-indexed），找不到返回 1
    """
    candidates = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        non_empty = sum(1 for c in row if c is not None and str(c).strip() != '')
        if non_empty >= 3:
            row_text = ' '.join(str(c) for c in row if c is not None)
            # 含关键词
            if any(kw in row_text for kw in ['区名', '材料', '规格', '单位', '价']):
                return row_idx
            candidates.append((row_idx, non_empty))
    if candidates:
        # 候选中取非空最多的
        return max(candidates, key=lambda x: x[1])[0]
    return 1


def classify_column(header: str, data_type: str, fill_rate: float) -> tuple:
    """
    分类单列
    返回 (role, confidence)
    role: district/material/spec/unit/origin/notes/price_*/unknown
    """
    if not header:
        return ('unknown', 'low')

    header_clean = str(header).strip()

    # 关键词匹配
    for role, patterns in HEADER_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, header_clean):
                # 价格列需校验：数据应可解析为数字
                if role.startswith('price_'):
                    if data_type == 'numeric' or fill_rate > 0:
                        return (role, 'high')
                    else:
                        return (role, 'medium')  # 标题像价格但数据空
                # 其他列：标题命中即高置信度
                return (role, 'high')

    return ('unknown', 'low')


def analyze_column_data(ws, col_idx: int, data_start_row: int) -> tuple:
    """
    分析列的数据特征
    返回 (data_type, fill_rate)
    data_type: numeric / string / mixed
    fill_rate: 0.0-1.0
    """
    numeric_count = 0
    string_count = 0
    total = 0

    for row in ws.iter_rows(min_row=data_start_row, max_row=data_start_row + 100, values_only=True):
        if col_idx >= len(row):
            continue
        v = row[col_idx]
        if v is None or (isinstance(v, str) and v.strip() == ''):
            continue
        total += 1
        # 检测数字（字符串形式的数字也算）
        if isinstance(v, (int, float)):
            numeric_count += 1
        elif isinstance(v, str):
            try:
                float(v.strip())
                numeric_count += 1
            except (ValueError, AttributeError):
                string_count += 1

    if total == 0:
        return ('empty', 0.0)

    fill_rate = total / 100.0  # 扫了 100 行
    if fill_rate > 1.0:
        fill_rate = 1.0

    if numeric_count > string_count * 2:
        return ('numeric', fill_rate)
    elif string_count > numeric_count * 2:
        return ('string', fill_rate)
    else:
        return ('mixed', fill_rate)


def process_file(file_path: str, sheet_name: str) -> dict:
    """处理单个 xlsx 文件，返回列映射"""
    print(f'\n处理: {Path(file_path).name}')

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    # 1. 检测表头行
    header_row = detect_header_row(ws)
    print(f'  表头行: {header_row}')

    # 2. 读表头
    headers = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        headers = list(row)
        break

    print(f'  列数: {len(headers)}')
    for i, h in enumerate(headers):
        print(f'    [{i+1}] {h}')

    # 3. 分类每列
    data_start = header_row + 1
    mapping = {
        'district_col': '',
        'material_col': '',
        'spec_col': '',
        'unit_col': '',
        'origin_col': '',
        'notes_col': '',
        'price_with_tax_col': '',
        'price_without_tax_col': '',
        'price_excl_tax_col': '',
    }
    price_filled = []
    unrecognized = []

    for col_idx, header in enumerate(headers):
        col_name = f'col_{col_idx + 1}'  # 1-indexed 列名
        data_type, fill_rate = analyze_column_data(ws, col_idx, data_start)

        role, confidence = classify_column(header, data_type, fill_rate)

        if role.startswith('price_'):
            # 价格列：检查是否实际有数据
            if fill_rate > 0:
                price_filled.append((role, col_name, fill_rate))
                mapping[f'{role}_col'] = col_name
                print(f'  ✓ {col_name} ({header}): {role} ({fill_rate:.0%})')
            else:
                # 价格列但无数据
                print(f'  ○ {col_name} ({header}): {role} (空, 不计入)')
        elif role != 'unknown':
            mapping[f'{role}_col'] = col_name
            print(f'  ✓ {col_name} ({header}): {role}')
        else:
            unrecognized.append(f'{col_name}({header})')
            print(f'  ? {col_name} ({header}): 未识别')

    wb.close()

    # 4. 置信度评估
    required_roles = ['material', 'spec', 'unit']
    required_filled = sum(1 for r in required_roles if mapping.get(f'{r}_col'))
    price_count = len(price_filled)

    if required_filled == 3 and price_count >= 1 and len(unrecognized) <= 1:
        confidence = 'high'
    elif required_filled >= 2 and price_count >= 1:
        confidence = 'medium'
    else:
        confidence = 'low'

    return {
        'file_path': file_path,
        'sheet_name': sheet_name,
        'header_row': header_row,
        **mapping,
        'price_count_filled': price_count,
        'confidence': confidence,
        'unrecognized': ','.join(unrecognized),
    }


def main():
    parser = argparse.ArgumentParser(
        description='Phase 1 智能列识别'
    )
    parser.add_argument(
        '--manifest',
        default=str(MANIFEST_PATH),
        help='manifest.csv 路径'
    )
    args = parser.parse_args()

    # 读 manifest
    with open(args.manifest, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    print(f'从 manifest 读 {len(rows)} 个文件')

    # 处理每个文件
    output_rows = []
    for row in rows:
        result = process_file(row['file_path'], row['sheet_name'])
        output_rows.append(result)

    # 写 column_mapping.csv
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for row in output_rows:
            writer.writerow(row)

    print(f'\n=== Phase 1 完成 ===')
    print(f'✓ column_mapping.csv 已写: {OUTPUT_PATH}')
    print(f'  处理: {len(output_rows)} 个文件')

    # 汇总
    high = sum(1 for r in output_rows if r['confidence'] == 'high')
    medium = sum(1 for r in output_rows if r['confidence'] == 'medium')
    low = sum(1 for r in output_rows if r['confidence'] == 'low')
    print(f'  置信度: high={high}, medium={medium}, low={low}')


if __name__ == '__main__':
    main()
