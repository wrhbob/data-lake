"""
phase0_5_manifest.py - Phase 0.5 期数识别 + manifest.csv 生成

按项目.md §6 三级降级：
  L1: 文件名正则（自动）
  L2: 手动标注（hash 命名 fallback，--interactive 或 --map）
  L3: 用户编辑 manifest.csv（escape hatch）
  Excel 元数据（出版期/数据月份）→ 仅写入 excel_*_ref 列，**不参与期号判断**

产出：
  3_中间产物/manifest.csv            ← 12 列
  3_中间产物/period_mapping.yaml    ← sha256 缓存（L2）
"""

import argparse
import csv
import glob
import hashlib
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml

# Windows GBK 编码修复
sys.stdout.reconfigure(encoding='utf-8')

# === 路径配置 ===
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = PROJECT_ROOT / "1_输入"
MANIFEST_PATH = PROJECT_ROOT / "3_中间产物" / "manifest.csv"
MAPPING_PATH = PROJECT_ROOT / "3_中间产物" / "period_mapping.yaml"

# === L1 文件名正则 ===
# 匹配: "2026年成都市工程造价信息价第4期.xlsx" 或 "...04期.xlsx"
L1_PATTERN = re.compile(r'(?P<year>\d{4})\s*年.*?第(?P<period>\d{1,2})\s*期')

# manifest.csv 列（12 列）
MANIFEST_COLUMNS = [
    'file_path',           # xlsx 全路径
    'file_sha256',         # sha256(file_content)
    'filename_period_source',  # L1/L2/L3
    'year',                # 唯一可信源
    'period',              # 唯一可信源
    'city',                # Excel 元数据
    'sheet_name',          # 主 sheet 名
    'row_count',           # 主 sheet 行数
    'excel_year_ref',      # Excel 元数据（仅供参考）
    'excel_period_ref',    # Excel 元数据（仅供参考）
    'excel_month_ref',     # Excel 元数据（仅供参考）
    'notes',               # 备注
]


def compute_sha256(path: Path) -> str:
    """计算文件 sha256"""
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            h.update(chunk)
    return h.hexdigest()


def extract_filename_period(filename: str):
    """
    L1: 文件名正则匹配
    返回 (year, period) 或 None
    """
    m = L1_PATTERN.search(filename)
    if m:
        return m.group('year'), m.group('period').zfill(2)
    return None


def read_excel_metadata(path: Path):
    """
    读 Excel 元数据 sheet
    返回 {city, year, period, month} - 仅作 excel_*_ref 用
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if '元数据' not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb['元数据']
        meta = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1] is not None:
                meta[str(row[0]).strip()] = str(row[1]).strip()
        wb.close()
        return {
            'city': meta.get('城市', ''),
            'year': meta.get('年份', ''),
            'period': meta.get('出版期', ''),
            'month': meta.get('数据月份', ''),
        }
    except Exception as e:
        print(f'  ⚠️  读 Excel 元数据失败: {e}')
        return {}


def read_excel_row_count(path: Path) -> tuple:
    """
    读主 sheet 行数
    返回 (sheet_name, row_count)
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        # 优先级: 市场信息价 > 第一个 sheet
        main_sheet = '市场信息价' if '市场信息价' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[main_sheet]
        # 数据从第 4 行开始（表头在第 3 行）
        row_count = max(0, ws.max_row - 3)
        wb.close()
        return main_sheet, row_count
    except Exception as e:
        print(f'  ⚠️  读主 sheet 失败: {e}')
        return '', 0


def load_period_mapping() -> dict:
    """读 period_mapping.yaml 缓存"""
    if not MAPPING_PATH.exists():
        return {}
    with open(MAPPING_PATH, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f) or {}
    return data.get('period_mapping', {})


def save_period_mapping(mapping: dict):
    """写 period_mapping.yaml 缓存"""
    MAPPING_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(MAPPING_PATH, 'w', encoding='utf-8') as f:
        yaml.dump(
            {'period_mapping': mapping},
            f,
            allow_unicode=True,
            sort_keys=False,
        )


def parse_map_arg(map_str: str) -> dict:
    """
    解析 --map 参数
    格式: "file1=period1,file2=period2"
    返回 {filename: period}
    """
    result = {}
    for pair in map_str.split(','):
        if '=' not in pair:
            continue
        fname, period = pair.split('=', 1)
        result[fname.strip()] = period.strip().zfill(2)
    return result


def prompt_user_for_period(filename: str, excel_meta: dict) -> str:
    """--interactive: 提示用户输入期号"""
    print(f'\n  [{filename}]')
    if excel_meta:
        print(f'    Excel 元数据（仅供参考）:')
        print(f'      声称出版期: {excel_meta.get("period", "?")}')
        print(f'      声称数据月份: {excel_meta.get("month", "?")}')
    while True:
        try:
            period = input(f'    请输入期号（2 位数字，如 05）: ').strip().zfill(2)
            if period.isdigit() and 1 <= int(period) <= 12:
                return period
            print('    ✗ 必须是 01-12 的数字')
        except (EOFError, KeyboardInterrupt):
            print('\n  中断')
            sys.exit(1)


def process_file(path: Path, args, mapping: dict) -> dict | None:
    """
    处理单个文件，返回 manifest 行 dict 或 None（待手动标注但未标）
    """
    filename = path.name
    print(f'\n处理: {filename}')

    sha256 = compute_sha256(path)
    print(f'  sha256: {sha256[:16]}...')

    # L1: 文件名正则
    fp_result = extract_filename_period(filename)
    if fp_result:
        year, period = fp_result
        source = 'L1'
        notes = '文件名匹配'
        print(f'  ✓ L1 自动识别: year={year}, period={period}')
    else:
        # L2: 查缓存
        if sha256 in mapping:
            cached = mapping[sha256]
            year = cached['year']
            period = cached['period']
            source = 'L2'
            notes = f"缓存命中 ({cached.get('user_confirmed_at', '?')})"
            print(f'  ✓ L2 缓存命中: year={year}, period={period}')
        else:
            # L2 待标
            excel_meta = read_excel_metadata(path)

            if args.interactive:
                period = prompt_user_for_period(filename, excel_meta)
                year_input = input(f'    请输入年份（默认 2026）: ').strip()
                year = year_input if year_input else '2026'
                source = 'L2'
                notes = '交互标注'
                # 写缓存
                mapping[sha256] = {
                    'filename': filename,
                    'period': period,
                    'year': int(year),
                    'source': 'L2_interactive',
                    'user_confirmed_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'excel_period_ref': excel_meta.get('period', ''),
                    'excel_month_ref': excel_meta.get('month', ''),
                }
                print(f'  ✓ L2 交互标注: year={year}, period={period}')
            elif args.map:
                # 从 --map 解析
                filename_map = parse_map_arg(args.map)
                if filename in filename_map:
                    period = filename_map[filename]
                    year = '2026'
                    source = 'L2'
                    notes = '命令行标注'
                    mapping[sha256] = {
                        'filename': filename,
                        'period': period,
                        'year': int(year),
                        'source': 'L2_cmdline',
                        'user_confirmed_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'excel_period_ref': excel_meta.get('period', ''),
                        'excel_month_ref': excel_meta.get('month', ''),
                    }
                    print(f'  ✓ L2 命令行标注: year={year}, period={period}')
                else:
                    print(f'  ✗ --map 中无 {filename} 的标注')
                    return None
            else:
                print(f'  ✗ L1 未匹配 + 无缓存 + 无 --interactive/--map')
                print(f'     解决: 重新跑 --interactive 或 --map "filename=period"')
                return None

    # 读 Excel 元数据 + 主 sheet
    excel_meta = read_excel_metadata(path)
    sheet_name, row_count = read_excel_row_count(path)

    return {
        'file_path': str(path),
        'file_sha256': sha256,
        'filename_period_source': source,
        'year': year,
        'period': period,
        'city': excel_meta.get('city', ''),
        'sheet_name': sheet_name,
        'row_count': row_count,
        'excel_year_ref': excel_meta.get('year', ''),
        'excel_period_ref': excel_meta.get('period', ''),
        'excel_month_ref': excel_meta.get('month', ''),
        'notes': notes,
    }


def write_manifest(rows: list, manifest_path: Path = MANIFEST_PATH):
    """写 manifest.csv（12 列）"""
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    with open(manifest_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=MANIFEST_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    print(f'\n✓ manifest.csv 已写: {manifest_path}')


def main():
    parser = argparse.ArgumentParser(
        description='Phase 0.5 期数识别 + manifest.csv 生成'
    )
    parser.add_argument(
        '--input',
        default=str(DEFAULT_INPUT / '*.xlsx'),
        help='xlsx 文件 glob 模式（默认 1_输入/*.xlsx）'
    )
    parser.add_argument(
        '--interactive',
        action='store_true',
        help='hash 文件进入交互式标注'
    )
    parser.add_argument(
        '--map',
        type=str,
        help='命令行标注，格式 "file1=period1,file2=period2"'
    )
    args = parser.parse_args()

    # 收集文件
    files = sorted(glob.glob(args.input))
    if not files:
        print(f'未找到文件: {args.input}')
        sys.exit(1)

    print(f'扫描 {len(files)} 个文件...')

    # 读缓存
    mapping = load_period_mapping()

    # 处理每个文件
    rows = []
    failed = []
    for f in files:
        path = Path(f)
        row = process_file(path, args, mapping)
        if row:
            rows.append(row)
        else:
            failed.append(path.name)

    # 写缓存 + manifest
    if any(r['filename_period_source'] == 'L2' for r in rows):
        save_period_mapping(mapping)
        print(f'✓ period_mapping.yaml 已更新: {MAPPING_PATH}')

    if rows:
        write_manifest(rows)

    # 总结
    print(f'\n=== Phase 0.5 完成 ===')
    print(f'成功: {len(rows)}/{len(files)}')
    if failed:
        print(f'失败: {len(failed)}')
        for f in failed:
            print(f'  - {f}')
        sys.exit(1)


if __name__ == '__main__':
    main()
