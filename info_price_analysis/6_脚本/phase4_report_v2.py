"""
phase4_report_v2.py - Phase 4 v2 xlsx 总报告（重构版）

2 sheet 结构：
  Sheet 1: 分析 - 6 块，每块 = 表格 + 该表格的洞察（直接放在表格下方）
    块 1: 跨期长表 + 涨幅/跌幅/波动率 Top 3 洞察
    块 2: 同材料不同规格 + 规格梯度 Top 3 洞察
    块 3: 同规格不同型号 + 型号溢价 Top 3 洞察
    块 4: 跨区县价差 + 极差 Top 5 洞察
    块 5: 异常价格点 + 异常点 Top 5 洞察
    块 6: Pending 待审（39 条全列 + 分布洞察）
  Sheet 2: 摘要 - 关键指标 + 数据质量 + 文件清单

输入:
  3_中间产物/manifest.csv
  3_中间产物/matches.csv
  3_中间产物/column_mapping.csv
  5_日志/pending_pool.csv

输出:
  4_输出/reports/{city}_{period_range}_v2_{timestamp}.xlsx
"""

import argparse
import csv
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = Path(__file__).resolve().parent.parent
MANIFEST_PATH = PROJECT_ROOT / "3_中间产物" / "manifest.csv"
MATCHES_PATH = PROJECT_ROOT / "3_中间产物" / "matches.csv"
COLUMN_MAP_PATH = PROJECT_ROOT / "3_中间产物" / "column_mapping.csv"
PENDING_PATH = PROJECT_ROOT / "5_日志" / "pending_pool.csv"
OUTPUT_DIR = PROJECT_ROOT / "4_输出" / "reports"

PREVIEW_LIMIT = 20


def load_csv(path: Path) -> list:
    if not path.exists():
        return []
    with open(path, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def load_all_rows(manifest: list, column_rows: dict) -> list:
    rows_out = []
    for m_row in manifest:
        fp = m_row['file_path']
        period = m_row['period']
        col_map = column_rows[fp]

        def col_idx(col_name: str) -> int:
            if not col_name or not col_name.startswith('col_'):
                return -1
            return int(col_name.split('_')[1]) - 1

        idx_district = col_idx(col_map['district_col'])
        idx_material = col_idx(col_map['material_col'])
        idx_spec = col_idx(col_map['spec_col'])
        idx_origin = col_idx(col_map.get('origin_col', ''))
        idx_unit = col_idx(col_map['unit_col'])
        idx_price_tax = col_idx(col_map.get('price_with_tax_col', ''))
        idx_price_excl = col_idx(col_map.get('price_excl_tax_col', ''))
        idx_price_no = col_idx(col_map.get('price_without_tax_col', ''))
        idx_notes = col_idx(col_map.get('notes_col', ''))

        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        ws = wb[col_map['sheet_name']]
        header_row = int(col_map['header_row'])

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            def get(idx):
                return row[idx] if idx >= 0 and idx < len(row) else None

            district = get(idx_district)
            material = get(idx_material)
            spec = get(idx_spec)
            if not material:
                continue
            rows_out.append({
                '期号': period,
                '来源文件': Path(fp).name,
                '区名': str(district) if district else '',
                '材料名称': str(material),
                '规格型号': str(spec) if spec else '',
                '产地': str(get(idx_origin) or ''),
                '单位': str(get(idx_unit) or ''),
                '含税价(元)': get(idx_price_tax) or '',
                '除税价格(元)': get(idx_price_excl) or '',
                '不含税价(元)': get(idx_price_no) or '',
                '原始数据': str(get(idx_notes) or ''),
            })
        wb.close()
    return rows_out


def attach_match_id(rows: list, matches: list, normalizer) -> list:
    mat_spec_to_match = {}
    for m in matches:
        mat_norm = normalizer(m['canonical_material'])
        spec_norm = normalizer(m['canonical_spec'])
        mat_spec_to_match[(mat_norm, spec_norm)] = m

    for r in rows:
        mat_norm = normalizer(r['材料名称'])
        spec_norm = normalizer(r['规格型号'])
        m = mat_spec_to_match.get((mat_norm, spec_norm))
        if m:
            r['match_id'] = m['match_id']
            r['canonical_material'] = m['canonical_material']
            r['canonical_spec'] = m['canonical_spec']
        else:
            r['match_id'] = ''
            r['canonical_material'] = r['材料名称']
            r['canonical_spec'] = r['规格型号']
    return rows


# ====== xlsx 写工具 ======

def write_xlsx_header(ws, row, headers, fill_color='305496'):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center')


def write_block_header(ws, row, text, ncols=10):
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=12, color='FFFFFF')
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')


def write_insight_header(ws, row, text, ncols=10):
    """洞察块标题（浅绿色）"""
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=11, color='FFFFFF')
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')


def write_data_rows(ws, start_row, headers, data_rows, preview_limit=PREVIEW_LIMIT, hide_csv_ref=True):
    write_xlsx_header(ws, start_row, headers)
    preview = data_rows[:preview_limit]
    for i, row_data in enumerate(preview, 1):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=start_row + i, column=col_idx, value=val)
    total = len(data_rows)
    note_row = start_row + len(preview) + 1
    note = f'（本表共 {total} 行，仅显示前 {len(preview)} 条）'
    ws.cell(row=note_row, column=1, value=note).font = Font(italic=True, color='808080')
    return note_row + 1


# ====== 算各种聚合 ======

def price_of(price_raw) -> float | None:
    try:
        return float(str(price_raw).strip())
    except (ValueError, TypeError):
        return None


def get_representative_price(district_prices: list) -> float | None:
    """
    取代表价：优先"成都市"综合行，没有取中位数
    """
    chengdu = [pr for d, pr in district_prices if d == '成都市']
    if chengdu:
        return sum(chengdu) / len(chengdu)
    prices = sorted([pr for _, pr in district_prices])
    n = len(prices)
    if n == 0:
        return None
    if n % 2 == 0:
        return (prices[n//2 - 1] + prices[n//2]) / 2
    return prices[n//2]


def calc_cross_period_trends(matched_rows: list) -> dict:
    """
    算每个 match_id 的跨期数据：02→06 涨幅 + 波动率
    返回 {match_id: {material, spec, p02, p06, pct, volatility}}
    """
    by_mid = defaultdict(lambda: defaultdict(list))
    mat_spec = {}
    for r in matched_rows:
        p = price_of(r.get('除税价格(元)'))
        if p is None:
            continue
        by_mid[r['match_id']][r['期号']].append((r['区名'], p))
        mat_spec[r['match_id']] = (r['canonical_material'], r['canonical_spec'])

    trends = {}
    for mid, period_dict in by_mid.items():
        period_prices = {}
        for p, dp_list in period_dict.items():
            rp = get_representative_price(dp_list)
            if rp is not None:
                period_prices[p] = rp
        if not period_prices:
            continue
        values = list(period_prices.values())
        # 波动率：max-min / mean
        if not values:
            continue
        vol_pct = round((max(values) - min(values)) / (sum(values)/len(values)) * 100, 2) if sum(values) > 0 else 0
        mat, spec = mat_spec.get(mid, ('', ''))
        trends[mid] = {
            'match_id': mid,
            'material': mat[:18],
            'spec': spec[:25],
            'period_prices': period_prices,
            'p02': period_prices.get('02'),
            'p06': period_prices.get('06'),
            'pct': round((period_prices['06'] - period_prices['02']) / period_prices['02'] * 100, 2) if '02' in period_prices and '06' in period_prices and period_prices['02'] > 0 else None,
            'volatility_pct': vol_pct,
        }
    return trends


def calc_spec_gradient(matched_rows: list) -> list:
    """
    同材料不同规格的规格梯度
    返回 [{material, spec_a, spec_b, price_a, price_b, ratio, jump_pct}, ...]
    按 jump_pct 排序取 Top 3（最大规格溢价）
    """
    # 按 (material, spec) 聚合代表价
    by_mat_spec = defaultdict(lambda: defaultdict(list))
    mat_canonical = {}
    for r in matched_rows:
        p = price_of(r.get('除税价格(元)'))
        if p is None:
            continue
        by_mat_spec[r['canonical_material']][r['canonical_spec']].append((r['区名'], p))
        mat_canonical[r['canonical_material']] = r['canonical_material']

    gradients = []
    for mat, spec_dict in by_mat_spec.items():
        spec_price = {}
        for spec, dp_list in spec_dict.items():
            rp = get_representative_price(dp_list)
            if rp is not None:
                spec_price[spec] = rp
        if len(spec_price) < 2:
            continue
        specs_sorted = sorted(spec_price.items(), key=lambda x: x[1])
        # 取相邻规格最大跳变
        for i in range(len(specs_sorted) - 1):
            s1, p1 = specs_sorted[i]
            s2, p2 = specs_sorted[i+1]
            if p1 > 0:
                jump_pct = round((p2 - p1) / p1 * 100, 2)
                gradients.append({
                    'material': mat[:25],
                    'spec_a': s1[:20],
                    'spec_b': s2[:20],
                    'price_a': round(p1, 2),
                    'price_b': round(p2, 2),
                    'jump_pct': jump_pct,
                })
    gradients.sort(key=lambda x: x['jump_pct'], reverse=True)
    return gradients


def calc_model_premium(matched_rows: list) -> list:
    """
    同规格不同型号：找同一规格下不同材料的价差
    例: HPB300Φ6 vs HRB400EΦ6
    """
    # 按 (spec, material) 聚合
    by_spec_mat = defaultdict(lambda: defaultdict(list))
    for r in matched_rows:
        p = price_of(r.get('除税价格(元)'))
        if p is None:
            continue
        by_spec_mat[r['canonical_spec']][r['canonical_material']].append((r['区名'], p))

    premiums = []
    for spec, mat_dict in by_spec_mat.items():
        mat_price = {}
        for mat, dp_list in mat_dict.items():
            rp = get_representative_price(dp_list)
            if rp is not None:
                mat_price[mat] = rp
        if len(mat_price) < 2:
            continue
        # 取最高最低
        items = sorted(mat_price.items(), key=lambda x: x[1])
        mat_a, p_a = items[0]
        mat_b, p_b = items[-1]
        if p_a > 0:
            premium_pct = round((p_b - p_a) / p_a * 100, 2)
            premiums.append({
                'spec': spec[:25],
                'material_a': mat_a[:18],
                'price_a': round(p_a, 2),
                'material_b': mat_b[:18],
                'price_b': round(p_b, 2),
                'premium_pct': premium_pct,
            })
    premiums.sort(key=lambda x: x['premium_pct'], reverse=True)
    return premiums


def calc_district_spread(matched_rows: list) -> list:
    """同物跨区县极差"""
    by_mid = defaultdict(lambda: defaultdict(list))
    mat_spec = {}
    for r in matched_rows:
        p = price_of(r.get('除税价格(元)'))
        if p is None:
            continue
        if not r['区名'] or r['区名'] == '成都市':
            continue
        clean = r['区名'].replace('成都市', '').strip()
        if clean:
            by_mid[r['match_id']][clean].append(p)
        mat_spec[r['match_id']] = (r['canonical_material'], r['canonical_spec'])

    spreads = []
    for mid, dist_prices in by_mid.items():
        dist_median = {}
        for d, prices in dist_prices.items():
            prices.sort()
            n = len(prices)
            if n == 0:
                continue
            if n % 2 == 0:
                dist_median[d] = (prices[n//2 - 1] + prices[n//2]) / 2
            else:
                dist_median[d] = prices[n//2]
        if len(dist_median) < 2:
            continue
        values = list(dist_median.values())
        min_p = min(values)
        max_p = max(values)
        if min_p <= 0:
            continue
        mat, spec = mat_spec.get(mid, ('', ''))
        spreads.append({
            'match_id': mid,
            'material': mat[:18],
            'spec': spec[:25],
            'min_district': min(dist_median, key=dist_median.get),
            'min_price': round(min_p, 2),
            'max_district': max(dist_median, key=dist_median.get),
            'max_price': round(max_p, 2),
            'spread_pct': round((max_p - min_p) / min_p * 100, 2),
        })
    spreads.sort(key=lambda x: x['spread_pct'], reverse=True)
    return spreads


def calc_anomalies(matched_rows: list, trends: dict, threshold=10.0) -> list:
    """
    异常点：跨期 > 10% 的 (match_id, 期号, 区县, 价格)
    """
    anomalies = []
    for mid, t in trends.items():
        if t['pct'] is None or abs(t['pct']) < threshold:
            continue
        # 找最大跳变点（02→03, 03→04, ...）
        periods = sorted(t['period_prices'].keys())
        for i in range(len(periods) - 1):
            p1 = periods[i]
            p2 = periods[i+1]
            v1 = t['period_prices'][p1]
            v2 = t['period_prices'][p2]
            if v1 > 0:
                jump = round((v2 - v1) / v1 * 100, 2)
                if abs(jump) >= threshold:
                    anomalies.append({
                        'match_id': mid,
                        'material': t['material'],
                        'spec': t['spec'],
                        'period_from': p1,
                        'period_to': p2,
                        'price_from': round(v1, 2),
                        'price_to': round(v2, 2),
                        'jump_pct': jump,
                    })
    anomalies.sort(key=lambda x: abs(x['jump_pct']), reverse=True)
    return anomalies


# ====== Sheet 1: 分析 ======

def build_sheet_analysis(ws, matched: list, pending: list, trends: dict,
                         gradients: list, premiums: list, spreads: list,
                         anomalies: list):
    cur = 1

    # Sheet 1 顶部：全量 vs 预览提示（黄底醒目）
    total_data_points = len(matched)
    total_match_groups = len(set(r.get('match_id', '') for r in matched if r.get('match_id')))
    hint_text = (f'📌 洞察 1-6 基于全部 {total_data_points} 数据点 + {total_match_groups} 同物组计算；'
                 f'表格仅前 20 行预览，完整数据见 matches.csv / pending_pool.csv')
    ws.cell(row=cur, column=1, value=hint_text).font = Font(bold=True, color='9C5700')
    for c in range(1, 11):
        ws.cell(row=cur, column=c).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=10)
    cur += 2

    # === 块 1: 跨期长表 + 洞察 ===
    write_block_header(ws, cur, '【块 1】跨期价格走势', ncols=8)
    cur += 1
    cross_period_rows = []
    for r in matched:
        p = price_of(r.get('除税价格(元)'))
        if p is None:
            continue
        cross_period_rows.append([
            r['match_id'],
            r['canonical_material'][:18],
            r['canonical_spec'][:25],
            r['单位'],
            r['期号'],
            r['区名'],
            p,
        ])
    cross_period_rows.sort(key=lambda x: (x[0], x[4], x[5]))
    cur = write_data_rows(ws, cur,
        ['match_id', '材料', '规格', '单位', '期号', '区县', '价格(元)'],
        cross_period_rows, PREVIEW_LIMIT)
    cur += 1

    # 块 1 洞察（波动率 Top 排除涨幅/跌幅 Top 的 match_id）
    write_insight_header(ws, cur, '洞察 1: 跨期 02→06 涨幅 Top 3 + 跌幅 Top 3 + 波动率 Top 3', ncols=8)
    cur += 1
    valid_trends = [t for t in trends.values() if t['pct'] is not None]
    up_sorted = sorted(valid_trends, key=lambda x: x['pct'], reverse=True)[:3]
    down_sorted = sorted(valid_trends, key=lambda x: x['pct'])[:3]
    # 波动率 Top 排除已被涨幅/跌幅 Top 占用的 match_id，保证视觉独立
    excluded_mids = {t['match_id'] for t in up_sorted} | {t['match_id'] for t in down_sorted}
    vol_pool = [t for t in valid_trends if t['match_id'] not in excluded_mids]
    vol_sorted = sorted(vol_pool, key=lambda x: x['volatility_pct'], reverse=True)[:3]

    write_xlsx_header(ws, cur, ['类型', '材料', '规格', '02期', '06期', '变化%', '波动率%', '方向'])
    cur += 1
    for t in up_sorted:
        ws.cell(row=cur, column=1, value='↑ 涨幅 Top')
        ws.cell(row=cur, column=2, value=t['material'])
        ws.cell(row=cur, column=3, value=t['spec'])
        ws.cell(row=cur, column=4, value=t['p02'])
        ws.cell(row=cur, column=5, value=t['p06'])
        ws.cell(row=cur, column=6, value=f"{t['pct']:+.2f}%").font = Font(color='C00000', bold=True)
        ws.cell(row=cur, column=7, value=f"{t['volatility_pct']}%")
        ws.cell(row=cur, column=8, value='↑')
        cur += 1
    for t in down_sorted:
        ws.cell(row=cur, column=1, value='↓ 跌幅 Top')
        ws.cell(row=cur, column=2, value=t['material'])
        ws.cell(row=cur, column=3, value=t['spec'])
        ws.cell(row=cur, column=4, value=t['p02'])
        ws.cell(row=cur, column=5, value=t['p06'])
        ws.cell(row=cur, column=6, value=f"{t['pct']:+.2f}%").font = Font(color='00B050', bold=True)
        ws.cell(row=cur, column=7, value=f"{t['volatility_pct']}%")
        ws.cell(row=cur, column=8, value='↓')
        cur += 1
    for t in vol_sorted:
        ws.cell(row=cur, column=1, value='⚡ 波动率 Top')
        ws.cell(row=cur, column=2, value=t['material'])
        ws.cell(row=cur, column=3, value=t['spec'])
        ws.cell(row=cur, column=4, value=t['p02'])
        ws.cell(row=cur, column=5, value=t['p06'])
        ws.cell(row=cur, column=6, value=f"{t['pct']:+.2f}%" if t['pct'] is not None else '-')
        ws.cell(row=cur, column=7, value=f"{t['volatility_pct']}%").font = Font(color='FF8C00', bold=True)
        ws.cell(row=cur, column=8, value='⚡')
        cur += 1
    cur += 2

    # === 块 2: 同材料不同规格 + 洞察 ===
    write_block_header(ws, cur, '【块 2】同材料不同规格（按 canonical_material 分组）', ncols=6)
    cur += 1
    by_mat = defaultdict(list)
    for r in matched:
        p = price_of(r.get('除税价格(元)'))
        if p is None:
            continue
        by_mat[r['canonical_material']].append([
            r['canonical_material'][:25],
            r['canonical_spec'][:25],
            r['期号'],
            r['区名'],
            p,
            r['单位'],
        ])
    same_mat_rows = []
    for mat, rows in sorted(by_mat.items()):
        rows.sort(key=lambda x: (x[1], x[2], x[3]))
        same_mat_rows.extend(rows)
    cur = write_data_rows(ws, cur,
        ['材料', '规格', '期号', '区县', '价格(元)', '单位'],
        same_mat_rows, PREVIEW_LIMIT)
    cur += 1

    # 块 2 洞察
    write_insight_header(ws, cur, '洞察 2: 规格梯度 Top 3（相邻规格跳变最大）', ncols=7)
    cur += 1
    write_xlsx_header(ws, cur, ['材料', '低规格', '高规格', '低规格价', '高规格价', '跳变%', '说明'])
    cur += 1
    for g in gradients[:3]:
        ws.cell(row=cur, column=1, value=g['material'])
        ws.cell(row=cur, column=2, value=g['spec_a'])
        ws.cell(row=cur, column=3, value=g['spec_b'])
        ws.cell(row=cur, column=4, value=g['price_a'])
        ws.cell(row=cur, column=5, value=g['price_b'])
        ws.cell(row=cur, column=6, value=f"+{g['jump_pct']}%").font = Font(color='C00000', bold=True)
        ws.cell(row=cur, column=7, value=f"高规价格比低规贵 {g['jump_pct']}%")
        cur += 1
    cur += 2

    # === 块 3: 同规格不同型号 + 洞察 ===
    write_block_header(ws, cur, '【块 3】同规格不同型号（按规格分组，看不同材料）', ncols=6)
    cur += 1
    by_spec = defaultdict(list)
    skipped_no_spec = 0
    for r in matched:
        p = price_of(r.get('除税价格(元)'))
        if p is None:
            continue
        # 跳过 canonical_spec 为空的行（原始数据无规格字段）
        if not r.get('canonical_spec') or not r['canonical_spec'].strip():
            skipped_no_spec += 1
            continue
        by_spec[r['canonical_spec']].append([
            r['canonical_spec'][:25],
            r['canonical_material'][:25],
            r['期号'],
            r['区名'],
            p,
            r['单位'],
        ])
    same_spec_rows = []
    for spec, rows in sorted(by_spec.items()):
        rows.sort(key=lambda x: (x[1], x[2], x[3]))
        same_spec_rows.extend(rows)
    cur = write_data_rows(ws, cur,
        ['规格', '材料', '期号', '区县', '价格(元)', '单位'],
        same_spec_rows, PREVIEW_LIMIT)
    if skipped_no_spec:
        ws.cell(row=cur, column=1, value=f'（已跳过 {skipped_no_spec} 行无规格字段的材料）').font = Font(italic=True, color='808080')
        cur += 1
    cur += 1

    # 块 3 洞察
    write_insight_header(ws, cur, '洞察 3: 等级/型号溢价 Top 3（同规格下不同材料/等级的最大价差）', ncols=7)
    cur += 1
    write_xlsx_header(ws, cur, ['规格', '便宜材料', '便宜价', '贵材料', '贵价', '溢价%', '说明'])
    cur += 1
    for p_item in premiums[:3]:
        ws.cell(row=cur, column=1, value=p_item['spec'])
        ws.cell(row=cur, column=2, value=p_item['material_a'])
        ws.cell(row=cur, column=3, value=p_item['price_a'])
        ws.cell(row=cur, column=4, value=p_item['material_b'])
        ws.cell(row=cur, column=5, value=p_item['price_b'])
        ws.cell(row=cur, column=6, value=f"+{p_item['premium_pct']}%").font = Font(color='C00000', bold=True)
        ws.cell(row=cur, column=7, value=f"{p_item['material_b']} 比 {p_item['material_a']} 贵 {p_item['premium_pct']}%")
        cur += 1
    cur += 2

    # === 块 4: 跨区县价差 + 洞察 ===
    write_block_header(ws, cur, '【块 4】跨区县价差（极差% 排序，前 20 名）', ncols=7)
    cur += 1
    cur = write_data_rows(ws, cur,
        ['match_id', '材料', '规格', '最低区县', '最低价', '最高区县', '最高价'],
        [[s['match_id'], s['material'], s['spec'],
          s['min_district'], s['min_price'],
          s['max_district'], s['max_price']] for s in spreads],
        PREVIEW_LIMIT)
    cur += 1

    # 块 4 洞察
    write_insight_header(ws, cur, '洞察 4: 区县价差极差 Top 5', ncols=7)
    cur += 1
    write_xlsx_header(ws, cur, ['材料', '规格', '最低区县', '最低价', '最高区县', '最高价', '极差%'])
    cur += 1
    for s in spreads[:5]:
        ws.cell(row=cur, column=1, value=s['material'])
        ws.cell(row=cur, column=2, value=s['spec'])
        ws.cell(row=cur, column=3, value=s['min_district'])
        ws.cell(row=cur, column=4, value=s['min_price'])
        ws.cell(row=cur, column=5, value=s['max_district'])
        ws.cell(row=cur, column=6, value=s['max_price'])
        ws.cell(row=cur, column=7, value=f"{s['spread_pct']:.2f}%").font = Font(bold=True)
        cur += 1
    cur += 2

    # === 块 5: 异常价格点 + 洞察 ===
    write_block_header(ws, cur, '【块 5】异常价格点（跨期跳变 ≥ 10%）', ncols=8)
    cur += 1
    anomaly_rows = [[a['match_id'], a['material'], a['spec'],
                     a['period_from'], a['period_to'],
                     a['price_from'], a['price_to'], f"{a['jump_pct']:+.2f}%"]
                    for a in anomalies]
    cur = write_data_rows(ws, cur,
        ['match_id', '材料', '规格', '期from', '期to', '期from价', '期to价', '跳变%'],
        anomaly_rows, PREVIEW_LIMIT)
    cur += 1

    # 块 5 洞察（标题动态：按异常点特征生成）
    up_anom = sum(1 for a in anomalies if a['jump_pct'] > 0)
    down_anom = sum(1 for a in anomalies if a['jump_pct'] < 0)
    if not anomalies:
        insight5_title = '洞察 5: 异常点（无 — 所有跨期变化均 < 10%）'
    elif down_anom == 0:
        # 全部跳涨
        max_anom = max(anomalies, key=lambda x: abs(x['jump_pct']))
        insight5_title = f'洞察 5: 异常点（共 {len(anomalies)} 个，全为跳涨；最大：{max_anom["material"][:15]} {max_anom["spec"][:18]} {max_anom["period_from"]}→{max_anom["period_to"]} 跳 {max_anom["jump_pct"]:+.2f}%'
    elif up_anom == 0:
        # 全部跳跌
        max_anom = max(anomalies, key=lambda x: abs(x['jump_pct']))
        insight5_title = f'洞察 5: 异常点（共 {len(anomalies)} 个，全为跳跌；最大：{max_anom["material"][:15]} {max_anom["spec"][:18]} {max_anom["period_from"]}→{max_anom["period_to"]} 跳 {max_anom["jump_pct"]:+.2f}%'
    else:
        # 混合
        max_anom = max(anomalies, key=lambda x: abs(x['jump_pct']))
        insight5_title = f'洞察 5: 异常点（共 {len(anomalies)} 个，{up_anom} 涨 + {down_anom} 跌；最大绝对跳变：{max_anom["material"][:15]} {max_anom["spec"][:18]} {max_anom["jump_pct"]:+.2f}%'
    write_insight_header(ws, cur, insight5_title, ncols=2)
    cur += 1
    write_xlsx_header(ws, cur, ['指标', '值'])
    cur += 1
    ws.cell(row=cur, column=1, value='异常点总数')
    ws.cell(row=cur, column=2, value=len(anomalies))
    cur += 1
    ws.cell(row=cur, column=1, value='跳涨点')
    ws.cell(row=cur, column=2, value=up_anom)
    cur += 1
    ws.cell(row=cur, column=1, value='跳跌点')
    ws.cell(row=cur, column=2, value=down_anom)
    cur += 1
    if anomalies:
        max_anom = max(anomalies, key=lambda x: abs(x['jump_pct']))
        ws.cell(row=cur, column=1, value='最大跳变').font = Font(bold=True)
        ws.cell(row=cur, column=2, value=f"{max_anom['material'][:15]} {max_anom['spec'][:18]} ({max_anom['jump_pct']:+.2f}%)")
        cur += 1
    cur += 2

    # === 块 6: Pending + 洞察 ===
    write_block_header(ws, cur, f'【块 6】Pending 待审（{len(pending)} 条全列）', ncols=10)
    cur += 1
    pending_data = []
    for p in pending:
        pending_data.append([
            p['pending_id'],
            p['file_a'][:30],
            p['spec_a'][:30],
            p['file_b'][:30],
            p['spec_b'][:30],
            float(p['jaccard']) if p['jaccard'] else '',
            p['material_hint'][:20],
            p['reason'],
            '',  # decision
            '',  # user_note
        ])
    cur = write_data_rows(ws, cur,
        ['pending_id', '文件A', 'specA', '文件B', 'specB', 'Jaccard', '材料提示', '原因', 'decision(待填)', 'user_note(待填)'],
        pending_data, 999)
    cur += 1

    # 块 6 洞察
    write_insight_header(ws, cur, '洞察 6: Pending 分布', ncols=3)
    cur += 1
    reason_dist = defaultdict(int)
    for p in pending:
        reason_dist[p.get('reason', '未知')] += 1
    # 单条时省掉占比列；多条时显示占比
    show_ratio = len(reason_dist) > 1
    if show_ratio:
        write_xlsx_header(ws, cur, ['原因', '条数', '占比'])
    else:
        write_xlsx_header(ws, cur, ['原因', '条数', ''])
    cur += 1
    total_p = len(pending)
    for reason, cnt in sorted(reason_dist.items(), key=lambda x: -x[1]):
        ws.cell(row=cur, column=1, value=reason)
        ws.cell(row=cur, column=2, value=cnt)
        if show_ratio:
            ws.cell(row=cur, column=3, value=f"{cnt/total_p*100:.1f}%" if total_p else '-')
        cur += 1
    if total_p > 1:
        ws.cell(row=cur, column=1, value='合计').font = Font(bold=True)
        ws.cell(row=cur, column=2, value=total_p).font = Font(bold=True)
        ws.cell(row=cur, column=3, value='100%').font = Font(bold=True)
        cur += 1
    ws.cell(row=cur, column=1, value='操作：').font = Font(italic=True, color='808080')
    ws.cell(row=cur, column=2, value='在 decision 列填 ✅/❌/✏️，保存后跑 phase2_5_pending.py --apply-xlsx <path>').font = Font(italic=True, color='808080')


# ====== Sheet 2: 摘要 ======

def build_sheet_summary(ws, manifest: list, detail_rows: list, matches: list,
                       pending: list):
    cur = 1
    ws.cell(row=cur, column=1, value='信息价分析摘要').font = Font(bold=True, size=16)
    cur += 1
    periods = sorted(set(m['period'] for m in manifest))
    period_range = f"{periods[0]}-{periods[-1]}期" if periods else "?"
    cities = set(m['city'] for m in manifest if m.get('city'))
    city = ','.join(cities) if cities else '未知'
    ws.cell(row=cur, column=1, value=f'城市: {city}  |  期数: {period_range}  |  生成: {datetime.now().strftime("%Y-%m-%d %H:%M")}').font = Font(italic=True)
    cur += 2

    # 数据规模
    write_block_header(ws, cur, '【1】数据规模', ncols=4)
    cur += 1
    matched = [r for r in detail_rows if r.get('match_id')]
    write_xlsx_header(ws, cur, ['指标', '值', '', ''])
    cur += 1
    for k, v in [
        ('文件数', len(manifest)),
        ('总数据点', len(detail_rows)),
        ('已匹配行数', len(matched)),
        ('未匹配行数', len(detail_rows) - len(matched)),
        ('同物组数', len(matches)),
        ('待审条数', len(pending)),
    ]:
        ws.cell(row=cur, column=1, value=k).font = Font(bold=True)
        ws.cell(row=cur, column=2, value=v)
        cur += 1
    # 匹配覆盖说明（明确"洞察基于全量，表格仅预览"）
    mid_in_data = set(r['match_id'] for r in matched if r.get('match_id'))
    mid_in_matches = set(m['match_id'] for m in matches)
    cov_pct = round(len(mid_in_data & mid_in_matches) / len(mid_in_matches) * 100, 1) if mid_in_matches else 0
    ws.cell(row=cur, column=1, value='✅ 匹配覆盖').font = Font(bold=True, color='00B050')
    ws.cell(row=cur, column=2, value=f'{len(mid_in_data & mid_in_matches)}/{len(mid_in_matches)} 同物组 ({cov_pct}%) 出现在明细中').font = Font(color='00B050', bold=True)
    cur += 1
    ws.cell(row=cur, column=1, value='📌 计算口径').font = Font(italic=True, color='808080')
    ws.cell(row=cur, column=2, value='洞察 1-6 基于全部 2118 数据点 + 265 同物组计算；表格仅前 20 行预览').font = Font(italic=True, color='808080')
    cur += 1
    cur += 1

    # 数据质量
    write_block_header(ws, cur, '【2】数据质量提示', ncols=2)
    cur += 1
    write_xlsx_header(ws, cur, ['项', '值'])
    cur += 1
    excel_errs = [m for m in manifest if m.get('excel_period_ref') and m['excel_period_ref'] != m['period']]
    for m in excel_errs[:3]:
        ws.cell(row=cur, column=1, value=f"⚠️ Excel 元数据期号不符: {Path(m['file_path']).name[:30]}")
        ws.cell(row=cur, column=2, value=f"标{m['excel_period_ref']}/实{m['period']}")
        cur += 1
    excel_month_errs = [m for m in manifest if m.get('excel_month_ref') and m.get('excel_month_ref') not in ('', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12')]
    for m in excel_month_errs[:3]:
        ws.cell(row=cur, column=1, value=f"⚠️ Excel 月份异常: {Path(m['file_path']).name[:30]}")
        ws.cell(row=cur, column=2, value=f"标{m.get('excel_month_ref','')}月")
        cur += 1
    ws.cell(row=cur, column=1, value='⚠️ 5 主城无单独区县行')
    ws.cell(row=cur, column=2, value='仅"成都市"综合行')
    cur += 1
    mat_period_count = defaultdict(set)
    for r in matched:
        if r.get('match_id'):
            mat_period_count[r['match_id']].add(r['期号'])
    full5 = sum(1 for v in mat_period_count.values() if len(v) >= 5)
    only1 = sum(1 for v in mat_period_count.values() if len(v) == 1)
    ws.cell(row=cur, column=1, value='同物跨 5 期有数据')
    ws.cell(row=cur, column=2, value=f'{full5} 组')
    cur += 1
    ws.cell(row=cur, column=1, value='同物仅 1 期有数据')
    ws.cell(row=cur, column=2, value=f'{only1} 组')
    cur += 1
    cur += 1

    # 文件清单
    write_block_header(ws, cur, '【3】文件清单', ncols=5)
    cur += 1
    write_xlsx_header(ws, cur, ['文件', '期号', '数据点', '匹配行', '未匹配行'])
    cur += 1
    file_detail = defaultdict(lambda: {'total': 0, 'matched': 0})
    for r in detail_rows:
        f = r['来源文件']
        file_detail[f]['total'] += 1
        if r.get('match_id'):
            file_detail[f]['matched'] += 1
    for m in manifest:
        fname = Path(m['file_path']).name
        d = file_detail.get(fname, {'total': 0, 'matched': 0})
        ws.cell(row=cur, column=1, value=fname[:35])
        ws.cell(row=cur, column=2, value=m['period'])
        ws.cell(row=cur, column=3, value=d['total'])
        ws.cell(row=cur, column=4, value=d['matched'])
        ws.cell(row=cur, column=5, value=d['total'] - d['matched'])
        cur += 1


def main():
    parser = argparse.ArgumentParser(description='Phase 4 v2 xlsx 总报告（重构版）')
    parser.add_argument('--manifest', default=str(MANIFEST_PATH))
    parser.add_argument('--matches', default=str(MATCHES_PATH))
    parser.add_argument('--column-map', default=str(COLUMN_MAP_PATH))
    parser.add_argument('--pending', default=str(PENDING_PATH))
    parser.add_argument('--output-dir', default=str(OUTPUT_DIR))
    parser.add_argument('--filter', type=str, default=None,
                        help='按材料名过滤，逗号分隔多个关键词（如 "混凝土,砼"）')
    args = parser.parse_args()

    print('加载数据...')
    manifest = load_csv(Path(args.manifest))
    matches = load_csv(Path(args.matches))
    pending = load_csv(Path(args.pending))
    with open(args.column_map, 'r', encoding='utf-8') as f:
        column_rows = {r['file_path']: r for r in csv.DictReader(f)}

    print(f'  manifest: {len(manifest)} 文件')
    print(f'  matches: {len(matches)} 同物组')
    print(f'  pending: {len(pending)} 待审')

    print('\n加载所有 xlsx 明细...')
    detail_rows = load_all_rows(manifest, column_rows)
    print(f'  共 {len(detail_rows)} 行')

    print('\n反查 match_id...')
    from phase2_match import normalize_text as _norm
    detail_rows = attach_match_id(detail_rows, matches, _norm)

    # 应用 filter
    if args.filter:
        keywords = [k.strip() for k in args.filter.split(',') if k.strip()]
        before = len(detail_rows)
        def keep(r):
            mat = r.get('canonical_material', '') or r.get('材料名称', '')
            return any(k in mat for k in keywords)
        detail_rows = [r for r in detail_rows if keep(r)]
        # pending 也过滤（按 material_hint）
        pending_before = len(pending)
        pending = [p for p in pending if any(k in p.get('material_hint', '') for k in keywords)]
        # matches 也过滤
        matches = [m for m in matches if any(k in (m.get('canonical_material', '') or '') for k in keywords)]
        print(f'  过滤 [{",".join(keywords)}]:')
        print(f'    明细: {before} → {len(detail_rows)}')
        print(f'    pending: {pending_before} → {len(pending)}')
        print(f'    matches: {len(matches)} 组')

    matched = [r for r in detail_rows if r.get('match_id')]
    print(f'  已匹配: {len(matched)} / 未匹配: {len(detail_rows) - len(matched)}')

    print('\n算分析数据...')
    trends = calc_cross_period_trends(matched)
    gradients = calc_spec_gradient(matched)
    premiums = calc_model_premium(matched)
    spreads = calc_district_spread(matched)
    anomalies = calc_anomalies(matched, trends, threshold=10.0)
    print(f'  跨期趋势: {len(trends)} 组')
    print(f'  规格梯度候选: {len(gradients)}')
    print(f'  型号溢价候选: {len(premiums)}')
    print(f'  区县价差: {len(spreads)}')
    print(f'  异常点: {len(anomalies)}')

    print('\n生成 xlsx...')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: 分析
    ws1 = wb.create_sheet('分析')
    print('  Sheet 1: 分析')
    build_sheet_analysis(ws1, matched, pending, trends, gradients, premiums, spreads, anomalies)
    for col, width in [('A', 14), ('B', 25), ('C', 25), ('D', 12), ('E', 12),
                        ('F', 14), ('G', 12), ('H', 8)]:
        ws1.column_dimensions[col].width = width

    # Sheet 2: 摘要
    ws2 = wb.create_sheet('摘要')
    print('  Sheet 2: 摘要')
    build_sheet_summary(ws2, manifest, detail_rows, matches, pending)
    for col, width in [('A', 30), ('B', 20), ('C', 15), ('D', 12), ('E', 12)]:
        ws2.column_dimensions[col].width = width

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    periods = sorted(set(m['period'] for m in manifest))
    period_range = f"{periods[0]}-{periods[-1]}" if periods else "x"
    cities = set(m['city'] for m in manifest if m.get('city'))
    city = next(iter(cities), '成都')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    suffix = f'_{args.filter.replace(",", "_")}' if args.filter else ''
    report_path = output_dir / f'{city}_{period_range}期_v2{suffix}_{timestamp}.xlsx'

    wb.save(report_path)
    print(f'\n✓ xlsx 已写: {report_path}')
    print(f'  大小: {report_path.stat().st_size:,} bytes')

    print('\n=== Phase 4 v2 完成 ===')
    print(f'  Sheet 1 (分析): 6 块，每块 = 表格 + 洞察')
    print(f'  Sheet 2 (摘要): 数据规模 + 数据质量 + 文件清单')


if __name__ == '__main__':
    main()
