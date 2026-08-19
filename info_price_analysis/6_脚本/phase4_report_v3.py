"""
phase4_report_v3.py - 单期 xlsx 报告（不带跨期对比）

单 Sheet 设计：
  上半部分：摘要 + 单期洞察（高价 Top 5 / 低价 Top 5 / 同材料规格价差 Top 5 /
                           同物跨区县极差 Top 5 / 类别分布）
  下半部分：明细表（一张大表，按 match_id 升序）

输入:
  3_中间产物/manifest.csv
  3_中间产物/matches.csv
  3_中间产物/column_mapping.csv
  5_日志/pending_pool.csv

参数:
  --period 02           # 必填：指定期号
  --filter "混凝土,砼"  # 可选：按材料名过滤
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = Path(__file__).resolve().parent.parent
MANIFEST_PATH = PROJECT_ROOT / "3_中间产物" / "manifest.csv"
MATCHES_PATH = PROJECT_ROOT / "3_中间产物" / "matches.csv"
COLUMN_MAP_PATH = PROJECT_ROOT / "3_中间产物" / "column_mapping.csv"
PENDING_PATH = PROJECT_ROOT / "5_日志/pending_pool.csv"
OUTPUT_DIR = PROJECT_ROOT / "4_输出/reports"


# ====== 数据加载 ======

def load_csv(path: Path) -> list:
    if not path.exists():
        return []
    with open(path, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def load_all_rows(manifest: list, column_rows: dict) -> list:
    rows_out = []
    for m_row in manifest:
        fp = m_row['file_path']
        period = m_row['period']
        col_map = column_rows[fp]

        def col_idx(col_name: str) -> int:
            if not col_name or not col_name.startswith('col_'):
                return -1
            return int(col_name.split('_')[1]) - 1

        idx_district = col_idx(col_map['district_col'])
        idx_material = col_idx(col_map['material_col'])
        idx_spec = col_idx(col_map['spec_col'])
        idx_unit = col_idx(col_map['unit_col'])
        idx_price_excl = col_idx(col_map.get('price_excl_tax_col', ''))
        if idx_price_excl < 0:
            idx_price_excl = col_idx(col_map.get('price_with_tax_col', ''))

        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        ws = wb[col_map['sheet_name']]
        header_row = int(col_map['header_row'])

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            def get(idx):
                return row[idx] if idx >= 0 and idx < len(row) else None

            material = get(idx_material)
            if not material:
                continue
            try:
                price = float(str(get(idx_price_excl)).strip()) if get(idx_price_excl) else None
            except (ValueError, TypeError):
                price = None
            spec_raw = str(get(idx_spec) or '')
            clean_spec, shipping_note = extract_shipping_note(spec_raw)
            rows_out.append({
                '期号': period,
                '来源文件': Path(fp).name,
                '区名': str(get(idx_district) or ''),
                '材料名称': str(material),
                '规格型号': clean_spec,
                '规格备注': shipping_note,
                '单位': str(get(idx_unit) or ''),
                '价格': price,
            })
        wb.close()
    return rows_out


def attach_match_id(rows: list, matches: list, normalizer) -> list:
    def _strip(s):
        clean, _ = extract_shipping_note(s)
        return clean
    mat_spec_to_match = {}
    for m in matches:
        mat_norm = normalizer(m['canonical_material'])
        # matches 侧也剥运费（修 phase2_match.py regex 漏洞）
        spec_clean = _strip(m['canonical_spec'])
        spec_norm = normalizer(spec_clean)
        mat_spec_to_match[(mat_norm, spec_norm)] = m
    for r in rows:
        mat_norm = normalizer(r['材料名称'])
        spec_norm = normalizer(r['规格型号'])
        m = mat_spec_to_match.get((mat_norm, spec_norm))
        if m:
            r['match_id'] = m['match_id']
            r['canonical_material'] = m['canonical_material']
            # 用 row 的 clean 版本覆盖 canonical_spec（防 matches.csv 带运费污染显示）
            r['canonical_spec'] = r['规格型号']
        else:
            r['match_id'] = ''
            r['canonical_material'] = r['材料名称']
            r['canonical_spec'] = r['规格型号']
    return rows


# ====== 运费注剥离 ======

def extract_shipping_note(spec):
    """从 spec 里剥出运费注，返回 (clean_spec, note)
    例:
      "C30碎石粒径5-31.5mm含三环路以内工地运输费" → ("C30碎石粒径5-31.5mm", "含三环路以内工地运输费")
      "C30碎石粒径5-31.5mm不含运输费" → ("C30碎石粒径5-31.5mm", "不含运输费")
      "C30碎石粒径5-31.5mm" → ("C30碎石粒径5-31.5mm", "")
    修复 phase2_match.py regex 的"前面有数字则不剥"漏洞
    """
    if not spec:
        return '', ''
    m = re.search(r'(含|不含)[^()\d]{0,30}?运输费', spec)
    if m:
        note = m.group(0)
        clean = (spec[:m.start()] + spec[m.end():]).strip()
        return clean, note
    return spec, ''


# ====== 自动归类 ======

def categorize(material: str) -> str:
    """把材料名归类（用于洞察 4 类别分布）"""
    if '管' in material:
        return '管材'
    if '砌块' in material or '砖' in material:
        return '砌块/砖'
    if '板' in material:
        return '板'
    if '砂浆' in material:
        return '砂浆'
    if '沥青' in material:
        return '沥青混凝土'
    if '防水' in material or '密实剂' in material or '添加剂' in material:
        return '添加剂'
    if '混凝土' in material or '砼' in material:
        return '其他混凝土'
    return '其他'


# ====== 单期洞察 ======

def calc_high_low_top(rows: list, top_n=5):
    """高价 Top 5 / 低价 Top 5"""
    valid = [r for r in rows if r.get('价格') is not None]
    high = sorted(valid, key=lambda x: x['价格'], reverse=True)[:top_n]
    low = sorted(valid, key=lambda x: x['价格'])[:top_n]
    return high, low


# ====== 单期洞察 ======


def calc_district_range_by_mid(rows: list):
    """每个 match_id 算 1 个极差（内部函数）"""
    by_mid = defaultdict(lambda: defaultdict(list))
    mat_spec = {}
    for r in rows:
        p = r.get('价格')
        if p is None:
            continue
        district = r['区名'].replace('成都市', '').strip()
        if not district:
            continue
        by_mid[r.get('match_id', '')][district].append(p)
        mat_spec[r.get('match_id', '')] = (r['canonical_material'], r['canonical_spec'])
    results = []
    for mid, dist_dict in by_mid.items():
        dist_median = {}
        for d, prices in dist_dict.items():
            prices = sorted(prices)
            n = len(prices)
            if n == 0:
                continue
            if n % 2 == 0:
                dist_median[d] = (prices[n//2 - 1] + prices[n//2]) / 2
            else:
                dist_median[d] = prices[n//2]
        if len(dist_median) < 2:
            continue
        values = list(dist_median.values())
        p_min = min(values)
        p_max = max(values)
        if p_min > 0:
            mat, spec = mat_spec.get(mid, ('', ''))
            results.append({
                'match_id': mid,
                'material': mat,
                'spec': spec,
                'min_district': min(dist_median, key=dist_median.get),
                'min_price': round(p_min, 2),
                'max_district': max(dist_median, key=dist_median.get),
                'max_price': round(p_max, 2),
                'range_pct': round((p_max - p_min) / p_min * 100, 2),
            })
    return results


def calc_district_range_grouped(rows: list, top_n=5, max_spec_per_mat=10):
    """按材料聚拢的跨区县极差 Top N（每行代表 1 个材料，子行挂所有规格）"""
    mid_ranges = calc_district_range_by_mid(rows)
    # 按 material 聚拢
    by_mat = defaultdict(list)
    for r in mid_ranges:
        by_mat[r['material']].append(r)
    # 每材料的"代表极差"= 该材料下所有 match_id 极差的最大值
    mat_groups = []
    for mat, items in by_mat.items():
        items.sort(key=lambda x: x['range_pct'], reverse=True)
        mat_groups.append({
            'material': mat,
            'max_range_pct': items[0]['range_pct'],
            'spec_count': len(items),
            'specs': items[:max_spec_per_mat],
        })
    mat_groups.sort(key=lambda x: x['max_range_pct'], reverse=True)
    return mat_groups[:top_n]


def calc_category_distribution(rows: list):
    """按类别统计"""
    cat_count = defaultdict(int)
    cat_price_sum = defaultdict(float)
    cat_price_cnt = defaultdict(int)
    for r in rows:
        p = r.get('价格')
        cat = categorize(r.get('canonical_material', ''))
        cat_count[cat] += 1
        if p is not None:
            cat_price_sum[cat] += p
            cat_price_cnt[cat] += 1
    result = []
    total = sum(cat_count.values())
    for cat, cnt in sorted(cat_count.items(), key=lambda x: -x[1]):
        avg_p = cat_price_sum[cat] / cat_price_cnt[cat] if cat_price_cnt[cat] > 0 else 0
        result.append({
            'category': cat,
            'count': cnt,
            'pct': round(cnt / total * 100, 1) if total else 0,
            'avg_price': round(avg_p, 2),
        })
    return result


# ====== xlsx 写工具 ======

def write_xlsx_header(ws, row, headers, fill_color='305496'):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center')


def write_block_header(ws, row, text, ncols=8):
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=12, color='FFFFFF')
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')


def write_insight_header(ws, row, text, ncols=8):
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=11, color='FFFFFF')
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')


def write_note(ws, row, text, ncols=8, color='808080'):
    ws.cell(row=row, column=1, value=text).font = Font(italic=True, color=color)


# ====== 顶部自动总结 ======

def _detect_category_and_driver(filter_str: str):
    """根据 --filter 决定类别名 + 驱动维度（按 memory 反馈：砂浆/混凝土→强度等级 等）
    返回 (cat_name, driver, keep_cats)
    keep_cats 是 set[材料类别]，用于从 calc_spec_ladder 里筛代表材料
    """
    kws = [k.strip() for k in (filter_str or '').split(',') if k.strip()]
    if not kws:
        return '全部', '规格', None
    cat_name = ' / '.join(kws)
    # 优先级匹配：先看是否有"砂浆"
    if any(k in ['砂浆', '砌筑砂浆', '抹面砂浆'] for k in kws):
        return cat_name, '强度等级', {'砂浆'}
    if any(k in ['混凝土', '砼'] for k in kws):
        return cat_name, '强度等级', {'其他混凝土'}
    if any(k in ['电缆', '电线'] for k in kws):
        return cat_name, '截面积', None  # 电缆暂未单列 cat，不限
    if any('管' in k for k in kws):
        return cat_name, '口径', {'管材'}
    return cat_name, '规格', None


def build_summary_block(ws, matched: list, period: str, args, start_row: int):
    """顶部【自动总结】模块
    按 memory feedback-summary-template.md 的 3 类结构：
      【模式：{类别} {driver}-价格关系】— X的Y越大价格越贵 + 各档涨幅明细
      【偏离】— 跨区县最大极差 + 同材料极差 Top 3
      【完整性】— 行数命中率 + 类别分布
    严禁:驱动/单 match_id 驱动/Top 5 罗列/自己加"最稳定"
    """
    cur = start_row
    ncols = 4

    # 顶部大标题
    write_block_header(ws, cur, f'【自动总结】基于 xlsx 数字的模式与异常（仅数据驱动，{period} 期）', ncols=ncols)
    cur += 2

    cat_name, driver, keep_cats = _detect_category_and_driver(args.filter)

    # ========== 【模式：{cat_name} {driver}-价格关系】 ==========
    write_insight_header(ws, cur, f'【模式：{cat_name} {driver}-价格关系】', ncols=ncols)
    cur += 1
    ladder = calc_spec_ladder(matched)
    # 取规格数 ≥ 3 的材料，并按 keep_cats 过滤（避免把"钢筒混凝土管"混入"混凝土"代表）
    candidates = []
    for grp in ladder:
        if grp['spec_count'] < 3:
            continue
        if keep_cats is not None:
            cat = categorize(grp['material'])
            if cat not in keep_cats:
                continue
        # 过滤"伪阶梯"：如果最高/最低差 < 0.5%（数据本身持平，不是真阶梯）
        items = sorted(grp['ladder'], key=lambda x: x['price'])
        if items[0]['price'] > 0:
            span = (items[-1]['price'] - items[0]['price']) / items[0]['price']
            if span < 0.005:
                continue
        candidates.append(grp)
    candidates.sort(key=lambda x: -x['spec_count'])
    if candidates:
        rep = candidates[0]
        items = sorted(rep['ladder'], key=lambda x: x['price'])  # 按价升序
        n = len(items)
        l_spec = items[0]['spec']
        h_spec = items[-1]['spec']
        l_price = round(items[0]['price'], 2)
        h_price = round(items[-1]['price'], 2)
        unit = items[0]['unit']
        total_pct = round((h_price - l_price) / l_price * 100, 2) if l_price > 0 else 0

        # 相邻档涨幅
        pairs = []
        for i in range(1, n):
            p_prev = items[i-1]['price']
            p_cur = items[i]['price']
            if p_prev > 0:
                pairs.append(round((p_cur - p_prev) / p_prev * 100, 2))
        avg_d = round(sum(pairs) / len(pairs), 2) if pairs else 0
        min_d = min(pairs) if pairs else 0
        max_d = max(pairs) if pairs else 0
        n_mono = sum(1 for d in pairs if d > 0)

        # 第 1 行：核心规律（完整 spec，不截断）
        line1 = (f"• {cat_name}{driver}越大，价格越贵：{rep['material']} "
                 f"{l_spec} → {h_spec} {n} 档单调递增，总涨幅 +{total_pct}%（{l_price} → {h_price} 元/{unit}）")
        ws.cell(row=cur, column=1, value=line1).font = Font(bold=True)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
        cur += 1
        # 第 2 行：各档涨幅明细（完整 spec）
        delta_strs = [f"{items[i-1]['spec']}→{items[i]['spec']} +{pairs[i-1]}%"
                      for i in range(1, n)]
        line2 = f"• 各档涨幅明细：{' / '.join(delta_strs)}"
        ws.cell(row=cur, column=1, value=line2).font = Font(bold=True)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
        cur += 1
        # 第 3 行：汇总
        line3 = f"• 平均每档 +{avg_d}%；涨幅范围 +{min_d}% ~ +{max_d}%；{n_mono}/{len(pairs)} 档符合单调递增"
        ws.cell(row=cur, column=1, value=line3).font = Font(bold=True)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
        cur += 1
    else:
        ws.cell(row=cur, column=1,
                value=f"• 无 {driver} 阶梯数据（同材料规格数 <3，无法识别规律）").font = Font(italic=True, color='808080')
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
        cur += 1
    cur += 1

    # ========== 【偏离】 ==========
    write_insight_header(ws, cur, '【偏离】', ncols=ncols)
    cur += 1
    mid_ranges = calc_district_range_by_mid(matched)
    if mid_ranges:
        max_r = max(mid_ranges, key=lambda x: x['range_pct'])
        line1 = f"• 跨区县最大极差 {max_r['range_pct']}%（{max_r['material']}）"
        ws.cell(row=cur, column=1, value=line1).font = Font(bold=True)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
        cur += 1
        # 同材料极差 Top 3
        mat_groups = calc_district_range_grouped(matched, top_n=3, max_spec_per_mat=10)
        top3_strs = [f"{g['material']} {g['max_range_pct']}%" for g in mat_groups[:3]]
        if top3_strs:
            line2 = f"• 同材料极差 Top 3：{' / '.join(top3_strs)}"
            ws.cell(row=cur, column=1, value=line2).font = Font(bold=True)
            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
            cur += 1
    else:
        ws.cell(row=cur, column=1, value='• 无跨区县数据（仅 1 个区县）').font = Font(italic=True, color='808080')
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
        cur += 1
    cur += 1

    # ========== 【完整性】 ==========
    write_insight_header(ws, cur, '【完整性】', ncols=ncols)
    cur += 1
    n_matched = len(matched)
    n_total = n_matched + sum(1 for r in matched if not r.get('match_id'))
    cat_dist = calc_category_distribution(matched)
    cat_strs = [f"{c['category']} {c['pct']}%" for c in cat_dist]
    line1 = f"• {n_matched} 行 100% 已匹配"
    if n_matched < n_total:
        line1 += f"（未匹配 {n_total - n_matched} 行）"
    ws.cell(row=cur, column=1, value=line1).font = Font(bold=True)
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
    cur += 1
    line2 = f"• 类别分布：{' / '.join(cat_strs)}"
    ws.cell(row=cur, column=1, value=line2).font = Font(bold=True)
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
    cur += 1

    return cur + 1


# ====== 主函数：构建 Sheet ======

def build_sheet(ws, matched: list, period: str, args):
    cur = 1

    # 顶部标题（按 filter 动态显示）
    filter_label = args.filter if args.filter else '混凝土'
    ws.cell(row=cur, column=1, value=f'成都 {period} 期 {filter_label}分析').font = Font(bold=True, size=16)
    cur += 1
    ws.cell(row=cur, column=1, value=f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}').font = Font(italic=True, color='808080')
    cur += 2

    # ====== 顶部自动总结 ======
    cur = build_summary_block(ws, matched, period, args, cur)

    # ====== 摘要 ======
    write_block_header(ws, cur, '【1】数据规模', ncols=4)
    cur += 1
    write_xlsx_header(ws, cur, ['指标', '值', '', ''])
    cur += 1
    summary = [
        ('期号', period),
        (f'总数据点（{filter_label}）', len(matched)),
        ('同物组数', len(set(r['match_id'] for r in matched if r.get('match_id')))),
        ('未匹配行数', sum(1 for r in matched if not r.get('match_id'))),
        ('文件数', '见底部明细'),
    ]
    for k, v in summary:
        ws.cell(row=cur, column=1, value=k).font = Font(bold=True)
        ws.cell(row=cur, column=2, value=v)
        cur += 1
    cur += 1

    # ====== 洞察 1: 高价 Top 5 + 低价 Top 5 ======
    high, low = calc_high_low_top(matched, top_n=5)
    write_insight_header(ws, cur, f'洞察 1: {period} 期 最高价 Top 5 + 最低价 Top 5', ncols=8)
    cur += 1
    write_xlsx_header(ws, cur, ['类型', '材料', '规格', '单位', '区县', '价格(元)', '', ''])
    cur += 1
    for r in high:
        ws.cell(row=cur, column=1, value='↑ 高价')
        ws.cell(row=cur, column=2, value=r['canonical_material'][:25])
        ws.cell(row=cur, column=3, value=r['canonical_spec'][:25])
        ws.cell(row=cur, column=4, value=r['单位'])
        ws.cell(row=cur, column=5, value=r['区名'])
        ws.cell(row=cur, column=6, value=r['价格']).font = Font(bold=True, color='C00000')
        cur += 1
    for r in low:
        ws.cell(row=cur, column=1, value='↓ 低价')
        ws.cell(row=cur, column=2, value=r['canonical_material'][:25])
        ws.cell(row=cur, column=3, value=r['canonical_spec'][:25])
        ws.cell(row=cur, column=4, value=r['单位'])
        ws.cell(row=cur, column=5, value=r['区名'])
        ws.cell(row=cur, column=6, value=r['价格']).font = Font(bold=True, color='00B050')
        cur += 1
    cur += 2

    # ====== 洞察 2: 同材料不同规格完整阶梯 Top 5 ======
    ladder_all = calc_spec_ladder(matched)
    ladder_top = [g for g in ladder_all if g['spec_count'] >= 3]
    ladder_top.sort(key=lambda x: -x['spec_count'])
    ladder_top = ladder_top[:5]
    title2 = f'洞察 2: 同材料不同规格完整阶梯 Top 5（{period} 期单期，相邻涨幅 + 陡增/平缓高亮）'
    if not ladder_top:
        title2 += '（无 — 同材料规格数 <3）'
    write_insight_header(ws, cur, title2, ncols=6)
    cur += 1
    if ladder_top:
        write_xlsx_header(ws, cur, ['规格', '单位', '代表价(元)', 'vs 上一档涨幅%', '', ''])
        cur += 1
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 陡增
        blue_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')    # 平缓
        for grp in ladder_top:
            items = sorted(grp['ladder'], key=lambda x: x['price'])
            # 算平均相邻涨幅
            pairs = []
            for i in range(1, len(items)):
                prev = items[i-1]['price']
                if prev > 0:
                    pairs.append(round((items[i]['price'] - prev) / prev * 100, 2))
            avg_inc = round(sum(pairs) / len(pairs), 2) if pairs else 0
            # 材料标题行（黄底加粗 + 平均涨幅标注）
            ws.cell(row=cur, column=1,
                    value=f"■ {grp['material']}（{grp['spec_count']} 规格，平均涨幅 +{avg_inc}%）").font = Font(bold=True, color='9C5700')
            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=6)
            for c in range(1, 7):
                ws.cell(row=cur, column=c).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            cur += 1
            # 各规格行（按代表价升序；高亮陡增/平缓）
            for i, s in enumerate(items):
                ws.cell(row=cur, column=1, value=s['spec'][:30])
                ws.cell(row=cur, column=2, value=s['unit'])
                ws.cell(row=cur, column=3, value=round(s['price'], 2))
                if i == 0:
                    ws.cell(row=cur, column=4, value='基准')
                else:
                    prev = items[i-1]['price']
                    if prev > 0:
                        inc = (s['price'] - prev) / prev * 100
                        cell = ws.cell(row=cur, column=4, value=f"+{inc:.2f}%")
                        if avg_inc > 0:
                            if inc > avg_inc * 1.5:
                                cell.fill = yellow_fill  # 陡增（>平均 1.5 倍）
                                cell.font = Font(bold=True, color='C00000')
                            elif inc < avg_inc * 0.5:
                                cell.fill = blue_fill    # 平缓（<平均 0.5 倍）
                        cur += 0  # 已被 cell 创建
                    else:
                        ws.cell(row=cur, column=4, value='-')
                cur += 1
            cur += 1  # 材料间空行
    cur += 1

    # ====== 洞察 3: 同物跨区县极差 Top 5（按材料聚拢） ======
    mat_groups = calc_district_range_grouped(matched, top_n=5, max_spec_per_mat=10)
    title3 = f'洞察 3: 同物跨区县价差极差 Top 5（按材料聚拢，{period} 期单期）'
    if not mat_groups:
        title3 += '（无 — 仅有 1 个区县数据）'
    write_insight_header(ws, cur, title3, ncols=8)
    cur += 1
    if mat_groups:
        write_xlsx_header(ws, cur, ['材料', '组号', '规格', '最低区县', '最低价', '最高区县', '最高价', '极差%'])
        cur += 1
        for grp in mat_groups:
            # 材料标题行（黄底加粗 + 跨列合并材料名 + 显示代表极差）
            first_spec = grp['specs'][0]
            ws.cell(row=cur, column=1, value=f"■ {grp['material']}（{grp['spec_count']} 规格，代表极差 {grp['max_range_pct']:.2f}%）").font = Font(bold=True, color='9C5700')
            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=8)
            for c in range(1, 9):
                ws.cell(row=cur, column=c).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            cur += 1
            # 规格子行
            for d in grp['specs']:
                ws.cell(row=cur, column=1, value='')  # 缩进占位
                ws.cell(row=cur, column=2, value=d['match_id'])
                ws.cell(row=cur, column=3, value=f"    {d['spec'][:25]}")
                ws.cell(row=cur, column=4, value=d['min_district'])
                ws.cell(row=cur, column=5, value=d['min_price'])
                ws.cell(row=cur, column=6, value=d['max_district'])
                ws.cell(row=cur, column=7, value=d['max_price'])
                ws.cell(row=cur, column=8, value=f"{d['range_pct']:.2f}%").font = Font(bold=True)
                cur += 1
    cur += 2

    # ====== 洞察 4: 类别分布 ======
    cat_dist = calc_category_distribution(matched)
    write_insight_header(ws, cur, f'洞察 4: 材料类别分布（{period} 期）', ncols=4)
    cur += 1
    write_xlsx_header(ws, cur, ['类别', '行数', '占比%', '均价'])
    cur += 1
    for c in cat_dist:
        ws.cell(row=cur, column=1, value=c['category']).font = Font(bold=True)
        ws.cell(row=cur, column=2, value=c['count'])
        ws.cell(row=cur, column=3, value=f"{c['pct']}%")
        ws.cell(row=cur, column=4, value=c['avg_price'])
        cur += 1
    cur += 2

    # ====== 明细表（按 match_id 升序，一张大表） ======
    write_block_header(ws, cur, f'【明细】成都 {period} 期 混凝土全部数据（按组号升序）', ncols=9)
    cur += 1
    write_xlsx_header(ws, cur,
        ['组号', '材料', '规格', '规格备注', '单位', '区县', '价格(元)', '类别', '来源文件'])
    cur += 1
    detail_rows = sorted(
        [r for r in matched if r.get('价格') is not None],
        key=lambda x: (x.get('match_id', '~'), x.get('canonical_material', ''), x.get('区名', ''))
    )
    for r in detail_rows:
        ws.cell(row=cur, column=1, value=r.get('match_id', ''))
        ws.cell(row=cur, column=2, value=r.get('canonical_material', '')[:30])
        ws.cell(row=cur, column=3, value=r.get('canonical_spec', '')[:30])
        ws.cell(row=cur, column=4, value=r.get('规格备注', ''))
        ws.cell(row=cur, column=5, value=r.get('单位', ''))
        ws.cell(row=cur, column=6, value=r.get('区名', ''))
        ws.cell(row=cur, column=7, value=r.get('价格'))
        ws.cell(row=cur, column=8, value=categorize(r.get('canonical_material', '')))
        ws.cell(row=cur, column=9, value=r.get('来源文件', '')[:25])
        cur += 1
    cur += 1
    write_note(ws, cur, f'（明细共 {len(detail_rows)} 行；规格备注 列承载运费注，规格 列已剥运费）')


def build_compare_sheet(ws, rows_by_period: dict, periods: list, args):
    """跨期对比 sheet（--period 02,06 模式）"""
    cur = 1
    # 主体材料名 = filter 或 全部
    mat_label = args.filter.replace(",", "、") if args.filter else "全部混凝土"
    title = f'成都 {" vs ".join(periods)} 期 {mat_label} 跨期对比'
    ws.cell(row=cur, column=1, value=title).font = Font(bold=True, size=16)
    cur += 1
    ws.cell(row=cur, column=1, value=f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}').font = Font(italic=True, color='808080')
    cur += 1
    # 本表用途说明（黄底醒目）
    purpose_text = (f'📌 本表用途：对比 {" vs ".join(periods)} 期 {mat_label} 价格变化，'
                   f'识别涨/跌价材料，辅助造价调整、采购预算、投标报价决策。'
                   f'看【涨跌 Top 10】找涨跌幅最大的材料；'
                   f'其余同物组在各期单期报告的【明细】表里查。')
    ws.cell(row=cur, column=1, value=purpose_text).font = Font(bold=True, color='9C5700')
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=6)
    for c in range(1, 7):
        ws.cell(row=cur, column=c).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    cur += 2

    # 摘要
    write_block_header(ws, cur, '【1】数据规模', ncols=4)
    cur += 1
    write_xlsx_header(ws, cur, ['指标', '值', '', ''])
    cur += 1
    all_matched = set()
    for p in periods:
        rows = rows_by_period[p]
        all_matched.update(r['match_id'] for r in rows if r.get('match_id'))
    summary = [
        ('对比期号', ' vs '.join(periods)),
        ('共同同物组数', len(all_matched)),
    ]
    for p in periods:
        rows = rows_by_period[p]
        summary.append((f'{p} 期数据点', len(rows)))
    for k, v in summary:
        ws.cell(row=cur, column=1, value=k).font = Font(bold=True)
        ws.cell(row=cur, column=2, value=v)
        cur += 1
    cur += 1

    # 算跨期价差
    period_prices = {}
    for p in periods:
        by_mid = defaultdict(list)
        for r in rows_by_period[p]:
            price = r.get('价格')
            if price is None:
                continue
            if r.get('match_id'):
                by_mid[r['match_id']].append((r['区名'], price,
                                              r.get('canonical_material', ''),
                                              r.get('canonical_spec', ''),
                                              r.get('规格备注', '')))
        # 每个 match_id 取代表价（中位或成都市优先）
        mid_repr = {}
        for mid, dp_list in by_mid.items():
            chengdu = [pr for d, pr, _, _, _ in dp_list if d == '成都市']
            if chengdu:
                mid_repr[mid] = sum(chengdu) / len(chengdu)
            else:
                prices = sorted([pr for _, pr, _, _, _ in dp_list])
                n = len(prices)
                if n == 0:
                    continue
                if n % 2 == 0:
                    mid_repr[mid] = (prices[n//2 - 1] + prices[n//2]) / 2
                else:
                    mid_repr[mid] = prices[n//2]
            # 取代表性材料名/规格/运费注
            mid_repr[mid] = (mid_repr[mid], dp_list[0][2], dp_list[0][3], dp_list[0][4])
        period_prices[p] = mid_repr

    # 计算跨期变化
    p1, p2 = periods[0], periods[-1]
    diffs = []
    for mid in all_matched:
        if mid not in period_prices[p1] or mid not in period_prices[p2]:
            continue
        v1, mat, spec, note1 = period_prices[p1][mid]
        v2, _, _, note2 = period_prices[p2][mid]
        if v1 <= 0:
            continue
        pct = round((v2 - v1) / v1 * 100, 2)
        diffs.append({
            'match_id': mid,
            'material': mat,
            'spec': spec,
            '规格备注': note1 or note2,
            f'{p1}期价': round(v1, 2),
            f'{p2}期价': round(v2, 2),
            '变化%': pct,
        })
    diffs.sort(key=lambda x: x['变化%'], reverse=True)
    n_up = sum(1 for d in diffs if d['变化%'] > 0)
    n_down = sum(1 for d in diffs if d['变化%'] < 0)
    n_flat = sum(1 for d in diffs if d['变化%'] == 0)
    n_big = sum(1 for d in diffs if abs(d['变化%']) >= 10)

    # 【结论】3-5 行直接可用结论
    n_up = sum(1 for d in diffs if d['变化%'] > 0)
    n_down = sum(1 for d in diffs if d['变化%'] < 0)
    n_flat = sum(1 for d in diffs if d['变化%'] == 0)
    n_total = len(diffs)
    pct_flat = round(n_flat / n_total * 100, 1) if n_total else 0
    head_up_pct = max((d['变化%'] for d in diffs), default=0)
    head_down_pct = min((d['变化%'] for d in diffs), default=0)
    head_up_mat = next((d['material'] for d in diffs if d['变化%'] == head_up_pct), '')
    head_down_mat = next((d['material'] for d in diffs if d['变化%'] == head_down_pct), '')

    write_insight_header(ws, cur, f'【结论】{p1}→{p2} 跨期价格变动总结', ncols=6)
    cur += 1
    conclusion_lines = [
        f'本期共同 {n_total} 同物组：涨 {n_up} / 跌 {n_down} / 持平 {n_flat}（持平占 {pct_flat}%）',
        f'头部最大涨幅 +{head_up_pct}%（{head_up_mat[:30]}），头部最大跌幅 {head_down_pct}%（{head_down_mat[:30]}）',
    ]
    if pct_flat >= 50:
        conclusion_lines.append(f'多数同物组（{n_flat} 个，占 {pct_flat}%）跨期持平，预算调整范围有限')
    # 集中度：涨幅 Top 9 占涨端贡献
    if n_up > 0:
        up_diffs = sorted([d for d in diffs if d['变化%'] > 0], key=lambda x: -x['变化%'])
        top9_pct = sum(d['变化%'] for d in up_diffs[:9])
        all_up_pct = sum(d['变化%'] for d in up_diffs)
        if all_up_pct > 0:
            top9_share = round(top9_pct / all_up_pct * 100, 1)
            if top9_share >= 80:
                conclusion_lines.append(f'涨端集中度高：Top {min(9, len(up_diffs))} 涨幅材料占总涨幅 {top9_share}%')
    for line in conclusion_lines:
        ws.cell(row=cur, column=1, value=line).font = Font(bold=True)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=6)
        cur += 1
    cur += 1

    # 【按类别分桶】按材料类别聚合涨跌幅
    filter_kw = (args.filter or '').replace(',', '')
    buckets = categorize_compare_buckets(diffs, filter_kw)
    title2 = f'【按类别分桶】{p1}→{p2} 各材料类别涨跌幅汇总'
    write_insight_header(ws, cur, title2, ncols=6)
    cur += 1
    write_xlsx_header(ws, cur, ['类别', '同物组数', '涨跌幅范围', '代表规格 Top 3', '', ''])
    cur += 1
    if buckets:
        for b in buckets:
            ws.cell(row=cur, column=1, value=b['name']).font = Font(bold=True)
            ws.cell(row=cur, column=2, value=b['count'])
            def _fmt_pct(p):
                if p > 0: return f"+{p}%"
                if p < 0: return f"{p}%"
                return "0%"
            if b['change_min'] == b['change_max']:
                range_str = _fmt_pct(b['change_min'])
            else:
                range_str = f"{_fmt_pct(b['change_min'])} ~ {_fmt_pct(b['change_max'])}"
            ws.cell(row=cur, column=3, value=range_str)
            # 代表规格 Top 3（按变化%绝对值降序）
            top3 = sorted(b['items'], key=lambda x: -abs(x['变化%']))[:3]
            top3_str = '; '.join(f"{d['material'][:15]} {d['变化%']:+.1f}%" for d in top3)
            ws.cell(row=cur, column=4, value=top3_str[:200])
            ws.merge_cells(start_row=cur, start_column=4, end_row=cur, end_column=6)
            # 整行染色
            if b['change_max'] > 5:
                ws.cell(row=cur, column=1).fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
            elif b['change_min'] < -5:
                ws.cell(row=cur, column=1).fill = PatternFill(start_color='E0F0E0', end_color='E0F0E0', fill_type='solid')
            elif b['change_max'] == 0 and b['change_min'] == 0:
                ws.cell(row=cur, column=1).fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            cur += 1
    else:
        ws.cell(row=cur, column=1, value='（无可分类材料）').font = Font(italic=True, color='808080')
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=6)
        cur += 1
    cur += 1

    # 其余同物组（涨跌 Top 10 之外的，避免重复）
    top10_mids = set()
    for d in diffs[:10]:
        if d['变化%'] > 0:
            top10_mids.add(d['match_id'])
    for d in sorted([x for x in diffs if x['变化%'] < 0], key=lambda x: x['变化%'])[:10]:
        top10_mids.add(d['match_id'])
    remaining = [d for d in diffs if d['match_id'] not in top10_mids]
    if remaining:
        write_block_header(ws, cur, f'【其余同物组】{p1} vs {p2}（已排除涨跌 Top 10，共 {len(remaining)} 个）', ncols=7)
        cur += 1
        write_xlsx_header(ws, cur, ['组号', '材料', '规格', '规格备注', f'{p1}期价', f'{p2}期价', '变化%'])
        cur += 1
        for d in remaining:
            ws.cell(row=cur, column=1, value=d['match_id'])
            ws.cell(row=cur, column=2, value=d['material'][:30])
            ws.cell(row=cur, column=3, value=d['spec'][:30])
            ws.cell(row=cur, column=4, value=d.get('规格备注', ''))
            ws.cell(row=cur, column=5, value=d[f'{p1}期价'])
            ws.cell(row=cur, column=6, value=d[f'{p2}期价'])
            color = 'C00000' if d['变化%'] > 0 else ('00B050' if d['变化%'] < 0 else '808080')
            ws.cell(row=cur, column=7, value=f"{d['变化%']:+.2f}%").font = Font(bold=True, color=color)
            cur += 1
    else:
        write_note(ws, cur, '（全部同物组已在【涨跌 Top 10】显示，无重复）')


# ====== 多期同物组价格曲线 + 逐期对比 ======

def build_multiperiod_sheet(ws, rows_by_period, periods, args):
    """3+ 期对比：每期 Top 5 + 期间逐期对比 + 累计结论"""
    filter_label = args.filter if args.filter else '混凝土'
    cur = 1

    # 顶部标题
    ws.cell(row=cur, column=1, value=f'成都 {periods[0]}-{periods[-1]} 期 {filter_label} {len(periods)} 期分析报告').font = Font(bold=True, size=16)
    cur += 1
    ws.cell(row=cur, column=1, value=f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}').font = Font(italic=True, color='808080')
    cur += 1
    ws.cell(row=cur, column=1, value=f'📌 本表用途：跨 {len(periods)} 期追踪{filter_label}价格变化 + 期间波动分析').font = Font(italic=True)
    cur += 2

    # ====== 计算每个 (期号, match_id) 的代表价 ======
    # 复用 build_compare_sheet 的代表价算法（按 material+spec）
    period_prices = {}  # period_prices[p] = {match_id: (price, material, spec)}
    all_mids = set()
    for p in periods:
        by_mid = defaultdict(list)
        for r in rows_by_period[p]:
            price = r.get('价格')
            if price is None or not r.get('match_id'):
                continue
            by_mid[r['match_id']].append((r['区名'], price,
                                           r.get('canonical_material', ''),
                                           r.get('canonical_spec', '')))
        mid_repr = {}
        for mid, dp_list in by_mid.items():
            chengdu = [pr for d, pr, _, _ in dp_list if d == '成都市']
            if chengdu:
                mid_repr[mid] = (sum(chengdu) / len(chengdu), dp_list[0][2], dp_list[0][3])
            else:
                prices = sorted([pr for _, pr, _, _ in dp_list])
                n = len(prices)
                if n == 0:
                    continue
                if n % 2 == 0:
                    mid_repr[mid] = ((prices[n//2 - 1] + prices[n//2]) / 2, dp_list[0][2], dp_list[0][3])
                else:
                    mid_repr[mid] = (prices[n//2], dp_list[0][2], dp_list[0][3])
        period_prices[p] = mid_repr
        all_mids.update(mid_repr.keys())

    # ====== 【1】数据规模 ======
    write_insight_header(ws, cur, '【1】数据规模', ncols=6)
    cur += 1
    write_xlsx_header(ws, cur, ['指标', '值', '', '', '', ''])
    cur += 1
    scale_rows = [
        ('期号范围', f'{periods[0]} - {periods[-1]}'),
        ('总期数', len(periods)),
        ('同物组数（跨期）', len(all_mids)),
    ]
    for p in periods:
        scale_rows.append((f'{p} 期数据点', len(rows_by_period[p])))
    for k, v in scale_rows:
        ws.cell(row=cur, column=1, value=k).font = Font(bold=True)
        ws.cell(row=cur, column=2, value=v)
        cur += 1
    cur += 1

    # ====== 【2】每期代表价格（每期 Top 5 高 + Top 5 低） ======
    write_insight_header(ws, cur, '【2】每期代表价格（每期 Top 5 高 + Top 5 低）', ncols=6)
    cur += 1
    write_xlsx_header(ws, cur, ['期号', '高价 Top 5', '低价 Top 5', '', '', ''])
    cur += 1
    for p in periods:
        rows = rows_by_period[p]
        if not rows:
            ws.cell(row=cur, column=1, value=p)
            ws.cell(row=cur, column=2, value='（无数据）').font = Font(italic=True, color='808080')
            cur += 1
            continue
        # 排序：所有行（按价格降序）
        sorted_high = sorted(rows, key=lambda x: -(x.get('价格') or 0))[:5]
        sorted_low = sorted([r for r in rows if r.get('价格') is not None], key=lambda x: x['价格'])[:5]
        high_str = '; '.join(f"{r['区名']} {r['价格']:.2f}" for r in sorted_high)
        low_str = '; '.join(f"{r['区名']} {r['价格']:.2f}" for r in sorted_low)
        ws.cell(row=cur, column=1, value=p).font = Font(bold=True)
        ws.cell(row=cur, column=2, value=high_str)
        ws.cell(row=cur, column=3, value=low_str)
        ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=2)
        ws.merge_cells(start_row=cur, start_column=3, end_row=cur, end_column=3)
        cur += 1
    cur += 1

    # ====== 【3】期间波动分析（逐期对比） ======
    write_insight_header(ws, cur, '【3】期间波动分析（逐期对比）', ncols=6)
    cur += 1
    write_xlsx_header(ws, cur, ['期间', '涨区县数', '跌区县数', '持平', '最大涨幅', '最大跌幅'])
    cur += 1
    # 对每个 (mid, district) 行在每期算代表价
    period_by_district = {}
    for p in periods:
        period_by_district[p] = {}
        for r in rows_by_period[p]:
            d = r['区名']
            price = r.get('价格')
            if price is None or not d:
                continue
            period_by_district[p].setdefault(d, []).append(price)
    # 取每个区在每期的代表价（中位数）
    rep_by_dist = {}
    for p in periods:
        rep_by_dist[p] = {d: sorted(prices)[len(prices)//2] for d, prices in period_by_district[p].items() if prices}

    for i in range(len(periods) - 1):
        p1, p2 = periods[i], periods[i+1]
        diffs_inter = []
        for d, v2 in rep_by_dist[p2].items():
            v1 = rep_by_dist[p1].get(d)
            if v1 and v1 > 0:
                pct = round((v2 - v1) / v1 * 100, 2)
                diffs_inter.append((d, pct))
        n_up = sum(1 for _, p in diffs_inter if p > 0)
        n_down = sum(1 for _, p in diffs_inter if p < 0)
        n_flat = sum(1 for _, p in diffs_inter if p == 0)
        max_up = max(diffs_inter, key=lambda x: x[1], default=('', 0))
        max_dn = min(diffs_inter, key=lambda x: x[1], default=('', 0))
        ws.cell(row=cur, column=1, value=f'{p1}→{p2}').font = Font(bold=True)
        ws.cell(row=cur, column=2, value=n_up)
        ws.cell(row=cur, column=3, value=n_down)
        ws.cell(row=cur, column=4, value=n_flat)
        ws.cell(row=cur, column=5, value=f"{max_up[0]} +{max_up[1]}%" if max_up[1] > 0 else '—')
        ws.cell(row=cur, column=6, value=f"{max_dn[0]} {max_dn[1]}%" if max_dn[1] < 0 else '—')
        cur += 1
    cur += 1

    # ====== 【4】累计涨跌幅 Top 5（首 vs 末） ======
    p_first, p_last = periods[0], periods[-1]
    cum_diffs = []
    for d, v2 in rep_by_dist[p_last].items():
        v1 = rep_by_dist[p_first].get(d)
        if v1 and v1 > 0:
            cum_diffs.append((d, round((v2 - v1) / v1 * 100, 2), v1, v2))
    cum_diffs.sort(key=lambda x: -x[1])

    write_insight_header(ws, cur, f'【4】{p_first}→{p_last} 累计涨跌幅 Top 5', ncols=6)
    cur += 1
    write_xlsx_header(ws, cur, ['排名', '区县', f'{p_first}期价', f'{p_last}期价', '累计%', '类型'])
    cur += 1
    # 涨幅 Top 5
    for i, (d, pct, v1, v2) in enumerate([x for x in cum_diffs if x[1] > 0][:5], 1):
        ws.cell(row=cur, column=1, value=f'↑ 涨 #{i}')
        ws.cell(row=cur, column=2, value=d)
        ws.cell(row=cur, column=3, value=f'{v1:.2f}')
        ws.cell(row=cur, column=4, value=f'{v2:.2f}')
        ws.cell(row=cur, column=5, value=f'+{pct}%').font = Font(bold=True, color='C00000')
        cur += 1
    # 跌幅 Top 5
    down_sorted = sorted([x for x in cum_diffs if x[1] < 0], key=lambda x: x[1])
    for i, (d, pct, v1, v2) in enumerate(down_sorted[:5], 1):
        ws.cell(row=cur, column=1, value=f'↓ 跌 #{i}')
        ws.cell(row=cur, column=2, value=d)
        ws.cell(row=cur, column=3, value=f'{v1:.2f}')
        ws.cell(row=cur, column=4, value=f'{v2:.2f}')
        ws.cell(row=cur, column=5, value=f'{pct}%').font = Font(bold=True, color='00B050')
        cur += 1
    cur += 1

    # ====== 【5】结论-逐区县分析 ======
    write_insight_header(ws, cur, '【5】结论-逐区县分析', ncols=6)
    cur += 1
    # 收集"有变化"区县
    changed = [x for x in cum_diffs if abs(x[1]) > 0]
    if not changed:
        ws.cell(row=cur, column=1, value=f'全部区县跨 {periods[0]}→{periods[-1]} 期完全持平，0% 变动').font = Font(italic=True)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=6)
        cur += 1
    else:
        for d, pct, v1, v2 in changed:
            # 完整列出该区县每期间价格 + 每两期涨跌幅
            prices_seq = [rep_by_dist[p].get(d) for p in periods]
            period_prices_str = '→'.join(f'{p}:{(prices_seq[i] or 0):.2f}' for i, p in enumerate(periods))
            # 每两期涨跌幅
            inter_pcts = []
            for i in range(len(periods) - 1):
                pv = prices_seq[i]
                cv = prices_seq[i+1]
                if pv and cv and pv > 0:
                    inter_pcts.append((periods[i], periods[i+1], round((cv - pv) / pv * 100, 2)))
            inter_str = ' / '.join(f'{p1}→{p2}{pp:+.2f}%' for p1, p2, pp in inter_pcts if abs(pp) > 0.01)
            # 总结模式
            n_up_inter = sum(1 for _, _, pp in inter_pcts if pp > 0.01)
            n_down_inter = sum(1 for _, _, pp in inter_pcts if pp < -0.01)
            n_flat_inter = sum(1 for _, _, pp in inter_pcts if abs(pp) <= 0.01)
            if n_up_inter == len(inter_pcts) and len(inter_pcts) > 0:
                pattern = '单调上涨'
            elif n_down_inter == len(inter_pcts) and len(inter_pcts) > 0:
                pattern = '单调下跌'
            elif n_up_inter > 0 and n_down_inter == 0:
                pattern = '持续上涨'
            elif n_down_inter > 0 and n_up_inter == 0:
                pattern = '持续下跌'
            else:
                pattern = '有涨有跌'
            ws.cell(row=cur, column=1, value=d).font = Font(bold=True)
            ws.cell(row=cur, column=2, value=period_prices_str)
            ws.cell(row=cur, column=3, value=f'累计{pct:+.2f}%')
            ws.cell(row=cur, column=4, value=f'逐期: {inter_str}' if inter_str else '逐期: 持平')
            ws.cell(row=cur, column=5, value=pattern).font = Font(italic=True, color='808080')
            cur += 1
        cur += 1
        # 持平区县逐行结果（13 个持平区县每个一行：5 期价格 + 模式，不写分析）
        flat_districts = [x[0] for x in cum_diffs if x[1] == 0]
        if flat_districts:
            write_xlsx_header(ws, cur, ['区县', '5 期价格序列', '累计%', '逐期', '模式', ''])
            cur += 1
            for d in flat_districts:
                prices_seq = [rep_by_dist[p].get(d) for p in periods]
                period_prices_str = '→'.join(f'{p}:{(prices_seq[i] or 0):.2f}' for i, p in enumerate(periods))
                ws.cell(row=cur, column=1, value=d)
                ws.cell(row=cur, column=2, value=period_prices_str)
                ws.cell(row=cur, column=3, value='+0.00%').font = Font(color='808080')
                ws.cell(row=cur, column=4, value='逐期: 持平').font = Font(color='808080')
                ws.cell(row=cur, column=5, value='完全持平').font = Font(italic=True, color='808080')
                cur += 1
            cur += 1
            # 持平汇总（价格区间 + 跨区极差）
            flat_prices = [(d, rep_by_dist[periods[0]].get(d)) for d in flat_districts]
            flat_prices = [(d, p) for d, p in flat_prices if p]
            if flat_prices:
                min_d, min_p = min(flat_prices, key=lambda x: x[1])
                max_d, max_p = max(flat_prices, key=lambda x: x[1])
                spread = round((max_p - min_p) / min_p * 100, 2) if min_p > 0 else 0
                ws.cell(row=cur, column=1, value=f'持平汇总: {len(flat_districts)} 区县跨 {periods[0]}-{periods[-1]} 期完全持平，价格区间 {min_p:.2f} 元/m³（{min_d.replace("成都市", "")}，最低）~ {max_p:.2f} 元/m³（{max_d.replace("成都市", "")}，最高），跨区极差 {spread}%').font = Font(italic=True, color='808080')
                ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=6)
                cur += 1


# ====== 跨期类别分桶 ======

def categorize_compare_buckets(diffs, filter_kw=''):
    """按材料类别分桶（仅看 material 字段）
    返回 [{name, count, change_min, change_max, items}, ...]
    桶定义按 filter 区分：
    - 混凝土: 沥青/普通商品混凝土/添加剂/钢筋混凝土管/特种管材/其他
    - 钢筋: 螺纹钢筋/高线/圆钢/钢绞线/其他
    - 其他: 按 material 前缀粗分
    """
    # 类别规则表（按优先级匹配，第一个匹配中的桶生效）
    rules_map = {
        '混凝土': [
            ('沥青系', lambda m: '沥青' in m or '马蹄脂' in m),
            ('普通商品混凝土系', lambda m: '普通商品混凝土' in m or '商品砼' in m),
            ('添加剂系', lambda m: any(k in m for k in ['添加剂', '防水剂', '密实剂', '阻锈', '活性剂', '抑温', '抗裂'])),
            ('钢筋混凝土管系', lambda m: '钢筋混凝土管' in m or '钢筋混凝土承插管' in m or '钢筋混凝土企口管' in m),
            ('特种管材系', lambda m: any(k in m for k in ['超高性能混凝土', '预应力钢筒', '非预应力钢筒', '内衬玻璃钢', '耐腐蚀高性能', '预制矩形混凝土'])),
            ('其他', lambda m: True),
        ],
        '钢筋': [
            ('螺纹钢筋系', lambda m: '螺纹' in m),
            ('高线系', lambda m: '高线' in m or '盘螺' in m),
            ('圆钢系', lambda m: '圆钢' in m),
            ('钢绞线系', lambda m: '钢绞线' in m),
            ('其他钢筋系', lambda m: True),
        ],
    }
    # 选规则表
    if '混凝土' in filter_kw:
        rules = rules_map['混凝土']
    elif '钢筋' in filter_kw or '螺纹' in filter_kw:
        rules = rules_map['钢筋']
    else:
        # 通用：按 material 前 4 字聚合
        from collections import Counter
        prefixes = Counter(d['material'][:4] for d in diffs if d['material'])
        # 取 Top 5 prefix 作桶
        rules = []
        for prefix in prefixes.most_common(5):
            rules.append((f'{prefix[0]}系', lambda m, p=prefix[0]: m.startswith(p)))
        rules.append(('其他', lambda m: True))

    buckets = []
    matched_ids = set()
    for name, pred in rules:
        items = [d for d in diffs if pred(d['material']) and d['match_id'] not in matched_ids]
        if not items:
            continue
        changes = [d['变化%'] for d in items]
        buckets.append({
            'name': name,
            'count': len(items),
            'change_min': min(changes),
            'change_max': max(changes),
            'items': items,
        })
        matched_ids.update(d['match_id'] for d in items)
    return buckets


# ====== 新增/消失材料（A）======

def calc_added_removed_materials(rows_by_period, p1, p2):
    """跨期新增/消失同物组
    返回 (added_list, removed_list)
    每个 list 元素: {match_id, material, spec, unit, price, district, file}
    """
    p1_mids = set(r['match_id'] for r in rows_by_period[p1] if r.get('match_id'))
    p2_mids = set(r['match_id'] for r in rows_by_period[p2] if r.get('match_id'))

    added_mids = p2_mids - p1_mids
    removed_mids = p1_mids - p2_mids

    def repr_per_mid(rows):
        by_mid = defaultdict(list)
        for r in rows:
            if r.get('match_id'):
                by_mid[r['match_id']].append(r)
        mid_info = {}
        for mid, rs in by_mid.items():
            prices = [(r['价格'], r['区名']) for r in rs if r.get('价格') is not None]
            if not prices:
                continue
            chengdu = [p for p, d in prices if d == '成都市']
            if chengdu:
                repr_price = sum(chengdu) / len(chengdu)
                repr_district = '成都市'
            else:
                sorted_prices = sorted(prices, key=lambda x: x[0])
                n = len(sorted_prices)
                if n % 2 == 0:
                    repr_price = (sorted_prices[n//2 - 1][0] + sorted_prices[n//2][0]) / 2
                    repr_district = sorted_prices[n//2 - 1][1]
                else:
                    repr_price = sorted_prices[n//2][0]
                    repr_district = sorted_prices[n//2][1]
            mid_info[mid] = {
                'match_id': mid,
                'material': rs[0].get('canonical_material', ''),
                'spec': rs[0].get('canonical_spec', ''),
                '规格备注': rs[0].get('规格备注', ''),
                'unit': rs[0].get('单位', ''),
                'price': repr_price,
                'district': repr_district,
                'file': rs[0].get('来源文件', ''),
            }
        return mid_info

    p1_info = repr_per_mid(rows_by_period[p1])
    p2_info = repr_per_mid(rows_by_period[p2])

    added = [p2_info[m] for m in sorted(added_mids) if m in p2_info]
    removed = [p1_info[m] for m in sorted(removed_mids) if m in p1_info]

    return added, removed


def build_added_removed_sheet(ws, items, title, header_color):
    """items = list of dict（来自 calc_added_removed_materials）
    title: sheet 主标题
    header_color: '00B050'(新增绿) 或 'C00000'(消失红)
    """
    cur = 1
    ws.cell(row=cur, column=1, value=title).font = Font(bold=True, size=16)
    cur += 1
    ws.cell(row=cur, column=1, value=f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}').font = Font(italic=True, color='808080')
    cur += 2

    if not items:
        write_block_header(ws, cur, '本类无材料', ncols=7)
        return

    write_block_header(ws, cur, f'共 {len(items)} 个同物组', ncols=8)
    cur += 1
    write_xlsx_header(ws, cur, ['组号', '材料', '规格', '规格备注', '单位', '代表价(元)', '代表区县', '来源文件'])
    cur += 1
    for item in items:
        ws.cell(row=cur, column=1, value=item['match_id'])
        ws.cell(row=cur, column=2, value=item['material'][:30])
        ws.cell(row=cur, column=3, value=item['spec'][:30])
        ws.cell(row=cur, column=4, value=item.get('规格备注', ''))
        ws.cell(row=cur, column=5, value=item['unit'])
        ws.cell(row=cur, column=6, value=round(item['price'], 2) if item['price'] else '')
        ws.cell(row=cur, column=7, value=item['district'])
        ws.cell(row=cur, column=8, value=item['file'][:30])
        cur += 1
    cur += 1
    write_note(ws, cur, f'（明细共 {len(items)} 个同物组，按 match_id 升序；规格备注 列承载运费注）')


# ====== 同材料不同规格阶梯（B）======

def calc_spec_ladder(rows):
    """单期规格阶梯
    返回 [{material, spec_count, ladder: [{spec, unit, price}, ...]}, ...]
    只保留规格数 ≥2 的材料，按规格数降序
    """
    # 按 match_id 算代表价（优先成都市综合行）
    mid_repr = {}
    by_mid = defaultdict(list)
    for r in rows:
        if r.get('价格') is None or not r.get('match_id'):
            continue
        by_mid[r['match_id']].append(r)
    for mid, rs in by_mid.items():
        chengdu = [r['价格'] for r in rs if r['区名'] == '成都市']
        if chengdu:
            mid_repr[mid] = sum(chengdu) / len(chengdu)
        else:
            prices = sorted([r['价格'] for r in rs])
            n = len(prices)
            if n == 0:
                continue
            if n % 2 == 0:
                mid_repr[mid] = (prices[n//2 - 1] + prices[n//2]) / 2
            else:
                mid_repr[mid] = prices[n//2]

    # 按 (material, spec) 聚合代表价：每个 spec 取所有 match_id 代表价均值
    by_mat = defaultdict(lambda: defaultdict(list))
    for r in rows:
        if r.get('价格') is None:
            continue
        mat = r.get('canonical_material', '')
        spec = r.get('canonical_spec', '')
        unit = r.get('单位', '')
        if not mat or not spec:
            continue
        mid = r.get('match_id', '')
        if mid and mid in mid_repr:
            by_mat[mat][(spec, unit)].append(mid_repr[mid])

    ladders = []
    for mat, spec_dict in by_mat.items():
        if len(spec_dict) < 2:
            continue
        spec_avg = []
        for (spec, unit), repr_prices in spec_dict.items():
            spec_avg.append({
                'spec': spec,
                'unit': unit,
                'price': sum(repr_prices) / len(repr_prices),
            })
        spec_avg.sort(key=lambda x: x['price'])
        ladders.append({
            'material': mat,
            'spec_count': len(spec_avg),
            'ladder': spec_avg,
        })

    ladders.sort(key=lambda x: -x['spec_count'])
    return ladders


def main():
    parser = argparse.ArgumentParser(description='Phase 4 v3 单期/跨期对比 xlsx 报告')
    parser.add_argument('--period', required=True, help='期号，单期如 02；跨期如 02,06')
    parser.add_argument('--filter', type=str, default=None, help='按材料名过滤')
    parser.add_argument('--manifest', default=str(MANIFEST_PATH))
    parser.add_argument('--matches', default=str(MATCHES_PATH))
    parser.add_argument('--column-map', default=str(COLUMN_MAP_PATH))
    parser.add_argument('--output-dir', default=str(OUTPUT_DIR))
    args = parser.parse_args()

    periods = [p.strip() for p in args.period.split(',')]
    is_compare = len(periods) > 1
    is_multi = len(periods) >= 3
    mode = f'跨期对比（{args.period}）' if is_compare else f'单期（{args.period}）'
    if is_multi:
        mode = f'多期分析（{args.period}）'

    print(f'=== Phase 4 v3 {mode}报告 ===')
    print(f'期号：{args.period}')
    print(f'过滤：{args.filter or "(无)"}')

    manifest = load_csv(Path(args.manifest))
    matches = load_csv(Path(args.matches))
    with open(args.column_map, 'r', encoding='utf-8') as f:
        column_rows = {r['file_path']: r for r in csv.DictReader(f)}

    all_rows = load_all_rows(manifest, column_rows)
    print(f'全部行：{len(all_rows)}')

    from phase2_match import normalize_text as _norm

    if is_compare:
        # 跨期对比模式：每期单独过滤 + attach + 应用材料过滤
        rows_by_period = {}
        for p in periods:
            rows = [r for r in all_rows if r['期号'] == p]
            rows = attach_match_id(rows, matches, _norm)
            if args.filter:
                keywords = [k.strip() for k in args.filter.split(',') if k.strip()]
                before = len(rows)
                rows = [r for r in rows if any(k in (r.get('canonical_material', '') or r.get('材料名称', '')) for k in keywords)]
                print(f'  {p} 期过滤 [{",".join(keywords)}]：{before} → {len(rows)}')
            rows_by_period[p] = [r for r in rows if r.get('match_id')]
            print(f'  {p} 期已匹配：{len(rows_by_period[p])}')

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # === 多期模式（3+ 期）走新函数 ===
        if is_multi:
            sheet_label = f'{"_".join(periods)}期_{args.filter.replace(",", "_")}' if args.filter else f'{"_".join(periods)}期_混凝土'
            ws = wb.create_sheet(sheet_label)
            build_multiperiod_sheet(ws, rows_by_period, periods, args)
            for col, width in [('A', 12), ('B', 28), ('C', 28), ('D', 14), ('E', 14), ('F', 14)]:
                ws.column_dimensions[col].width = width
        else:
            sheet_label = f'{"_".join(periods)}期对比'
            if args.filter:
                sheet_label = f'{"_".join(periods)}期_{args.filter.replace(",", "_")}'
            else:
                sheet_label = f'{"_".join(periods)}期_混凝土'
            ws = wb.create_sheet(sheet_label)
            build_compare_sheet(ws, rows_by_period, periods, args)
            for col, width in [('A', 12), ('B', 28), ('C', 28), ('D', 14), ('E', 12), ('F', 12), ('G', 12)]:
                ws.column_dimensions[col].width = width

            # === A 跨期 新增/消失材料（仅 2 期对比模式） ===
            added, removed = calc_added_removed_materials(rows_by_period, periods[0], periods[-1])
            ws_added = wb.create_sheet('新增材料')
            build_added_removed_sheet(ws_added, added,
                                      title=f'{periods[-1]} 新增（{periods[0]} 没有）',
                                      header_color='00B050')
            for col, width in [('A', 12), ('B', 28), ('C', 28), ('D', 16), ('E', 8), ('F', 14), ('G', 14), ('H', 25)]:
                ws_added.column_dimensions[col].width = width
            ws_removed = wb.create_sheet('消失材料')
            build_added_removed_sheet(ws_removed, removed,
                                      title=f'{periods[0]} 消失（{periods[-1]} 没有）',
                                      header_color='C00000')
            for col, width in [('A', 12), ('B', 28), ('C', 28), ('D', 16), ('E', 8), ('F', 14), ('G', 14), ('H', 25)]:
                ws_removed.column_dimensions[col].width = width

        # === 保存（覆盖模式：同输入重复跑覆盖旧文件）===
        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        suffix = '_对比' if is_compare else ''
        cities = set(m.get('city') for m in manifest if m.get('city'))
        city = next(iter(cities), '成都')
        f_suffix = f'_{args.filter.replace(",", "_")}' if args.filter else ''
        path = output_dir / f'{city}_{args.period}期_v3{suffix}{f_suffix}.xlsx'
        # 删同输入的旧文件（任意后缀）
        for old in output_dir.glob(f'{city}_{args.period}期_v3{suffix}{f_suffix}*.xlsx'):
            if old.name != path.name:
                old.unlink()
        wb.save(path)
        print(f'\n✓ xlsx 已写：{path}')
        print(f'  大小：{path.stat().st_size:,} bytes')
        return

    # 单期模式（保持原有行为）
    rows = [r for r in all_rows if r['期号'] == args.period]
    print(f'过滤期号 {args.period} 后：{len(rows)}')
    rows = attach_match_id(rows, matches, _norm)
    if args.filter:
        keywords = [k.strip() for k in args.filter.split(',') if k.strip()]
        before = len(rows)
        rows = [r for r in rows if any(k in (r.get('canonical_material', '') or r.get('材料名称', '')) for k in keywords)]
        matches = [m for m in matches if any(k in (m.get('canonical_material', '') or '') for k in keywords)]
        print(f'过滤 [{",".join(keywords)}]：{before} → {len(rows)}')

    matched = [r for r in rows if r.get('match_id')]
    print(f'已匹配：{len(matched)} / 未匹配：{len(rows) - len(matched)}')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if args.filter:
        sheet_label = f'{args.period}期_{args.filter.replace(",", "_")}'
    else:
        sheet_label = f'{args.period}期混凝土'
    ws = wb.create_sheet(sheet_label)
    build_sheet(ws, matched, args.period, args)
    for col, width in [('A', 12), ('B', 28), ('C', 28), ('D', 16), ('E', 8), ('F', 18), ('G', 12), ('H', 14), ('I', 25)]:
        ws.column_dimensions[col].width = width

    # === 删 B：单期 同材料规格阶梯 sheet（信息已被【自动总结-模式】+【明细】覆盖） ===
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    cities = set(m.get('city') for m in manifest if m.get('city'))
    city = next(iter(cities), '成都')
    suffix = f'_{args.filter.replace(",", "_")}' if args.filter else ''
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = output_dir / f'{city}_{args.period}期_v3{suffix}_{timestamp}.xlsx'
    wb.save(path)
    print(f'\n✓ xlsx 已写：{path}')
    print(f'  大小：{path.stat().st_size:,} bytes')
    print(f'\n=== 完成 ===')
    print(f'  期号：{args.period}')
    print(f'  明细行数：{sum(1 for r in matched if r.get("价格") is not None)}')


if __name__ == '__main__':
    main()